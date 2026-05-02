"""
Full-Auto Financial Template Populator
========================================
Enter NSE symbol + upload template — the app handles everything:

  1. Reads template  — exact field labels + year columns needed
  2. Fetches web data — Screener.in + yfinance (structured, fast)
  3. Auto-downloads annual report PDF from NSE/BSE (no manual upload)
  4. Extracts tables from the PDF using pdfplumber (granular line items)
  5. LLM reconciles web data + PDF data → maps to template labels
  6. Writes to template — download the populated Excel

Web sources cover 70–80 % of fields cleanly.
PDF extraction fills the rest (granular notes, sub-line items, etc.)
Together they get you very close to 100 % coverage.

Run:
    streamlit run web_only_app.py
"""

from __future__ import annotations

import io
import json
import os
import re
import sys
import subprocess
import time
import traceback
from pathlib import Path
from typing import Optional

sys.path.insert(0, str(Path(__file__).parent))

import logging
logger = logging.getLogger(__name__)

# ── Auto-install dependencies ──────────────────────────────────────────────────
def _pip(*pkgs):
    subprocess.check_call([sys.executable, "-m", "pip", "install", "--quiet", *pkgs])

try:
    import streamlit as st
except ImportError:
    _pip("streamlit>=1.35.0"); import streamlit as st

try:
    import openpyxl
except ImportError:
    _pip("openpyxl>=3.1.2"); import openpyxl

try:
    import pandas as pd
except ImportError:
    _pip("pandas>=2.2.0"); import pandas as pd

try:
    import requests as _requests_mod
except ImportError:
    _pip("requests>=2.31.0"); import requests as _requests_mod

try:
    from bs4 import BeautifulSoup as _BS4
except ImportError:
    _pip("beautifulsoup4>=4.12.0"); from bs4 import BeautifulSoup as _BS4

try:
    import yfinance as yf
except ImportError:
    _pip("yfinance>=0.2.38"); import yfinance as yf

try:
    from openai import OpenAI
except ImportError:
    _pip("openai>=1.30.0"); from openai import OpenAI

try:
    from playwright.sync_api import sync_playwright as _sync_playwright
    _PLAYWRIGHT_AVAILABLE = True
except ImportError:
    _pip("playwright>=1.44.0")
    try:
        from playwright.sync_api import sync_playwright as _sync_playwright
        _PLAYWRIGHT_AVAILABLE = True
    except ImportError:
        _PLAYWRIGHT_AVAILABLE = False


# ── Page config ────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Auto Financial Extractor",
    page_icon="🚀",
    layout="wide",
)

st.title("🚀 Full-Auto Financial Template Populator")
st.caption(
    "Enter NSE symbol → auto-downloads annual report PDF + scrapes web → "
    "LLM maps everything to your template."
)


# ═══════════════════════════════════════════════════════════════════════════════
# Template structure reader
# ═══════════════════════════════════════════════════════════════════════════════

def read_template_structure(template_bytes: bytes) -> dict:
    from utils.helpers import normalise_year, clean_label
    wb  = openpyxl.load_workbook(io.BytesIO(template_bytes), data_only=True)
    out = {"sheets": {}}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # ── Smarter year-header detection: scan rows 1-30, pick row with most years ──
        best_year_row: Optional[int] = None
        best_year_cols: dict[int, str] = {}   # col_idx → F-year string
        for ri in range(1, 31):
            row_years: dict[int, str] = {}
            for ci in range(1, ws.max_column + 1):
                v = ws.cell(row=ri, column=ci).value
                if v is None:
                    continue
                yr = normalise_year(str(v).strip())
                if yr:
                    row_years[ci] = yr
            if len(row_years) > len(best_year_cols):
                best_year_cols = row_years
                best_year_row  = ri

        if not best_year_cols or best_year_row is None:
            continue

        header_row = best_year_row

        # ── Multi-column label scan: try col 1 and col 2, pick whichever has more labels ──
        def _count_labels(col_idx: int) -> int:
            count = 0
            for ri in range(header_row + 1, ws.max_row + 1):
                v = ws.cell(row=ri, column=col_idx).value
                if v and str(v).strip() and not str(v).strip().startswith("#"):
                    count += 1
            return count

        col1_count = _count_labels(1)
        col2_count = _count_labels(2)
        label_col  = 2 if col2_count > col1_count else 1

        labels: list[str] = []
        seen:   set[str]  = set()
        for ri in range(header_row + 1, ws.max_row + 1):
            v = ws.cell(row=ri, column=label_col).value
            if v is None:
                continue
            lbl = str(v).strip()
            if not lbl or lbl.startswith("#"):
                continue
            c = clean_label(lbl)
            if c and c not in seen:
                seen.add(c)
                labels.append(lbl)

        # Always include a sheet that has year headers — even if it has no label
        # rows yet.  The empty-sheet path in map_raw_excel_to_template detects
        # `labels == []` and dumps ALL matching raw data as new rows.
        if best_year_cols:
            out["sheets"][sheet_name] = {
                "years":     sorted(set(best_year_cols.values())),
                "labels":    labels,   # may be [] for blank sheets
                "label_col": label_col,
            }

    return out


# ═══════════════════════════════════════════════════════════════════════════════
# Web data: Screener.in (raw, no pre-mapping)
# ═══════════════════════════════════════════════════════════════════════════════

_SCREENER_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Referer": "https://www.screener.in/",
}
_SCREENER_SECTIONS = {
    "profit-loss":   "P&L",
    "balance-sheet": "Balance Sheet",
    "cash-flow":     "Cash Flow",
    "ratios":        "Operating Metrics",   # ROE, ROA, NIM, GNPA%, CRAR etc.
    "quarters":      "Quarterly",           # last 8 quarters of P&L
}

# Sections that carry quarterly (not annual) data
_SCREENER_QUARTERLY_SECTIONS = {"quarters"}


def _scr_year(text: str) -> Optional[str]:
    m = re.search(r"(\d{4})", text.strip())
    return f"F{m.group(1)}" if m else None


def _scr_qtr_year(text: str) -> Optional[str]:
    """
    Convert Screener quarterly column header to Indian-FY quarter key.
    Handles all formats Screener uses:
      'Jun 2024'  → '1QF2025'   (4-digit year with space)
      'Jun2024'   → '1QF2025'   (4-digit year no space)
      "Jun '24"   → '1QF2025'   (apostrophe 2-digit year, Screener default)
      'Sep 2024'  → '2QF2025'
      'Dec 2024'  → '3QF2025'
      'Mar 2025'  → '4QF2025'
    """
    _MON = {
        "jan": (4, 0), "feb": (4, 0), "mar": (4, 0),
        "apr": (1, 1), "may": (1, 1), "jun": (1, 1),
        "jul": (2, 1), "aug": (2, 1), "sep": (2, 1),
        "oct": (3, 1), "nov": (3, 1), "dec": (3, 1),
    }
    t = text.strip()
    # Try 4-digit year first: "Jun 2024" or "Jun2024"
    m = re.search(r"([A-Za-z]{3})\s*(\d{4})", t)
    if m:
        mon = m.group(1).lower()
        yr  = int(m.group(2))
    else:
        # Fallback: apostrophe/short 2-digit year: "Jun '24" or "Jun'24" or "Jun 24"
        m = re.search(r"([A-Za-z]{3})\s*['‘’]?\s*(\d{2})\b", t)
        if not m:
            return None
        mon = m.group(1).lower()
        yy  = int(m.group(2))
        yr  = (2000 + yy) if yy < 50 else (1900 + yy)
    if mon not in _MON:
        return None
    q_num, add_yr = _MON[mon]
    return f"{q_num}QF{yr + add_yr}"


def _scr_val(text: str) -> Optional[float]:
    t = text.strip().replace(",", "").replace(" ", "").replace("%", "")
    if not t or t in ("-", "—", "N/A"):
        return None
    neg = t.startswith("(") and t.endswith(")")
    t = t.strip("()")
    try:
        v = float(t)
        return -v if neg else v
    except ValueError:
        return None


def fetch_screener_raw(symbol: str, consolidated: bool = True) -> dict:
    suffix = "/consolidated/" if consolidated else "/"
    url    = f"https://www.screener.in/company/{symbol.upper()}{suffix}"
    try:
        r = _requests_mod.get(url, headers=_SCREENER_HEADERS, timeout=25)
        if r.status_code == 404:
            return {}
        r.raise_for_status()
    except Exception:
        return {}

    soup   = _BS4(r.text, "html.parser")
    result = {}

    for sec_id, sec_name in _SCREENER_SECTIONS.items():
        el = soup.find("section", {"id": sec_id}) or soup.find(id=sec_id)
        if not el:
            continue
        tbl = el.find("table")
        if not tbl:
            continue
        thead = tbl.find("thead")
        if not thead:
            continue
        ths = thead.find_all("th")
        # Use quarterly year parser for the quarters section
        _yr_parser = _scr_qtr_year if sec_id in _SCREENER_QUARTERLY_SECTIONS else _scr_year
        years = [_yr_parser(th.get_text(separator=" ", strip=True)) for th in ths[1:]]

        tbody = tbl.find("tbody")
        if not tbody:
            continue

        data: dict[str, dict[str, float]] = {}
        for tr in tbody.find_all("tr"):
            cells = tr.find_all("td")
            if not cells:
                continue
            lbl = re.sub(r"\s*[+\-]\s*$", "", cells[0].get_text(separator=" ", strip=True)).strip()
            if not lbl:
                continue
            row: dict[str, float] = {}
            for ci, yr in enumerate(years):
                if yr is None:
                    continue
                v = _scr_val(cells[ci + 1].get_text(strip=True)) if ci + 1 < len(cells) else None
                if v is not None:
                    row[yr] = v
            if row:
                data[lbl] = row

        if data:
            result[sec_name] = {
                "years": sorted({y for y in years if y}),
                "data":  data,
            }

    return result


# ═══════════════════════════════════════════════════════════════════════════════
# Web data: InstaFinancials (historical P&L / BS for Indian companies)
# ═══════════════════════════════════════════════════════════════════════════════

_INSTA_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml",
    "Accept-Language": "en-US,en;q=0.9",
    "Referer": "https://www.instafinancials.com/",
}


def _insta_parse_table(table_el) -> dict[str, dict[str, float]]:
    """Parse an InstaFinancials HTML financial table → {label: {year: value}}."""
    from utils.helpers import normalise_year
    ths  = table_el.find_all("th")
    # First th is label column — rest are years
    year_cols: list[Optional[str]] = []
    for th in ths[1:]:
        raw = th.get_text(strip=True)
        year_cols.append(normalise_year(raw))

    data: dict[str, dict[str, float]] = {}
    tbody = table_el.find("tbody") or table_el
    for tr in tbody.find_all("tr"):
        tds = tr.find_all("td")
        if not tds:
            continue
        lbl = tds[0].get_text(separator=" ", strip=True)
        lbl = re.sub(r"\s*[+\-]\s*$", "", lbl).strip()
        if not lbl:
            continue
        row: dict[str, float] = {}
        for ci, yr in enumerate(year_cols):
            if yr is None or ci + 1 >= len(tds):
                continue
            raw_val = tds[ci + 1].get_text(strip=True).replace(",", "").replace(" ", "")
            neg = raw_val.startswith("(") and raw_val.endswith(")")
            raw_val = raw_val.strip("()")
            try:
                v = float(raw_val)
                row[yr] = -v if neg else v
            except ValueError:
                pass
        if row:
            data[lbl] = row
    return data


def _instafinancials_search_slug(symbol: str) -> Optional[str]:
    """Search InstaFinancials to find the company slug for a given NSE symbol."""
    try:
        search_url = f"https://www.instafinancials.com/search?q={symbol}"
        r = _requests_mod.get(search_url, headers=_INSTA_HEADERS, timeout=15)
        if r.status_code != 200:
            return None
        soup = _BS4(r.text, "html.parser")
        # Look for a company link containing the symbol
        for a in soup.find_all("a", href=True):
            href = a["href"]
            if "/company/" in href:
                slug = href.split("/company/")[-1].split("/")[0].strip("/")
                if slug and len(slug) > 2:
                    return slug
    except Exception:
        pass
    return None


def fetch_instafinancials_raw(symbol: str) -> tuple[dict, list[dict]]:
    """
    Scrape InstaFinancials for P&L + Balance Sheet data.

    Returns:
        (result_dict, debug_log)
        result_dict : same structure as fetch_screener_raw
        debug_log   : list of dicts with per-URL attempt details for UI display
    """
    base = symbol.lower()
    slug_candidates = [
        base,
        base.replace("bnk", "-bank"),
        base.replace("bnk", "-small-finance-bank-limited"),
        base.replace("bnk", "-small-finance-bank"),
        base + "-limited",
        base + "-bank-limited",
        base.replace("bank", "-bank-limited"),
        base.replace("ltd", "-limited"),
        re.sub(r"(bank|bnk)$", "", base).rstrip("-") + "-small-finance-bank-limited",
        re.sub(r"(bank|bnk)$", "", base).rstrip("-") + "-bank-limited",
    ]
    seen: set = set()
    slug_candidates = [s for s in slug_candidates if s and not (s in seen or seen.add(s))]

    result: dict = {}
    debug_log: list[dict] = []

    def _try_fetch(slug: str) -> dict:
        import json as _j

        def _extract_fin_from_json(obj, depth=0) -> dict:
            """Recursively search JSON blob for {label: {year: value}} financial tables."""
            if depth > 6:
                return {}
            if isinstance(obj, dict):
                yr_re = re.compile(r"^F?\d{4}$|^20\d{2}-\d{2}$")
                candidate: dict = {}
                for k, v in obj.items():
                    if isinstance(v, dict):
                        yr_vals = {kk: vv for kk, vv in v.items()
                                   if yr_re.match(str(kk)) and isinstance(vv, (int, float))}
                        if yr_vals:
                            candidate[str(k)] = {
                                (f"F{yy}" if not str(yy).startswith("F") else str(yy)): float(vv)
                                for yy, vv in yr_vals.items()
                            }
                if len(candidate) >= 3:
                    return candidate
                best: dict = {}
                for v in obj.values():
                    sub = _extract_fin_from_json(v, depth + 1)
                    if len(sub) > len(best):
                        best = sub
                return best
            elif isinstance(obj, list):
                best: dict = {}
                for item in obj[:20]:
                    sub = _extract_fin_from_json(item, depth + 1)
                    if len(sub) > len(best):
                        best = sub
                return best
            return {}

        out: dict = {}
        for stmt, sec_name in [("profitloss", "P&L"), ("balancesheet", "Balance Sheet")]:
            for stmt_type in ("standalone", "consolidated"):
                url = f"https://www.instafinancials.com/company/{slug}/financials/{stmt_type}/{stmt}"
                entry: dict = {"slug": slug, "url": url, "section": sec_name,
                               "stmt_type": stmt_type, "status": None,
                               "tables_found": 0, "rows_parsed": 0,
                               "years_found": [], "raw_headers": [], "error": None}
                try:
                    r = _requests_mod.get(url, headers=_INSTA_HEADERS, timeout=15)
                    entry["status"] = r.status_code
                    if r.status_code not in (200, 201):
                        debug_log.append(entry)
                        continue
                    if "404" in r.text[:500] or "not found" in r.text[:500].lower():
                        entry["error"] = "404 page content"
                        debug_log.append(entry)
                        continue
                    soup = _BS4(r.text, "html.parser")
                    tables = soup.find_all("table")
                    entry["tables_found"] = len(tables)
                    best_data: dict = {}
                    best_headers: list = []

                    # ── Strategy A: HTML <table> elements ────────────────────
                    for tbl in tables:
                        ths = tbl.find_all("th")
                        raw_hdrs = [th.get_text(strip=True) for th in ths[1:]]
                        parsed = _insta_parse_table(tbl)
                        if len(parsed) > len(best_data):
                            best_data = parsed
                            best_headers = raw_hdrs

                    # ── Strategy B: JSON embedded in <script> tags ────────────
                    # InstaFinancials and similar React/Next.js sites embed data in
                    # window.__NEXT_DATA__, window.__INITIAL_STATE__, or similar blobs.
                    if not best_data:
                        for script in soup.find_all("script"):
                            sc = script.string or ""
                            if len(sc) < 100:
                                continue
                            # Priority patterns: Next.js data, common SPA patterns
                            for pat in [
                                r'id="__NEXT_DATA__"[^>]*>(\{[\s\S]+?\})\s*</script>',
                                r'window\.__NEXT_DATA__\s*=\s*(\{[\s\S]{100,}\})\s*;',
                                r'window\.__INITIAL_STATE__\s*=\s*(\{[\s\S]{100,}\})\s*;',
                                r'window\.pageData\s*=\s*(\{[\s\S]{100,}\})\s*;',
                                r'window\.appData\s*=\s*(\{[\s\S]{100,}\})\s*;',
                                r'(?:var|const|let)\s+financialData\s*=\s*(\{[\s\S]{100,}\})\s*;',
                                r'(?:var|const|let)\s+tableData\s*=\s*(\[[\s\S]{100,}\])\s*;',
                            ]:
                                for m in re.finditer(pat, sc, re.S):
                                    try:
                                        obj = _j.loads(m.group(1))
                                        candidate = _extract_fin_from_json(obj)
                                        if len(candidate) > len(best_data):
                                            best_data = candidate
                                            entry["raw_headers"] = [f"[script JSON: {len(candidate)} rows extracted]"]
                                    except Exception:
                                        pass
                            if best_data:
                                break

                    # ── Strategy C: Try InstaFinancials JSON API endpoints ─────
                    if not best_data:
                        for api_url in [
                            f"https://www.instafinancials.com/api/company/{slug}/financials/{stmt_type}/{stmt}",
                            f"https://www.instafinancials.com/company/{slug}/financials/{stmt_type}/{stmt}/data",
                            f"https://www.instafinancials.com/api/financials/{slug}/{stmt}",
                            f"https://www.instafinancials.com/api/{slug}/{stmt_type}/{stmt}",
                        ]:
                            try:
                                ra = _requests_mod.get(
                                    api_url,
                                    headers={**_INSTA_HEADERS, "Accept": "application/json",
                                             "X-Requested-With": "XMLHttpRequest"},
                                    timeout=10,
                                )
                                if ra.status_code == 200:
                                    try:
                                        jd = ra.json()
                                        candidate = _extract_fin_from_json(jd)
                                        if len(candidate) > len(best_data):
                                            best_data = candidate
                                            entry["raw_headers"] = [f"[API JSON: {len(candidate)} rows from {api_url.split('/')[-1]}]"]
                                    except Exception:
                                        pass
                            except Exception:
                                pass

                    entry["rows_parsed"] = len(best_data)
                    entry["raw_headers"] = best_headers or entry.get("raw_headers", [])
                    if best_data:
                        years = sorted({yr for row in best_data.values() for yr in row})
                        entry["years_found"] = years
                        out[sec_name] = {"years": years, "data": best_data}
                        debug_log.append(entry)
                        break   # found data for this section
                    else:
                        entry["error"] = "no parseable rows in any table (JS-rendered?)"
                        debug_log.append(entry)
                except Exception as e:
                    entry["error"] = str(e)
                    debug_log.append(entry)
                    logger.debug(f"InstaFinancials {slug}/{sec_name}: {e}")
        return out

    # Try slug candidates
    for slug in slug_candidates:
        r = _try_fetch(slug)
        if r:
            result = r
            logger.info(f"InstaFinancials: found data for {symbol} using slug '{slug}'")
            break

    # Search fallback
    if not result:
        found_slug = _instafinancials_search_slug(symbol)
        if found_slug:
            r = _try_fetch(found_slug)
            if r:
                result = r
                logger.info(f"InstaFinancials: found data for {symbol} via search slug '{found_slug}'")

    if not result:
        logger.info(f"InstaFinancials: no data found for {symbol}")
    return result, debug_log


# ═══════════════════════════════════════════════════════════════════════════════
# Web data: Tickertape (additional historical data — JSON API)
# ═══════════════════════════════════════════════════════════════════════════════

_TT_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "application/json, text/html,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Referer": "https://www.tickertape.in/",
}


def _tickertape_search_slug(symbol: str) -> Optional[str]:
    """Find Tickertape stock slug (sid) for a given NSE symbol via their search API."""
    try:
        url = f"https://api.tickertape.in/search?text={symbol}&type=stock"
        r = _requests_mod.get(url, headers=_TT_HEADERS, timeout=10)
        if r.status_code != 200:
            return None
        data = r.json()
        hits = data.get("data", {}).get("stocks", []) or data.get("data", []) or []
        sym_up = symbol.upper()
        for hit in hits[:10]:
            ticker = str(hit.get("ticker", "") or hit.get("slug", "") or "").upper()
            if ticker == sym_up:
                return hit.get("sid") or hit.get("slug") or ticker
        return (hits[0].get("sid") or hits[0].get("slug")) if hits else None
    except Exception:
        return None


def fetch_tickertape_raw(symbol: str) -> tuple[dict, list[dict]]:
    """
    Fetch P&L + Balance Sheet data from Tickertape's public API.

    Tickertape's /financials endpoint returns structured JSON containing
    annual P&L and BS rows for up to 10 years — far more history than XBRL.

    Returns (result_dict, debug_log) — same structure as fetch_screener_raw.
    """
    from utils.helpers import normalise_year
    result: dict = {}
    debug_log: list[dict] = []

    slug = _tickertape_search_slug(symbol) or symbol.upper()
    debug_log.append({"step": "slug", "slug": slug})

    # Map Tickertape statement types to our section names
    stmt_map = [
        ("incomeStatement", "P&L"),
        ("balanceSheet",    "Balance Sheet"),
        ("cashFlow",        "Cash Flow"),
    ]

    for tt_stmt, sec_name in stmt_map:
        for period in ("annual",):
            url = (
                f"https://api.tickertape.in/stocks/{slug}/financials"
                f"?statement={tt_stmt}&period={period}&consolidated=true"
            )
            entry = {"step": "fetch", "url": url, "section": sec_name,
                     "status": None, "rows": 0, "years": [], "error": None}
            try:
                r = _requests_mod.get(url, headers=_TT_HEADERS, timeout=15)
                entry["status"] = r.status_code
                if r.status_code != 200:
                    debug_log.append(entry)
                    continue
                payload = r.json()
                # Tickertape response: {"data": {"columns": [...years...], "rows": [{"label": ..., "values": [...]}]}}
                inner = payload.get("data", payload)
                if isinstance(inner, list) and inner:
                    inner = inner[0]
                columns = inner.get("columns", [])
                rows    = inner.get("rows", [])
                if not columns or not rows:
                    entry["error"] = "no columns/rows in response"
                    debug_log.append(entry)
                    continue

                # Parse year columns
                yr_keys: list[Optional[str]] = []
                for col in columns:
                    raw_yr = str(col.get("label", "") or col.get("period", "") or col).strip()
                    yr_keys.append(normalise_year(raw_yr))

                data: dict[str, dict[str, float]] = {}
                for row in rows:
                    lbl = str(row.get("label", "") or row.get("name", "")).strip()
                    if not lbl:
                        continue
                    vals = row.get("values", []) or row.get("data", [])
                    row_data: dict[str, float] = {}
                    for ci, yr in enumerate(yr_keys):
                        if yr is None or ci >= len(vals):
                            continue
                        raw_val = vals[ci]
                        if raw_val is None:
                            continue
                        try:
                            v = float(raw_val)
                            row_data[yr] = v
                        except (TypeError, ValueError):
                            pass
                    if row_data:
                        data[lbl] = row_data

                if data:
                    years_found = sorted({yr for yr in yr_keys if yr})
                    entry["rows"] = len(data)
                    entry["years"] = years_found
                    result[sec_name] = {"years": years_found, "data": data}
                    debug_log.append(entry)
                    break   # got data for this section — don't retry
                else:
                    entry["error"] = "parsed 0 rows"
                    debug_log.append(entry)
            except Exception as e:
                entry["error"] = str(e)
                debug_log.append(entry)
                logger.debug(f"Tickertape {slug}/{sec_name}: {e}")

    if not result:
        logger.info(f"Tickertape: no data found for {symbol} (slug={slug})")
    return result, debug_log


# ═══════════════════════════════════════════════════════════════════════════════
# Web data: StockAnalysis.com (clean HTML tables, 10-year history)
# ═══════════════════════════════════════════════════════════════════════════════

_SA_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Referer": "https://stockanalysis.com/",
}


def _sa_parse_table(table_el) -> dict[str, dict[str, float]]:
    """Parse a StockAnalysis.com financial table → {label: {year: value}}."""
    from utils.helpers import normalise_year
    thead = table_el.find("thead")
    tbody = table_el.find("tbody")
    if not thead or not tbody:
        return {}

    # Year headers are in thead th cells (skip first which is the label column)
    ths = thead.find_all("th")
    yr_keys: list[Optional[str]] = []
    for th in ths[1:]:
        raw = th.get_text(strip=True)
        yr_keys.append(normalise_year(raw))

    data: dict[str, dict[str, float]] = {}
    for tr in tbody.find_all("tr"):
        tds = tr.find_all("td")
        if not tds:
            continue
        lbl = tds[0].get_text(separator=" ", strip=True)
        lbl = re.sub(r"\s*[+\-]\s*$", "", lbl).strip()
        if not lbl:
            continue
        row_data: dict[str, float] = {}
        for ci, yr in enumerate(yr_keys):
            if yr is None or ci + 1 >= len(tds):
                continue
            raw_val = tds[ci + 1].get_text(strip=True).replace(",", "").replace(" ", "")
            if not raw_val or raw_val in ("-", "—", "N/A", "NA", "TTM"):
                continue
            neg = raw_val.startswith("(") and raw_val.endswith(")")
            raw_val = raw_val.strip("()")
            try:
                v = float(raw_val)
                row_data[yr] = -v if neg else v
            except ValueError:
                pass
        if row_data:
            data[lbl] = row_data
    return data


def fetch_stockanalysis_raw(symbol: str) -> tuple[dict, list[dict]]:
    """
    Fetch financial data from StockAnalysis.com for Indian NSE stocks.
    Tries /in/stocks/{symbol}/financials/ paths for P&L, BS, CF.

    Returns (result_dict, debug_log).
    """
    result: dict = {}
    debug_log: list[dict] = []

    sym_l = symbol.lower()
    # StockAnalysis uses lowercase NSE ticker for Indian stocks under /in/stocks/
    page_map = [
        ("income-statement", "P&L",           f"https://stockanalysis.com/in/stocks/{sym_l}/financials/"),
        ("balance-sheet",    "Balance Sheet",  f"https://stockanalysis.com/in/stocks/{sym_l}/financials/balance-sheet/"),
        ("cash-flow",        "Cash Flow",      f"https://stockanalysis.com/in/stocks/{sym_l}/financials/cash-flow-statement/"),
    ]

    for page_key, sec_name, url in page_map:
        entry = {"step": "fetch", "url": url, "section": sec_name,
                 "status": None, "rows": 0, "years": [], "error": None}
        try:
            r = _requests_mod.get(url, headers=_SA_HEADERS, timeout=15)
            entry["status"] = r.status_code
            if r.status_code != 200:
                debug_log.append(entry)
                continue
            soup = _BS4(r.text, "html.parser")
            # StockAnalysis renders a single large table or a div with id="financial-table"
            tbl = (soup.find("table", {"id": "financial-table"})
                   or soup.find("table", class_=re.compile(r"financial|table", re.I))
                   or soup.find("table"))
            if not tbl:
                entry["error"] = "no table found"
                debug_log.append(entry)
                continue
            data = _sa_parse_table(tbl)
            if data:
                years_found = sorted({yr for row in data.values() for yr in row})
                entry["rows"] = len(data)
                entry["years"] = years_found
                result[sec_name] = {"years": years_found, "data": data}
            else:
                entry["error"] = "table found but 0 rows parsed"
            debug_log.append(entry)
        except Exception as e:
            entry["error"] = str(e)
            debug_log.append(entry)
            logger.debug(f"StockAnalysis {sym_l}/{sec_name}: {e}")

    if not result:
        logger.info(f"StockAnalysis: no data for {symbol}")
    return result, debug_log


# ═══════════════════════════════════════════════════════════════════════════════
# Web data: Moneycontrol (replaces Tofler — free public financial tables)
# ═══════════════════════════════════════════════════════════════════════════════
# Tofler requires login for financial data and returns nothing without auth.
# Moneycontrol has free, publicly accessible financial tables for all NSE companies.

_MC_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Referer": "https://www.moneycontrol.com/",
}


def _mc_search_stock_url(symbol: str) -> Optional[str]:
    """
    Search Moneycontrol to find the stock page URL for a given NSE symbol.
    Tries multiple autocomplete endpoints and search strategies.
    Returns an absolute URL like:
      https://www.moneycontrol.com/india/stockpricequote/banks-private-sector/equitas-small-finance-bank/ESB
    """
    import json as _json, urllib.parse as _up

    def _extract_link(hit: dict) -> Optional[str]:
        link = hit.get("link_src") or hit.get("url") or hit.get("link") or ""
        if link:
            if not link.startswith("http"):
                link = "https://www.moneycontrol.com" + link
            return link
        return None

    def _try_endpoint(query: str) -> Optional[str]:
        # MC uses several autocomplete endpoint formats over time
        endpoints = [
            f"https://www.moneycontrol.com/mccode/common/autosuggestion/getSearchresult.php"
            f"?classic=true&query={_up.quote(query)}&type=1&format=json&callback=mcontext",
            f"https://www.moneycontrol.com/mccode/common/autosuggestion/getSearchresult.php"
            f"?query={_up.quote(query)}&type=1&format=json",
            f"https://www.moneycontrol.com/mccode/common/autosuggestion/auto_suggestion.php"
            f"?classic=true&query={_up.quote(query)}&type=1&format=json",
            f"https://www.moneycontrol.com/mccode/common/autosuggestion/auto_suggestion.php"
            f"?query={_up.quote(query)}&type=1&format=json",
        ]
        for url in endpoints:
            try:
                r = _requests_mod.get(url, headers=_MC_HEADERS, timeout=12)
                if r.status_code != 200:
                    continue
                text = r.text.strip()
                # Strip JSONP wrapper
                if text.startswith("mcontext("):
                    text = text[9:].rstrip(");")
                try:
                    data = _json.loads(text)
                except Exception:
                    continue
                hits = (data.get("data", []) if isinstance(data, dict)
                        else (data if isinstance(data, list) else []))
                # Prefer exact NSE symbol match
                sym_up = symbol.upper()
                for hit in hits[:10]:
                    nse_id = str(hit.get("nse_id", "") or hit.get("NSE", "") or "").upper()
                    link   = _extract_link(hit)
                    if link and (nse_id == sym_up or link.upper().endswith(f"/{sym_up}")):
                        return link
                # Fallback: first hit
                for hit in hits[:3]:
                    link = _extract_link(hit)
                    if link:
                        return link
            except Exception as e:
                logger.debug(f"Moneycontrol endpoint {url[:60]}: {e}")
        return None

    # Strategy 1: search by NSE symbol directly
    link = _try_endpoint(symbol)
    if link:
        return link

    # Strategy 2: search by common company name variants
    # e.g. EQUITASBNK → "Equitas Small Finance Bank"
    sym_lower = symbol.lower()
    name_variants = []
    if "bnk" in sym_lower:
        base = re.sub(r"bnk$", "", sym_lower).strip()
        name_variants = [
            base + " small finance bank",
            base + " bank",
            base,
        ]
    elif "bank" in sym_lower:
        base = re.sub(r"bank$", "", sym_lower).strip()
        name_variants = [base + " bank", base]
    elif "fin" in sym_lower:
        name_variants = [sym_lower.replace("fin", " finance ").strip()]
    else:
        name_variants = [sym_lower]

    for name in name_variants:
        link = _try_endpoint(name)
        if link:
            return link

    # Strategy 3: try NSE's own data to get the ISIN, then search MC by ISIN
    try:
        nse_meta_url = f"https://www.nseindia.com/api/quote-equity?symbol={symbol}"
        r = _requests_mod.get(nse_meta_url, headers=_SCREENER_HEADERS, timeout=10)
        if r.status_code == 200:
            meta = r.json()
            isin = meta.get("info", {}).get("isin", "")
            if isin:
                link = _try_endpoint(isin)
                if link:
                    return link
    except Exception:
        pass

    # Strategy 4: Direct URL construction from NSE symbol
    # Moneycontrol stock URLs follow: /india/stockpricequote/{sector}/{slug}/{SC_ID}
    # We guess the slug and SC_ID from the symbol and probe up to 5 variants.
    sym_l = symbol.lower()
    # Build MC-style slug: remove common suffixes, join with hyphens
    mc_slug_base = re.sub(r"(bnk|bank|fin|ltd|limited|sfb)$", "", sym_l).strip("-")
    mc_slug_variants = [
        mc_slug_base + "-small-finance-bank",
        mc_slug_base + "-bank",
        mc_slug_base,
        sym_l.replace("bnk", "bank"),
        sym_l,
    ]
    # SC_ID variants: Moneycontrol uses 3-7 char codes, often derived from ticker
    sc_id_variants = [
        symbol[:3].upper(),
        symbol[:4].upper(),
        symbol[:5].upper(),
        re.sub(r"BNK$", "", symbol).upper()[:3] + "05",
        re.sub(r"BNK$", "", symbol).upper()[:4],
    ]
    sector_variants = [
        "banks-private-sector",
        "banks-public-sector",
        "finance-nbfc",
        "finance",
        "banks",
    ]
    for _slug in mc_slug_variants[:3]:
        for _sc in sc_id_variants[:3]:
            for _sec in sector_variants[:2]:
                probe_url = f"https://www.moneycontrol.com/india/stockpricequote/{_sec}/{_slug}/{_sc}"
                try:
                    _pr = _requests_mod.get(probe_url, headers=_MC_HEADERS, timeout=8, allow_redirects=True)
                    if _pr.status_code == 200 and "financial" in _pr.text.lower():
                        logger.debug(f"Moneycontrol direct probe hit: {probe_url}")
                        return _pr.url.rstrip("/")
                except Exception:
                    pass

    logger.debug(f"Moneycontrol: could not resolve stock page for {symbol}")
    return None


def _mc_parse_table(table_el) -> dict[str, dict[str, float]]:
    """
    Moneycontrol-specific table parser.

    Moneycontrol financial tables often use:
    - First <tr> with <td> cells for column headers (not <th>)
    - Year headers like "Mar '25", "Mar-25", "Mar 2025", "TTM"
    - Label in first cell, values in subsequent cells

    Falls back gracefully — returns {} if table doesn't look financial.
    """
    from utils.helpers import normalise_year
    rows = table_el.find_all("tr")
    if not rows:
        return {}

    # Find header row: first row where cells contain year-like strings
    year_cols: list[Optional[str]] = []
    data_start_row = 0
    for ri, row in enumerate(rows[:5]):
        cells = row.find_all(["td", "th"])
        if len(cells) < 2:
            continue
        # Try to parse cells 1+ as years
        _candidate_years = []
        for cell in cells[1:]:
            raw = cell.get_text(strip=True)
            # Normalise MC-specific apostrophe year format: "Mar '25" → "Mar 25"
            raw_clean = re.sub(r"['''`]", "", raw).strip()
            yr = normalise_year(raw_clean)
            _candidate_years.append(yr)
        if any(y is not None for y in _candidate_years):
            year_cols = _candidate_years
            data_start_row = ri + 1
            break

    if not year_cols:
        return {}

    data: dict[str, dict[str, float]] = {}
    for row in rows[data_start_row:]:
        cells = row.find_all(["td", "th"])
        if not cells:
            continue
        lbl = cells[0].get_text(separator=" ", strip=True)
        lbl = re.sub(r"\s*[+\-]\s*$", "", lbl).strip()
        if not lbl or lbl.lower() in ("", "particulars", "description"):
            continue
        row_data: dict[str, float] = {}
        for ci, yr in enumerate(year_cols):
            if yr is None or ci + 1 >= len(cells):
                continue
            raw_val = cells[ci + 1].get_text(strip=True).replace(",", "").replace(" ", "")
            if not raw_val or raw_val in ("-", "—", "N/A", "NA"):
                continue
            neg = raw_val.startswith("(") and raw_val.endswith(")")
            raw_val = raw_val.strip("()")
            try:
                v = float(raw_val)
                row_data[yr] = -v if neg else v
            except ValueError:
                pass
        if row_data:
            data[lbl] = row_data
    return data


def fetch_tofler_raw(symbol: str) -> tuple[dict, list[dict]]:
    """
    Fetch financial data from Moneycontrol (Tofler replacement).

    Returns:
        (result_dict, debug_log)
    """
    result: dict = {}
    debug_log: list[dict] = []

    try:
        # ── Step 1: Resolve stock page URL ───────────────────────────────────
        stock_url = _mc_search_stock_url(symbol)
        if not stock_url:
            debug_log.append({"step": "search", "error": f"Could not find Moneycontrol stock page for {symbol}"})
            logger.info(f"Moneycontrol: could not find stock page for {symbol}")
            return {}, debug_log

        sc_id = stock_url.rstrip("/").split("/")[-1]
        debug_log.append({"step": "search", "stock_url": stock_url, "sc_id": sc_id})
        logger.debug(f"Moneycontrol: stock_url={stock_url}, sc_id={sc_id}")

        # ── Step 2: Scrape stock page to find /financials/ href ───────────────
        # Extract company name slug from stock URL:
        # https://www.moneycontrol.com/india/stockpricequote/banks.../equitas-small-finance-bank/EQU
        # company_slug = "equitas-small-finance-bank"
        _url_parts = stock_url.rstrip("/").split("/")
        _company_slug = _url_parts[-2] if len(_url_parts) >= 2 else sc_id.lower()

        fin_base: Optional[str] = None
        try:
            r0 = _requests_mod.get(stock_url, headers=_MC_HEADERS, timeout=15)
            debug_log.append({"step": "stock_page", "url": stock_url, "status": r0.status_code})
            if r0.status_code == 200:
                soup0 = _BS4(r0.text, "html.parser")
                for a in soup0.find_all("a", href=True):
                    href = a["href"]
                    # STRICT: the URL must end with /{sc_id} (SC_ID at tail, not just anywhere)
                    # and must contain /financials/ and profit-loss
                    if ("/financials/" in href
                            and "profit-loss" in href.lower()
                            and href.rstrip("/").split("/")[-1].upper() == sc_id.upper()):
                        parts = href.split("/")
                        for i, p in enumerate(parts):
                            if "profit-loss" in p.lower() or "profitloss" in p.lower():
                                fin_base = "/".join(parts[:i])
                                break
                        if fin_base:
                            if not fin_base.startswith("http"):
                                fin_base = "https://www.moneycontrol.com" + fin_base
                            debug_log.append({"step": "fin_base_found", "fin_base": fin_base})
                            break
        except Exception as e:
            debug_log.append({"step": "stock_page", "error": str(e)})
            logger.debug(f"Moneycontrol: stock page scrape failed: {e}")

        # ── Step 3: Fallback — construct fin_base from company slug ──────────
        if not fin_base:
            # Use the company name slug from the stock URL, not the SC_ID
            # e.g. "equitas-small-finance-bank" → /financials/equitassmallfinancebank
            _slug_clean = re.sub(r"[^a-z0-9]", "", _company_slug.lower())
            fin_base = f"https://www.moneycontrol.com/financials/{_slug_clean}"
            debug_log.append({"step": "fin_base_guessed", "fin_base": fin_base})
            logger.debug(f"Moneycontrol: using guessed fin_base={fin_base}")

        # ── Step 4: Fetch P&L, Balance Sheet, Cash Flow and Key Ratios ──────
        fin_pages = [
            (f"{fin_base}/profit-lossVI/{sc_id}",              "P&L"),
            (f"{fin_base}/balance-sheetVI/{sc_id}",            "Balance Sheet"),
            (f"{fin_base}/cash-flowVI/{sc_id}",                "Cash Flow"),
            (f"{fin_base}/key-ratiosVI/{sc_id}",               "Operating Metrics"),
            (f"{fin_base}/profit-lossVI/{sc_id}#Annual",       "P&L"),
            (f"{fin_base}/consolidated-profit-lossVI/{sc_id}", "P&L"),
            (f"{fin_base}/consolidated-balance-sheetVI/{sc_id}", "Balance Sheet"),
            (f"{fin_base}/consolidated-key-ratiosVI/{sc_id}",  "Operating Metrics"),
        ]

        fetched_sections: set = set()
        for fin_url, sec_name in fin_pages:
            if sec_name in fetched_sections:
                continue
            entry: dict = {"step": "fetch", "url": fin_url, "section": sec_name,
                           "status": None, "tables_found": 0, "rows_parsed": 0,
                           "years_found": [], "raw_headers": [], "error": None}
            try:
                r = _requests_mod.get(fin_url, headers=_MC_HEADERS, timeout=18)
                entry["status"] = r.status_code
                if r.status_code != 200:
                    debug_log.append(entry)
                    continue
                soup = _BS4(r.text, "html.parser")
                tbls = soup.find_all("table")
                entry["tables_found"] = len(tbls)
                best_data: dict = {}
                best_headers: list = []
                for tbl in tbls:
                    # Try standard insta parser first
                    raw_hdrs = [th.get_text(strip=True) for th in tbl.find_all("th")[1:]]
                    parsed = _insta_parse_table(tbl)
                    # Moneycontrol-specific fallback: headers may be in first <tr> as <td>
                    # and years are formatted as "Mar '25", "Mar-25", or "Mar 2025"
                    if not parsed:
                        parsed = _mc_parse_table(tbl)
                        raw_hdrs = [td.get_text(strip=True)
                                    for td in (tbl.find("tr") or _BS4("","html.parser")).find_all("td")[1:]]
                    if len(parsed) > len(best_data):
                        best_data = parsed
                        best_headers = raw_hdrs
                entry["rows_parsed"] = len(best_data)
                entry["raw_headers"] = best_headers
                if best_data:
                    years = sorted({yr for row in best_data.values() for yr in row})
                    entry["years_found"] = years
                    result[sec_name] = {"years": years, "data": best_data}
                    fetched_sections.add(sec_name)
                    logger.debug(f"Moneycontrol: {sec_name} OK via {fin_url} ({len(best_data)} rows)")
                else:
                    entry["error"] = "no parseable rows"
                debug_log.append(entry)
            except Exception as e:
                entry["error"] = str(e)
                debug_log.append(entry)
                logger.debug(f"Moneycontrol {sec_name} at {fin_url}: {e}")

    except Exception as e:
        debug_log.append({"step": "outer", "error": str(e)})
        logger.debug(f"Moneycontrol fetch_tofler_raw: {e}")

    if result:
        logger.info(f"Moneycontrol: found {sum(len(s['data']) for s in result.values())} fields for {symbol}")
    else:
        logger.info(f"Moneycontrol: no data found for {symbol}")
    return result, debug_log


# ═══════════════════════════════════════════════════════════════════════════════
# Web data: yfinance
# ═══════════════════════════════════════════════════════════════════════════════
# Web data: Company's own IR website (financial highlights / key metrics pages)
# ═══════════════════════════════════════════════════════════════════════════════

def _extract_urls_from_text(text: str, prefer_domain: Optional[str] = None) -> list[str]:
    """
    Extract HTTP/HTTPS URLs from a block of text (e.g. annual report cover page).
    Filters out known third-party domains (NSE, BSE, SEBI, stock exchanges).
    Returns deduplicated list, preferred domain first if provided.
    """
    _SKIP = re.compile(
        r"nseindia|bseindia|sebi\.gov|mca\.gov|rbi\.org|stockexchange"
        r"|linkedin|twitter|facebook|instagram|youtube|google|microsoft"
        r"|adobe|acrobat|kfintech|registrar|karvy|link.*intime"
        # Third-party aggregators — not the company's own site
        r"|annualreports\.com|moneycontrol\.com|screener\.in|tickertape"
        r"|tofler\.in|instafinancials|capitaline|prowessdx|cmie"
        r"|valueresearch|morningstar|investing\.com|tradingview"
        r"|economictimes|livemint|businessstandard|thehindu|reuters|bloomberg",
        re.I,
    )
    found: list[str] = []
    seen: set[str] = set()
    for m in re.finditer(r"https?://[^\s\"'<>)\]]+", text):
        url = m.group(0).rstrip(".,;:)")
        from urllib.parse import urlparse
        parsed = urlparse(url)
        host = (parsed.hostname or "").lower()
        if not host or _SKIP.search(host):
            continue
        # Normalise to just scheme + host (strip path for website lookup)
        base = f"{parsed.scheme}://{host}"
        if base not in seen:
            seen.add(base)
            found.append(base)
    # Also scan for bare domain patterns like "www.equitasbank.com"
    for m in re.finditer(r"\bwww\.[a-z0-9][-a-z0-9]*\.[a-z]{2,6}\b", text, re.I):
        base = "https://" + m.group(0).lower()
        if base not in seen:
            seen.add(base)
            found.append(base)
    if prefer_domain:
        found.sort(key=lambda u: 0 if prefer_domain.lower() in u.lower() else 1)
    return found


def _playwright_fetch(url: str, wait_selector: str = "table", timeout_ms: int = 20_000) -> Optional[str]:
    """
    Fetch a JS-rendered page using a headless Chromium browser via Playwright.
    Returns the fully-rendered HTML string, or None if Playwright is unavailable
    or the fetch fails.

    One-time browser install (run once on first use):
        python -m playwright install chromium --with-deps
    """
    if not _PLAYWRIGHT_AVAILABLE:
        return None
    try:
        with _sync_playwright() as pw:
            browser = pw.chromium.launch(headless=True)
            ctx = browser.new_context(
                user_agent=(
                    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/124.0.0.0 Safari/537.36"
                ),
                locale="en-US",
            )
            page = ctx.new_page()
            page.goto(url, wait_until="networkidle", timeout=timeout_ms)
            # Wait for either a table or the body — whichever the caller prefers
            try:
                page.wait_for_selector(wait_selector, timeout=8_000)
            except Exception:
                pass   # selector didn't appear — still return whatever rendered
            html = page.content()
            browser.close()
            return html
    except Exception as _pw_err:
        logger.debug(f"_playwright_fetch({url}): {_pw_err}")
        return None


def fetch_company_website_financials(
    symbol: str,
    bse_code: Optional[str] = None,
    pdf_text_hint: str = "",
    report_urls_hint: list[str] | None = None,
) -> tuple[dict, list[dict]]:
    """
    Scrape the company's own Investor Relations website for multi-year financial
    tables (Key Financials, Financial Highlights, Financial Summary pages).

    pdf_text_hint     : First few thousand chars of any annual report / filing PDF.
                        URLs extracted from it are tried as base URL candidates.
    report_urls_hint  : List of PDF/filing URLs already known (from NSE/BSE download
                        metadata). The domain portion of these URLs is tried as a
                        base URL candidate when the yfinance/BSE lookup fails.

    Returns:
        (result_dict, debug_log)
        result_dict : {"P&L": {"years": [...], "data": {...}},
                       "Balance Sheet": {...}, "Operating Metrics": {...}}
    """
    from web.pdf_fetcher import _get_company_website, _get_requests
    import json as _json

    requests = _get_requests()
    debug_log: list[dict] = []
    result: dict = {}

    # ── Step 1: Find company website ─────────────────────────────────────────
    # Build a ranked list of base URL candidates from multiple sources:
    # A. Standard lookup (yfinance / BSE / NSE / Screener)
    # B. URLs extracted from annual report PDF text (cover page always has company site)
    # C. Domain portion of filing URLs from NSE/BSE (the filing server host ≠ company site,
    #    but the cover PDF text mentions the company website)
    from urllib.parse import urlparse as _urlparse

    _base_candidates: list[str] = []

    # A. Standard lookup
    _std = _get_company_website(symbol, bse_code)
    if _std:
        _base_candidates.append(_std.rstrip("/"))

    # B. Extract URLs from annual report PDF text / corporate announcement text
    if pdf_text_hint:
        _pdf_urls = _extract_urls_from_text(pdf_text_hint[:8_000])
        _base_candidates.extend(u for u in _pdf_urls if u not in _base_candidates)
        if _pdf_urls:
            debug_log.append({"step": "pdf_url_hint", "urls_found": _pdf_urls[:5]})

    # C. Filing URL hints — extract company name from path to guess domain
    for _furl in (report_urls_hint or []):
        try:
            _ph = _urlparse(_furl)
            # Filing URLs are like bsecorporate.bseindia.com/... — skip exchange domains
            _fhost = (_ph.hostname or "").lower()
            if any(d in _fhost for d in ("nseindia", "bseindia", "sebi", "mca", "rbi")):
                # Try to extract company name from the path segments
                for seg in reversed(_ph.path.split("/")):
                    seg = seg.strip().lower()
                    if len(seg) > 5 and not seg.endswith((".pdf", ".xml", ".zip")):
                        # Guess: company slug → try as www.{slug}.com / www.{slug}.in
                        _slug = re.sub(r"[^a-z0-9]", "", seg)
                        if len(_slug) > 4:
                            for _tld in (".com", ".in", ".co.in"):
                                _c = f"https://www.{_slug}{_tld}"
                                if _c not in _base_candidates:
                                    _base_candidates.append(_c)
                        break
        except Exception:
            pass

    if not _base_candidates:
        debug_log.append({"step": "website_lookup", "error": f"Could not resolve website for {symbol}"})
        return {}, debug_log

    # Probe each candidate and use the first that responds with HTTP 200
    base_url = _base_candidates[0]
    for _cand in _base_candidates:
        try:
            _probe = requests.get(_cand, headers={"User-Agent": "Mozilla/5.0"}, timeout=8, allow_redirects=True)
            if _probe.status_code < 400:
                base_url = _probe.url.rstrip("/")
                debug_log.append({"step": "website_lookup", "url": base_url,
                                  "source": "probed", "status": _probe.status_code})
                break
            else:
                debug_log.append({"step": "website_probe_failed", "url": _cand, "status": _probe.status_code})
        except Exception as _pe:
            debug_log.append({"step": "website_probe_failed", "url": _cand, "error": str(_pe)})
    else:
        # All probes failed — fall back to first candidate anyway
        debug_log.append({"step": "website_lookup", "url": base_url, "warning": "all probes failed, using first candidate"})

    base_url = base_url.rstrip("/")

    # ── Step 2: Try common IR financial data paths ────────────────────────────
    # Indian company IR pages commonly use these patterns
    ir_paths = [
        "/investor-relations/financial-highlights",
        "/investor-relations/key-financials",
        "/investor-relations/financial-information",
        "/investor-relations/financial-results",
        "/investor-relations/financials",
        "/investor-relations/annual-report",
        "/investor-relations",
        "/investors/financial-highlights",
        "/investors/financial-information",
        "/investors/key-financials",
        "/investors/financials",
        "/investors/financial-results",
        "/investors",
        "/investor-corner/financial-results",
        "/investor-corner",
        "/financial-highlights",
        "/key-financials",
        "/financials",
        "/about-us/financial-highlights",
        "/about-us/key-financials",
        "/about-us/investor-relations",
        "/media-room/financial-results",
    ]

    _headers = {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,*/*",
        "Accept-Language": "en-US,en;q=0.9",
        "Referer": base_url,
    }

    # Financial keyword indicators — high-density on the right page
    _FIN_KW_RE = re.compile(
        r"net\s+(?:revenue|income|profit|interest|npa|advances|deposits)"
        r"|gross\s+(?:advances|npa|income)"
        r"|total\s+(?:assets|deposits|income|revenue)"
        r"|(?:capital\s+adequacy|crar|casa|nim|roa|roe|gnpa|nnpa|pat|pbdt|ebitda)"
        r"|return\s+on\s+(?:assets|equity|net\s+worth)"
        r"|net\s+interest\s+margin"
        r"|(?:201[5-9]|202[0-9])\s*[-–]\s*(?:201[6-9]|202[0-9])",  # year ranges
        re.I,
    )

    best_page_text: str = ""
    best_page_url:  str = ""
    best_score:     int = 0
    all_tables_data: dict = {}   # merged from all pages

    for path in ir_paths:
        url = base_url + path
        entry: dict = {"step": "ir_page", "url": url, "status": None,
                       "tables": 0, "rows": 0, "score": 0, "error": None}
        try:
            r = requests.get(url, headers=_headers, timeout=12, allow_redirects=True)
            entry["status"] = r.status_code
            if r.status_code not in (200, 301, 302):
                debug_log.append(entry)
                continue

            soup = _BS4(r.text, "html.parser")
            page_text = soup.get_text(" ", strip=True)
            score = len(_FIN_KW_RE.findall(page_text[:10_000]))
            entry["score"] = score

            # Parse all tables on the page
            tables = soup.find_all("table")
            entry["tables"] = len(tables)
            page_data: dict = {}
            for tbl in tables:
                parsed = _insta_parse_table(tbl)
                for lbl, yr_vals in parsed.items():
                    if lbl not in page_data:
                        page_data[lbl] = yr_vals
                    else:
                        page_data[lbl].update(yr_vals)
            entry["rows"] = len(page_data)

            # Also check for embedded JSON (some sites use React/Vue)
            _js_detected = False
            if not page_data:
                for script in soup.find_all("script"):
                    sc = script.string or ""
                    if score > 0 and len(sc) > 200:
                        for m in re.finditer(r'"(F?\d{4}|20\d{2}-\d{2,4})"\s*:\s*\{', sc):
                            entry["error"] = entry.get("error", "") + "[JS data detected] "
                            _js_detected = True
                            break

            # ── Playwright fallback: JS-rendered content, no static tables ─────
            if _js_detected and not page_data and _PLAYWRIGHT_AVAILABLE:
                entry["playwright"] = "attempted"
                _pw_html = _playwright_fetch(url)
                if _pw_html:
                    _pw_soup = _BS4(_pw_html, "html.parser")
                    _pw_tables = _pw_soup.find_all("table")
                    for _tbl in _pw_tables:
                        _parsed = _insta_parse_table(_tbl)
                        for _lbl, _yvs in _parsed.items():
                            page_data.setdefault(_lbl, {}).update(_yvs)
                    if page_data:
                        entry["playwright"] = f"success ({len(page_data)} rows from {len(_pw_tables)} tables)"
                        # Also re-score from rendered text
                        _pw_text = _pw_soup.get_text(" ", strip=True)
                        score = max(score, len(_FIN_KW_RE.findall(_pw_text[:10_000])))
                        if score > best_score:
                            best_score = score
                            best_page_url = url
                            best_page_text = _pw_text[:30_000]
                    else:
                        entry["playwright"] = "rendered but no tables parsed"
                else:
                    entry["playwright"] = "fetch failed (browser not installed?)"

            if page_data and score > 0:
                all_tables_data.update(page_data)

            if score > best_score:
                best_score = score
                best_page_url = url
                best_page_text = page_text[:30_000]

            debug_log.append(entry)

            if score >= 5 and page_data:
                break   # found a good page with data — stop

        except Exception as e:
            entry["error"] = str(e)
            debug_log.append(entry)
            continue

    # ── Step 3: Classify rows into sheets ────────────────────────────────────
    _BS_KEYWORDS = re.compile(
        r"advance|deposit|asset|borrowing|investment|capital|reserve|equity"
        r"|npa|liabilit|net\s+worth|fixed\s+asset",
        re.I,
    )
    _PL_KEYWORDS = re.compile(
        r"income|revenue|profit|loss|tax|interest\s+earn|interest\s+expend"
        r"|provision|operating\s+expense|employee|nii|ppop",
        re.I,
    )
    _OP_KEYWORDS = re.compile(
        r"nim|casa|crar|roa|roe|gnpa|nnpa|coverage|disbursement|branch"
        r"|customer|yield|cost\s+of\s+fund|spread|net\s+interest\s+margin"
        r"|return\s+on|capital\s+adequacy",
        re.I,
    )

    pl_data: dict = {}
    bs_data: dict = {}
    op_data: dict = {}

    for lbl, yr_vals in all_tables_data.items():
        if not yr_vals:
            continue
        if _OP_KEYWORDS.search(lbl):
            op_data[lbl] = yr_vals
        elif _BS_KEYWORDS.search(lbl):
            bs_data[lbl] = yr_vals
        else:
            pl_data[lbl] = yr_vals  # default to P&L if unclear

    for sec_name, sec_data in [("P&L", pl_data), ("Balance Sheet", bs_data), ("Operating Metrics", op_data)]:
        if sec_data:
            years = sorted({yr for row in sec_data.values() for yr in row})
            result[sec_name] = {"years": years, "data": sec_data}

    # ── Step 3b: Link discovery — crawl IR/investors root page for fin links ───
    # If all fixed paths 404'd and we have no data, try discovering links
    if not all_tables_data and best_score == 0:
        _crawl_roots = [base_url + "/investor-relations", base_url + "/investors",
                        base_url + "/investor-corner", base_url]
        _fin_link_re = re.compile(
            r"financial|highlight|result|annual|key.metric|performance|summary",
            re.I,
        )
        for _root in _crawl_roots:
            try:
                _cr = requests.get(_root, headers=_headers, timeout=10, allow_redirects=True)
                if _cr.status_code != 200:
                    continue
                _csoup = _BS4(_cr.text, "html.parser")
                # Collect all <a href> links that look financial
                _cands: list[str] = []
                for _a in _csoup.find_all("a", href=True):
                    _href = _a["href"]
                    _text = _a.get_text(strip=True)
                    if _fin_link_re.search(_href) or _fin_link_re.search(_text):
                        if _href.startswith("http"):
                            if base_url in _href:
                                _cands.append(_href)
                        elif _href.startswith("/"):
                            _cands.append(base_url + _href)
                # Deduplicate and try top 5 unique candidates
                _seen_c: set = set()
                for _curl in _cands:
                    _curl = _curl.rstrip("/")
                    if _curl in _seen_c or _curl in [base_url + p for p in ir_paths]:
                        continue
                    _seen_c.add(_curl)
                    if len(_seen_c) > 5:
                        break
                    try:
                        _cr2 = requests.get(_curl, headers=_headers, timeout=10, allow_redirects=True)
                        if _cr2.status_code != 200:
                            continue
                        _s2 = _BS4(_cr2.text, "html.parser")
                        _pt2 = _s2.get_text(" ", strip=True)
                        _sc2 = len(_FIN_KW_RE.findall(_pt2[:10_000]))
                        _tables2 = _s2.find_all("table")
                        _pd2: dict = {}
                        for _t2 in _tables2:
                            for _lbl, _yvs in _insta_parse_table(_t2).items():
                                _pd2.setdefault(_lbl, {}).update(_yvs)
                        _entry2 = {"step": "discovered_link", "url": _curl,
                                   "status": 200, "tables": len(_tables2),
                                   "rows": len(_pd2), "score": _sc2}
                        debug_log.append(_entry2)
                        if _pd2 and _sc2 > 0:
                            all_tables_data.update(_pd2)
                        if _sc2 > best_score:
                            best_score = _sc2
                            best_page_url = _curl
                            best_page_text = _pt2[:30_000]
                    except Exception:
                        pass
                if all_tables_data:
                    break   # found something — stop crawling
            except Exception:
                pass

        # ── Playwright fallback for Step 3b: try rendering the best-scored URL ─
        if not all_tables_data and best_page_url and _PLAYWRIGHT_AVAILABLE:
            _pw_entry = {"step": "playwright_crawl_fallback", "url": best_page_url,
                         "playwright": "attempted"}
            _pw_html = _playwright_fetch(best_page_url)
            if _pw_html:
                _pw_soup = _BS4(_pw_html, "html.parser")
                _pw_tables = _pw_soup.find_all("table")
                _pw_data: dict = {}
                for _tbl in _pw_tables:
                    for _lbl, _yvs in _insta_parse_table(_tbl).items():
                        _pw_data.setdefault(_lbl, {}).update(_yvs)
                if _pw_data:
                    all_tables_data.update(_pw_data)
                    _pw_entry["playwright"] = f"success ({len(_pw_data)} rows)"
                    _pw_entry["tables"] = len(_pw_tables)
                else:
                    # No tables, but grab rendered text for LLM fallback
                    _pw_text = _pw_soup.get_text(" ", strip=True)
                    _pw_score = len(_FIN_KW_RE.findall(_pw_text[:10_000]))
                    if _pw_score > best_score:
                        best_score = _pw_score
                        best_page_text = _pw_text[:30_000]
                    _pw_entry["playwright"] = f"rendered, no tables — {_pw_score} kw score"
            else:
                _pw_entry["playwright"] = "fetch failed (browser not installed?)"
            debug_log.append(_pw_entry)

    # Re-classify after discovery step
    if all_tables_data:
        for lbl, yr_vals in all_tables_data.items():
            if not yr_vals:
                continue
            sec = "Operating Metrics" if _OP_KEYWORDS.search(lbl) else (
                  "Balance Sheet" if _BS_KEYWORDS.search(lbl) else "P&L")
            if sec not in result:
                result[sec] = {"years": [], "data": {}}
            result[sec]["data"][lbl] = yr_vals
        for sec in result:
            if isinstance(result[sec], dict) and "data" in result[sec]:
                result[sec]["years"] = sorted({yr for row in result[sec]["data"].values() for yr in row})

    # ── Step 4: If no tables found, store page text for LLM ──────────────────
    if not all_tables_data and best_page_text:
        result["__ir_page_text__"] = best_page_text
        debug_log.append({"step": "fallback_text", "url": best_page_url,
                          "chars": len(best_page_text), "score": best_score})

    total = sum(len(s.get("data", {})) for s in result.values() if isinstance(s, dict) and "data" in s)
    logger.info(f"Company IR website: {total} fields from {base_url} for {symbol}")
    return result, debug_log


# ═══════════════════════════════════════════════════════════════════════════════

def fetch_yfinance_raw(symbol: str, years: list[str]) -> dict:
    """Fetch yfinance data with a hard 20-second timeout to prevent hanging."""
    from concurrent.futures import ThreadPoolExecutor, TimeoutError as FuturesTimeout

    def _fetch():
        tk = yf.Ticker(f"{symbol.upper()}.NS")
        year_set = set(years)
        _CRORE   = 1e7
        result   = {}
        for sec_name, attr in [
            ("Income Statement", "income_stmt"),
            ("Balance Sheet",    "balance_sheet"),
            ("Cash Flow",        "cashflow"),
        ]:
            try:
                df = getattr(tk, attr)
                if df is None or df.empty:
                    continue
                sec: dict[str, dict[str, float]] = {}
                for col in df.columns:
                    try:
                        yr_label = f"F{col.year if col.month <= 3 else col.year + 1}"
                    except AttributeError:
                        continue
                    if yr_label not in year_set:
                        continue
                    vals = {}
                    for metric in df.index:
                        try:
                            v = float(df.loc[metric, col])
                            if v == v:
                                vals[str(metric)] = round(v / _CRORE, 4)
                        except Exception:
                            pass
                    if vals:
                        sec[yr_label] = vals
                if sec:
                    result[sec_name] = sec
            except Exception:
                pass
        return result

    try:
        with ThreadPoolExecutor(max_workers=1) as ex:
            fut = ex.submit(_fetch)
            return fut.result(timeout=20)
    except FuturesTimeout:
        logger.warning(f"yfinance: timed out after 20s for {symbol} — skipping")
        return {}
    except Exception as e:
        logger.warning(f"yfinance: failed for {symbol} — {e}")
        return {}


# ═══════════════════════════════════════════════════════════════════════════════
# PDF fetcher: auto-download annual report from NSE / BSE
# ═══════════════════════════════════════════════════════════════════════════════

from web.pdf_fetcher import (
    fetch_nse_report_list,
    fetch_bse_report_list,
    fetch_screener_report_links,
    fetch_company_annual_reports,
    download_pdf,
    get_available_presentations,
    get_available_quarterly_results,
    fetch_xbrl_from_url,
)
from web.nse_connector import NSEConnector
from web.bse_connector import BSEConnector
from web.base import CompanyIdentifier

# ═══════════════════════════════════════════════════════════════════════════════
# NSE / BSE structured financial data (quarterly aggregated → annual)
# ═══════════════════════════════════════════════════════════════════════════════

def fetch_exchange_raw(symbol: str, bse_code: Optional[str], years: list[str]) -> dict:
    """
    Fetch structured financial data from NSE/BSE/Tickertape in parallel.
    Each connector runs in its own thread with a 25-second timeout so a
    slow or hung source never blocks the whole pipeline.
    """
    from concurrent.futures import ThreadPoolExecutor, as_completed, TimeoutError as FutTimeout
    from web.base import WebDataResult

    company = CompanyIdentifier(
        name=symbol,
        nse_symbol=symbol.upper(),
        bse_code=bse_code or "",
    )

    def _fetch_tickertape():
        from web.tickertape_connector import TickertapeConnector
        return TickertapeConnector().fetch(company, years)

    def _fetch_nse():
        return NSEConnector().fetch(company, years)

    def _fetch_bse():
        return BSEConnector().fetch(company, years) if bse_code else []

    tasks = {
        "tickertape": _fetch_tickertape,
        "nse":        _fetch_nse,
        "bse":        _fetch_bse,
    }

    all_results: list[WebDataResult] = []
    _CONNECTOR_TIMEOUT = 25  # seconds per connector

    with ThreadPoolExecutor(max_workers=3) as ex:
        futures = {ex.submit(fn): name for name, fn in tasks.items()}
        for fut in as_completed(futures, timeout=_CONNECTOR_TIMEOUT + 5):
            name = futures[fut]
            try:
                results = fut.result(timeout=_CONNECTOR_TIMEOUT)
                all_results.extend(results)
                logger.info(f"exchange_raw: {name} returned {len(results)} data points")
            except FutTimeout:
                logger.warning(f"exchange_raw: {name} timed out — skipping")
            except Exception as e:
                logger.warning(f"exchange_raw: {name} failed — {e}")

    if not all_results:
        return {}

    # Bucket into the same section structure as Screener
    _SECTION_BY_FIELD = {
        "Revenue from Operations": "P&L",
        "Total Revenue":           "P&L",
        "Other Income":            "P&L",
        "Total Expenses":          "P&L",
        "Operating Profit":        "P&L",
        "Finance Costs":           "P&L",
        "Depreciation and Amortisation": "P&L",
        "Profit Before Tax":       "P&L",
        "Tax Expense":             "P&L",
        "Profit After Tax":        "P&L",
        "EPS (Basic)":             "P&L",
        "EPS (Diluted)":           "P&L",
        "Interest Income":         "P&L",
        "Interest Expended":       "P&L",
        "Net Interest Income":     "P&L",
        "Provisions and Contingencies": "P&L",
        "Operating Expenses":      "P&L",
        "Profit Before Depreciation & Tax": "P&L",
        # Balance Sheet
        "Share Capital":           "Balance Sheet",
        "Reserves and Surplus":    "Balance Sheet",
        "Total Borrowings":        "Balance Sheet",
        "Total Liabilities":       "Balance Sheet",
        "Fixed Assets":            "Balance Sheet",
        "Investments":             "Balance Sheet",
        "Total Assets":            "Balance Sheet",
        "Cash and Cash Equivalents": "Balance Sheet",
        "Total Equity":            "Balance Sheet",
        "Deposits":                "Balance Sheet",
        "Advances":                "Balance Sheet",
        "Net NPA":                 "Balance Sheet",
        "Gross NPA":               "Balance Sheet",
        # Cash Flow (Tickertape)
        "Net Cash from Operating Activities": "Cash Flow",
        "Net Cash from Investing Activities": "Cash Flow",
        "Net Cash from Financing Activities": "Cash Flow",
        "Capital Expenditure":     "Cash Flow",
        "Free Cash Flow":          "Cash Flow",
    }

    sections: dict[str, dict] = {}
    for res in all_results:
        sec = _SECTION_BY_FIELD.get(res.template_field, "P&L")
        if sec not in sections:
            sections[sec] = {"years": set(), "data": {}}
        field_data = sections[sec]["data"].setdefault(res.template_field, {})
        # NSE > BSE — don't overwrite NSE value with BSE value for same field+year
        if res.year not in field_data:
            field_data[res.year] = res.value
        sections[sec]["years"].add(res.year)

    # Serialise sets to sorted lists
    return {
        sec: {"years": sorted(sd["years"]), "data": sd["data"]}
        for sec, sd in sections.items()
    }


def _fmt_exchange_for_llm(exchange_raw: dict, years: list[str], section: str) -> str:
    """Format exchange data for a given section as JSON lines for the LLM prompt."""
    sd = exchange_raw.get(section, {})
    if not sd:
        return ""
    d = {
        lbl: {yr: v for yr, v in yr_vals.items() if yr in years}
        for lbl, yr_vals in sd.get("data", {}).items()
    }
    d = {k: v for k, v in d.items() if v}
    if not d:
        return ""
    return json.dumps(d, indent=2)


# Sheets that should be populated from investor presentations rather than
# (or in addition to) the annual report.  Generic keywords — works for any
# company template, not just banks.
_PRESENTATION_SHEET_RE = re.compile(
    r"operat|asset.?qual|npa|nim|yield|spread|kpi|quarter|metrics?",
    re.I,
)
# Sheets that benefit from quarterly result PDFs as primary context
_QUARTERLY_SHEET_RE = re.compile(
    r"asset.?qual|npa|gnpa|nnpa|quarter|operat|metrics?|kpi",
    re.I,
)


def get_available_reports(symbol: str, bse_code: Optional[str] = None) -> list[dict]:
    """Return all available annual report records from company IR + NSE + BSE + Screener.

    Company IR page is checked first so that direct PDF links (not zip-packaged)
    appear with highest priority. NSE/BSE are tried next for any years the company
    site missed, followed by Screener as a broad fallback.
    """
    seen_urls: set[str] = set()
    all_reports: list[dict] = []

    # ── 1. Company's own IR page (highest quality — direct PDFs, not zips) ──
    try:
        for rep in fetch_company_annual_reports(symbol, bse_code):
            if rep["url"] not in seen_urls:
                rep["source"] = "Company IR"
                all_reports.append(rep)
                seen_urls.add(rep["url"])
    except Exception:
        pass

    # ── 2. NSE ──────────────────────────────────────────────────────────────
    for rep in fetch_nse_report_list(symbol):
        if rep["url"] not in seen_urls:
            rep["source"] = "NSE"
            all_reports.append(rep)
            seen_urls.add(rep["url"])

    # ── 3. BSE ──────────────────────────────────────────────────────────────
    if bse_code:
        for rep in fetch_bse_report_list(bse_code):
            if rep["url"] not in seen_urls:
                rep["source"] = "BSE"
                all_reports.append(rep)
                seen_urls.add(rep["url"])

    # ── 4. Screener ─────────────────────────────────────────────────────────
    for rep in fetch_screener_report_links(symbol):
        if rep["url"] not in seen_urls:
            rep["source"] = "Screener"
            all_reports.append(rep)
            seen_urls.add(rep["url"])

    all_reports.sort(key=lambda r: r.get("year") or "", reverse=True)
    return all_reports


# ═══════════════════════════════════════════════════════════════════════════════
# PDF text extraction — raw text from financial pages for LLM context
# ═══════════════════════════════════════════════════════════════════════════════

# ── Page heading classification ───────────────────────────────────────────────
#
# We read only the first 5 non-empty lines of each page (the heading area) to
# decide what kind of content that page contains.  This is fast and reliable:
# financial statement pages always start with their section title, and
# presentation slides always start with the slide title.
#
# Three-tier decision per page:
#   KEEP_FINANCIAL   — core financial statement page (P&L, BS, CF, NPA, metrics)
#   SKIP             — notes, disclosures, governance, ESG — never useful for LLM
#   KEEP_ALL         — everything else (only used for short docs where we take all)

_HEADING_KEEP = re.compile(
    # Annual report financial statements
    r"balance\s+sheet|profit\s+(?:and|&)\s+loss|profit\s+and\s+loss\s+account"
    r"|cash\s+flow|income\s+statement|statement\s+of\s+(?:profit|financial\s+position)"
    r"|revenue\s+from\s+operations|interest\s+(?:earned|income|expended)"
    # ── Indian bank RBI Schedules 13–17: these MUST be included ─────────────────
    # Schedule 13 = Interest earned breakdown, 14 = Other income,
    # 15 = Interest expended, 16 = Operating expenses (employee + other sub-lines),
    # 17 = Provisions and contingencies (credit cost).
    # Blanket-skipping schedule pages causes employee/opex sub-lines to be missed.
    r"|schedule\s+1[3-7]\b"             # Schedule 13, 14, 15, 16, 17
    r"|schedule\s+[1-9]\b(?!\d)"        # Schedule 1-9 (BS schedules)
    r"|schedule\s+1[0-2]\b"             # Schedule 10, 11, 12
    r"|operating\s+expenses?\s+schedule"
    r"|payments?\s+to\s+(?:and\s+)?(?:provisions?\s+for\s+)?employees?"
    r"|other\s+operating\s+expenses?"
    # Quarterly / presentation operating metrics
    r"|asset\s+quality|npa|gnpa|nnpa|net\s+npa|gross\s+npa"
    r"|net\s+interest\s+(?:margin|income)|nim\b"
    r"|yield\s+on|cost\s+of\s+(?:funds|deposits|borrowings)"
    r"|loan\s+book|advances\s+(?:portfolio|mix|break)"
    r"|key\s+(?:financial|operating|performance)\s+(?:highlights?|metrics?|indicators?|data)"
    r"|financial\s+highlights?|operating\s+highlights?"
    r"|five.?year\s+(?:financial\s+)?(?:summary|highlights?|data|trend)"
    r"|ten.?year\s+(?:financial\s+)?(?:summary|highlights?|data|trend)"
    r"|at\s+a\s+glance|annual\s+(?:snapshot|highlights?|summary)"
    r"|casa|deposit\s+(?:mix|composition|break)"
    r"|capital\s+(?:adequacy|position)\s+(?:ratio)?"
    r"|return\s+on\s+(?:assets?|equity|capital)"
    r"|earnings\s+per\s+share|eps\b"
    r"|provisions?\s+(?:and\s+)?contingenc"
    r"|net\s+(?:profit|revenue|income)"
    r"|total\s+(?:assets?|income|revenue|deposits?|advances?)"
    # ── Investor presentation slide headings (broad) ──────────────────────────
    # Many slides use short headings that don't appear above — include them so
    # they're caught as 'keep' rather than 'unclear'
    r"|^advances$|gross\s+advances|advance\s+(?:mix|book|portfolio|break)"
    r"|disbursement|portfolio\s+(?:mix|overview|quality|break)"
    r"|product\s+(?:mix|portfolio|breakdown|wise)|segment\s+(?:mix|wise|break)"
    r"|loan\s+(?:mix|portfolio|book|breakup)|book\s+quality"
    r"|micro\s+(?:finance|loan|credit)|mfi\b|sbl\b|msme?\b|nbfc\b"
    r"|vehicle\s+finance|housing\s+finance|gold\s+loan"
    r"|liability\s+(?:profile|mix|franchise)|branch\s+banking"
    r"|collection\s+efficiency|credit\s+cost|write.?off"
    r"|restructur|slippage|recovery|resolution|collection"
    r"|key\s+highlights?|snapshot|financial\s+performance"
    r"|liabilities?|borrowing\s+(?:mix|profile)|funding\s+(?:mix|profile)"
    r"|spread|margin|roe|roa|roae|roaa"
    r"|business\s+(?:update|highlight|overview|performance)"
    r"|^financials?$|financials?\s+(?:summary|highlight|overview|performance)"
    r"|other\s+updates?|beyond\s+banking|branch\s+network|operational",
    re.I,
)

_HEADING_SKIP = re.compile(
    r"notes?\s+(?:to|on|forming\s+part\s+of)\s+(?:the\s+)?(?:financial|accounts?)"
    r"|significant\s+accounting\s+polic"
    r"|basis\s+of\s+(?:preparation|consolidation)"
    r"|independent\s+auditor|statutory\s+auditor|auditor.s\s+report"
    r"|related\s+party"
    r"|directors.{0,2}\s*report|board\s+of\s+directors"
    r"|corporate\s+governance"
    r"|management\s+discussion\s+(?:and|&)\s+analysis"   # MD&A narrative — skip
    r"|business\s+responsibility"
    r"|\bbrsr\b|\besg\b"
    r"|sustainability\s+report"
    r"|csr\s+(?:activit|spend|committee)"
    # AGM / EGM notices and e-voting pages
    r"|notice\s+of\s+(?:agm|annual\s+general|egm|extra.?ordinary)"
    r"|annual\s+general\s+meeting"
    r"|^notice\s+annual\s+report"        # "Notice ANNUAL REPORT 2024-25" cover page
    r"|(?:remote\s+)?e.?voting|postal\s+ballot|cut.?off\s+date"
    r"|(?:ninth|tenth|eleventh|twelfth|thirteenth|special)\s+annual\s+(?:general|report)"
    # Exchange submission cover letters (the wrapper around the PDF)
    r"|national\s+stock\s+exchange\s+of\s+india|bombay\s+stock\s+exchange"
    r"|bse\s+limited|nse\s+(?:limited|india)|exchange\s+plaza"
    r"|phiroze\s+jeejeebhoy|bandra\s+kurla\s+complex"
    r"|dear\s+sir\s*/\s*madam|sub:\s+notice|scrip\s+code:"
    r"|listing\s+obligations?\s+and\s+disclosure"
    r"|sebi\s+\(listing"
    # Other non-financial sections
    r"|secretarial\s+(?:audit|report)"
    r"|dividend\s+(?:distribution|history)"
    # Note: ten-year / five-year financial summaries are now KEPT (moved to _HEADING_KEEP)
    # NOTE: do NOT skip schedule pages generically — Schedule 13-17 are KEEP (see _HEADING_KEEP).
    # Only skip schedules that are pure notes / disclosures, not financial data schedules.
    r"|schedule\s+18\b"                               # Schedule 18 = significant accounting policies
    r"|schedule\s+1[9-9]\b"                           # Schedule 19+ = notes to accounts
    r"|schedule\s+2\d\b",                             # Schedule 20+ = notes to accounts
    re.I,
)


def _page_heading(raw_text: str, n_lines: int = 10) -> str:
    """Return the first n non-empty lines joined — the heading area of a page.

    10 lines (up from 5) because Indian annual reports often have 4-6 lines of
    company header/footer before the actual section title (e.g. "Schedule 16").
    """
    lines = [l.strip() for l in raw_text.splitlines() if l.strip()]
    return " | ".join(lines[:n_lines])


# Secondary deep-scan pattern: check full page body when heading scan is unclear.
# Catches schedule pages where the schedule number appears mid-page.
_BODY_KEEP = re.compile(
    r"payments\s+to\s+and\s+provisions\s+for\s+employees"   # Schedule 16 (i)
    r"|other\s+operating\s+expenses"                         # Schedule 16 (ii)
    r"|schedule\s+1[3-8]\b"                                  # P&L schedules
    r"|schedule\s+[6-9]\b(?!\d)"                             # BS asset schedules
    r"|schedule\s+1[0-2]\b"                                  # BS schedules 10-12
    r"|schedules\s+forming\s+part\s+of"                      # "Schedules forming part of..."
    r"|provisions\s+and\s+contingencies"                     # Schedule 17/18
    r"|interest\s+on\s+deposits"                             # Schedule 15 sub-item
    r"|interest\/discount\s+on\s+advances"                   # Schedule 13 sub-item
    r"|cash\s+and\s+balances\s+with\s+reserve\s+bank"        # Schedule 6 heading
    # Financial highlights / KPI pages (for operating metrics historical data)
    r"|five.?year\s+(?:financial\s+)?(?:summary|highlight|data|trend)"
    r"|at\s+a\s+glance"
    r"|key\s+(?:financial|operating|performance)\s+(?:highlight|metric|indicator|data)"
    r"|return\s+on\s+(?:average\s+)?(?:assets?|equity)\s*[:\|]"  # ROA/ROE table rows
    r"|net\s+interest\s+margin\s*[:\|%]|\bnim\b\s*[:\|%]"        # NIM in table
    r"|cost.to.income\s+ratio\s*[:\|%]|\bctir?\b\s*[:\|%]"
    r"|capital\s+adequacy\s+ratio\s*[:\|%]|\bcrar?\b\s*[:\|%]",
    re.I,
)


def _classify_heading(heading: str, full_text: str = "") -> str:
    """
    Classify a page heading as 'keep', 'skip', or 'unclear'.
    'keep'    → include page in LLM context
    'skip'    → exclude (notes, governance, ESG, etc.)
    'unclear' → caller decides based on doc_type

    Two-pass: fast heading check first; if unclear, do a body scan for
    financial schedule keywords (catches schedule pages whose heading is
    dominated by the company letterhead).
    """
    h = heading.lower()
    if _HEADING_KEEP.search(h):
        return "keep"
    if _HEADING_SKIP.search(h):
        return "skip"
    # Secondary pass: scan up to first 3000 chars of body for schedule markers
    if full_text and _BODY_KEEP.search(full_text[:3000]):
        return "keep"
    return "unclear"


# ── TOC extractor ─────────────────────────────────────────────────────────────
#
# Many Indian annual reports and some presentations include a Table of Contents
# that maps section names → page numbers.  If we find it, we can jump directly
# to the right pages without scanning every page.
#
# Patterns we recognise:
#   "Balance Sheet ......... 45"   (dot leaders)
#   "Balance Sheet         45"     (large gap)
#   "Balance Sheet  45"            (moderate gap, 2+ spaces)

_TOC_LINE_RE = re.compile(
    r"^(.{4,60?}?)\s*\.{2,}\s*(\d{1,3})\s*$"       # dot leaders
    r"|^(.{4,60?}?)\s{3,}(\d{1,3})\s*$",            # large whitespace gap
    re.MULTILINE,
)

# TOC entry heading → which category it maps to
_TOC_CATEGORY_MAP = [
    (re.compile(r"balance\s+sheet|financial\s+position",           re.I), "balance_sheet"),
    (re.compile(r"profit\s+(?:and|&)\s+loss|income\s+statement",   re.I), "pl"),
    (re.compile(r"cash\s+flow",                                     re.I), "cash_flow"),
    (re.compile(r"schedule[s]?\s+to\s+(?:the\s+)?balance",         re.I), "bs_schedules"),
    (re.compile(r"schedule[s]?\s+(?:1[0-2]|[1-9])\b",              re.I), "bs_schedules"),
    (re.compile(r"schedule[s]?\s+to\s+(?:the\s+)?profit",          re.I), "pl_schedules"),
    (re.compile(r"schedule[s]?\s+1[3-7]\b",                        re.I), "pl_schedules"),
    (re.compile(r"asset\s+quality|npa",                             re.I), "asset_quality"),
    (re.compile(r"key\s+(?:financial|operating|performance)|financial\s+highlight", re.I), "highlights"),
    (re.compile(r"loan\s+book|advances|portfolio",                  re.I), "loan_book"),
    (re.compile(r"nim|net\s+interest\s+margin|yield|cost\s+of\s+fund", re.I), "nim_yields"),
    (re.compile(r"operational\s+(?:highlights?|metrics?)|operat",   re.I), "operating"),
]


def _extract_toc(pdf, max_toc_pages: int = 15) -> dict[str, list[int]]:
    """
    Scan the first max_toc_pages pages for a Table of Contents.
    Returns {category: [page_numbers]} if found, else {}.
    """
    from utils.helpers import clean_for_llm

    toc: dict[str, list[int]] = {}
    n_pages = len(pdf.pages)

    for pg_idx in range(min(max_toc_pages, n_pages)):
        raw = pdf.pages[pg_idx].extract_text() or ""
        if not raw.strip():
            continue

        # Look for a concentration of TOC-like lines on this page
        matches = list(_TOC_LINE_RE.finditer(raw))
        if len(matches) < 4:          # not a TOC page if fewer than 4 entries
            continue

        for m in matches:
            # Groups: (dot-leader title, dot-leader page) or (space title, space page)
            title = (m.group(1) or m.group(3) or "").strip()
            pg_str = (m.group(2) or m.group(4) or "").strip()
            if not title or not pg_str:
                continue
            try:
                target_pg = int(pg_str)
            except ValueError:
                continue
            if not (1 <= target_pg <= n_pages):
                continue

            # Map title to a known category
            for pattern, category in _TOC_CATEGORY_MAP:
                if pattern.search(title):
                    toc.setdefault(category, []).append(target_pg)
                    break

        if toc:
            logger.info(f"pdf_text: TOC found on page {pg_idx + 1} — "
                        f"{sum(len(v) for v in toc.values())} entries across "
                        f"{len(toc)} categories")
            return toc

    return {}


def _page_to_text_with_tables(page, max_chars: int, doc_type: str) -> str:
    """
    Extract text from a single pdfplumber page.
    For financial pages ('annual'/'quarterly'), tries structured table extraction first
    — this preserves column alignment for Balance Sheet / P&L / Schedule tables.
    Falls back to raw extract_text() if no tables found.
    Returns up to max_chars characters.
    """
    from utils.helpers import clean_for_llm

    # ── Strategy A: structured table extraction ────────────────────────────
    # For annual/quarterly reports, financial tables are the primary content.
    # pdfplumber's extract_tables() returns a list of 2-D cell arrays,
    # preserving multi-column layout much better than extract_text().
    if doc_type in ("annual", "quarterly"):
        try:
            tables = page.extract_tables() or []
            table_parts: list[str] = []
            for tbl in tables:
                if not tbl:
                    continue
                # Convert each row to pipe-separated text; skip fully-empty rows
                for row in tbl:
                    cells = [str(c).replace("\n", " ").strip() if c else "" for c in row]
                    if any(cells):
                        table_parts.append("  |  ".join(cells))
            if table_parts:
                table_text = "\n".join(table_parts)
                return table_text[:max_chars]
        except Exception:
            pass  # fall through to raw text

    # ── Strategy B: raw text (presentations, or when no tables found) ─────
    raw = page.extract_text() or ""
    return clean_for_llm(raw)[:max_chars]


def _extract_oper_highlights(pdf_text: str, max_chars: int = 20_000) -> str:
    """
    Search annual-report pdf_text for pages that contain operating metrics /
    financial-highlights / KPI data.  Returns the most relevant page blocks
    (up to max_chars) so the LLM can fill historical operating-metrics years.

    Each annual report contributes pages in this format:
        === ANNUAL REPORT: report_N ===
        === Page M [heading...] ===
        <page text>

    We score each page block and pick the highest-signal ones.
    """
    _KPI_RE = re.compile(
        r"financial\s+highlight|at\s+a\s+glance|key\s+(?:financial|operating|performance)"
        r"|five.?year\s+(?:financial|summary|data|highlights?)"
        r"|annual\s+(?:summary|highlights?|snapshot)"
        r"|gross\s+advances|total\s+deposits"
        r"|cost.to.income\s+ratio|\bctir?\b"
        r"|return\s+on\s+(?:assets?|equity)\s*[:\(]|\broa\b|\broe\b|\broaa?\b|\broae?\b"
        r"|net\s+interest\s+margin|\bnim\b"
        r"|\bgnpa\b|\bnnpa\b|gross\s+npa|net\s+npa"
        r"|capital\s+adequacy\s+ratio|\bcar\b|\bcrar\b"
        r"|casa\s+ratio|td\s+ratio|cost\s+of\s+(?:funds|deposits|borrowings)"
        r"|yield\s+on\s+advances"
        r"|disbursements?\s*(?:for\s+(?:q\d|fy|the\s+year))?"
        r"|credit\s+cost|provision\s+coverage",
        re.I,
    )

    # Split on page/report boundary markers
    # Pattern: === Page N [...] === or === ANNUAL REPORT: ... ===
    blocks = re.split(r"(?=\n=== (?:Page \d|ANNUAL REPORT))", pdf_text)

    scored: list[tuple[int, str]] = []
    for block in blocks:
        if len(block) < 50:
            continue
        # Score by number of KPI keyword matches (more = more relevant)
        score = len(_KPI_RE.findall(block[:3000]))
        if score > 0:
            scored.append((score, block))

    # Sort by score descending, take top pages
    scored.sort(key=lambda x: x[0], reverse=True)

    out_parts: list[str] = []
    total = 0
    for _, block in scored:
        if total >= max_chars:
            break
        out_parts.append(block)
        total += len(block)

    return "\n".join(out_parts)[:max_chars]


def extract_pdf_text_for_llm(
    pdf_bytes: bytes,
    max_chars_per_page: int = 4500,
    doc_type: str = "annual",       # "annual" | "presentation" | "quarterly"
    label: str = "",                 # optional year label, e.g. "FY2024"
) -> str:
    """
    Extract cleaned plain text from a financial PDF for the LLM.

    Strategy:
      1. Scan the first 15 pages for a Table of Contents (page-number map).
         If found, those pages are marked 'keep' and we also include the next
         3 pages after each TOC target (schedules often follow immediately).
      2. For every page NOT covered by the TOC:
         a. Extract the first 5 lines as a heading.
         b. Classify as 'keep', 'skip', or 'unclear'.
         c. If 'keep': extract using table-aware extraction (preserves columns).
            If 'skip': skip entirely.
            If 'unclear': include for annual/quarterly; skip for presentations.
      3. For annual/quarterly: also force-include pages immediately following
         any 'keep' page (schedules often continue onto the next page).
      4. Char budget: 120 k for annual, 80 k for presentations/quarterly.
         'keep' pages get full budget; 'unclear' pages get up to 2 k chars.
    """
    try:
        import pdfplumber
    except ImportError:
        return ""

    from utils.helpers import clean_for_llm

    # Char budgets by doc type
    MAX_TOTAL = 120_000 if doc_type == "annual" else 80_000
    # 'unclear' pages get a smaller slice to save budget for critical data pages
    MAX_UNCLEAR_CHARS = 2_000
    parts: list[str] = []
    total_chars = 0

    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            n_pages = len(pdf.pages)

            # ── Step 1: TOC extraction ─────────────────────────────────────────
            toc = _extract_toc(pdf)

            toc_pages: set[int] = set()
            for section_key, pg_list in toc.items():
                toc_pages.update(pg_list)
                # Schedule sections can span many pages — extend window by type
                lookahead = 10 if section_key in ("bs_schedules", "pl_schedules") else 5
                for pg in pg_list:
                    toc_pages.update(p for p in range(pg + 1, pg + 1 + lookahead) if p <= n_pages)

            # ── Step 2: Heading-based classification for all pages ─────────────
            page_decisions: dict[int, str] = {}
            raw_cache: dict[int, str] = {}   # cache raw text so we don't re-read

            for pg_idx in range(n_pages):
                pg_num = pg_idx + 1
                if pg_num in toc_pages:
                    page_decisions[pg_num] = "keep"
                    continue
                raw = pdf.pages[pg_idx].extract_text() or ""
                raw_cache[pg_num] = raw
                if not raw.strip():
                    page_decisions[pg_num] = "skip"
                    continue
                heading = _page_heading(raw)
                # Pass full raw text for body scan (catches mid-page schedule headings)
                page_decisions[pg_num] = _classify_heading(heading, full_text=raw)

            # ── Step 3: For annual/quarterly, cascade 'keep' to next 5 pages ──
            # Schedules (Balance Sheet sub-schedules 6-12, Schedule 16 etc.)
            # usually appear immediately after the main statement page and won't
            # match KEEP by heading alone. Cascade 5 pages (was 2) because
            # some schedules span many pages in multi-year annual reports.
            if doc_type in ("annual", "quarterly"):
                # Two-pass cascade so newly-promoted pages also cascade forward
                for _pass in range(2):
                    keep_pages = sorted(
                        pg for pg, dec in page_decisions.items() if dec == "keep"
                    )
                    for kp in keep_pages:
                        for nextp in range(kp + 1, kp + 6):   # 5 pages (was 2)
                            if nextp > n_pages:
                                break
                            if page_decisions.get(nextp) in ("unclear", "skip"):
                                page_decisions[nextp] = "keep"   # promote → keep

            # ── Step 4: Select pages to extract ───────────────────────────────
            selected_keep    = [pg for pg, dec in sorted(page_decisions.items()) if dec == "keep"]
            selected_unclear = [pg for pg, dec in sorted(page_decisions.items()) if dec == "unclear"]

            # For presentations: only include 'keep' + nearby 'unclear' slides
            # For annual/quarterly: include all 'keep' + all 'unclear' (may be schedules)
            selected = sorted(set(selected_keep + selected_unclear))
            if not selected:
                selected = [pg for pg, dec in sorted(page_decisions.items()) if dec != "skip"]

            logger.info(
                f"pdf_text [{doc_type}{' ' + label if label else ''}]: "
                f"{n_pages} total | TOC: {len(toc_pages)} | "
                f"keep: {len(selected_keep)} | unclear: {len(selected_unclear)} | "
                f"selected: {len(selected)}"
            )

            # ── Step 5: Extract text from selected pages ───────────────────────
            for pg_num in selected:
                if total_chars >= MAX_TOTAL:
                    break
                page = pdf.pages[pg_num - 1]
                is_keep = page_decisions.get(pg_num) == "keep"
                char_limit = max_chars_per_page if is_keep else MAX_UNCLEAR_CHARS

                # Use table-aware extraction for annual/quarterly keep pages
                extracted = _page_to_text_with_tables(page, char_limit, doc_type)
                if not extracted.strip():
                    continue

                # Build heading from raw text cache or re-read
                raw_text = raw_cache.get(pg_num) or page.extract_text() or ""
                heading = _page_heading(raw_text)
                parts.append(f"\n=== Page {pg_num} [{heading[:80]}] ===\n{extracted}")
                total_chars += len(extracted)

    except Exception as e:
        logger.warning(f"extract_pdf_text_for_llm: {e}")

    logger.info(f"pdf_text: extracted {total_chars:,} chars from {len(parts)} pages")
    return "\n".join(parts)


# ═══════════════════════════════════════════════════════════════════════════════
# PDF table extraction (pdfplumber)
# ═══════════════════════════════════════════════════════════════════════════════

def extract_pdf_data(pdf_bytes: bytes, years: list[str], statement_type: str = "consolidated") -> dict:
    """
    Run pdfplumber table extraction on a PDF and return structured data.

    Returns:
        {
          "P&L":            {"F2025": {"Revenue from Operations": 45000, ...}, ...},
          "Balance Sheet":  {...},
          "Cash Flow":      {...},
        }
    """
    from extractor.pdf_parser import parse_pdf
    from extractor.table_extractor import extract_tables_from_pdf

    try:
        doc = parse_pdf(io.BytesIO(pdf_bytes))
        doc._pdf_bytes = pdf_bytes
    except Exception as e:
        return {"error": str(e)}

    try:
        raw_results = extract_tables_from_pdf(
            pdf_source=pdf_bytes,
            parsed_doc=doc,
            statement_type=statement_type,
        )
    except Exception as e:
        return {"error": str(e)}

    year_set = set(years)
    _SECTION_MAP = {
        "Annual P&L":           "P&L",
        "Annual Balance Sheet":  "Balance Sheet",
        "Annual Cash Flow":      "Cash Flow",
        "Operating Metrics":     "Operating Metrics",
        "Quarterly":             "Quarterly",
    }

    combined: dict[str, dict[str, dict[str, float]]] = {}

    for result in raw_results:
        if result.year not in year_set:
            continue
        section = _SECTION_MAP.get(result.section, result.section)
        combined.setdefault(section, {}).setdefault(result.year, {})
        for label, value in result.data.items():
            if label not in combined[section][result.year]:
                combined[section][result.year][label] = value

    return combined


# ═══════════════════════════════════════════════════════════════════════════════
# LLM: maps template labels → merged web + PDF data
# ═══════════════════════════════════════════════════════════════════════════════

_SYSTEM_PROMPT = """\
You are a financial data extraction expert for Indian company annual reports.

You receive:
  1. EXACT row labels from an Excel template that need to be filled
  2. Structured data from Screener.in, Yahoo Finance, and PDF table parsing
  3. RAW PDF TEXT from the actual financial statement pages of the annual report

Your job: for every template label, find the correct value for each year.

════ ANNUAL vs QUARTERLY — READ THIS FIRST ════
The template contains ANNUAL fiscal-year figures (F2017, F2018 … F2026 = full 12-month year).

  ⚠ NEVER use a single-quarter number as an annual total.
  ⚠ Any data labelled "QUARTERLY XBRL CONTEXT" below contains per-quarter values — ignore for annual P&L / BS rows.
  ⚠ For annual sheets, read the ANNUAL REPORT PDF first. Screener.in shows annual aggregates.

If you see both quarterly data (from XBRL) and annual data (from PDF / Screener) for the same field:
  → ALWAYS prefer the annual figure.

════ CRITICAL — INDIAN BANK BALANCE SHEET (RBI SCHEDULE FORMAT) ════
Indian bank Balance Sheets follow RBI format. The main BS page shows summary values with a
"Schedule" number column — the detail is in the referenced schedule page.

LIABILITIES side (Schedules 1–5):
  Schedule 1  → Capital (paid-up share capital)
  Schedule 2  → Reserves and Surplus (retained earnings + revaluation reserve + share premium)
  Schedule 3  → Deposits (savings + current + term deposits, total)
  Schedule 4  → Borrowings (RBI refinance + NCD + sub-debt)
  Schedule 5  → Other liabilities and provisions

ASSETS side (Schedules 6–11):
  Schedule 6  → Cash and balances with Reserve Bank of India
  Schedule 7  → Balances with banks and money at call and short notice
  Schedule 8  → Investments
  Schedule 9  → Advances (gross advances less provisions = net advances shown on BS)
  Schedule 10 → Fixed Assets
  Schedule 11 → Other Assets
  Schedule 12 → Contingent Liabilities (off-balance-sheet)

MAPPING RULES FOR BALANCE SHEET TEMPLATE LABELS:
  "Capital"                                    → Schedule 1 total value (share capital, typically 100-2000 Cr)
  "Reserves & surplus"                         → Schedule 2 total value (retained earnings + premium, typically larger than Capital)
  "Deposits"                                   → Schedule 3 total value (total deposits, largest liability for a bank)
  "Borrowings"                                 → Schedule 4 total value (RBI + NCD + bonds)
  "Other liabilities & provision"              → Schedule 5 total value (small: 100-2000 Cr)
  "Cash & balance with RBI"                    → Schedule 6 total value (cash + RBI CRR balance)
  "Balances with banks and money at call..."   → Schedule 7 total value (interbank deposits, call money)
  "Investments"                                → Schedule 8 total value (G-sec + other securities portfolio)
  "Advances"                                   → Schedule 9 total value (net advances = gross loans minus NPA provisions)
  "Fixed assets"                               → Schedule 10 total value (small: 100-700 Cr for most banks)
  "Other assets"                               → Schedule 11 total value ONLY (typically 200-1000 Cr — small)
  "Total Assets" / "Total Liabilities"        → bottom-line TOTAL on the main Balance Sheet page

⚠ CRITICAL: Do NOT confuse these labels:
  "Other assets"  = Schedule 11 ONLY = small residual assets (200-1000 Cr range for a mid-size bank)
                    ≠ Total Assets (which is the grand sum of all assets, much larger)
  "Total Assets"  = sum of Schedules 6+7+8+9+10+11 = the GRAND TOTAL bottom line
  If Screener or yfinance provides "Total Assets", map it to the "Total Assets" template row,
  NEVER to "Other assets".

⚠ CRITICAL: The main Balance Sheet page has a SCHEDULE column (1, 2, 3…) between
  Particulars and the financial values. Do NOT treat the schedule number as a value.
  Pattern:  "Particulars  |  Schedule  |  FY_Current  |  FY_Prior"
  Example:  "Deposits  |  3  |  1,23,45,678  |  98,76,543"
  Here 3 is the schedule reference — the values are 1,23,45,678 and 98,76,543.

BALANCE SHEET SANITY CHECKS:
  Total Assets ≈ Total Liabilities (they must balance)
  Total Assets = Cash + Investments + Advances + Fixed Assets + Other Assets ± small items
  Deposits should be the LARGEST single line (60-80% of total liabilities for a retail bank)
  "Other assets" should be SMALL relative to Advances and Investments

════ CRITICAL — INDIAN BANK P&L (RBI SCHEDULE FORMAT) ════
Indian banks file under RBI's Schedule 13–18 format. These items are DISTINCT:

  Schedule 13  →  "Interest earned" = GROSS interest income (loans + investments + RBI balances)
  Schedule 14  →  "Other income" = fee income, forex, treasury profit
  Schedule 15  →  "Interest expended" = interest paid on deposits & borrowings
  Schedule 16  →  Operating expenses — has exactly TWO separate sub-lines:
      (i)  Payments to and provisions for employees  →  template label = "Employee expenses"
      (ii) Other operating expenses (rent, IT, depreciation, etc.)  →  template label = "Operating expenses"

  Screener.in / yfinance "Operating Expenses" for banks = TOTAL of (i)+(ii).
  ⚠ NEVER assign Screener "Operating Expenses" to the "Operating expenses" template row.
  ⚠ NEVER assign Screener "Operating Expenses" to the "Employee expenses" template row.
  For these two rows, use only the annual report PDF or Schedule 16 breakdown.

  "Net Interest Income" (NII) = Interest earned − Interest expended (COMPUTED, may be a formula cell)
  "Pre-provisioning operating profit" = Total income − Employee expenses − Operating expenses
  "Credit cost" = Provisions and Contingencies (Schedule 17/18)
  "Profit before tax" = PPOP − Credit cost
  "Profit after tax" = PBT − Tax provision

════ SCHEDULE 16 EXTRACTION — STEP BY STEP ════
The PDF contains TWO representations of operating expenses. You MUST use BOTH:

  A) P&L SUMMARY page (high-level):
       "Operating expenses  [Note ref]  [TOTAL_FY]  [TOTAL_PRIOR]"
       ⚠ This TOTAL = line I + line II combined. DO NOT assign this value to any template row.
       Use it only to sanity-check: TOTAL should equal Employee + Operating.

  B) "Schedules to Profit and Loss Account" page — Schedule 16 BREAKDOWN:
       Look for a page titled "Schedules to Profit and Loss Account" or "Schedule 16".
       It will list individual line items ending in a TOTAL row:
         "I   Payments to and provisions for employees  [EMP_FY]    [EMP_PRIOR]"
         "II  Rent, taxes and lighting                  [RENT_FY]   ..."
         "III Printing and stationery                   ..."
         ...  (items II through XIV or more)
         "    TOTAL                                     [TOTAL_FY]  [TOTAL_PRIOR]"

  MAPPING RULES (MANDATORY):
    "Employee expenses"  = Schedule 16 line I value  =  [EMP_FY]
    "Operating expenses" = [TOTAL_FY] − [EMP_FY]     (ALL non-employee items combined)

  WORKED EXAMPLE (Equitas SFB FY2025, units = 000's of ₹):
    Schedule 16 line I  = 1,62,10,067  →  1621.0 Cr
    Schedule 16 TOTAL   = 2,82,88,345  →  2828.8 Cr
    Employee expenses   = 1621.0 Cr   ← assign to "Employee expenses"
    Operating expenses  = 2828.8 − 1621.0 = 1207.8 Cr  ← assign to "Operating expenses"

════ CREDIT COST / PROVISIONS — EXTRACTION RULE ════
  In the P&L summary, look for the line:
    "Provisions and contingencies  [optional: (refer note. X)]  [VALUE_FY]  [VALUE_PRIOR]"
  The text "(refer note. X)" is a page cross-reference — IGNORE IT and use the numeric values.
    "Credit cost" = Provisions and contingencies value = [VALUE_FY]

  EXAMPLE (Equitas SFB FY2025):
    "Provisions and contingencies (refer note. 18)  1,05,10,000  89,50,000"
    Credit cost FY2025 = 1,05,10,000 / 10,000 = 1051.0 Cr
    Credit cost FY2024 = 89,50,000 / 10,000 = 895.0 Cr

════ UNIT CONVERSION REMINDER ════
  Page header "All amounts in 000's of ₹" or "₹ in Thousands" → divide by 10,000 to get Crores.
  Example: 1,62,10,067 in 000's = 1,621.0 Crores.

════ SCREENER.IN BANK-SPECIFIC LABEL ALIASES ════
Screener uses non-standard labels for banks. Map these exactly:
  "Financing Profit"       → Net Interest Income (NII = Sch13 − Sch15)
  "Financing Margin %"     → NIM (net interest margin %)
  "Operating Profit"       → Pre-provisioning operating profit (PPOP)
  "OPM %"                  → PPOP margin % — SKIP (ratio row)
  "Other Income"           → Other income (Sch 14)
  "Interest Earned"        → Interest earned (Sch 13)
  "Interest Expended"      → Interest expended (Sch 15)
  "Operating Expenses"     → TOTAL opex (do NOT map to any single template row for banks)
  "Net Profit"             → Profit after tax
  "Provisions"             → Credit cost / Provisions and Contingencies
  "Tax %"                  → Tax rate % — SKIP (ratio row)
  "Provision for Taxation" → Tax expense / Tax provision
  "Income Tax"             → Tax expense / Tax provision
  "Current Tax"            → Tax expense (if no separate deferred tax line)

════ PDF TEXT TABLE READING ════
  Rows in annual report P&L appear as: Label | [Note ref 1-200, SKIP] | Value_CurrentFY | Value_PriorFY
  ⚠ FIRST numeric column = CURRENT (most recent) year; SECOND = prior year.
  ⚠ Do NOT swap columns — confirm by checking the year headers printed above the table.
  Values in brackets like (123.45) are NEGATIVE → use -123.45.
  Convert units: "₹ in Crores" → ÷1; "₹ in Lakhs" → ÷100; "₹ in Thousands / 000's" → ÷10000.
  Percentage rows (e.g., "NIM %", "OPM %") → set null.

════ SOURCE PRIORITY (annual P&L / BS sheets) ════
  For EACH (field, year) combination, use the HIGHEST-priority source that has a non-null value:

    ① ANNUAL XBRL DATA (Q4 full-year)  ← highest priority: machine-readable, audited
    ② Annual Report PDF text            ← high priority: actual filing
    ③ PDF structured table extraction   ← medium priority
    ④ Screener.in annual data           ← good for historical years where XBRL/PDF is absent
    ⑤ Yahoo Finance annual data         ← fallback

  ⭐ XBRL FIRST: If XBRL has a value for a field+year → use it, do NOT override with PDF/Screener.
  ⭐ SCREENER FOR HISTORY: For years where XBRL/PDF have no value (e.g. F2017-F2021), USE Screener.
     Screener provides all years going back to IPO — always fill from Screener when XBRL/PDF is missing.
  ⭐ NEVER leave a cell null if ANY source has the value — cascade down the priority list.

  KEY ALIASES — any source field name → template row labels (for banking templates):
    "Advances" / "Net Loans" / "Loans"       → "Advances" template row
    "Cash and Cash Equivalents"              → "Cash & balance with RBI" template row
    "Cash Equivalents" / "Cash"              → "Cash & balance with RBI" template row
    "Total Assets"                           → "Total Assets" template row
    "Other Assets"                           → "Other assets" template row (Schedule 11)
    "Net Interest Income"                    → "Net interest income" template row
    "Deposits"                               → "Deposits" template row
    "Net Income" / "Profit For Period"       → "Profit after tax" template row
    "Pretax Income" / "Income Before Tax"    → "Profit before tax" template row
    "Tax Provision" / "Income Tax Expense"   → "Less: Tax" template row
    "Net Loans" / "Loans And Advances"       → "Advances" template row
    "Fixed Assets" / "Net PPE" / "Properties"→ "Fixed assets" template row

  PER-QUARTER XBRL / quarterly PDFs → use ONLY for quarterly/operating-metrics sheets, never for annual totals.

════ GENERAL LABEL ALIASES (non-banking) ════
  "Net Profit" = "Profit After Tax" = "PAT" = "Net Income"
  "Sales" = "Revenue from Operations" = "Net Sales" = "Total Revenue"
  "Finance Costs" = "Interest Expense" = "Interest Paid"
  "Depreciation" = "D&A" = "Depreciation and Amortisation"
  "EPS" = "Earnings per share" = "Basic EPS"

════ EPS / EARNINGS PER SHARE — UNIT RULE ════
  EPS is expressed in ₹ per share (NOT in Crores). DO NOT convert EPS to Crores.
  Report EPS exactly as stated in the annual report or XBRL data (e.g., 14.35 means ₹14.35 per share).
  ⚠ If the annual report page says "units = 000's of ₹", that unit conversion applies ONLY to
    rupee-amount line items (interest, profit, etc.), NOT to EPS which is always per-share.
  Example: "Basic EPS: ₹ 14.35" → output 14.35 (not 0.0014, not 14350).

════ OPERATING METRICS SHEET — SPECIAL RULES ════
When the sheet name contains "Operating" or "Metrics" or "KPI":

  1. DATA SOURCES (in priority order for each year):
     ① XBRL structured data  — use "Advances", "Deposits", "Total Assets" for all available years
     ② Screener.in data      — use for historical years (F2017–F2022) where XBRL/PDF is absent
     ③ Investor presentation PDFs — primary source for product mix, NIM, GNPA %, disbursements
     ④ Annual Report FINANCIAL HIGHLIGHTS section — key for F2020–F2024 historical data
        ⭐ IMPORTANT: The context includes "ANNUAL REPORT FINANCIAL HIGHLIGHTS" sections from
           MULTIPLE annual reports (e.g. F2025 AR has F2021–F2025 data; F2024 AR has F2020–F2024).
           Look for tables/rows labelled "Financial Highlights", "At a Glance", "Five Year Summary",
           "Key Performance Indicators" — these contain year-by-year rows for Gross Advances,
           Total Deposits, NIM, GNPA%, NNPA%, ROA, ROE, CASA ratio, EPS, etc.
           Use this data to fill historical years (F2020, F2021, F2022, F2023, F2024) where
           presentation data is unavailable.
     ⑤ Quarterly result PDFs — for asset quality (GNPA/NNPA amounts and %)
     ⚠ Many presentation slides are IMAGE-BASED charts — you will NOT find tabular numbers for
       those rows. It is OK to output null for product-mix breakdowns if no text data is present.

  2. GROSS ADVANCES / TOTAL DEPOSITS / TOTAL ASSETS — ALL YEARS:
     These three are available in structured data for ALL years. Use them:
     - XBRL "Advances"         → "Gross Advances" (or "Total Advances", "Loan Book")
     - XBRL "Deposits"         → "Total Deposits" template row
     - XBRL "Total Assets"     → "Total Assets" template row
     - Yahoo Finance "Net Loans"        → "Gross Advances"  ← USE THIS for historical years
     - Yahoo Finance "Total Deposits"   → "Total Deposits"  ← USE THIS for historical years
     - Yahoo Finance "Total Assets"     → "Total Assets"    ← USE THIS for historical years
     - Screener "Advances"     → "Gross Advances" for years XBRL/yfinance don't cover
     ⭐ DO NOT leave Gross Advances / Total Deposits / Total Assets null for any year if
        structured data (XBRL, yfinance, Screener) has those values — map them directly.

  3. ASSET QUALITY — ALL YEARS:
     - XBRL "Gross NPA" / "Net NPA"   → GNPA Cr / NNPA Cr rows
     - XBRL "Gross NPA %" / "Net NPA %" (or compute from amount / advances)
     - Presentation: "GNPA: X.X%" or "Gross NPA ratio: X.X%"
     - Annual Report Highlights: GNPA% / NNPA% columns in five-year table

  4. KEY RATIOS — ALL YEARS (from Annual Report Highlights and yfinance):
     - NIM / Net Interest Margin %  — in highlights table, look for "NIM" or "Net Interest Margin" column
     - CASA Ratio %                 — in highlights table, look for "CASA" or "CASA ratio" column
     - ROA / Return on Assets %     — in highlights table, look for "ROA" or "Return on Assets" column
     - ROE / Return on Equity %     — in highlights table, look for "ROE" or "Return on Equity" column
     - Cost to Income Ratio %       — in highlights table
     - Capital Adequacy / CRAR %    — in highlights table

  5. PRODUCT MIX (SBL, MFI, VF, HF, MSE, NBFC, Gold, Others):
     These are ONLY available from presentation text or quarterly result PDFs.
     If found in text as a table or list like:
       "SBL  40%  |  MFI  12%  |  HF  13%  ..."  → extract the Crore amounts (not %)
       "Micro Finance  |  5,423  |  4,876" → extract 5423 (current) and 4876 (prior year)
     If the data is in charts/images (no text), output null.

  6. DISBURSEMENTS: look for "Disbursements for Q4FY25: Rs. X,XXX Cr" in presentation text.

  7. YEAR ASSIGNMENT:
     - Presentation data: year = slide header (e.g., "Q4FY26" = F2026, "Q4FY25" = F2025).
     - Annual Report Financial Highlights: columns labelled "FY25", "FY24", "FY23" → map directly.
     - "As at March 31, 2025" = F2025; "As at March 31, 2024" = F2024, etc.
     For annual operating metrics rows, use Q4 / full-year figures (not single-quarter).

════ NET INTEREST INCOME / SPREADS SHEET — SPECIAL RULES ════
  "Net interest income" = Interest earned − Interest expended
  "Yield on advances"   = from investor presentation NIM slide or annual report
  "Cost of deposits"    = from investor presentation or annual report
  "NIM" / "Net interest margin" = from investor presentation; expressed as % → output as %, not Crores

If no match exists set the value to null.  Values must be in INR Crores (except EPS ₹/share, NIM/yields in %, count-based rows as integers).
════ QUARTERLY SHEETS — SPECIAL RULES ════
When the sheet name contains "Quarterly" or "Quarter":

  1. COLUMN KEYS are NQF format: "4QF2025" = Q4 (Jan–Mar 2025), "1QF2025" = Q1 (Apr–Jun 2024), etc.
     - 1QF = Apr–Jun  (Q1 of Indian FY)
     - 2QF = Jul–Sep  (Q2)
     - 3QF = Oct–Dec  (Q3)
     - 4QF = Jan–Mar  (Q4)

  2. PRIMARY DATA SOURCE: "SCREENER.IN QUARTERLY DATA" block (if present) → use DIRECTLY.
     Screener already uses the same NQF key format — just map values to matching template labels.

  3. DO NOT sum up quarterly figures to create annual totals. Each cell must be a SINGLE QUARTER value.

  4. LABEL MAPPING for Screener Quarterly (bank labels):
     "Revenue"          → template "Revenue" or "Interest Earned" or "Total Income"
     "Interest"         → template "Interest" or "Interest Expended"
     "Expenses"         → template "Expenses" or "Total Expenses" or "Operating Expenses"
     "Financing Profit" → template "Financing Profit" or "NII" or "PPOP"
     "Financing Margin %"→ template "Financing Margin %" or "NIM %"
     "Other Income"     → template "Other Income"
     "Depreciation"     → template "Depreciation"
     "Profit before tax"→ template "Profit before tax" or "PBT"
     "Tax %"            → template "Tax %" (output as number, e.g. 27.0 for 27%)
     "Net Profit"       → template "Net Profit" or "PAT"
     "EPS in Rs"        → template "EPS in Rs" or "EPS" (in ₹, NOT crores)
     "Gross NPA %"      → template "Gross NPA %" or "GNPA %"
     "Net NPA %"        → template "Net NPA %" or "NNPA %"

  5. For quarters not covered by Screener: check XBRL quarterly data.

  6. If the template has columns like "Mar-25", "Dec-24" — these map to "4QF2025", "3QF2025" etc.
     The NQF keys in the data are already aligned after period normalisation.

Return ONLY valid JSON — no markdown fences, no explanations.

Output format (annual sheets):
{
  "Template Label 1": {"F2025": 1234.5, "F2024": 1100.0},
  "Template Label 2": {"F2025": null,   "F2024": 567.8},
  ...
}

Output format (quarterly sheets):
{
  "Revenue":           {"4QF2026": 1836.0, "3QF2026": 1692.0, "1QF2026": 1649.0},
  "Net Profit":        {"4QF2026":  213.0, "3QF2026":   90.0, "1QF2026":   42.0},
  "EPS in Rs":         {"4QF2026":    1.86, "3QF2026":   0.79},
  ...
}
"""


_HARVEST_SYSTEM = """\
You are a financial data parser for Indian company reports.

Your ONLY job: read the text below and output every financial number you find as structured JSON.

Output format (strict):
{
  "label": {"F2025": value_as_number, "F2024": value_as_number, ...},
  ...
}

Rules:
- Label = the exact financial item name as it appears in the text (clean it up minimally)
- Year keys: use "F" + 4-digit year for annual (F2025, F2024, …) OR "NQF YYYY" for quarterly (4QF2025, 1QF2025, …)
  Detect quarterly data from column headers like "Mar-25", "Jun-24", "Sep 2024" → use NQF format
  Detect annual data from headers like "FY25", "2024-25", "Year ended March 2025" → use F-year format
- Values: always plain numbers in INR Crores (convert if unit header says Lakhs or Thousands)
- Negative values: plain negative numbers (e.g. -123.4)
- If you see a table: extract every row, every column
- If you see key: value pairs in text: extract them
- Include ALL metrics found: P&L, Balance Sheet, ratios, KPIs, operating metrics — everything
- Do NOT map to any template labels. Output raw labels as they appear.
- Respond with ONLY valid JSON, no explanation, no markdown fences.
"""


def llm_harvest_from_text(
    text: str,
    years: list[str],
    client,
    model: str,
) -> dict[str, dict[str, float]]:
    """
    Phase 1 of two-phase extraction: ask LLM to convert raw financial text into
    a structured {raw_label: {year: value}} dict without any template involvement.

    Returns empty dict on failure (non-fatal — falls through to existing pipeline).
    """
    if not text or not text.strip():
        return {}

    # Cap text to keep prompt manageable — harvest prompt is smaller so we can fit more text
    MAX = 80_000
    sample = text[:MAX] + ("\n[...truncated...]" if len(text) > MAX else "")
    yr_hint = f"Years of interest (use these F-format keys): {years}" if years else ""

    prompt = f"""{yr_hint}

Extract ALL financial data from the text below into JSON {{label: {{year: value}}}}.

{sample}
"""
    try:
        resp = client.chat.completions.create(
            model=model,
            messages=[
                {"role": "system", "content": _HARVEST_SYSTEM},
                {"role": "user",   "content": prompt},
            ],
            temperature=0,
            max_tokens=4096,
            response_format={"type": "json_object"},
            timeout=45,
        )
        raw = resp.choices[0].message.content or "{}"
        import json as _j
        parsed = _j.loads(raw)

        # Normalise: ensure year keys are F-format, values are float
        result: dict[str, dict[str, float]] = {}
        from utils.helpers import normalise_year, parse_number
        for lbl, yr_map in parsed.items():
            if not isinstance(yr_map, dict):
                continue
            row: dict[str, float] = {}
            for yr_raw, val in yr_map.items():
                yr = normalise_year(str(yr_raw))
                if yr is None:
                    continue
                v = parse_number(str(val)) if not isinstance(val, (int, float)) else float(val)
                if v is not None:
                    row[yr] = v
            if row:
                result[str(lbl).strip()] = row
        logger.info(f"llm_harvest_from_text: extracted {len(result)} labels from {len(text):,} chars")
        return result
    except Exception as e:
        logger.debug(f"llm_harvest_from_text failed: {e}")
        return {}


def map_harvested_to_template(
    template_labels: list[str],
    years: list[str],
    harvested: dict[str, dict[str, float]],
    threshold: float = 0.72,
) -> dict[str, dict[str, Optional[float]]]:
    """
    Phase 2 of two-phase extraction: for each template label, find the best
    matching key in the harvested dict using financial semantic similarity.

    This is pure Python — no LLM call needed.
    Returns {template_label: {year: value}} for successfully matched rows.
    """
    from utils.helpers import clean_label
    result: dict[str, dict[str, Optional[float]]] = {}
    if not harvested:
        return result

    # Pre-clean harvested labels for faster comparison
    harvested_clean = {clean_label(k).lower(): (k, v) for k, v in harvested.items()}

    for tmpl_lbl in template_labels:
        ct = clean_label(tmpl_lbl).lower()
        if not ct:
            continue

        best_score = 0.0
        best_raw_lbl: Optional[str] = None
        best_yr_map: dict[str, float] = {}

        for h_clean, (h_orig, h_yr_map) in harvested_clean.items():
            # Exact / substring
            if ct == h_clean:
                score = 1.0
            elif ct in h_clean or h_clean in ct:
                score = 0.88
            else:
                score = _financial_sim(ct, h_clean)

            if score > best_score:
                best_score = score
                best_raw_lbl = h_orig
                best_yr_map = h_yr_map

        if best_score >= threshold and best_raw_lbl is not None:
            yr_vals: dict[str, Optional[float]] = {}
            for yr in years:
                yr_vals[yr] = best_yr_map.get(yr)
            result[tmpl_lbl] = yr_vals

    logger.info(f"map_harvested_to_template: matched {len(result)}/{len(template_labels)} labels "
                f"at threshold {threshold}")
    return result


def llm_map_fields(
    sheet_name: str,
    template_labels: list[str],
    years: list[str],
    screener_raw: dict,
    yfinance_raw: dict,
    pdf_data: dict,
    client: OpenAI,
    model: str = "deepseek-chat",
    pdf_text: str = "",
    exchange_raw: dict | None = None,
    xbrl_data: dict | None = None,
    xbrl_qtrly: dict | None = None,
    mc_raw: dict | None = None,
    ir_structured_raw: dict | None = None,
) -> dict[str, dict[str, Optional[float]]]:
    """
    xbrl_data  : For annual sheets → Q4-only XBRL (full-year annual figures, SAFE to use).
                 For quarterly/operating sheets → combined all-quarters XBRL.
    xbrl_qtrly : Per-quarter XBRL data (Q1/Q2/Q3 filings) — only passed for quarterly sheets.
    """

    sl = sheet_name.lower()
    if "p&l" in sl or "profit" in sl or "income" in sl:
        s_secs  = ["P&L"]
        yf_secs = ["Income Statement"]
        p_secs  = ["P&L"]
        exc_secs = ["P&L"]
    elif "balance" in sl or "bs" in sl:
        s_secs  = ["Balance Sheet"]
        yf_secs = ["Balance Sheet"]
        p_secs  = ["Balance Sheet"]
        exc_secs = ["Balance Sheet"]
    elif "cash" in sl or "cf" in sl:
        s_secs  = ["Cash Flow"]
        yf_secs = ["Cash Flow"]
        p_secs  = ["Cash Flow"]
        exc_secs = ["Cash Flow"]   # Tickertape covers CF
    else:
        s_secs   = list(screener_raw.keys())
        yf_secs  = list(yfinance_raw.keys())
        p_secs   = list(pdf_data.keys())
        exc_secs = list((exchange_raw or {}).keys())

    lines = [
        f"EXCEL SHEET: '{sheet_name}'",
        f"YEARS NEEDED: {years}",
        "",
        "TEMPLATE LABELS (fill EVERY one of these — use null only if truly not found):",
        json.dumps(template_labels, indent=2),
        "",
        "=== STRUCTURED DATA (pre-parsed, INR Crores) ===",
        "",
    ]

    _is_annual_sheet = any(
        k in sl for k in ("p&l", "profit", "income", "balance", "bs", "cash", "cf")
    )

    # ── XBRL data ──────────────────────────────────────────────────────────────
    # For ANNUAL sheets: xbrl_data = Q4-only annual figures → label as ANNUAL, high priority
    # For quarterly/operating sheets: xbrl_data = combined all-quarters → label accordingly
    if xbrl_data:
        filtered_xbrl = {
            lbl: {yr: v for yr, v in yr_map.items() if yr in years}
            for lbl, yr_map in xbrl_data.items()
        }
        filtered_xbrl = {k: v for k, v in filtered_xbrl.items() if v}
        if filtered_xbrl:
            if _is_annual_sheet:
                lines.append(
                    "ANNUAL XBRL DATA (full-year figures from Q4 filings — PRIORITY ①, highest):\n"
                    "⭐ Use these values directly for any field+year they cover. Do NOT override with PDF or Screener.\n"
                    "⭐ For years/fields NOT covered by XBRL, fall through to PDF → Screener → yfinance."
                )
            else:
                lines.append("XBRL STRUCTURED DATA (use for KPIs, ratios, asset quality, operating metrics):")
            lines.append(json.dumps(filtered_xbrl, indent=2))
            lines.append("")

    # For quarterly/presentation sheets: also include per-quarter XBRL separately
    if xbrl_qtrly and not _is_annual_sheet:
        filtered_qtr = {
            lbl: {yr: v for yr, v in yr_map.items() if yr in years}
            for lbl, yr_map in xbrl_qtrly.items()
        }
        filtered_qtr = {k: v for k, v in filtered_qtr.items() if v}
        if filtered_qtr:
            lines.append("PER-QUARTER XBRL DATA (individual quarter figures — use for quarterly trend rows):")
            lines.append(json.dumps(filtered_qtr, indent=2))
            lines.append("")

    # ── Moneycontrol data (mc_raw) — historical 5-year P&L / BS ───────────────────
    # Passed separately from screener_raw so the LLM can distinguish the source and priority.
    if mc_raw:
        for sec in s_secs:
            if sec not in mc_raw:
                continue
            d = {
                lbl: {yr: v for yr, v in yr_vals.items() if yr in years}
                for lbl, yr_vals in mc_raw[sec].get("data", {}).items()
            }
            d = {k: v for k, v in d.items() if v}
            if d:
                lines.append(
                    f"MONEYCONTROL DATA ({sec}) — historical 5-year data; "
                    "use for years Screener / XBRL doesn't cover:"
                )
                lines.append(json.dumps(d, indent=2))
                lines.append("")

    # ── Company IR website structured tables (ir_structured_raw) ──────────────
    # Tables scraped directly from the company's own investor-relations page.
    if ir_structured_raw:
        for sec, sec_data in ir_structured_raw.items():
            if not isinstance(sec_data, dict):
                continue
            d = {
                lbl: {yr: v for yr, v in yr_vals.items() if yr in years}
                for lbl, yr_vals in sec_data.get("data", {}).items()
            }
            d = {k: v for k, v in d.items() if v}
            if d:
                lines.append(f"COMPANY IR WEBSITE TABLES ({sec}) — company's own investor-relations site:")
                lines.append(json.dumps(d, indent=2))
                lines.append("")

    # PDF data — highest priority structured source
    for sec in p_secs:
        if sec not in pdf_data or not pdf_data[sec]:
            continue
        filtered = {yr: pdf_data[sec][yr] for yr in years if yr in pdf_data[sec]}
        if filtered:
            lines.append(f"PDF TABLE EXTRACTION ({sec}):")
            lines.append(json.dumps(filtered, indent=2))
            lines.append("")

    # Exchange data (NSE/BSE quarterly XBRL aggregated) — higher quality than Screener
    if exchange_raw:
        for sec in exc_secs:
            exc_block = _fmt_exchange_for_llm(exchange_raw, years, sec)
            if exc_block:
                lines.append(f"NSE/BSE EXCHANGE DATA ({sec}) — quarterly XBRL aggregated, high confidence:")
                lines.append(exc_block)
                lines.append("")

    # ── Screener Quarterly section — ALWAYS render first for quarterly sheets ──
    # Screener quarterly uses NQF-keyed columns (1QF2025, 4QF2024 …) which
    # directly map to the template quarterly column headers after normalise_year().
    _is_qtr_sheet_llm = any(k in sl for k in ("quarter", "qtrl"))
    if _is_qtr_sheet_llm and "Quarterly" in screener_raw:
        qtr_d = screener_raw["Quarterly"]
        d = {
            lbl: {yr: v for yr, v in yr_vals.items() if yr in years}
            for lbl, yr_vals in qtr_d.get("data", {}).items()
        }
        d = {k: v for k, v in d.items() if v}
        if d:
            lines.append(
                "SCREENER.IN QUARTERLY DATA (per-quarter granularity — PRIMARY SOURCE for quarterly columns):\n"
                f"Column keys are in NQF format matching the template (e.g. 1QF2025 = Q1 of FY2025, Apr-Jun 2024)."
            )
            lines.append(json.dumps(d, indent=2))
            lines.append("")

    # Also render exchange quarterly data prominently for quarterly sheets
    if _is_qtr_sheet_llm and exchange_raw:
        for sec in (exc_secs or list(exchange_raw.keys())):
            exc_block = _fmt_exchange_for_llm(exchange_raw, years, sec)
            if exc_block:
                lines.append(f"NSE/BSE EXCHANGE QUARTERLY DATA ({sec}):")
                lines.append(exc_block)
                lines.append("")

    # Screener
    for sec in s_secs:
        if sec not in screener_raw:
            continue
        # Already rendered Quarterly above — skip duplicate
        if _is_qtr_sheet_llm and sec == "Quarterly":
            continue
        d = {
            lbl: {yr: v for yr, v in yr_vals.items() if yr in years}
            for lbl, yr_vals in screener_raw[sec].get("data", {}).items()
        }
        d = {k: v for k, v in d.items() if v}
        if d:
            lines.append(f"SCREENER.IN ({sec}):")
            lines.append(json.dumps(d, indent=2))
            lines.append("")

    # yfinance — filter to ~30 most relevant metrics to keep prompt compact
    # (yfinance returns 100+ internal metrics most of which are useless for the LLM)
    _YF_KEEP = re.compile(
        r"interest|revenue|income|profit|loss|tax|eps|earning|deposit|advance"
        r"|borrow|debt|asset|equity|capital|reserve|cash|npa|provision|expense"
        r"|total|net|gross|operating|dividend|return|ebit|fund|fee|liabilit"
        r"|loan|invest|fixed|tangible|payable|receivable|depreci|amort|employ",
        re.I,
    )
    for sec in yf_secs:
        if sec not in yfinance_raw:
            continue
        d = {}
        for yr in years:
            if yr not in yfinance_raw[sec]:
                continue
            yr_metrics = {
                k: v for k, v in yfinance_raw[sec][yr].items()
                if _YF_KEEP.search(str(k))
            }
            if yr_metrics:
                d[yr] = yr_metrics
        if d:
            lines.append(f"YAHOO FINANCE ({sec}):")
            lines.append(json.dumps(d, indent=2))
            lines.append("")

    # Raw PDF text — highest priority, lets LLM find anything the parser missed
    if pdf_text:
        if "p&l" in sl or "profit" in sl or "income" in sl:
            kw = ["profit", "loss", "income", "revenue", "interest earned",
                  "schedule 13", "schedule 14", "schedule 15", "schedule 16", "schedule 17",
                  "operating", "provision", "tax", "earnings per share",
                  "payments to and provisions for employees", "other operating expenses"]
        elif "balance" in sl or "bs" in sl:
            kw = ["balance sheet", "assets", "liabilities", "capital", "reserves",
                  "deposits", "borrowings", "investments", "advances",
                  "schedule 6", "schedule 7", "schedule 8", "schedule 9", "schedule 10", "schedule 11",
                  "cash and balances", "fixed assets", "other assets"]
        elif "cash" in sl or "cf" in sl:
            kw = ["cash flow", "operating activities", "investing", "financing",
                  "net cash", "profit before tax"]
        elif "operat" in sl or "metric" in sl or "kpi" in sl:
            # Operating metrics sheet: loan book breakdown, disbursements, KPIs
            # Include financial-highlight keywords so annual report KPI pages are kept
            kw = ["gross advances", "disbursement", "loan book", "portfolio",
                  "micro finance", "vehicle finance", "housing finance", "small business",
                  "msme", "mse", "mfi", "sbl", "nbfc", "gold",
                  "branch", "customer", "employee", "casa",
                  "yield on advances", "cost of funds", "nim", "net interest margin",
                  "gnpa", "nnpa", "npa", "provision coverage",
                  "key highlight", "snapshot", "financial performance",
                  "total deposits", "pat", "quarter",
                  # Annual report financial-highlights page keywords
                  "financial highlight", "five year", "at a glance",
                  "key performance", "annual summary", "return on asset",
                  "return on equity", "cost to income", "capital adequacy",
                  "crar", "roa", "roe"]
        else:
            # For Spreads / NIM / other presentation sheets
            kw = ["nim", "yield", "cost of fund", "spread", "margin",
                  "net interest", "interest income", "interest expense",
                  "casa", "deposit", "advance", "loan"]

        relevant_text = pdf_text
        if kw:
            kept = []
            for chunk in pdf_text.split("\n=== Page "):
                cl = chunk.lower()
                if any(k in cl for k in kw):
                    kept.append(chunk)
            if kept:
                relevant_text = "\n=== Page ".join(kept)

        # Generous limit — DeepSeek has 64k context; system prompt + structured data ~10k
        # leaving ~50k for PDF text. More text = better extraction of Schedule 16 etc.
        MAX_PDF_TEXT = 45_000
        if len(relevant_text) > MAX_PDF_TEXT:
            relevant_text = relevant_text[:MAX_PDF_TEXT] + "\n[...truncated for length...]"

        lines.append("=== RAW PDF TEXT (financial statement pages — highest priority) ===")
        lines.append("Use this to find ANY field not covered by structured data above.")
        lines.append("Row format: Label | [note ref 1-200, skip] | value_recent_year | value_prior_year")
        lines.append("Note: many investor presentation tables are image-based; extract what text is present.")
        lines.append("")
        lines.append(relevant_text)
        lines.append("")

    # ── Dynamic synonym hints for this sheet's template labels ────────────────
    # Build a compact mapping hint: for each template label find all synonym
    # cluster members and tell the LLM "if you see X in source data → map to Y".
    # This is much cheaper than running sentence-transformers and covers all
    # Indian financial terminology variations reliably.
    _hint_lines: list[str] = []
    for tmpl_lbl in template_labels:
        ct = tmpl_lbl.lower().strip()
        ci = _FIN_SYN_LOOKUP.get(ct)
        if ci is None:
            # Try partial match against synonym cluster terms
            for _ci, _cluster in enumerate(_FIN_SYNONYMS):
                if any(ct in t or t in ct for t in _cluster):
                    ci = _ci
                    break
        if ci is not None:
            synonyms = [t for t in _FIN_SYNONYMS[ci] if t.lower() != ct]
            if synonyms:
                _hint_lines.append(f"  {tmpl_lbl!r} ← also called: {', '.join(repr(s) for s in synonyms[:6])}")
    if _hint_lines:
        lines.append("")
        lines.append("=== FIELD SYNONYM HINTS FOR THIS SHEET ===")
        lines.append("If you see any of these alternative names in source data, map them to the template label shown:")
        lines.extend(_hint_lines)
        lines.append("")

    user_prompt = "\n".join(lines)

    last_error: str = "unknown error"
    for attempt in range(3):
        try:
            resp = client.chat.completions.create(
                model=model,
                messages=[
                    {"role": "system", "content": _SYSTEM_PROMPT},
                    {"role": "user",   "content": user_prompt},
                ],
                temperature=0.0,
                max_tokens=8192,
            )
            raw = resp.choices[0].message.content.strip()
            raw = re.sub(r"^```[a-z]*\n?", "", raw, flags=re.M)
            raw = re.sub(r"\n?```$", "", raw, flags=re.M).strip()

            # If response is truncated mid-JSON, try to recover by appending "}}"
            if raw and not raw.endswith("}"):
                for suffix in ["}", "}}", "}}}"]:
                    try:
                        json.loads(raw + suffix)
                        raw = raw + suffix
                        logger.warning(f"llm_map_fields: recovered truncated JSON with suffix {suffix!r}")
                        break
                    except json.JSONDecodeError:
                        pass

            parsed = json.loads(raw)
            result: dict[str, dict[str, Optional[float]]] = {}
            for k, v in parsed.items():
                if isinstance(v, dict):
                    result[k] = {yr: (float(val) if val is not None else None) for yr, val in v.items()}
                else:
                    result[k] = {yr: None for yr in years}
            return result

        except json.JSONDecodeError as e:
            last_error = f"JSON parse error: {e} | Raw response started: {raw[:300]!r}"
            logger.error(f"llm_map_fields attempt {attempt+1}/3: {last_error}")
            if attempt < 2:
                time.sleep(1)
        except Exception as e:
            last_error = f"{type(e).__name__}: {e}"
            logger.error(f"llm_map_fields attempt {attempt+1}/3: {last_error}")
            if attempt < 2:
                time.sleep(2)

    # All retries exhausted — raise so caller can surface the real error
    raise RuntimeError(f"LLM failed after 3 attempts: {last_error}")


# ═══════════════════════════════════════════════════════════════════════════════
# Excel writer
# ═══════════════════════════════════════════════════════════════════════════════

def _jaccard(a: str, b: str) -> float:
    sa, sb = set(a.split()), set(b.split())
    return len(sa & sb) / len(sa | sb) if (sa and sb) else 0.0


# ── Financial-domain semantic synonym groups ───────────────────────────────────
# Each tuple = a cluster of terms that mean the same thing.
# When matching source labels to template labels, any two terms in the same
# cluster get a high similarity score (0.92) regardless of word overlap.
_FIN_SYNONYMS: list[tuple[str, ...]] = [
    # Income statement — top line
    # "Revenue" (Screener quarterly) = total income for a bank
    ("net sales", "net revenue", "revenue from operations", "total revenue",
     "total income from operations", "sales", "turnover", "operating revenue",
     "revenue", "total income", "total revenues", "gross income",
     "net total income", "income from operations"),
    ("other income", "non-operating income", "other operating income",
     "other revenues", "miscellaneous income", "non interest income",
     "non-interest income"),

    # Interest / banking income
    ("interest earned", "interest income", "total interest income",
     "interest and similar income", "income from advances",
     "interest and discount", "interest on advances"),
    # "Interest" alone in Screener quarterly P&L = interest expended (expense side)
    ("interest expended", "interest paid", "interest expense",
     "interest and finance charges", "finance costs", "interest"),
    ("net interest income", "nii", "net interest margin income",
     "net interest earned"),

    # Profit lines — include BSE quarterly result format variations
    ("profit before tax", "pbt", "profit before taxation",
     "earnings before tax", "ebt", "pre-tax profit",
     "profit / (loss) from ordinary activities before tax",
     "profit/(loss) from ordinary activities before tax",
     "profit / (loss) before tax", "profit/(loss) before tax",
     "profit before provisions and tax"),
    ("profit after tax", "pat", "net income", "net profit",
     "profit for the year", "profit for the period",
     "net earnings", "bottom line profit",
     "net profit / (loss) for the period",
     "net profit / (loss) from ordinary activities after tax",
     "net profit for the period", "net profit for the year",
     "profit / (loss) for the period"),
    # Pre-provision operating profit — "Financing Profit" in Screener quarterly
    ("operating profit before provisions", "ppop",
     "pre-provision operating profit", "preprovision operating profit",
     "operating profit before provisions & contingencies",
     "operating profit before provisions and contingencies",
     "financing profit", "net interest income after opex",
     "profit before provisions and contingencies"),
    ("profit before depreciation and tax", "pbdt",
     "profit before depreciation interest and tax", "pbdit"),
    ("ebitda", "earnings before interest tax depreciation amortisation",
     "operating ebitda"),
    ("operating profit", "pbit", "ebit",
     "earnings before interest and tax", "profit before interest and tax"),

    # Cost lines
    ("employee costs", "staff costs", "personnel expenses",
     "employee benefit expense", "salaries wages", "manpower costs"),
    ("depreciation", "depreciation and amortisation", "d&a",
     "amortisation", "depreciation amortisation impairment"),
    ("provisions and contingencies", "provisions", "loan loss provisions",
     "credit costs", "provision for bad debts", "expected credit loss",
     "provision for npa",
     "provisions (other than tax) and contingencies",
     "provision for non performing advances"),
    # Tax expense — MUST be its own cluster so "less: tax expenses" doesn't
    # accidentally match "total expenses" via the generic expenses cluster.
    ("tax expense", "tax expenses", "less: tax expenses",
     "income tax expense", "provision for taxation",
     "provision for income tax", "current tax", "deferred tax",
     "tax on profits", "income tax", "provision for tax"),
    ("operating expenses", "opex", "total operating expenses",
     "other operating expenses", "other expenses", "total expenditure",
     "expenses", "total expenses"),

    # Balance sheet — assets
    ("gross advances", "net loans", "loan book", "loans and advances",
     "total advances", "gross loans", "total loan book",
     "advances to customers", "advances"),
    ("net advances", "net loan book", "loans net of provisions"),
    ("total assets", "total balance sheet", "balance sheet size",
     "total assets employed"),
    ("fixed assets", "net fixed assets", "property plant equipment",
     "ppe net", "tangible fixed assets", "net block"),
    ("cash and cash equivalents", "cash and bank balances",
     "cash equivalents", "cash and short term investments",
     "cash balances", "cash & balance with rbi",
     "cash and balances with rbi"),
    ("investments", "total investments", "investment book",
     "securities portfolio", "treasury investments"),
    ("goodwill", "goodwill and intangibles", "intangible assets"),

    # Balance sheet — liabilities
    ("total deposits", "customer deposits", "deposit base",
     "deposits from customers", "total customer deposits", "deposits"),
    ("borrowings", "total borrowings", "debt", "total debt",
     "loans borrowed", "market borrowings"),
    ("shareholders funds", "shareholders equity", "net worth",
     "total equity", "book value of equity", "stockholders equity",
     "reserves & surplus", "reserves and surplus", "networth"),
    # Tier 1 — absolute capital amount (INR Crs, large numbers like 1100+)
    ("tier 1 capital", "core capital", "tier i capital",
     "tier-1 capital", "tier 1 capital amount",
     "common equity tier 1", "cet1 capital"),
    # Tier 1 — ratio / percentage (stored as decimal 0–1, e.g. 0.18 = 18%)
    # NOTE: do NOT include "tier 1 capital" here — that belongs to the absolute cluster above.
    ("tier - 1", "tier 1 ratio", "tier-1 ratio",
     "tier i capital ratio", "tier 1 car", "tier 1 crar",
     "tier - 1 capital ratio", "tier 1 capital adequacy ratio"),
    ("tier 2 capital", "tier ii capital", "tier - 2", "tier 2"),
    ("total capital", "total regulatory capital", "tier 1 and tier 2 capital"),

    # Ratios — banking
    ("net interest margin", "nim", "nims", "net interest margin %",
     "financing margin", "financing margin %", "spread on assets",
     "net interest margin ratio", "nim %", "nim ratio",
     "net interest margin (nim)", "nim - annualised"),
    ("casa ratio", "casa %", "current account savings account ratio",
     "low cost deposit ratio", "casa"),
    ("return on assets", "roa", "return on average assets", "roaa"),
    ("return on equity", "roe", "return on average equity", "roae",
     "return on net worth", "ronw", "roe %"),
    ("capital adequacy ratio", "car", "crar",
     "capital to risk weighted assets ratio",
     "total capital adequacy ratio", "basel iii capital adequacy ratio",
     "capital adequacy ratio (crar)", "pillar 3"),
    # Gross NPA — ratio forms (% labels)
    ("gross npa ratio", "gnpa ratio", "gnpa %", "gross npa %",
     "gross non-performing assets ratio",
     "gross non-performing loans ratio", "gross npl ratio",
     "% of gross npa to gross advances",
     "% gross npa to gross advances",
     "gross npa to gross advances",
     "gross npa as % of gross advances"),
    # Net NPA — ratio forms
    ("net npa ratio", "nnpa ratio", "nnpa %", "net npa %",
     "net non-performing assets ratio", "net npl ratio",
     "% of net npa to gross advances",
     "% net npa to net advances",
     "net npa to net advances",
     "net npa as % of net advances"),
    # Gross NPA — absolute amount
    ("gross npa", "gross non-performing assets", "gnpa",
     "gross non-performing loans", "gross npl", "absolute gnpa",
     "gross npa amount"),
    # Net NPA — absolute amount
    ("net npa", "net non-performing assets", "nnpa",
     "net non-performing loans", "net npl", "absolute nnpa"),
    ("provision coverage ratio", "pcr", "provision coverage %",
     "pcr %", "provision coverage"),
    ("credit cost", "credit cost %", "loan loss ratio",
     "provision to advances ratio"),
    ("cost to income ratio", "cost income ratio", "efficiency ratio",
     "opex ratio"),
    # EPS — "EPS in Rs" is Screener's label for diluted EPS
    ("earnings per share", "eps", "basic eps", "diluted eps",
     "profit per share", "eps in rs", "eps (in rs)",
     "diluted earnings per share", "basic earnings per share"),
    ("book value per share", "bvps", "nav per share",
     "net asset value per share"),
    ("dividend per share", "dps", "dividend paid per share"),
    ("net worth per share", "net book value per share"),

    # ── Cash Flow synonyms ──────────────────────────────────────────────────
    ("cash from operations", "net cash from operating activities",
     "cash flow from operations", "operating cash flow",
     "net cash generated from operating activities",
     "cash generated from operations", "cash flows from operations"),
    ("cash from investing", "net cash from investing activities",
     "cash used in investing activities", "investing cash flow",
     "net cash used in investing activities",
     "cash flows from investing activities"),
    ("cash from financing", "net cash from financing activities",
     "cash used in financing activities", "financing cash flow",
     "net cash used in financing activities",
     "cash flows from financing activities"),
    ("capital expenditure", "capex", "purchase of fixed assets",
     "additions to property plant and equipment", "purchase of ppe",
     "acquisition of fixed assets", "purchases of property plant and equipment",
     "purchase of property plant and equipment",
     "payment for purchase of fixed assets"),
    ("free cash flow", "fcf", "operating cash flow less capex"),
    ("net increase in cash", "net change in cash",
     "increase decrease in cash", "change in cash and cash equivalents",
     "net increase decrease in cash and cash equivalents"),

    # ── Balance Sheet synonyms ──────────────────────────────────────────────
    ("total borrowings", "borrowings", "total debt",
     "debt", "long term borrowings", "short term borrowings",
     "total liabilities borrowings", "total financial liabilities"),
    ("cash and cash equivalents", "cash and bank balances",
     "cash balance", "cash at bank", "cash and balances with rbi",
     "cash & balances with rbi", "cash and balances with reserve bank of india"),
    ("trade payables", "accounts payable", "creditors",
     "trade creditors", "sundry creditors", "trade and other payables"),
    ("trade receivables", "accounts receivable", "debtors",
     "trade debtors", "sundry debtors", "trade and other receivables"),
    ("reserves and surplus", "reserves", "retained earnings",
     "other equity", "shareholders equity excluding share capital",
     "equity reserves", "surplus in profit and loss account"),
    ("share capital", "paid up capital", "paid-up equity share capital",
     "equity share capital"),
    ("total equity", "shareholders equity", "net worth",
     "total stockholders equity"),
    ("investments", "total investments", "investment portfolio"),
    ("fixed assets", "property plant and equipment", "ppe",
     "net block", "net fixed assets", "tangible assets"),

    # ── Quarterly / Banking P&L specific ────────────────────────────────────
    ("total expenses", "total expenditure", "total operating expenditure",
     "expenses", "total costs", "operating expenditure"),
    # Note: "financing margin %" / "nim" are already covered by the NIM ratio cluster above.
    # Duplicate removed to prevent lookup-dict collision (last write wins in _FIN_SYN_LOOKUP).
]

# Build reverse-lookup: normalised phrase → cluster_id
_FIN_SYN_LOOKUP: dict[str, int] = {}
for _ci, _cluster in enumerate(_FIN_SYNONYMS):
    for _term in _cluster:
        _FIN_SYN_LOOKUP[_term.lower().strip()] = _ci


def _financial_sim(a: str, b: str) -> float:
    """
    Compute semantic similarity between two financial label strings.

    Returns a float in [0, 1]:
      0.95  — both labels belong to the same financial synonym cluster
      0.0–0.8 — Jaccard word-overlap score as fallback

    This is much faster and more domain-accurate than sentence-transformers
    for Indian financial statement terminology.
    """
    if not a or not b:
        return 0.0
    a_key = a.lower().strip()
    b_key = b.lower().strip()
    if a_key == b_key:
        return 1.0
    # Check synonym clusters
    ca = _FIN_SYN_LOOKUP.get(a_key)
    cb = _FIN_SYN_LOOKUP.get(b_key)
    if ca is not None and ca == cb:
        return 0.95
    # Partial cluster match — the key must appear as a WHOLE-WORD prefix/stem of a cluster term,
    # or a cluster term must be a WHOLE-WORD prefix/stem of the key.
    # This prevents single generic words like "interest" or "income" from matching long phrases
    # like "net interest margin" just because they're substrings.
    # Generic single-word financial terms that are too ambiguous to use for
    # partial cluster matching — they only match via exact _FIN_SYN_LOOKUP.
    _GENERIC_FIN_WORDS = frozenset({
        "expenses", "income", "profit", "loss", "assets", "liabilities",
        "capital", "interest", "tax", "revenue", "sales", "earnings",
        "provisions", "advances", "deposits", "investments", "borrowings",
        "reserves", "equity", "debt", "cost", "costs", "margin",
    })

    def _meaningful_partial(key: str, term: str) -> bool:
        """True if key and term meaningfully overlap for cluster membership purposes.

        Rules:
        1. Pluralisation / minor suffix: key or term is a prefix of the other,
           the shorter must be >= 4 chars AND the leftover suffix must be short
           (<=3 chars, e.g. 's', 'ed', '%') — prevents "tier 1 capital" from
           matching "tier 1 capital ratio" (suffix 7 chars).
        2. Stem inside longer key: cluster term is a whole-word substring of key,
           but term must be >= 60% of key length so "expenses" (8) does not match
           "less: tax expenses" (18) — ratio 0.44 < 0.6.
        3. Key inside term: key must be >= 60% of term length (symmetric guard).
        Generic single standalone words are blocked from partial matching.
        """
        if len(key) < 4:
            return False
        # Block generic ambiguous single-word keys from partial cluster matching
        if key in _GENERIC_FIN_WORDS:
            return False
        import re as _re2
        # Rule 1: prefix/suffix match with short leftover
        if term.startswith(key) and len(key) >= 4:
            return len(term) - len(key) <= 3  # only short suffixes: 's', '%', 'ed'
        if key.startswith(term) and len(term) >= 4:
            return len(key) - len(term) <= 3
        # Rule 2: cluster term (shorter) as whole-word substring of key (longer)
        if term in key and len(term) >= 4 and len(term) / len(key) >= 0.60:
            return bool(_re2.search(r"\b" + _re2.escape(term) + r"\b", key))
        # Rule 3: key (shorter) as whole-word substring of term (longer)
        if key in term and len(key) >= 5 and len(key) / len(term) >= 0.60:
            return bool(_re2.search(r"\b" + _re2.escape(key) + r"\b", term))
        return False

    for _ci, _cluster in enumerate(_FIN_SYNONYMS):
        _a_match = any(_meaningful_partial(a_key, t) for t in _cluster)
        _b_match = any(_meaningful_partial(b_key, t) for t in _cluster)
        if _a_match and _b_match:
            return 0.88
    # Fallback: Jaccard word-overlap + SequenceMatcher + n-gram cosine
    j = _jaccard(a_key, b_key)
    s = _seq_sim(a_key, b_key)
    n = _ngram_cos(a_key, b_key)
    return max(j, s, n)


# ── AI-like label similarity helpers ──────────────────────────────────────────

import difflib as _difflib

def _seq_sim(a: str, b: str) -> float:
    """SequenceMatcher character-level similarity — catches abbreviations like
    'gnpa' ↔ 'gross npa ratio', 'nim' ↔ 'net interest margin'."""
    return _difflib.SequenceMatcher(None, a, b).ratio()


def _ngram_cos(a: str, b: str, n: int = 3) -> float:
    """
    Character n-gram cosine similarity using numpy.
    Works well for financial abbreviations and partial matches:
      'net npa ratio' ↔ 'nnpa %'  → shares 'npa' trigram
      'cost of funds' ↔ 'cost of deposits'  → shares 'cost of' bigrams
    """
    try:
        import numpy as _np
        def _ngrams(s):
            s = s.replace(" ", "_")  # keep word boundaries
            return {s[i:i+n] for i in range(len(s) - n + 1)} if len(s) >= n else {s}
        ag = _ngrams(a)
        bg = _ngrams(b)
        all_g = list(ag | bg)
        va = _np.array([1.0 if g in ag else 0.0 for g in all_g])
        vb = _np.array([1.0 if g in bg else 0.0 for g in all_g])
        denom = (_np.linalg.norm(va) * _np.linalg.norm(vb))
        return float(_np.dot(va, vb) / denom) if denom else 0.0
    except Exception:
        return 0.0


def build_output_excel(
    company_name: str,
    screener_raw: dict,
    exchange_raw: dict,
    yfinance_raw: dict,
    tofler_raw: dict,
    ir_raw: dict,
    pdf_data: dict,
    xbrl_annual: dict,
    xbrl_qtrly: dict,
    llm_harvest: dict | None = None,
) -> bytes:
    """
    Build a clean 5-sheet Excel from ALL extracted data — no mapping required.
    Every label found in every source is printed as-is.

    Sheets:
      1. Annual P&L
      2. Annual Balance Sheet
      3. Annual Cash Flow
      4. Operating Metrics
      5. Quarterly

    Data priority (first value found wins per label+year):
      PDF tables > XBRL > Screener > Moneycontrol > IR tables > yfinance > LLM harvest
    """
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  numbers as xl_numbers)
    from openpyxl.utils import get_column_letter
    import re as _re

    # ── Colour palette (matches Indian banking analyst convention) ─────────────
    _HDR_BG   = "1F3864"   # dark navy
    _HDR_FG   = "FFFFFF"   # white
    _SEC_BG   = "D6E4F7"   # light blue  — section/sub-header rows
    _ALT_BG   = "F2F7FF"   # very light blue — alternate data rows
    _SRC_BG   = "FFF2CC"   # light yellow — source-tag row
    _TITLE_FG = "1F3864"
    _NUM_FMT  = '#,##0.00;(#,##0.00);"-"'
    _PCT_FMT  = '0.00%;(0.00%);"-"'

    def _hdr_font(bold=True, size=10, color=_HDR_FG):
        return Font(name="Calibri", bold=bold, size=size, color=color)

    def _fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def _thin_border():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    # ── Merge all sources into {section: {label: {year: value}}} ───────────────
    # Priority order — earlier entries win (do NOT overwrite existing values)
    _ANNUAL_QRE = _re.compile(r"^F\d{4}$")
    _QTRLY_RE   = _re.compile(r"^\dQF\d{4}$")

    def _is_pct_label(lbl: str) -> bool:
        """Heuristic: label looks like a ratio/percentage field."""
        _PCT_KW = _re.compile(
            r"%|ratio|margin|return|roa\b|roe\b|nim\b|crar\b|gnpa|nnpa|casa|yield|cost\s+of|coverage|spread",
            _re.I,
        )
        return bool(_PCT_KW.search(lbl))

    # ── Junk-label patterns (PDF schedule sub-totals, OCR artifacts, etc.) ──────
    _JUNK_RE = _re.compile(
        r"^Total\s*[-–]\s*\([A-Z]{1,3}\)\s*$"            # Total - (A), Total - (B) …
        r"|^[A-Z]\s+Balance\s+in\b"                       # "X Balance in Surplus…"
        r"|^March\s+March\b"                               # duplicate header row
        r"|^otal\s+"                                       # OCR drop of "T" in "Total"
        r"|^ess:\s+"                                       # OCR drop of "L" in "Less"
        r"|\bREPORTS\s*$"                                  # "…REPORTS" — leaked page hdr
        r"|\bSTATUTORY\s*$"                                # "…STATUTORY"
        r"|\bSTATEMENTS\s*$"                               # "…STATEMENTS"
        r"|^No\.\s+of\s+valid\s+votes\b"                  # AGM voting row
        r"|^In\s+favour\s*$"                               # AGM voting row
        r"|^Against\s*$"                                   # AGM voting row
        r"|^Percentage\s*\(%\)\s+(in\s+favour|against)\b" # AGM %
        # Share capital schedule sub-items (contain share-count numbers, not INR Crores)
        r"|equity\s+shares\s+of\s+[₹\d]"                 # "…Equity Shares of ₹ each"
        r"|\bprevious\s+year\s*[:]\s*[\d,]"              # "(Previous year: 1,70,00,…)"
        # OCR-broken numbered list items from PDF schedules — "a) Opening …"
        r"|^\s*[a-e]\)\s+"                                 # a) b) c) d) e) list items
        r"|^\s*\([a-e]\)\s+"                               # (a) (b) (c) list items
        r"|^\s*[ivxlIVXL]+\)\s+"                          # i) ii) iii) roman-numeral items
        # Long garbled OCR sentences (> 80 chars with spaces are schedule footnotes)
        r"|^.{80,}\s+crore\b"                             # "…crore" footnote sentence
        r"|^the\s+bank\s+had\b",                          # narrative sentence leaked in
        _re.I,
    )
    # Separate all-caps check — only filter long page-header words (≥7 chars).
    # Short ALL-CAPS abbreviations like "NII", "ROE", "GNPA", "CRAR" are valid metric names.
    _ALLCAPS_RE = _re.compile(r"^[A-Z]{7,}(?:[\s\-][A-Z]{2,})*\s*$")

    def _is_junk_label(lbl: str) -> bool:
        """Return True for meaningless schedule sub-total / OCR artefact labels."""
        lbl = lbl.strip()
        if not lbl or len(lbl) < 3:
            return True
        if _JUNK_RE.search(lbl):
            return True
        # All-caps words like "SCHEDULE", "ANNEXURE", "OTAL" (page headers)
        if _ALLCAPS_RE.match(lbl):
            return True
        return False

    def _clean_label(lbl: str) -> str:
        """Light cosmetic cleanup of a label string."""
        # Remove stray back-ticks used for ₹ in some PDFs
        lbl = lbl.replace("`", "₹")
        # Collapse multiple spaces / newlines
        lbl = _re.sub(r"\s{2,}", " ", lbl).strip()
        # Remove trailing schedule refs like "(Refer Note 3.2)" at end
        lbl = _re.sub(r"\s*\(Refer\s+Note[^)]*\)\s*$", "", lbl, flags=_re.I).strip()
        # Remove page-header suffixes leaked in by PDF parser
        lbl = _re.sub(r"\s+(REPORTS|STATUTORY|STATEMENTS)\s*$", "", lbl, flags=_re.I).strip()
        return lbl

    def _merge_sources(section_keys: list[str], year_filter=None) -> dict[str, dict[str, float]]:
        """
        Merge multiple source dicts for the given section names.
        year_filter: callable(year_str) → bool, or None for no filter.
        Returns {label: {year: value}}.
        """
        merged: dict[str, dict[str, float]] = {}

        def _absorb(src_dict: dict) -> None:
            for sec, sd in src_dict.items():
                if not isinstance(sd, dict):
                    continue
                match = any(sk.lower() in sec.lower() or sec.lower() in sk.lower()
                            for sk in section_keys)
                if not match:
                    continue
                for raw_lbl, yr_map in sd.get("data", {}).items():
                    if not isinstance(yr_map, dict):
                        continue
                    lbl = _clean_label(raw_lbl)
                    if _is_junk_label(lbl):
                        continue
                    row = merged.setdefault(lbl, {})
                    for yr, val in yr_map.items():
                        if year_filter and not year_filter(yr):
                            continue
                        if yr not in row and val is not None:
                            row[yr] = val

        def _absorb_yf(sec_name: str) -> None:
            """yfinance has {year: {label: val}} structure."""
            for sec, yr_dict in yfinance_raw.items():
                if sec_name.lower() not in sec.lower():
                    continue
                if not isinstance(yr_dict, dict):
                    continue
                for yr, lbl_map in yr_dict.items():
                    if year_filter and not year_filter(yr):
                        continue
                    if not isinstance(lbl_map, dict):
                        continue
                    for raw_lbl, val in lbl_map.items():
                        if val is None:
                            continue
                        lbl = _clean_label(raw_lbl)
                        if _is_junk_label(lbl):
                            continue
                        row = merged.setdefault(lbl, {})
                        if yr not in row:
                            row[yr] = val

        def _absorb_pdf(sec_name: str) -> None:
            """pdf_data has {section: {year: {label: val}}} structure."""
            for sec, yr_dict in pdf_data.items():
                if sec_name.lower() not in sec.lower():
                    continue
                if not isinstance(yr_dict, dict):
                    continue
                for yr, lbl_map in yr_dict.items():
                    if year_filter and not year_filter(yr):
                        continue
                    if not isinstance(lbl_map, dict):
                        continue
                    for raw_lbl, val in lbl_map.items():
                        if val is None:
                            continue
                        lbl = _clean_label(raw_lbl)
                        if _is_junk_label(lbl):
                            continue
                        row = merged.setdefault(lbl, {})
                        if yr not in row:
                            row[yr] = val

        def _absorb_xbrl(xbrl: dict) -> None:
            """XBRL has {label: {year: val}} — already flat."""
            for raw_lbl, yr_map in (xbrl or {}).items():
                lbl = _clean_label(raw_lbl)
                if _is_junk_label(lbl):
                    continue
                row = merged.setdefault(lbl, {})
                for yr, val in yr_map.items():
                    if year_filter and not year_filter(yr):
                        continue
                    if yr not in row and val is not None:
                        row[yr] = val

        # PDF tables first (highest fidelity for annual statements)
        for sk in section_keys:
            _absorb_pdf(sk)

        # For Operating Metrics / Quarterly sheets: also absorb PDF sections whose
        # names don't match the standard three (P&L / Balance Sheet / Cash Flow).
        # These "other" sections come from KPI pages, Asset Quality tables, etc.
        _STD_PDF_SECS = {"p&l", "balance sheet", "cash flow",
                         "profit", "income statement", "financial position"}
        _is_op_or_qtr_sheet = any(
            sk.lower() in ("operating metrics", "key metrics", "kpi",
                           "asset quality", "ratios", "quarterly")
            for sk in section_keys
        )
        if _is_op_or_qtr_sheet:
            for sec, yr_dict in pdf_data.items():
                if any(s in sec.lower() for s in _STD_PDF_SECS):
                    continue   # skip standard statements
                if not isinstance(yr_dict, dict):
                    continue
                for yr, lbl_map in yr_dict.items():
                    if year_filter and not year_filter(yr):
                        continue
                    if not isinstance(lbl_map, dict):
                        continue
                    for raw_lbl, val in lbl_map.items():
                        if val is None:
                            continue
                        lbl = _clean_label(raw_lbl)
                        if _is_junk_label(lbl):
                            continue
                        row = merged.setdefault(lbl, {})
                        if yr not in row:
                            row[yr] = val

        # XBRL next
        _absorb_xbrl(xbrl_annual)
        # Then structured web sources
        for src in [screener_raw, tofler_raw, ir_raw, exchange_raw]:
            _absorb(src)
        # yfinance — useful for ratios
        for sk in section_keys:
            _absorb_yf(sk)
        # LLM harvest as last resort
        if llm_harvest:
            for raw_lbl, yr_map in llm_harvest.items():
                lbl = _clean_label(raw_lbl)
                if _is_junk_label(lbl):
                    continue
                row = merged.setdefault(lbl, {})
                for yr, val in yr_map.items():
                    if year_filter and not year_filter(yr):
                        continue
                    if yr not in row and val is not None:
                        row[yr] = val

        # Drop rows with no values at all, and strip any surviving junk labels
        return {
            lbl: yv
            for lbl, yv in merged.items()
            if yv and not _is_junk_label(lbl)
        }

    # ── Write one sheet ────────────────────────────────────────────────────────
    def _write_sheet(ws, title: str, data: dict[str, dict[str, float]],
                     years: list[str], has_uom: bool = False) -> None:
        """Write a clean financial table to worksheet ws."""
        if not data or not years:
            ws["A1"] = title
            ws["A2"] = "No data extracted for this sheet."
            return

        # ── Row 1: Title ──────────────────────────────────────────────────────
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=len(years) + (2 if has_uom else 1))
        c = ws.cell(1, 1, title)
        c.font      = Font(name="Calibri", bold=True, size=13, color=_TITLE_FG)
        c.fill      = _fill("FFFFFF")
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 22

        # ── Row 2: Units note ─────────────────────────────────────────────────
        ws.cell(2, 1, "INR Crores (unless noted)").font = Font(
            name="Calibri", italic=True, size=9, color="808080"
        )

        # ── Row 3: Header ─────────────────────────────────────────────────────
        hdr_row = 3
        col_offset = 2 if has_uom else 1
        ws.cell(hdr_row, 1, "Particulars").font      = _hdr_font()
        ws.cell(hdr_row, 1).fill                     = _fill(_HDR_BG)
        ws.cell(hdr_row, 1).alignment                = Alignment(horizontal="left",  vertical="center", wrap_text=True)
        if has_uom:
            ws.cell(hdr_row, 2, "UoM").font          = _hdr_font()
            ws.cell(hdr_row, 2).fill                 = _fill(_HDR_BG)
            ws.cell(hdr_row, 2).alignment            = Alignment(horizontal="center", vertical="center")
        for ci, yr in enumerate(years, start=col_offset + 1):
            c = ws.cell(hdr_row, ci, yr)
            c.font      = _hdr_font()
            c.fill      = _fill(_HDR_BG)
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[hdr_row].height = 18

        # ── Data rows ─────────────────────────────────────────────────────────
        for ri, (lbl, yr_map) in enumerate(data.items(), start=hdr_row + 1):
            is_alt = (ri - hdr_row) % 2 == 0
            bg = _ALT_BG if is_alt else "FFFFFF"

            c = ws.cell(ri, 1, lbl)
            c.font      = Font(name="Calibri", size=9)
            c.fill      = _fill(bg)
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

            if has_uom:
                uom_cell = ws.cell(ri, 2, "%" if _is_pct_label(lbl) else "INR Crs.")
                uom_cell.font      = Font(name="Calibri", size=8, color="808080")
                uom_cell.fill      = _fill(bg)
                uom_cell.alignment = Alignment(horizontal="center", vertical="center")

            for ci, yr in enumerate(years, start=col_offset + 1):
                val = yr_map.get(yr)
                cell = ws.cell(ri, ci, val)
                cell.fill      = _fill(bg)
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.font      = Font(name="Calibri", size=9,
                                      color="0000FF" if val is not None else "BBBBBB")
                if val is not None:
                    cell.number_format = _PCT_FMT if _is_pct_label(lbl) else _NUM_FMT
                else:
                    cell.value = "-"
                    cell.font  = Font(name="Calibri", size=9, color="BBBBBB")

        # ── Column widths ─────────────────────────────────────────────────────
        ws.column_dimensions["A"].width = 42
        if has_uom:
            ws.column_dimensions["B"].width = 10
        for ci in range(col_offset + 1, col_offset + 1 + len(years)):
            ws.column_dimensions[get_column_letter(ci)].width = 12

        # Freeze pane: keep label column and header row fixed while scrolling
        ws.freeze_panes = ws.cell(hdr_row + 1, col_offset + 1)

    # ── Assemble workbook ──────────────────────────────────────────────────────
    wb = Workbook()
    wb.remove(wb.active)   # remove default blank sheet

    # Helper: get sorted annual years from a merged dataset
    def _annual_years(data: dict) -> list[str]:
        yrs = {yr for yv in data.values() for yr in yv if _ANNUAL_QRE.match(str(yr))}
        return sorted(yrs, key=lambda k: k or "", reverse=True)

    def _qtrly_years(data: dict) -> list[str]:
        yrs = {yr for yv in data.values() for yr in yv if _QTRLY_RE.match(str(yr))}
        # Sort quarterly: F-year descending, then Q descending within year
        def _qsort(q):
            m = _re.match(r"(\d)QF(\d{4})", q)
            return (int(m.group(2)), int(m.group(1))) if m else (0, 0)
        return sorted(yrs, key=_qsort, reverse=True)

    # 1. Annual P&L
    pl_data = _merge_sources(["P&L", "Income Statement", "Profit"], _ANNUAL_QRE.match)
    ws1 = wb.create_sheet("Annual P&L")
    _write_sheet(ws1, f"{company_name} — Annual P&L", pl_data, _annual_years(pl_data))

    # 2. Annual Balance Sheet
    bs_data = _merge_sources(["Balance Sheet", "Financial Position"], _ANNUAL_QRE.match)
    ws2 = wb.create_sheet("Annual Balance Sheet")
    _write_sheet(ws2, f"{company_name} — Annual Balance Sheet", bs_data, _annual_years(bs_data))

    # 3. Annual Cash Flow
    cf_data = _merge_sources(["Cash Flow"], _ANNUAL_QRE.match)
    ws3 = wb.create_sheet("Annual Cash Flow")
    _write_sheet(ws3, f"{company_name} — Annual Cash Flow", cf_data, _annual_years(cf_data))

    # 4. Operating Metrics (KPIs, ratios, segment data — everything not in P&L/BS/CF)
    op_data = _merge_sources(
        ["Operating Metrics", "Key Metrics", "KPI", "Asset Quality",
         "Highlights", "Ratios", "Loan Book", "Advances", "Deposits"],
    )
    # Also pull ALL XBRL operating/ratio/asset-quality fields into Operating Metrics.
    # This includes both absolute amounts (Gross NPA in Crores) AND ratio fields (GNPA%).
    _XBRL_OP_LABELS = {
        # Asset Quality — absolute amounts
        "Gross NPA", "Net NPA", "Pre-provisioning Operating Profit",
        # Asset Quality — ratios
        "Gross NPA %", "Net NPA %", "PCR %",
        # Profitability ratios
        "ROA", "ROE", "NIM",
        # Capital
        "CRAR", "Tier 1 Ratio", "Tier 2 Ratio",
    }
    for raw_lbl, yr_map in (xbrl_annual or {}).items():
        lbl = _clean_label(raw_lbl)
        if not _is_junk_label(lbl) and (lbl in _XBRL_OP_LABELS or _is_pct_label(lbl)):
            if lbl not in op_data:
                op_data[lbl] = {yr: v for yr, v in yr_map.items() if _ANNUAL_QRE.match(str(yr))}
    ws4 = wb.create_sheet("Operating Metrics")
    _write_sheet(ws4, f"{company_name} — Operating Metrics & KPIs",
                 op_data, _annual_years(op_data), has_uom=True)

    # 5. Quarterly (all quarterly data from exchange + XBRL + screener quarterly section)
    qt_data = _merge_sources(
        ["Quarterly", "P&L", "Income Statement", "Asset Quality", "Operating Metrics"],
        _QTRLY_RE.match,
    )
    # Also pull XBRL quarterly — accept both re-keyed quarter format (1QF2025)
    # and fallback FY format (F2025) when quarter metadata wasn't available.
    for raw_lbl, yr_map in (xbrl_qtrly or {}).items():
        lbl = _clean_label(raw_lbl)
        if _is_junk_label(lbl):
            continue
        row = qt_data.setdefault(lbl, {})
        for yr, val in yr_map.items():
            if _QTRLY_RE.match(str(yr)) and yr not in row:
                row[yr] = val
            # Fallback: if yr is in "F2025" annual format, skip (don't mix annual into quarterly)
    # Remove labels that ended up with no quarterly year data.
    # This can happen when annual P&L labels are absorbed by _merge_sources
    # (they match the "P&L" section key) but all their year keys are annual
    # (F2025 etc.) which get filtered out by _QTRLY_RE — leaving empty entries.
    qt_data = {lbl: yv for lbl, yv in qt_data.items() if yv}
    ws5 = wb.create_sheet("Quarterly")
    _write_sheet(ws5, f"{company_name} — Quarterly Results",
                 qt_data, _qtrly_years(qt_data))

    # ── Coverage summary sheet ─────────────────────────────────────────────────
    ws_cov = wb.create_sheet("Data Coverage")
    ws_cov.column_dimensions["A"].width = 30
    ws_cov.column_dimensions["B"].width = 15
    ws_cov.column_dimensions["C"].width = 40
    ws_cov.cell(1, 1, "Sheet").font          = _hdr_font(size=10, color=_HDR_FG)
    ws_cov.cell(1, 1).fill                   = _fill(_HDR_BG)
    ws_cov.cell(1, 2, "Rows").font           = _hdr_font(size=10, color=_HDR_FG)
    ws_cov.cell(1, 2).fill                   = _fill(_HDR_BG)
    ws_cov.cell(1, 3, "Years covered").font  = _hdr_font(size=10, color=_HDR_FG)
    ws_cov.cell(1, 3).fill                   = _fill(_HDR_BG)
    _cov_rows = [
        ("Annual P&L",        pl_data, _annual_years(pl_data)),
        ("Annual Balance Sheet", bs_data, _annual_years(bs_data)),
        ("Annual Cash Flow",  cf_data, _annual_years(cf_data)),
        ("Operating Metrics", op_data, _annual_years(op_data)),
        ("Quarterly",         qt_data, _qtrly_years(qt_data)),
    ]
    for ri, (nm, dat, yrs) in enumerate(_cov_rows, start=2):
        ws_cov.cell(ri, 1, nm).font   = Font(name="Calibri", size=9)
        ws_cov.cell(ri, 2, len(dat)).font  = Font(name="Calibri", size=9, bold=True)
        ws_cov.cell(ri, 3, ", ".join(yrs[:12])).font = Font(name="Calibri", size=9, color="404040")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════════
# Raw-Excel → Template mapping  (bypasses LLM — uses clean labels from raw Excel)
# ═══════════════════════════════════════════════════════════════════════════════

def read_raw_excel(excel_bytes: bytes) -> dict[str, dict[str, dict[str, float]]]:
    """
    Parse a raw-data Excel produced by build_output_excel() back into a nested dict.

    Returns:
        {
          "P&L":              {label: {year: value}, ...},
          "Balance Sheet":    {label: {year: value}, ...},
          "Cash Flow":        {label: {year: value}, ...},
          "Operating Metrics":{label: {year: value}, ...},
          "Quarterly":        {label: {year: value}, ...},
        }

    Sheet classification is by name pattern so users can rename sheets as long as
    the key words are still present.
    """
    import re as _re
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)

    # ── Sheet-type classifier ─────────────────────────────────────────────────
    _SHEET_TYPE: list[tuple[str, list[str]]] = [
        ("P&L",              ["p&l", "profit", "income statement"]),
        ("Cash Flow",        ["cash flow", "cash"]),
        ("Balance Sheet",    ["balance sheet", "balance"]),
        ("Operating Metrics",["operating metric", "kpi", "ratio", "metric"]),
        ("Quarterly",        ["quarterly", "quarter"]),
    ]

    def _classify_sheet(name: str) -> Optional[str]:
        nl = name.lower()
        for sheet_type, kws in _SHEET_TYPE:
            if any(kw in nl for kw in kws):
                return sheet_type
        return None

    # Year pattern: F2024 (annual) or 1QF2025 (quarterly)
    # Must start with F or digits+QF — never bare numbers like 9007 (deposit values)
    _YR_RE = re.compile(r"^(F\d{4}|\d+QF\d{4})$")

    result: dict[str, dict[str, dict[str, float]]] = {}

    for sh_name in wb.sheetnames:
        sheet_type = _classify_sheet(sh_name)
        if sheet_type is None:
            continue   # skip Data Coverage and unknown sheets

        ws = wb[sh_name]

        # ── Find the header row: row where most cells look like year strings ──
        hdr_row: int = 3           # default (raw Excel always writes header at row 3)
        year_cols: dict[int, str] = {}   # col_idx → year_string
        for ri in range(1, 10):
            cands: dict[int, str] = {}
            for ci in range(1, ws.max_column + 1):
                v = ws.cell(row=ri, column=ci).value
                if v and _YR_RE.match(str(v).strip()):
                    cands[ci] = str(v).strip()
            if len(cands) > len(year_cols):
                year_cols = cands
                hdr_row   = ri

        if not year_cols:
            continue   # no year columns found — skip

        # ── Detect label column: col A (1) or col B (2) ──────────────────────
        # The raw Excel puts "Particulars" in col A; if a UoM column is present
        # col B is "UoM" — labels are still in col A.  We look for which of A/B
        # has more non-empty text cells below the header row.
        def _text_count(col: int) -> int:
            n = 0
            for ri in range(hdr_row + 1, ws.max_row + 1):
                v = ws.cell(row=ri, column=col).value
                if v and str(v).strip() and not _YR_RE.match(str(v).strip()):
                    n += 1
            return n

        lbl_col = 1 if _text_count(1) >= _text_count(2) else 2

        # ── Read data rows ────────────────────────────────────────────────────
        sheet_data: dict[str, dict[str, float]] = {}
        for ri in range(hdr_row + 1, ws.max_row + 1):
            raw_lbl = ws.cell(row=ri, column=lbl_col).value
            if raw_lbl is None:
                continue
            lbl = str(raw_lbl).strip()
            if not lbl or lbl == "Particulars" or lbl == "UoM":
                continue
            # Skip purely numeric rows (stray numbers in label column)
            try:
                float(lbl.replace(",", ""))
                continue
            except ValueError:
                pass

            yr_map: dict[str, float] = {}
            for ci, yr in year_cols.items():
                v = ws.cell(row=ri, column=ci).value
                if v is None or str(v).strip() in ("-", "—", ""):
                    continue
                try:
                    yr_map[yr] = float(str(v).replace(",", "").strip())
                except (ValueError, TypeError):
                    pass

            if yr_map:
                sheet_data[lbl] = yr_map

        if sheet_data:
            # Merge if multiple sheets of same type (e.g. standalone + consolidated)
            existing = result.get(sheet_type, {})
            for lbl, yr_map in sheet_data.items():
                if lbl not in existing:
                    existing[lbl] = yr_map
                else:
                    for yr, v in yr_map.items():
                        if yr not in existing[lbl]:
                            existing[lbl][yr] = v
            result[sheet_type] = existing

    return result


def map_raw_excel_to_template(
    raw_data: dict[str, dict[str, dict[str, float]]],
    sheets_data: dict,
    threshold: float = 0.68,
) -> dict[str, dict[str, dict[str, Optional[float]]]]:
    """
    Map a raw-data Excel (from read_raw_excel) to template labels for every
    template sheet.

    Strategy per template sheet:
      1. Detect which raw sheet(s) are relevant (P&L, Balance Sheet, etc.)
      2. Pool the relevant raw labels into a single harvested dict
      3. Run map_harvested_to_template (fuzzy + financial semantic similarity)
      4. Fill any remaining template label gaps by searching ALL raw sheets
         (catches cross-sheet items like "GNPA%" which may appear in any sheet)

    Returns:
        {template_sheet_name: {template_label: {year: value_or_None}}}
    """
    # Map keyword → raw sheet types (ordered by relevance).
    # Quarterly template sheets search ALL raw data because the BSE quarterly
    # result format contains P&L + Balance Sheet + Cash Flow items all in one sheet.
    # Operating/Asset Quality sheets also need the Quarterly raw sheet because
    # NPA%, GNPA%, NIM etc. come from the Screener quarterly section.
    _KW_TO_RAW: list[tuple[list[str], list[str]]] = [
        (["p&l", "income statement"],          ["P&L"]),
        (["balance sheet", " bs"],             ["Balance Sheet"]),
        (["cash flow"],                        ["Cash Flow"]),
        (["operating", "kpi", "metric", "ratio", "spread", "asset quality"],
                                               ["Operating Metrics", "Quarterly",
                                                "P&L", "Balance Sheet"]),
        (["quarterly", "quarter"],             ["Quarterly", "P&L",
                                                "Balance Sheet", "Operating Metrics"]),
    ]

    def _pick_raw_sheets(tmpl_name: str) -> list[str]:
        nl = tmpl_name.lower()
        for kws, raw_types in _KW_TO_RAW:
            if any(kw in nl for kw in kws):
                return raw_types
        return list(raw_data.keys())   # unknown sheet → search everything

    def _pool(raw_types: list[str]) -> dict[str, dict[str, float]]:
        pooled: dict[str, dict[str, float]] = {}
        for rt in raw_types:
            for lbl, yr_map in raw_data.get(rt, {}).items():
                if lbl not in pooled:
                    pooled[lbl] = yr_map
                else:
                    for yr, v in yr_map.items():
                        if yr not in pooled[lbl]:
                            pooled[lbl][yr] = v
        return pooled

    # Full pool (all sheets) for gap filling
    _all_raw: dict[str, dict[str, float]] = _pool(list(raw_data.keys()))

    filled: dict[str, dict[str, dict[str, Optional[float]]]] = {}

    for sheet_name, sheet_data in sheets_data.items():
        labels: list[str] = sheet_data["labels"]
        years:  list[str] = sheet_data["years"]

        if not years:
            filled[sheet_name] = {}
            continue

        # ── Empty-sheet mode ─────────────────────────────────────────────────
        # If the template sheet has year headers but NO label rows (e.g. a Cash
        # Flow sheet with only "Note: tabulate from annual reports" or an asset-
        # quality sheet the analyst left blank), dump ALL matching raw data
        # directly — no fuzzy label matching needed, just use the raw label names.
        # write_to_template detects the empty label column and appends new rows.
        if not labels:
            primary_pool = _pool(_pick_raw_sheets(sheet_name))
            dump: dict[str, dict[str, Optional[float]]] = {}
            for lbl, yr_map in primary_pool.items():
                row = {yr: yr_map.get(yr) for yr in years}
                if any(v is not None for v in row.values()):
                    dump[lbl] = row
            filled[sheet_name] = dump
            continue

        # ── Primary match: relevant sheet pool ───────────────────────────────
        primary_pool = _pool(_pick_raw_sheets(sheet_name))
        _sl = sheet_name.lower()
        _is_fin   = any(k in _sl for k in ("p&l", "profit", "income", "balance", "cash"))
        _thresh   = threshold if _is_fin else max(threshold - 0.05, 0.55)
        primary   = map_harvested_to_template(labels, years, primary_pool, threshold=_thresh)

        # ── Gap fill: search ALL sheets for labels not matched in primary ────
        unmatched = [lbl for lbl in labels if lbl not in primary
                     or all(v is None for v in primary[lbl].values())]
        if unmatched:
            fallback = map_harvested_to_template(
                unmatched, years, _all_raw, threshold=max(_thresh - 0.05, 0.50)
            )
            for lbl, yr_map in fallback.items():
                if lbl not in primary:
                    primary[lbl] = yr_map
                else:
                    # Merge: fill None slots
                    for yr, v in yr_map.items():
                        if primary[lbl].get(yr) is None and v is not None:
                            primary[lbl][yr] = v

        filled[sheet_name] = primary

    return filled


def write_to_template(template_bytes: bytes, filled: dict, overwrite: bool = False) -> tuple[bytes, dict]:
    from utils.helpers import normalise_year, clean_label

    wb    = openpyxl.load_workbook(io.BytesIO(template_bytes))
    stats = {"written": 0, "skipped_filled": 0, "skipped_formula": 0, "not_found": 0, "audit": [],
             "formula_by_year": {}, "written_by_year": {}}

    for sheet_name, label_year_vals in filled.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]

        # ── Year-header: scan ALL rows 1-30, pick the row with the MOST valid year cells ──
        # This fixes Issue 3: multi-row merged headers where a stray year in row 2 caused
        # the scanner to lock onto the wrong row (the real year headers were in row 3).
        _cand_yr_rows: dict[int, dict[str, int]] = {}  # {row: {year: col}}
        for ri in range(1, 31):
            _row_yrs: dict[str, int] = {}
            for ci in range(1, ws.max_column + 1):
                v = ws.cell(row=ri, column=ci).value
                if v is None:
                    continue
                yr = normalise_year(str(v).strip())
                if yr:
                    _row_yrs[yr] = ci
            if _row_yrs:
                _cand_yr_rows[ri] = _row_yrs

        year_cols:  dict[str, int] = {}
        header_row: Optional[int]  = None
        if _cand_yr_rows:
            # Pick the row with the greatest number of distinct year values
            header_row = max(_cand_yr_rows, key=lambda r: len(_cand_yr_rows[r]))
            year_cols  = _cand_yr_rows[header_row]

        if not year_cols or header_row is None:
            continue

        # ── Label column: use stored label_col or auto-detect ──────────────────
        _sheet_meta = label_year_vals.get("__meta__", {})
        _label_col  = _sheet_meta.get("label_col") if isinstance(_sheet_meta, dict) else None
        if _label_col is None:
            # Auto-detect: count non-empty cells in col 1 vs col 2
            _c1 = sum(1 for ri in range(header_row + 1, ws.max_row + 1)
                      if ws.cell(row=ri, column=1).value not in (None, ""))
            _c2 = sum(1 for ri in range(header_row + 1, ws.max_row + 1)
                      if ws.cell(row=ri, column=2).value not in (None, ""))
            _label_col = 2 if _c2 > _c1 else 1

        label_rows: dict[str, int] = {}
        for ri in range(header_row + 1, ws.max_row + 1):
            v = ws.cell(row=ri, column=_label_col).value
            if v is None:
                continue
            c = clean_label(str(v).strip())
            if c and c not in label_rows:   # FIRST occurrence wins (avoids schedule rows overwriting main section)
                label_rows[c] = ri

        # ── Append mode: sheet had year headers but NO existing label rows ────
        # This handles: empty Cash Flow sheets, blank template sections, etc.
        # Write each raw-label row directly below the header row.
        if not label_rows:
            _next_row = header_row + 1
            for tmpl_label, yr_vals in label_year_vals.items():
                if tmpl_label == "__meta__":
                    continue
                if not any(v is not None for v in yr_vals.values()):
                    continue
                ws.cell(row=_next_row, column=_label_col).value = tmpl_label
                for yr, val in yr_vals.items():
                    if val is None:
                        continue
                    col = year_cols.get(yr)
                    if col is None:
                        continue
                    ws.cell(row=_next_row, column=col).value = round(float(val), 2)
                    stats["written"] += 1
                    stats["written_by_year"][yr] = stats["written_by_year"].get(yr, 0) + 1
                stats["audit"].append({
                    "Sheet": sheet_name, "Label": tmpl_label,
                    "Year": "—", "Value (Cr)": "(appended row)",
                })
                _next_row += 1
            continue   # skip normal label-matching loop for this sheet

        for tmpl_label, yr_vals in label_year_vals.items():
            ct = clean_label(tmpl_label)
            row = label_rows.get(ct)
            if row is None:
                # Find best matching template row using financial semantic similarity
                best_score, best_ri = 0.0, None
                for lbl, ri in label_rows.items():
                    if not ct or not lbl:
                        continue
                    # Fast exact substring check first
                    if ct in lbl or lbl in ct:
                        score = 0.85
                    else:
                        score = _financial_sim(ct, lbl)
                    if score > best_score:
                        best_score, best_ri = score, ri
                if best_score >= 0.6:
                    row = best_ri
            if row is None:
                stats["not_found"] += 1
                continue

            for yr, val in yr_vals.items():
                if val is None:
                    continue
                col = year_cols.get(yr)
                if col is None:
                    continue
                cell = ws.cell(row=row, column=col)
                if cell.data_type == "f" or (isinstance(cell.value, str) and cell.value.startswith("=")):
                    stats["skipped_formula"] += 1
                    stats["formula_by_year"][yr] = stats["formula_by_year"].get(yr, 0) + 1
                    continue
                # Treat dash placeholders as empty — Indian templates use "-" / "—" for blank cells
                _is_empty = cell.value in (None, "", 0, "-", "—", "–", "—")
                if not overwrite and not _is_empty:
                    stats["skipped_filled"] += 1
                    continue
                # Auto-scale: Excel stores % values as decimals (7% → 0.07).
                # If the cell has a % number format but our value looks like a
                # percentage (>1) rather than a decimal ratio (≤1), divide by 100.
                _write_val = float(val)
                _nf = cell.number_format or ""
                if "%" in _nf and abs(_write_val) > 1.0:
                    _write_val = _write_val / 100.0
                cell.value = round(_write_val, 6)
                stats["written"] += 1
                stats["written_by_year"][yr] = stats["written_by_year"].get(yr, 0) + 1
                stats["audit"].append({"Sheet": sheet_name, "Label": tmpl_label, "Year": yr, "Value (Cr)": round(_write_val, 6)})

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), stats


# ═══════════════════════════════════════════════════════════════════════════════
# Streamlit UI
# ═══════════════════════════════════════════════════════════════════════════════

with st.sidebar:
    st.header("⚙️ Configuration")

    use_llm = st.checkbox(
        "🤖 AI Enhance mapping (optional)",
        value=False,
        key="use_llm",
        help=(
            "Uses DeepSeek LLM to fill any label gaps the fuzzy matcher misses. "
            "Slower and requires an API key, but handles unusual label wording better."
        ),
    )
    if use_llm:
        api_key   = st.text_input("DeepSeek API Key", type="password")
        api_model = st.selectbox("LLM Model", ["deepseek-chat", "deepseek-reasoner"])
    else:
        api_key   = ""
        api_model = "deepseek-chat"

    st.markdown("---")
    st.header("📥 Pipeline Inputs")

    # ════════════════════════════════════════════════════════════════════════════
    # Quick Map expander — only for users who already have a raw data Excel
    # ════════════════════════════════════════════════════════════════════════════
    with st.expander("⚡ Re-map existing raw data Excel to a new template"):
        st.caption(
            "Already ran the pipeline before and have a `*_raw_data.xlsx`? "
            "Skip re-scraping — upload both files and map in seconds."
        )
        _col_a, _col_b = st.columns(2)
        with _col_a:
            raw_excel_file = st.file_uploader(
                "Raw Data Excel  (*_raw_data.xlsx)",
                type=["xlsx"],
                key="raw_excel_upload",
                help="The *_raw_data.xlsx produced by a previous run of this app.",
            )
        with _col_b:
            template_file_quick = st.file_uploader(
                "Your Template  (.xlsx)",
                type=["xlsx"],
                key="template_quick",
                help="Your Excel template with row labels and F2025/F2024/… column headers.",
            )
        overwrite_quick = st.checkbox("Overwrite existing cell values", value=False, key="ow_quick")
        _raw_excel_mode = raw_excel_file is not None and template_file_quick is not None
        if _raw_excel_mode:
            st.success("✅ Both files ready.")
            quick_run_btn = st.button("🔗 Map → Template", type="primary", use_container_width=True)
        else:
            quick_run_btn = False
    template_file = st.file_uploader(
        "Excel Template (.xlsx)",
        type=["xlsx"],
        key="template_full",
        help="Upload your template — the app will match labels and fill it automatically. Without a template only the raw data Excel is produced.",
    )

    nse_symbol    = st.text_input("NSE Symbol", placeholder="e.g. HDFCBANK, RELIANCE").strip().upper()
    bse_code      = st.text_input("BSE Code (optional)", placeholder="e.g. 500180").strip()

    stmt_type  = st.radio("Statement type", ["Consolidated", "Standalone"], horizontal=True)
    overwrite  = st.checkbox("Overwrite existing cell values", value=False)

    st.markdown("---")
    st.header("🌐 Company Website (optional)")
    st.caption("Auto-detected from NSE/BSE. Paste here if auto-detection fails.")
    manual_company_site = st.text_input(
        "Company website URL",
        placeholder="e.g. https://www.equitasbnk.com",
    ).strip().rstrip("/")

    # ── Source modes ───────────────────────────────────────────────────────────
    st.markdown("---")
    st.header("📄 Annual Report PDF")
    pdf_mode = st.radio(
        "PDF source",
        ["Auto-download from NSE/BSE", "Upload manually", "Skip PDF (web only)"],
        help="Auto-download fetches the official annual report for you.",
    )
    manual_pdf = None
    if pdf_mode == "Upload manually":
        manual_pdf = st.file_uploader("Upload Annual Report PDF", type=["pdf"])

    st.markdown("---")
    st.header("📊 Investor Presentation")
    st.caption("Operating metrics & asset quality — sourced from quarterly investor decks.")
    pres_mode = st.radio(
        "Presentation source",
        ["Auto-download from NSE/BSE", "Upload manually", "Skip"],
        key="pres_mode",
        help="Investor presentations contain operating metrics, NIM, yields, asset quality data.",
    )
    manual_pres_pdf = None
    if pres_mode == "Upload manually":
        manual_pres_pdf = st.file_uploader(
            "Upload Investor Presentation PDF", type=["pdf"], key="pres_upload"
        )

    st.markdown("---")
    st.header("📋 Quarterly Results PDF")
    st.caption("Asset quality, NPA, operating metrics — sourced from quarterly result filings.")
    qres_mode = st.radio(
        "Quarterly result source",
        ["Auto-fetch from NSE/BSE/Company", "Upload manually", "Skip"],
        key="qres_mode",
        help="Q4 result PDFs include full-year P&L figures. All quarters have asset quality data.",
    )
    manual_qres_pdf = None
    if qres_mode == "Upload manually":
        manual_qres_pdf = st.file_uploader(
            "Upload Quarterly Result PDF", type=["pdf"], key="qres_upload"
        )

    # ── Step A: Discover available files (runs separately, stores in session_state)
    # This lets the user browse and choose BEFORE clicking Run, so no mid-run
    # widget interactions cause restarts.
    st.markdown("---")
    st.header("🔍 Step 1 — Discover Files")
    st.caption("Fetch the list of available reports/presentations from NSE/BSE so you can choose which ones to use.")

    discover_btn = st.button(
        "🔍 Discover Available Files",
        use_container_width=True,
        disabled=not nse_symbol,
        help="Queries NSE/BSE for available annual reports, presentations, and quarterly results.",
    )

    if discover_btn and nse_symbol:
        if manual_company_site:
            from web.pdf_fetcher import _website_cache
            _website_cache[f"{nse_symbol}:{bse_code or ''}"] = manual_company_site
        with st.spinner("Fetching file lists from NSE/BSE…"):
            st.session_state["_disc_reports"]  = get_available_reports(nse_symbol, bse_code or None) if pdf_mode == "Auto-download from NSE/BSE" else []
            st.session_state["_disc_pres"]     = get_available_presentations(nse_symbol, bse_code or None) if pres_mode == "Auto-download from NSE/BSE" else []
            st.session_state["_disc_qres"]     = get_available_quarterly_results(nse_symbol, bse_code or None) if qres_mode == "Auto-fetch from NSE/BSE/Company" else []
            st.session_state["_disc_sym"]      = nse_symbol
        st.success("✅ File lists fetched — choose below, then click Run.")

    # ── Show multiselects once lists are available ─────────────────────────────
    _selected_report_urls: list[str] = []
    _selected_pres_urls:   list[str] = []

    if pdf_mode == "Auto-download from NSE/BSE" and st.session_state.get("_disc_sym") == nse_symbol:
        _disc_reps = st.session_state.get("_disc_reports", [])
        if _disc_reps:
            def _rep_lbl(r): return f"{r.get('year','?')} — {r.get('filename','?')} ({r.get('size_mb','?')} MB) [{r.get('source','')}]"
            _rep_lbl_map = {_rep_lbl(r): r["url"] for r in _disc_reps}
            _all_lbls    = list(_rep_lbl_map.keys())
            _chosen_lbls = st.multiselect(
                "Annual reports to download:",
                _all_lbls,
                default=_all_lbls[:1],
                help="Select all years you need. Each report covers 2 fiscal years.",
                key="sel_reports",
            )
            _selected_report_urls = [_rep_lbl_map[l] for l in _chosen_lbls]
            # store friendly labels for display during the run
            st.session_state["_sel_report_labels"] = {_rep_lbl_map[l]: l for l in _chosen_lbls}
            st.session_state["_sel_report_meta"]   = {r["url"]: r for r in _disc_reps}
        else:
            st.info("No annual reports found. Run Discover first.")

    if pres_mode == "Auto-download from NSE/BSE" and st.session_state.get("_disc_sym") == nse_symbol:
        _disc_pres = st.session_state.get("_disc_pres", [])
        if _disc_pres:
            def _pres_lbl(p): return f"{p.get('year','?')} — {p.get('filename','?')} [{p.get('source','')}]"
            _pres_lbl_map = {_pres_lbl(p): p["url"] for p in _disc_pres[:20]}
            _all_plbls    = list(_pres_lbl_map.keys())
            _chosen_plbls = st.multiselect(
                "Investor presentations to download:",
                _all_plbls,
                default=_all_plbls[:1],
                help="Select multiple decks to fill more historical quarters.",
                key="sel_pres",
            )
            _selected_pres_urls = [_pres_lbl_map[l] for l in _chosen_plbls]
            st.session_state["_sel_pres_labels"] = {_pres_lbl_map[l]: l for l in _chosen_plbls}
            st.session_state["_sel_pres_meta"]   = {p["url"]: p for p in _disc_pres[:20]}
        else:
            st.info("No presentations found. Run Discover first.")

    # Quarterly result selection — multiselect so all results can be used together
    _selected_qres_urls: list[str] = []
    if qres_mode == "Auto-fetch from NSE/BSE/Company" and st.session_state.get("_disc_sym") == nse_symbol:
        _disc_qres = st.session_state.get("_disc_qres", [])
        if _disc_qres:
            def _qrl(r):
                q = r.get("quarter","?"); yr = r.get("year","?"); src = r.get("source","")
                ann = " ★Annual" if r.get("is_annual") else ""
                xbrl = " [XBRL✓]" if r.get("xbrl_url") else ""
                return f"{yr} {q}{ann}{xbrl} — {r.get('filename','?')} [{src}]"
            _disc_qres_limited = _disc_qres[:20]
            _qres_lbl_map  = {_qrl(r): r["url"] for r in _disc_qres_limited}
            _qres_meta_map = {r["url"]: r for r in _disc_qres_limited}
            all_qres_lbls  = list(_qres_lbl_map.keys())
            # Default: all Q4/annual results pre-selected; others unchecked
            _q4_lbls = [_qrl(r) for r in _disc_qres_limited if r.get("is_annual")]
            _def_lbls = _q4_lbls if _q4_lbls else all_qres_lbls[:3]
            _chosen_qlbls = st.multiselect(
                "Quarterly results to use for PDF text:",
                all_qres_lbls,
                default=_def_lbls,
                help="★Annual (Q4) results contain full-year P&L figures. Select multiple to combine text from all of them — useful for extracting historical NPA, CASA, NIM data across quarters.",
                key="sel_qres",
            )
            _selected_qres_urls = [_qres_lbl_map[l] for l in _chosen_qlbls if l in _qres_lbl_map]
            st.session_state["_sel_qres_meta"] = _qres_meta_map
            if not _chosen_qlbls:
                st.caption("ℹ️ No quarterly results selected — only XBRL data will be used.")
        else:
            st.info("No quarterly results found. Run Discover first.")

    # ── Full pipeline run button ───────────────────────────────────────────────
    st.markdown("---")
    st.header("🚀 Step 2 — Run Full Pipeline")
    run_btn = st.button("🚀 Run Full Pipeline", type="primary", use_container_width=True)


# ── Main pipeline ──────────────────────────────────────────════════════════════

# ── FAST PATH: quick_run_btn from Mode A ─────────────────────────────────────
if quick_run_btn and _raw_excel_mode:
    progress = st.progress(0, text="Reading raw data Excel…")
    status   = st.empty()
    try:
        status.info("📂 Reading raw data Excel…")
        progress.progress(10, text="Parsing raw Excel…")
        raw_excel_file.seek(0)
        raw_data = read_raw_excel(raw_excel_file.read())
        total_raw_labels = sum(len(v) for v in raw_data.values())
        total_raw_pts    = sum(len(yv) for d in raw_data.values() for yv in d.values())
        st.success(
            f"✅ Raw Excel loaded — **{total_raw_labels}** labels, "
            f"**{total_raw_pts:,}** data points across "
            f"{len(raw_data)} sheet(s): {', '.join(raw_data.keys())}"
        )

        # Show what was read (expandable)
        with st.expander("📊 Raw Excel contents"):
            for sheet_type, sheet_d in raw_data.items():
                st.markdown(f"**{sheet_type}** — {len(sheet_d)} labels")
                years_in_sheet = sorted({yr for yv in sheet_d.values() for yr in yv}, reverse=True)
                rows = [{"Label": lbl, **{yr: yv.get(yr) for yr in years_in_sheet[:8]}}
                        for lbl, yv in list(sheet_d.items())[:50]]
                if rows:
                    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

        progress.progress(30, text="Reading template…")
        status.info("📋 Reading template structure…")
        template_file_quick.seek(0)
        template_bytes = template_file_quick.read()
        needs       = read_template_structure(template_bytes)
        sheets_data = needs["sheets"]
        if not sheets_data:
            st.error("No year columns found in template. Make sure columns are labelled F2025, F2024 etc.")
            st.stop()
        st.info(
            f"📋 Template: **{len(sheets_data)}** sheet(s) — "
            + ", ".join(f"**{sn}** ({len(sd['labels'])} labels, {len(sd['years'])} years)"
                        for sn, sd in sheets_data.items())
        )

        progress.progress(55, text="Matching labels…")
        status.info("🔗 Matching raw labels to template labels…")
        filled = map_raw_excel_to_template(raw_data, sheets_data, threshold=0.68)

        # Summary
        total_matched = sum(
            1 for sf in filled.values()
            for yv in sf.values() if any(v is not None for v in yv.values())
        )
        st.success(f"✅ Label matching complete — **{total_matched}** template rows filled")

        with st.expander("🔗 Mapping preview (expand to check matches)"):
            for sn, sf in filled.items():
                st.markdown(f"**{sn}**")
                rows = []
                for lbl, yv in sf.items():
                    if any(v is not None for v in yv.values()):
                        row = {"Template Label": lbl}
                        row.update({yr: (f"{v:,.2f}" if v is not None else "—")
                                    for yr, v in yv.items()})
                        rows.append(row)
                if rows:
                    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
                else:
                    st.caption(f"No matches for {sn}")

        progress.progress(85, text="Writing to template…")
        status.info("✍️ Writing values to template…")
        output_bytes, stats = write_to_template(
            template_bytes, filled, overwrite=overwrite_quick
        )
        progress.progress(100, text="Done ✓")
        status.success(
            f"🎉 Done! Wrote **{stats['written']}** values to template "
            f"(skipped {stats['skipped_filled']} already-filled, "
            f"{stats['skipped_formula']} formula cells)."
        )

        st.download_button(
            "⬇️ Download Populated Template",
            data=output_bytes,
            file_name=f"{(template_file_quick.name or 'template').replace('.xlsx','')}_filled.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary",
        )
    except Exception as _re_err:
        st.error(f"❌ Raw-Excel mapping failed: {type(_re_err).__name__}: {_re_err}")
        import traceback as _tb
        with st.expander("Error details"):
            st.code(_tb.format_exc())


# ── Full pipeline ──────────────────────────────────────────────────────────────
if run_btn:
    # ════════════════════════════════════════════════════════════════════════════
    # FULL PIPELINE — scraping + PDF + XBRL + LLM mapping
    # ════════════════════════════════════════════════════════════════════════════
    if not nse_symbol:
        st.error("Please enter an NSE symbol.")
        st.stop()
    client = None
    if use_llm:
        if not api_key:
            st.error("Please enter your DeepSeek API key (required when AI Enhance is enabled).")
            st.stop()
        # ── Quick API key check before running anything expensive ─────────────
        try:
            _test_client = OpenAI(api_key=api_key, base_url="https://api.deepseek.com")
            _test_resp   = _test_client.chat.completions.create(
                model=api_model,
                messages=[{"role": "user", "content": "Reply with the single word: OK"}],
                max_tokens=5, temperature=0,
            )
            if "ok" not in _test_resp.choices[0].message.content.strip().lower():
                st.warning("⚠️ LLM API responded but gave unexpected output — proceeding anyway.")
        except Exception as _api_err:
            st.error(f"❌ Cannot reach DeepSeek API: {type(_api_err).__name__}: {_api_err}\n\nCheck your API key and try again.")
            st.stop()
        client = OpenAI(api_key=api_key, base_url="https://api.deepseek.com")
    consolidated = stmt_type == "Consolidated"
    progress   = st.progress(0, text="Starting…")
    status     = st.empty()

    # Inject manually-provided company website into the cache so all fetchers
    # (presentations, quarterly results, annual reports) use it automatically.
    if manual_company_site:
        from web.pdf_fetcher import _website_cache
        cache_key = f"{nse_symbol}:{bse_code or ''}"
        _website_cache[cache_key] = manual_company_site
        st.toast(f"Using company website: {manual_company_site}", icon="🌐")

    try:
        # ── 1. Read template (optional) ───────────────────────────────────────
        _has_template = bool(template_file)
        template_bytes: bytes = b""
        sheets_data: dict     = {}

        if _has_template:
            status.info("📋 Step 1 / 5 — Reading template…")
            progress.progress(4, text="Reading template…")

            template_file.seek(0)
            template_bytes = template_file.read()
            needs          = read_template_structure(template_bytes)
            sheets_data    = needs["sheets"]

            if not sheets_data:
                st.error("No year columns detected. Ensure columns are labelled F2025, F2024, etc.")
                st.stop()

            all_years = sorted({yr for sd in sheets_data.values() for yr in sd["years"] if yr}, key=lambda k: k or "")
            st.success(f"✅ Template: {len(sheets_data)} sheet(s) · Years: {', '.join(all_years)}")

            with st.expander("📊 Template fields detected"):
                for sn, sd in sheets_data.items():
                    _lbl_count = len(sd["labels"])
                    _lbl_note  = "empty — will auto-fill from extracted data" if _lbl_count == 0 else f"{_lbl_count} labels"
                    st.markdown(f"**{sn}** — {_lbl_note} · {sd['years']}")
                    if sd["labels"]:
                        st.caption(", ".join(sd["labels"][:25]) + ("…" if _lbl_count > 25 else ""))
        else:
            # No template — default to last 10 fiscal years; LLM mapping is skipped.
            # The raw data Excel will be the only output.
            status.info("📋 Step 1 / 5 — No template uploaded. Will produce Raw Data Excel only.")
            progress.progress(4, text="No template — raw data mode…")
            _cur_fy = 2025   # update if needed
            all_years = [f"F{y}" for y in range(_cur_fy, _cur_fy - 10, -1)]
            st.info(
                "ℹ️ **No template uploaded** — running in Raw Data mode. "
                "All extracted data will be written to a clean 5-sheet Excel. "
                "Upload a template to also get the mapped/filled version."
            )

        # ── 2. Fetch web data (Screener + yfinance) ───────────────────────────
        status.info(f"🌐 Step 2 / 5 — Fetching web data for **{nse_symbol}**…")
        progress.progress(12, text="Fetching Screener.in…")

        screener_raw = fetch_screener_raw(nse_symbol, consolidated=consolidated)
        if not screener_raw and consolidated:
            screener_raw = fetch_screener_raw(nse_symbol, consolidated=False)

        progress.progress(18, text="Fetching NSE/BSE quarterly results…")
        exchange_raw = fetch_exchange_raw(nse_symbol, bse_code or None, all_years)

        progress.progress(22, text="Fetching yfinance…")
        yfinance_raw = fetch_yfinance_raw(nse_symbol, all_years)

        progress.progress(26, text="Fetching Moneycontrol (historical data)…")
        tofler_raw, tofler_debug_log = fetch_tofler_raw(nse_symbol)

        progress.progress(28, text="Fetching Tickertape (historical data)…")
        tickertape_raw, tickertape_debug_log = fetch_tickertape_raw(nse_symbol)

        progress.progress(30, text="Fetching StockAnalysis (historical data)…")
        stockanalysis_raw, stockanalysis_debug_log = fetch_stockanalysis_raw(nse_symbol)

        def _merge_into_screener(source_raw: dict) -> None:
            """Merge a source dict into screener_raw without overwriting existing values."""
            for sec, sd in source_raw.items():
                if sec not in screener_raw:
                    screener_raw[sec] = sd
                else:
                    existing_data = screener_raw[sec].setdefault("data", {})
                    for lbl, yr_vals in sd.get("data", {}).items():
                        existing_row = existing_data.setdefault(lbl, {})
                        for yr, v in yr_vals.items():
                            if yr not in existing_row:
                                existing_row[yr] = v
                    screener_raw[sec]["years"] = sorted(
                        (y for y in set(screener_raw[sec].get("years", [])) | set(sd.get("years", [])) if y),
                        key=lambda k: k or "",
                    )

        _merge_into_screener(tofler_raw)
        _merge_into_screener(tickertape_raw)
        _merge_into_screener(stockanalysis_raw)

        # ── Company's own IR website — financial highlights / key metrics ─────
        # Pass pre-selected filing URLs from the sidebar as URL hints — the domain
        # embedded in those filing URLs (via PDF cover-page text) helps discover the
        # company's actual website when yfinance / BSE lookup returns a stale URL.
        progress.progress(30, text="Scraping company IR website for financial tables…")
        ir_raw, ir_debug_log = fetch_company_website_financials(
            nse_symbol, bse_code or None,
            report_urls_hint=list(_selected_report_urls) if _selected_report_urls else None,
        )
        ir_page_text: str = ir_raw.pop("__ir_page_text__", "")   # text fallback (no tables)
        if ir_raw:
            for sec, sd in ir_raw.items():
                if sec not in screener_raw:
                    screener_raw[sec] = sd
                else:
                    existing_data = screener_raw[sec].setdefault("data", {})
                    for lbl, yr_vals in sd.get("data", {}).items():
                        existing_row = existing_data.setdefault(lbl, {})
                        for yr, v in yr_vals.items():
                            if yr not in existing_row:
                                existing_row[yr] = v
                    screener_raw[sec]["years"] = sorted(
                        (y for y in set(screener_raw[sec].get("years", [])) | set(sd.get("years", [])) if y),
                        key=lambda k: k or "",
                    )

        s_count      = sum(len(s["data"]) for s in screener_raw.values() if isinstance(s, dict) and "data" in s)
        exc_count    = sum(len(s["data"]) for s in exchange_raw.values() if isinstance(s, dict) and "data" in s)
        y_count      = sum(len(v) for s in yfinance_raw.values() for v in (s.values() if isinstance(s, dict) else []))
        tofler_count = sum(len(s["data"]) for s in tofler_raw.values() if isinstance(s, dict) and "data" in s)
        tt_count     = sum(len(s["data"]) for s in tickertape_raw.values() if isinstance(s, dict) and "data" in s)
        sa_count     = sum(len(s["data"]) for s in stockanalysis_raw.values() if isinstance(s, dict) and "data" in s)
        ir_count     = sum(len(s["data"]) for s in ir_raw.values()     if isinstance(s, dict) and "data" in s)
        c1, c2, c3, c4, c5, c6, c7 = st.columns(7)
        c1.metric("Screener",      s_count)
        c2.metric("NSE/BSE",       exc_count)
        c3.metric("yfinance",      y_count)
        c4.metric("Moneycontrol",  tofler_count)
        c5.metric("Tickertape",    tt_count)
        c6.metric("StockAnalysis", sa_count)
        c7.metric("Company IR",    ir_count, help="Fields from company's own IR page")


        # ── Tofler / Moneycontrol debug expander ──────────────────────────────
        with st.expander(f"🔬 Moneycontrol (Tofler) fetch log ({tofler_count} fields found)"):
            if not tofler_debug_log:
                st.caption("No fetch attempts logged.")
            else:
                for entry in tofler_debug_log:
                    step = entry.get("step", "?")
                    if step == "search":
                        if "error" in entry and entry["error"]:
                            st.error(f"❌ Search: {entry['error']}")
                        else:
                            st.success(f"✅ Found stock page: `{entry.get('stock_url')}` → SC_ID: `{entry.get('sc_id')}`")
                    elif step in ("stock_page",):
                        st.caption(f"Stock page HTTP {entry.get('status')} · {entry.get('url', '')[:80]}")
                        if entry.get("error"):
                            st.caption(f"  ⚠ {entry['error']}")
                    elif step in ("fin_base_found", "fin_base_guessed"):
                        st.caption(f"{'📍 Found' if step == 'fin_base_found' else '🔮 Guessed'} fin_base: `{entry.get('fin_base')}`")
                    elif step == "fetch":
                        status_icon = "✅" if entry.get("rows_parsed", 0) > 0 else ("⚠️" if entry.get("status") == 200 else "❌")
                        st.markdown(
                            f"{status_icon} **{entry.get('section')}** — HTTP {entry.get('status')} · "
                            f"{entry.get('tables_found', 0)} tables · {entry.get('rows_parsed', 0)} rows · "
                            f"years: `{', '.join(entry.get('years_found', [])) or 'none'}`"
                        )
                        if entry.get("raw_headers"):
                            st.caption(f"  Column headers: {entry['raw_headers']}")
                        if entry.get("error"):
                            st.caption(f"  ⚠ {entry['error']}")
                    elif entry.get("error"):
                        st.error(f"Error in step '{step}': {entry['error']}")

                if tofler_count == 0:
                    st.error(
                        "Moneycontrol returned 0 fields. Common causes:\n"
                        "1. **Autocomplete failed** — NSE symbol not recognised by Moneycontrol search.\n"
                        "2. **Stock page scraped but /financials/ link not found** — Moneycontrol page structure may have changed.\n"
                        "3. **Guessed fin_base wrong** — the fallback SC_ID-based URL didn't work.\n"
                        "4. **Tables found but 0 rows parsed** — column headers don't parse to year format (check 'Column headers' above)."
                    )
                else:
                    for sec, sd in tofler_raw.items():
                        st.markdown(f"**{sec}** — {len(sd.get('data', {}))} labels, years: `{', '.join(sd.get('years', []))}`")
                        sample = list(sd.get("data", {}).items())[:5]
                        if sample:
                            rows = [{"Label": lbl, **yv} for lbl, yv in sample]
                            st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

        # ── Tickertape / StockAnalysis debug expanders ────────────────────────
        if tt_count > 0 or tickertape_debug_log:
            with st.expander(f"📈 Tickertape ({tt_count} fields)"):
                for entry in tickertape_debug_log:
                    if entry.get("step") == "search":
                        icon = "✅" if entry.get("slug") else "❌"
                        st.caption(f"{icon} Slug resolved: `{entry.get('slug') or 'not found'}`")
                    elif entry.get("step") == "fetch":
                        icon = "✅" if entry.get("rows_parsed", 0) > 0 else ("⚠️" if entry.get("status") == 200 else "❌")
                        st.markdown(
                            f"{icon} **{entry.get('section')}** — HTTP {entry.get('status')} · "
                            f"{entry.get('rows_parsed', 0)} rows · years: `{', '.join(entry.get('years_found', []))}`"
                        )
                        if entry.get("error"):
                            st.caption(f"  ⚠ {entry['error']}")
                for sec, sd in tickertape_raw.items():
                    st.markdown(f"**{sec}** — {len(sd.get('data', {}))} labels · years: `{', '.join(sd.get('years', []))}`")
                    sample = list(sd.get("data", {}).items())[:5]
                    if sample:
                        st.dataframe(pd.DataFrame([{"Label": l, **yv} for l, yv in sample]),
                                     use_container_width=True, hide_index=True)

        if sa_count > 0:
            with st.expander(f"📊 StockAnalysis.com ({sa_count} fields)"):
                for sec, sd in stockanalysis_raw.items():
                    st.markdown(f"**{sec}** — {len(sd.get('data', {}))} labels · years: `{', '.join(sd.get('years', []))}`")
                    sample = list(sd.get("data", {}).items())[:5]
                    if sample:
                        st.dataframe(pd.DataFrame([{"Label": l, **yv} for l, yv in sample]),
                                     use_container_width=True, hide_index=True)

        # ── Company IR website debug expander ──────────────────────────────────
        with st.expander(f"🌐 Company IR website ({ir_count} fields scraped)"):

            _ir_site_url = next((e.get("url","") for e in ir_debug_log if e.get("step")=="website_lookup"), "")
            if _ir_site_url and "error" not in ir_debug_log[0]:
                st.markdown(f"**Base URL:** [{_ir_site_url}]({_ir_site_url})")
            # Playwright availability badge
            if _PLAYWRIGHT_AVAILABLE:
                st.caption("🎭 Playwright available — JS-rendered pages will be retried with headless Chromium.")
            else:
                st.caption("⚠️ Playwright not installed — JS-rendered IR pages may return 0 fields. Run: `python -m playwright install chromium --with-deps`")
            for e in ir_debug_log:
                step = e.get("step")
                if step == "ir_page":
                    icon = "✅" if e.get("rows",0) > 0 else ("⚠️" if e.get("status")==200 else "❌")
                    pw_note = f" · 🎭 {e['playwright']}" if e.get("playwright") else ""
                    st.markdown(
                        f"{icon} HTTP {e.get('status')} · score={e.get('score',0)} · "
                        f"{e.get('tables',0)} tables · {e.get('rows',0)} rows{pw_note} — "
                        f"`{e.get('url','')}`"
                    )
                    if e.get("error"):
                        st.caption(f"  ⚠ {e['error']}")
                elif step in ("playwright_crawl_fallback",):
                    pw_icon = "✅" if "success" in (e.get("playwright") or "") else "🎭"
                    st.markdown(f"{pw_icon} Playwright crawl fallback: {e.get('playwright')} — `{e.get('url','')}`")
                elif step == "discovered_link":
                    icon = "✅" if e.get("rows",0) > 0 else "🔗"
                    st.markdown(f"{icon} Discovered link · score={e.get('score',0)} · {e.get('rows',0)} rows — `{e.get('url','')}`")
                elif e.get("error"):
                    st.error(f"Step '{step}': {e['error']}")
            if ir_count == 0 and not ir_page_text:
                st.warning(
                    "No financial tables found on the company's IR pages. "
                    "The page may load data via JavaScript (common for large companies). "
                    "Check the URLs above and the score — pages with score=0 had no financial keywords."
                )
            elif ir_page_text:
                st.info(f"No parseable tables found but captured {len(ir_page_text):,} chars of page text — will be sent to LLM.")
                with st.expander("IR page text sample"):
                    st.text(ir_page_text[:2000])
            else:
                for sec, sd in ir_raw.items():
                    st.markdown(f"**{sec}** — {len(sd.get('data',{}))} labels · years: `{', '.join(sd.get('years',[]))}`")
                    sample = list(sd.get("data",{}).items())[:5]
                    if sample:
                        st.dataframe(pd.DataFrame([{"Label": l, **yv} for l, yv in sample]),
                                     use_container_width=True, hide_index=True)

        if exchange_raw:
            with st.expander("📡 NSE/BSE exchange data (quarterly → annual)"):
                for sn, sd in exchange_raw.items():
                    st.markdown(f"**{sn}**")
                    rows = [{"Label": lbl, **{yr: v for yr, v in yv.items() if yr in all_years}}
                            for lbl, yv in sd.get("data", {}).items()]
                    if rows:
                        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

        with st.expander("🔍 Raw Screener data"):
            for sn, sd in screener_raw.items():
                st.markdown(f"**{sn}**")
                rows = [{"Label": lbl, **{yr: v for yr, v in yv.items() if yr in all_years}}
                        for lbl, yv in sd.get("data", {}).items()]
                if rows:
                    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

        # ── 3. Download PDF(s) — uses pre-selected URLs from sidebar Discover step ──
        # No widgets here. Selections were made in the sidebar before Run was clicked.
        pdf_bytes: Optional[bytes] = None
        pdf_label = ""
        pdf_texts_by_year: dict[str, str] = {}

        if pdf_mode == "Upload manually" and manual_pdf:
            manual_pdf.seek(0)
            pdf_bytes = manual_pdf.read()
            pdf_label = manual_pdf.name

        elif pdf_mode == "Auto-download from NSE/BSE":
            # Use pre-selected URLs from sidebar (set via Discover + multiselect)
            _rep_urls   = _selected_report_urls   # set in sidebar
            _rep_meta   = st.session_state.get("_sel_report_meta", {})

            if not _rep_urls:
                # Fallback: nothing selected via Discover — auto-pick most recent
                status.info("📄 Step 3 / 5 — Fetching report list (no pre-selection)…")
                _fallback_reps = get_available_reports(nse_symbol, bse_code or None)
                if _fallback_reps:
                    # Filter to template years, default to most recent
                    _fallback_reps = [r for r in _fallback_reps if r.get("year") in all_years] or _fallback_reps[:1]
                    _rep_urls  = [r["url"] for r in _fallback_reps]
                    _rep_meta  = {r["url"]: r for r in _fallback_reps}

            if _rep_urls:
                status.info(f"📄 Step 3 / 5 — Downloading {len(_rep_urls)} annual report(s)…")
                progress.progress(30, text=f"Downloading {len(_rep_urls)} report(s)…")
                ar_dl_bar = st.progress(0, text="Downloading annual reports…")
                for ri, url in enumerate(_rep_urls):
                    rep = _rep_meta.get(url, {"url": url, "filename": url.split("/")[-1], "year": f"report_{ri}"})
                    ar_dl_bar.progress(
                        int((ri + 1) / len(_rep_urls) * 100),
                        text=f"Downloading {rep.get('filename','?')} ({ri+1}/{len(_rep_urls)})…",
                    )
                    rep_bytes = download_pdf(url)
                    if not rep_bytes:
                        st.warning(f"⚠️ Failed to download {rep.get('filename','?')} — skipping.")
                        continue
                    yr_label = rep.get("year") or f"report_{ri}"
                    size_mb  = len(rep_bytes) / 1_048_576
                    st.success(f"✅ Downloaded **{rep.get('filename','?')}** ({size_mb:.1f} MB)")
                    if pdf_bytes is None:
                        pdf_bytes = rep_bytes
                        pdf_label = rep.get("filename", "")
                    rep_text = extract_pdf_text_for_llm(rep_bytes, doc_type="annual", label=yr_label)
                    pdf_texts_by_year[yr_label] = rep_text
                ar_dl_bar.empty()
            else:
                st.warning("⚠️ No annual report links found. Use Discover in sidebar or upload manually.")

        else:
            status.info("📄 Step 3 / 5 — Skipping PDF (web-only mode).")
            progress.progress(35)

        # Combine all annual report texts (most recent first)
        # Guard: some year keys may be None if _infer_fy returned nothing
        if pdf_texts_by_year:
            combined_parts = []
            for yr in sorted(pdf_texts_by_year.keys(), key=lambda k: k or "", reverse=True):
                combined_parts.append(
                    f"\n\n{'='*60}\n=== ANNUAL REPORT: {yr or 'unknown'} ===\n{'='*60}\n"
                    + pdf_texts_by_year[yr]
                )
            pdf_text_combined = "\n".join(combined_parts)
        else:
            pdf_text_combined = ""

        # ── 3a-retry. Company IR website — second pass with PDF text hint ─────
        # If the first IR scrape (step 2) found 0 fields AND we now have annual
        # report PDF text, re-try using URLs extracted from the PDF cover page.
        # Annual report cover pages always mention the company website.
        _ir_retry_needed = (ir_count == 0 and pdf_text_combined)
        if _ir_retry_needed:
            progress.progress(38, text="Re-trying company IR website using PDF cover page URL…")
            _ir2_raw, _ir2_log = fetch_company_website_financials(
                nse_symbol, bse_code or None,
                pdf_text_hint=pdf_text_combined[:8_000],
                report_urls_hint=list(_selected_report_urls) if _selected_report_urls else None,
            )
            ir_debug_log.extend(_ir2_log)
            if _ir2_raw:
                _ir2_page_text = _ir2_raw.pop("__ir_page_text__", "")
                if _ir2_page_text and not ir_page_text:
                    ir_page_text = _ir2_page_text
                for sec, sd in _ir2_raw.items():
                    if sec not in screener_raw:
                        screener_raw[sec] = sd
                    else:
                        _ed = screener_raw[sec].setdefault("data", {})
                        for lbl, yr_vals in sd.get("data", {}).items():
                            _er = _ed.setdefault(lbl, {})
                            for yr, v in yr_vals.items():
                                if yr not in _er:
                                    _er[yr] = v
                        screener_raw[sec]["years"] = sorted(
                            (y for y in set(screener_raw[sec].get("years", [])) | set(sd.get("years", [])) if y),
                            key=lambda k: k or "",
                        )
                # Recompute ir_raw and ir_count with merged data
                ir_raw.update({k: v for k, v in _ir2_raw.items() if k not in ir_raw})
                ir_count = sum(len(s["data"]) for s in ir_raw.values() if isinstance(s, dict) and "data" in s)
                # Update the website cache so presentation/quarterly fetchers use the correct URL too
                _resolved_url = next(
                    (e.get("url") for e in reversed(_ir2_log) if e.get("step") == "website_lookup" and e.get("url")),
                    None,
                )
                if _resolved_url:
                    try:
                        from web.pdf_fetcher import _website_cache
                        _cache_key = f"{nse_symbol}:{bse_code or ''}"
                        _website_cache[_cache_key] = _resolved_url
                        logger.info(f"IR retry: updated website cache to {_resolved_url}")
                    except Exception:
                        pass
                if ir_count > 0:
                    st.toast(f"✅ IR website retry found {ir_count} fields via PDF cover URL", icon="🌐")

        # ── 3b. Download investor presentation(s) — uses pre-selected URLs ────
        pres_bytes: Optional[bytes] = None
        pres_label = ""
        pres_texts: list[str] = []

        if pres_mode == "Upload manually" and manual_pres_pdf:
            manual_pres_pdf.seek(0)
            pres_bytes = manual_pres_pdf.read()
            pres_label = manual_pres_pdf.name

        elif pres_mode == "Auto-download from NSE/BSE":
            _pres_urls  = _selected_pres_urls   # set in sidebar
            _pres_meta  = st.session_state.get("_sel_pres_meta", {})

            if not _pres_urls:
                # Fallback: nothing selected — auto-pick most recent presentation
                _fallback_pres = get_available_presentations(nse_symbol, bse_code or None)
                if _fallback_pres:
                    _pres_urls = [_fallback_pres[0]["url"]]
                    _pres_meta = {_fallback_pres[0]["url"]: _fallback_pres[0]}

            if _pres_urls:
                status.info(f"📊 Step 3b — Downloading {len(_pres_urls)} presentation(s)…")
                pres_dl_bar = st.progress(0, text="Downloading presentations…")
                for pi, url in enumerate(_pres_urls):
                    p = _pres_meta.get(url, {"url": url, "filename": url.split("/")[-1], "year": f"pres_{pi}"})
                    pres_dl_bar.progress(
                        int((pi + 1) / len(_pres_urls) * 100),
                        text=f"Downloading presentation {pi+1}/{len(_pres_urls)}…",
                    )
                    pb = download_pdf(url)
                    if not pb:
                        st.warning(f"⚠️ Failed to download {p.get('filename','?')} — skipping.")
                        continue
                    if pres_bytes is None:
                        pres_bytes = pb
                        pres_label = p.get("filename", "")
                    size_mb = len(pb) / 1_048_576
                    st.success(f"✅ Downloaded **{p.get('filename','?')}** ({size_mb:.1f} MB)")
                    pt = extract_pdf_text_for_llm(pb, max_chars_per_page=5000, doc_type="presentation", label=p.get("year", ""))
                    pres_texts.append(pt)
                pres_dl_bar.empty()
            else:
                st.info("ℹ️ No presentations found — presentation sheets will use annual report only.")

        # Combine all presentation texts; if uploaded manually, extract now
        if pres_bytes and not pres_texts:
            pt = extract_pdf_text_for_llm(
                pres_bytes, max_chars_per_page=5000, doc_type="presentation",
                label=pres_label,
            )
            pres_texts.append(pt)

        pres_text: str = "\n\n".join(pres_texts)
        if pres_text:
            st.metric("Presentation text chars", f"{len(pres_text):,}",
                      delta=f"{len(pres_texts)} deck(s)" if len(pres_texts) > 1 else None)
            with st.expander("📊 Investor presentation text (sample)"):
                st.text(pres_text[:3000] + ("…" if len(pres_text) > 3000 else ""))

        # ── 3c. Get / download quarterly result PDF (+ XBRL bulk fetch) ──────────
        qres_bytes: Optional[bytes] = None
        qres_text: str = ""          # populated either here (multi-download) or in step 4c (single)
        qres_label = ""
        qres_is_annual = False   # True when Q4 result chosen (has full-year figures)
        qres_xbrl_data: dict = {}     # {template_label: {fy: value}} — ALL quarters merged
        xbrl_annual_data: dict = {}   # {template_label: {fy: value}} — Q4 filings ONLY (full-year figures)
        xbrl_qtrly_data: dict = {}    # {template_label: {fy: value}} — Q1/Q2/Q3 filings (per-quarter)

        if qres_mode == "Upload manually" and manual_qres_pdf:
            manual_qres_pdf.seek(0)
            qres_bytes     = manual_qres_pdf.read()
            qres_label     = manual_qres_pdf.name
            qres_is_annual = bool(re.search(r"\bq4\b|annual|full.?year", qres_label, re.I))

        elif qres_mode == "Auto-fetch from NSE/BSE/Company":
            status.info("📋 Step 3c — Fetching quarterly results + XBRL from NSE/BSE…")
            progress.progress(34, text="Fetching quarterly result list…")

            qresults = get_available_quarterly_results(nse_symbol, bse_code or None)

            if not qresults:
                st.info("ℹ️ No quarterly result PDFs found — using other sources for asset quality.")
            else:
                st.success(f"📋 Found {len(qresults)} quarterly result(s)")

                # ── BULK XBRL fetch: download ALL XBRL files and merge ────────────
                # We split into two buckets:
                #   xbrl_annual_data : Q4 filings only  → full-year annual figures → safe for annual P&L
                #   xbrl_qtrly_data  : Q1/Q2/Q3 filings → per-quarter figures → only for quarterly sheets
                #   qres_xbrl_data   : combined (for backward compat + operating metrics)
                xbrl_results_with_xbrl = [r for r in qresults if r.get("xbrl_url")]
                if xbrl_results_with_xbrl:
                    xbrl_bar = st.progress(0, text="Fetching XBRL from all quarters…")
                    total_xbrl_pts = 0
                    for xi, qr in enumerate(xbrl_results_with_xbrl):
                        xbrl_bar.progress(
                            int((xi + 1) / len(xbrl_results_with_xbrl) * 100),
                            text=f"XBRL: {qr.get('year','?')} {qr.get('quarter','?')} ({xi+1}/{len(xbrl_results_with_xbrl)})…",
                        )
                        try:
                            xbrl_chunk = fetch_xbrl_from_url(qr["xbrl_url"])
                            if not xbrl_chunk:
                                continue
                            is_q4 = bool(qr.get("is_annual"))   # Q4 → full-year figures

                            # For Q1/Q2/Q3, build a proper quarter key like "2QF2025"
                            # so the Quarterly sheet filter (_QTRLY_RE = ^\dQF\d{4}$) matches.
                            _Q_NUM_MAP = {"Q1": 1, "Q2": 2, "Q3": 3, "Q4": 4}
                            _q_str  = str(qr.get("quarter", "")).upper().strip()
                            _q_num  = _Q_NUM_MAP.get(_q_str)
                            _fy_raw = str(qr.get("year", "")).strip()
                            if _fy_raw.startswith("F"):
                                _fy_num = _fy_raw[1:]   # "2025"
                            elif re.match(r"^\d{4}$", _fy_raw):
                                _fy_num = _fy_raw
                            else:
                                _fy_num = None

                            # Merge into all-quarters combined dict (uses original FY keys)
                            for lbl, yr_map in xbrl_chunk.items():
                                existing_all = qres_xbrl_data.setdefault(lbl, {})
                                for yr, v in yr_map.items():
                                    if yr not in existing_all:
                                        existing_all[yr] = v
                                        total_xbrl_pts += 1

                                # Split into annual vs quarterly buckets
                                if is_q4:
                                    existing_ann = xbrl_annual_data.setdefault(lbl, {})
                                    for yr, v in yr_map.items():
                                        if yr not in existing_ann:
                                            existing_ann[yr] = v
                                else:
                                    existing_qtr = xbrl_qtrly_data.setdefault(lbl, {})
                                    for yr, v in yr_map.items():
                                        # Re-key to quarter format if metadata available
                                        # e.g. "F2025" → "2QF2025" for Q2 filing
                                        if _q_num and _fy_num:
                                            store_key = f"{_q_num}QF{_fy_num}"
                                        else:
                                            store_key = yr   # fallback: keep FY key
                                        if store_key not in existing_qtr:
                                            existing_qtr[store_key] = v
                        except Exception as _xe:
                            logger.debug(f"XBRL bulk fetch error for {qr.get('filename')}: {_xe}")

                    xbrl_bar.empty()
                    if qres_xbrl_data:
                        ann_pts = sum(len(v) for v in xbrl_annual_data.values())
                        qtr_pts = sum(len(v) for v in xbrl_qtrly_data.values())
                        st.success(
                            f"✅ XBRL bulk fetch: **{total_xbrl_pts}** data points — "
                            f"**{ann_pts}** annual (Q4) + **{qtr_pts}** quarterly"
                        )
                        with st.expander("📊 XBRL bulk data (all quarters merged)"):
                            rows = [
                                {"Label": lbl, **{yr: v for yr, v in sorted(yr_map.items(), key=lambda x: x[0] or "")}}
                                for lbl, yr_map in sorted(qres_xbrl_data.items(), key=lambda x: x[0] or "")
                            ]
                            if rows:
                                st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

                # ── Download ALL selected quarterly results and concatenate text ─
                _sel_qr_urls = list(_selected_qres_urls)   # from sidebar multiselect
                _sel_qr_meta = st.session_state.get("_sel_qres_meta", {})

                if not _sel_qr_urls:
                    # Fallback: auto-pick most recent Q4 + up to 3 most recent quarters
                    _auto_q4 = [r for r in qresults if r.get("is_annual")]
                    _auto_fallback = (_auto_q4 or qresults)[:4]
                    _sel_qr_urls = [r["url"] for r in _auto_fallback]
                    _sel_qr_meta = {r["url"]: r for r in _auto_fallback}

                # Download and concatenate text from all selected quarterly PDFs
                _qres_parts: list[str] = []
                _qres_labels: list[str] = []
                _any_annual = False
                for _qurl in _sel_qr_urls:
                    _qmeta   = _sel_qr_meta.get(_qurl, {"url": _qurl, "filename": _qurl.split("/")[-1]})
                    _qbytes  = download_pdf(_qurl)
                    if _qbytes:
                        _qlabel = _qmeta.get("filename", _qurl.split("/")[-1])
                        _qann   = _qmeta.get("is_annual", False)
                        if _qann:
                            _any_annual = True
                        _qres_labels.append(_qlabel)
                        # Extract text from this quarterly PDF
                        try:
                            _qt = extract_pdf_text_for_llm(
                                _qbytes, max_chars_per_page=3000, doc_type="quarterly"
                            )[:12_000]   # cap per-file so combined stays under budget
                            _qres_parts.append(f"=== {_qlabel} ===\n{_qt}")
                        except Exception:
                            pass

                if _qres_parts:
                    qres_bytes     = b"combined"   # sentinel — text already extracted
                    qres_text      = "\n\n".join(_qres_parts)
                    qres_label     = " + ".join(_qres_labels[:3]) + ("…" if len(_qres_labels) > 3 else "")
                    qres_is_annual = _any_annual
                    ann_note = " (includes full-year figures)" if qres_is_annual else ""
                    st.success(f"✅ Downloaded **{len(_qres_parts)}** quarterly result(s) — {qres_label}{ann_note}")
                else:
                    st.warning("⚠️ All quarterly result PDF downloads failed — using XBRL only.")

        # ── 4. Extract PDF data (structured tables from most recent report) ──
        pdf_data: dict = {}
        pdf_field_count = 0

        if pdf_bytes:
            status.info(f"🔍 Step 4 / 5 — Extracting data from **{pdf_label}**…")
            progress.progress(42, text="Extracting PDF tables…")

            # 4a. Structured table extraction (pdfplumber, most recent report only)
            with st.spinner("Running pdfplumber table extraction…"):
                pdf_data = extract_pdf_data(
                    pdf_bytes,
                    years=all_years,
                    statement_type="consolidated" if consolidated else "standalone",
                )

            if "error" in pdf_data:
                st.warning(f"⚠️ PDF table extraction error: {pdf_data['error']}")
                pdf_data = {}
            else:
                pdf_field_count = sum(
                    len(yr_data)
                    for sec_data in pdf_data.values()
                    for yr_data in sec_data.values()
                )

            # 4b. pdf_text_combined already built above during multi-report download.
            #     If uploaded manually, extract now.
            if not pdf_text_combined and pdf_bytes:
                with st.spinner("Extracting raw text from financial pages…"):
                    pdf_text_combined = extract_pdf_text_for_llm(
                        pdf_bytes, doc_type="annual", label=pdf_label
                    )

            c1, c2 = st.columns(2)
            c1.metric("PDF structured values", pdf_field_count)
            n_reports_used = len(pdf_texts_by_year) if pdf_texts_by_year else (1 if pdf_bytes else 0)
            c2.metric(
                "PDF text chars",
                f"{len(pdf_text_combined):,}",
                delta=f"{n_reports_used} report(s)" if n_reports_used > 1 else None,
            )

            with st.expander("📑 PDF structured data (table extraction)"):
                for sec, yr_map in pdf_data.items():
                    st.markdown(f"**{sec}**")
                    rows = []
                    for yr, fvals in yr_map.items():
                        for lbl, val in fvals.items():
                            rows.append({"Year": yr, "Label": lbl, "Value (Cr)": val})
                    if rows:
                        st.dataframe(pd.DataFrame(rows).head(200), use_container_width=True, hide_index=True)

            with st.expander("📄 Raw PDF text (sent to LLM)"):
                st.text(pdf_text_combined[:4000] + ("…" if len(pdf_text_combined) > 4000 else ""))
        else:
            status.info("📄 Step 4 / 5 — No PDF available, using web data only.")
            progress.progress(60)

        # pdf_text is an alias used below for routing to LLM
        pdf_text: str = pdf_text_combined

        # pres_text already set above in section 3b (multi-deck extraction)

        # ── 4c. Extract quarterly result text ─────────────────────────────────
        # qres_text initialised alongside qres_bytes above; may already be set
        # by the multi-download block (b"combined" sentinel). Only extract here
        # for the single-file case (manual upload or single auto-selected PDF).
        if qres_bytes and qres_bytes != b"combined":
            # Single file (manual upload or single auto-selected) — extract now
            progress.progress(59, text="Extracting quarterly result text…")
            with st.spinner(f"Extracting text from **{qres_label}**…"):
                qres_text = extract_pdf_text_for_llm(
                    qres_bytes, max_chars_per_page=4500, doc_type="quarterly"
                )
        elif qres_bytes == b"combined":
            # Text already extracted and concatenated during multi-file download above
            progress.progress(59, text="Using combined quarterly text…")
        ann_note = " (Q4 — includes annual figures)" if qres_is_annual else ""
        if qres_text:
            st.metric(f"Quarterly result text chars{ann_note}", f"{len(qres_text):,}")
            with st.expander("📋 Quarterly result text (sample)"):
                st.text(qres_text[:2000] + ("…" if len(qres_text) > 2000 else ""))

        # ── 5. Build raw-data Excel (always) ─────────────────────────────────────
        progress.progress(68, text="Building raw data Excel…")
        status.info("📊 Step 5 / 5 — Building raw data Excel…")
        raw_excel_bytes = build_output_excel(
            company_name   = nse_symbol,
            screener_raw   = screener_raw,
            exchange_raw   = exchange_raw,
            yfinance_raw   = yfinance_raw,
            tofler_raw     = tofler_raw,
            ir_raw         = {k: v for k, v in ir_raw.items() if isinstance(v, dict) and "data" in v},
            pdf_data       = pdf_data,
            xbrl_annual    = xbrl_annual_data or {},
            xbrl_qtrly     = qres_xbrl_data or {},
            llm_harvest    = None,
        )
        progress.progress(75, text="Raw data Excel ready.")

        # ── 6. Map raw Excel → template (fuzzy, no LLM) ───────────────────────
        filled: dict[str, dict] = {}
        total_matched = 0
        _sheet_diag: dict = {}

        if _has_template:
            status.info("🔗 Step 5 / 5 — Matching labels to template (fuzzy)…")
            progress.progress(78, text="Matching labels…")
            _raw_for_map = read_raw_excel(raw_excel_bytes)
            filled = map_raw_excel_to_template(_raw_for_map, sheets_data, threshold=0.68)
            total_matched = sum(
                1 for sf in filled.values()
                for yv in sf.values() if any(v is not None for v in yv.values())
            )
            st.success(f"✅ Fuzzy label matching complete — **{total_matched}** template rows filled")
            with st.expander("🔗 Fuzzy mapping preview"):
                for sn, sf in filled.items():
                    st.markdown(f"**{sn}**")
                    rows = []
                    for lbl, yv in sf.items():
                        if any(v is not None for v in yv.values()):
                            row = {"Label": lbl}
                            row.update({yr: (f"{v:,.2f}" if v is not None else "—") for yr, v in yv.items()})
                            rows.append(row)
                    if rows:
                        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
                    else:
                        st.caption("No matches for this sheet.")
        else:
            status.info("🔗 Step 5 / 5 — No template uploaded. Skipping mapping.")
            progress.progress(85, text="Building raw data Excel…")

        # ── 6b. Optional AI enhance — LLM fills gaps the fuzzy matcher missed ──
        if use_llm and _has_template and client is not None:
            status.info("🧠 AI Enhance — LLM filling remaining label gaps…")
            progress.progress(82, text="AI: filling gaps…")
            _lm_filled: dict[str, dict] = {}


            for i, (sheet_name, sheet_data) in enumerate(sheets_data.items()):
                pct = 65 + int((i / len(sheets_data)) * 22)
                progress.progress(pct, text=f"LLM: mapping '{sheet_name}'…")

                # ── PDF text routing per sheet type ───────────────────────────────
                # Priority layers (all trimmed to keep token budget sane):
                #
                #   Asset Quality / NPA / Operating Metrics / KPI sheets:
                #     Primary   → quarterly result text  (most granular NPA breakdown)
                #     Secondary → investor presentation text  (may have summary tables)
                #     Supplement → annual report text (cross-check)
                #
                #   Presentation-type sheets (NIM, yield, spread, etc.):
                #     Primary   → investor presentation
                #     Secondary → quarterly result
                #     Supplement → annual report
                #
                #   P&L / BS / CF / other (annual-report sheets):
                #     Primary   → annual report text
                #     +Q4 supplement → quarterly result text if Q4 (has annual figures)

                is_qsheet = bool(_QUARTERLY_SHEET_RE.search(sheet_name))
                is_psheet = bool(_PRESENTATION_SHEET_RE.search(sheet_name)) and not is_qsheet

                if is_qsheet:
                    # Operating Metrics / Asset Quality sheet.
                    # Context budget strategy (45k total after keyword filter):
                    #   ① Annual report Financial Highlights  → up to 20k (FIRST — historical years)
                    #   ② Quarterly result PDF                → up to 12k (NPA/AQ precision)
                    #   ③ Investor presentation               → up to 18k capped (recent year detail)
                    # Highlights go FIRST so they aren't truncated by the 45k cap.
                    parts = []

                    # ① Annual Report Financial Highlights (historical F2020–F2024)
                    if pdf_text:
                        _ar_highlights = _extract_oper_highlights(pdf_text, max_chars=20_000)
                        if _ar_highlights:
                            parts.append(
                                "=== ANNUAL REPORT FINANCIAL HIGHLIGHTS "
                                "(use for historical years F2017–F2024) ===\n"
                                + _ar_highlights
                            )

                    # ② Quarterly result — most granular NPA / asset quality data
                    if qres_text:
                        parts.append("\n=== QUARTERLY RESULT PDF ===\n" + qres_text[:12_000])

                    # ③ Investor presentations — recent years (F2025/F2026) detail
                    if pres_text:
                        parts.append("\n=== INVESTOR PRESENTATION ===\n" + pres_text[:18_000])

                    # ④ Company IR website text — supplementary historical data
                    if ir_page_text:
                        parts.append(
                            "\n=== COMPANY IR WEBSITE (use for multi-year operating metrics) ===\n"
                            + ir_page_text[:8_000]
                        )

                    effective_pdf_text = "\n".join(parts) if parts else pdf_text

                elif is_psheet:
                    # Spreads / NIM / presentation-type sheets:
                    # Primary = full presentation text; supplement with quarterly + annual
                    parts = []
                    if pres_text:
                        parts.append(pres_text)
                    if qres_text:
                        parts.append("\n\n=== QUARTERLY RESULT SUPPLEMENT ===\n" + qres_text[:8000])
                    if pdf_text:
                        _ar_highlights = _extract_oper_highlights(pdf_text, max_chars=15_000)
                        if _ar_highlights:
                            parts.append(
                                "\n\n=== ANNUAL REPORT FINANCIAL HIGHLIGHTS ===\n" + _ar_highlights
                            )
                        else:
                            parts.append("\n\n=== ANNUAL REPORT SUPPLEMENT ===\n" + pdf_text[:6000])
                    # Company IR website — multi-year figures for NIM / spreads
                    if ir_page_text:
                        parts.append(
                            "\n\n=== COMPANY IR WEBSITE (supplementary historical data) ===\n"
                            + ir_page_text[:6_000]
                        )
                    effective_pdf_text = "\n".join(parts) if parts else pdf_text

                else:
                    # Annual-report sheet (P&L, BS, CF):
                    # Primary = multi-year annual report PDF text (all downloaded reports)
                    # Supplement = Q4 quarterly result (full-year figures cross-check)
                    # Supplement = Company IR website (historical year tables)
                    effective_pdf_text = pdf_text
                    if qres_is_annual and qres_text:
                        effective_pdf_text += (
                            "\n\n=== Q4 QUARTERLY RESULT (contains full-year figures) ===\n"
                            + qres_text[:10000]
                        )
                    if ir_page_text:
                        effective_pdf_text += (
                            "\n\n=== COMPANY IR WEBSITE (historical financial data — "
                            "use for years not covered by annual report PDF) ===\n"
                            + ir_page_text[:10_000]
                        )

                # Collect years that have real source data — section-aware so that
                # a Balance Sheet sheet only considers years with actual BS data,
                # not years borrowed from P&L (which screener provides for more years).
                _data_years: set[str] = set()
                _fy_re  = re.compile(r"^F\d{4}$")
                _qtr_re = re.compile(r"^\dQF\d{4}$")

                # Detect whether this template sheet uses quarterly column keys
                _tmpl_yrs_all = [y for y in sheet_data["years"] if y]
                _is_qtr_sheet = (
                    is_qsheet or
                    bool(any(_qtr_re.match(str(y)) for y in _tmpl_yrs_all))
                )

                # Determine which screener/exchange sections are relevant to THIS sheet
                _sl = sheet_name.lower()
                if "balance" in _sl or " bs" in _sl:
                    _rel_secs = {"Balance Sheet"}
                elif "p&l" in _sl or "profit" in _sl or "income" in _sl:
                    _rel_secs = {"P&L"}
                elif "cash" in _sl or " cf" in _sl:
                    _rel_secs = {"Cash Flow"}
                else:
                    _rel_secs = None   # None = all sections (operating metrics, spreads etc.)

                def _add_sec_years(raw_dict: dict, allowed: Optional[set]) -> None:
                    for sec_nm, _sec in raw_dict.items():
                        if not isinstance(_sec, dict):
                            continue
                        if allowed and sec_nm not in allowed:
                            continue
                        for y in _sec.get("years", []):
                            ys = str(y)
                            if _fy_re.match(ys) or _qtr_re.match(ys):
                                _data_years.add(y)

                # Screener: section-aware (Quarterly section contains quarter keys)
                _add_sec_years(screener_raw, _rel_secs)
                # If section-filtered result is empty, fall back to all sections
                if not _data_years:
                    _add_sec_years(screener_raw, None)

                # Exchange: same layout as Screener
                _add_sec_years(exchange_raw or {}, _rel_secs)

                # yfinance: {"section_name": {"F2025": {...}, ...}}
                for _yf_sec, _yf_data in yfinance_raw.items():
                    if not isinstance(_yf_data, dict):
                        continue
                    if _rel_secs and _yf_sec not in _rel_secs:
                        continue
                    _data_years.update(k for k in _yf_data if _fy_re.match(str(k)))

                # If yfinance also yielded nothing after filtering, add all yfinance years
                if not _data_years:
                    for _yf_data in yfinance_raw.values():
                        if isinstance(_yf_data, dict):
                            _data_years.update(k for k in _yf_data if _fy_re.match(str(k)))

                # PDF structured data (all sections — one PDF per year)
                for _sec in pdf_data.values():
                    if isinstance(_sec, dict):
                        _data_years.update(k for k in _sec if _fy_re.match(str(k)))

                # XBRL data (all buckets)
                for _xbrl_bucket in (qres_xbrl_data, xbrl_annual_data, xbrl_qtrly_data):
                    for _yr_map in _xbrl_bucket.values():
                        if isinstance(_yr_map, dict):
                            _data_years.update(
                                k for k in _yr_map
                                if _fy_re.match(str(k)) or _qtr_re.match(str(k))
                            )

                # For operating/KPI sheets: also scan tofler_raw years directly
                if _rel_secs is None:
                    for _sec_name, _sec in tofler_raw.items():
                        if isinstance(_sec, dict):
                            _data_years.update(y for y in _sec.get("years", []) if _fy_re.match(str(y)))

                # ── _ask_years: which years to send to LLM ─────────────────────────────
                # For quarterly sheets: use ALL available quarter keys from Screener Quarterly section,
                # clipped to what the template actually asks for.
                _tmpl_years = sorted(_tmpl_yrs_all, key=lambda k: k or "", reverse=True)

                if _is_qtr_sheet:
                    # Collect all quarter keys available in any source
                    _avail_qtrs: set[str] = set()
                    _scr_q = screener_raw.get("Quarterly", {})
                    for _ym in _scr_q.get("data", {}).values():
                        _avail_qtrs.update(k for k in _ym if _qtr_re.match(str(k)))
                    for _yr_map in xbrl_qtrly_data.values():
                        if isinstance(_yr_map, dict):
                            _avail_qtrs.update(k for k in _yr_map if _qtr_re.match(str(k)))

                    # Template quarterly years (may be quarter keys OR annual keys if template
                    # columns were not yet parseable as quarters — use all in that case)
                    _tmpl_qtrs = {y for y in _tmpl_years if _qtr_re.match(str(y))}
                    if _tmpl_qtrs:
                        # Only ask for quarters present in both template and source data
                        _ask_years = sorted(
                            (_tmpl_qtrs & _avail_qtrs) or _tmpl_qtrs or _avail_qtrs,
                            key=lambda k: k or "", reverse=True,
                        )
                    else:
                        # Template columns are annual-format but this is a quarterly sheet
                        # (shouldn't happen after Fix A, but safety fallback)
                        _ask_years = sorted(_avail_qtrs or set(_tmpl_years[:8]),
                                            key=lambda k: k or "", reverse=True)
                else:
                    _base_years = {y for y in _tmpl_years if y in _data_years}
                    if _data_years and effective_pdf_text:
                        _safety = set(_tmpl_years[:2])
                    elif _data_years:
                        _safety = set(_tmpl_years[:1])
                    else:
                        _safety = set()
                    _ask_years = sorted(
                        _base_years | _safety,
                        key=lambda k: k or "",
                        reverse=True,
                    )
                _ask_years = _ask_years or _tmpl_years or sheet_data["years"]

                try:
                    # ── Phase 1: Harvest raw label→year→value dict from PDF text ─────
                    # The LLM reads the text freely and outputs whatever it finds —
                    # no template labels involved. This is fast and very accurate because
                    # the model just needs to "read" without also needing to "map".
                    _harvested: dict[str, dict[str, float]] = {}
                    if effective_pdf_text:
                        _harvested = llm_harvest_from_text(
                            effective_pdf_text, _ask_years, client, api_model
                        )

                    # ── Phase 2: Map harvested data to template labels (pure Python) ─
                    # _financial_sim synonym clusters do the matching — no LLM call needed.
                    # Lower threshold for non-standard sheets (operating/KPI) where label wording
                    # diverges more from template labels (e.g. "GNPA Ratio" vs "Gross NPA %").
                    _sl_lower = sheet_name.lower()
                    _is_fin_sheet = any(k in _sl_lower for k in ("p&l","profit","income","balance","bs","cash","cf"))
                    _p2_threshold = 0.65 if _is_fin_sheet else 0.60
                    _phase2_filled = map_harvested_to_template(
                        sheet_data["labels"], _ask_years, _harvested, threshold=_p2_threshold
                    )

                    # ── Phase 3: Standard LLM pass (structured data + template labels) ─
                    # This handles screener/yfinance/XBRL data which Phase 1 doesn't see.
                    # Phase 3 also fills any gaps Phase 2 left (unmatched labels or years
                    # where harvested text had no data).
                    _xbrl_for_sheet = (
                        qres_xbrl_data   # combined (all quarters) for quarterly/operating sheets
                        if (is_qsheet or is_psheet)
                        else xbrl_annual_data   # Q4 annual figures only for P&L/BS/CF sheets
                    ) or None
                    sheet_filled = llm_map_fields(
                        sheet_name=sheet_name,
                        template_labels=sheet_data["labels"],
                        years=_ask_years,
                        screener_raw=screener_raw,
                        yfinance_raw=yfinance_raw,
                        pdf_data=pdf_data,
                        client=client,
                        model=api_model,
                        pdf_text=effective_pdf_text,
                        exchange_raw=exchange_raw,
                        xbrl_data=_xbrl_for_sheet,
                        xbrl_qtrly=xbrl_qtrly_data if (is_qsheet or is_psheet) else None,
                        # Issues 8 & 9 fix: pass MC and IR data as SEPARATE labeled sources
                        # so the LLM can apply correct priority (not buried in screener_raw blend).
                        mc_raw=tofler_raw if tofler_raw else None,
                        ir_structured_raw=(
                            {k: v for k, v in ir_raw.items() if isinstance(v, dict)}
                            if ir_raw else None
                        ),
                    )

                    # ── Merge: Phase 2 harvest fills gaps left by Phase 3 ────────────
                    # Phase 3 (structured LLM) wins when it has a value; Phase 2 (harvest)
                    # fills any label+year slot that Phase 3 left as None.
                    _p2_wins = 0
                    for tmpl_lbl, p2_yr_vals in _phase2_filled.items():
                        p3_yr_vals = sheet_filled.get(tmpl_lbl, {})
                        merged: dict[str, Optional[float]] = {}
                        for yr in _ask_years:
                            p3_val = p3_yr_vals.get(yr)
                            p2_val = p2_yr_vals.get(yr)
                            if p3_val is not None:
                                merged[yr] = p3_val
                            elif p2_val is not None:
                                merged[yr] = p2_val
                                _p2_wins += 1
                            else:
                                merged[yr] = None
                        sheet_filled[tmpl_lbl] = merged
                    if _p2_wins:
                        logger.info(f"{sheet_name}: Phase-2 harvest filled {_p2_wins} additional values")

                except RuntimeError as _llm_err:
                    st.error(f"❌ **{sheet_name}** — LLM error: {_llm_err}")
                    sheet_filled = {}
                except Exception as _llm_err:
                    st.error(f"❌ **{sheet_name}** — unexpected error: {type(_llm_err).__name__}: {_llm_err}")
                    sheet_filled = {}

                n = sum(1 for yv in sheet_filled.values() for v in yv.values() if v is not None)
                total_matched += n
                _lm_filled[sheet_name] = sheet_filled

                # Count LLM-returned values per year for diagnostics
                _yr_counts = {}
                for yv in sheet_filled.values():
                    for yr, v in yv.items():
                        if v is not None:
                            _yr_counts[yr] = _yr_counts.get(yr, 0) + 1

                _sheet_diag[sheet_name] = {
                    "data_years": sorted(_data_years),
                    "ask_years": _ask_years,
                    "llm_values_per_year": _yr_counts,
                    "tmpl_years_total": len(sheet_data["years"]),
                    "harvest_labels": len(_harvested) if "_harvested" in dir() else 0,
                    "phase2_fills": _p2_wins if "_p2_wins" in dir() else 0,
                }
                _harvest_note = (f" · harvest: {_sheet_diag[sheet_name]['harvest_labels']} raw labels"
                                 f" → {_sheet_diag[sheet_name]['phase2_fills']} extra fills"
                                 if _sheet_diag[sheet_name]["harvest_labels"] else "")
                st.info(f"  **{sheet_name}**: {n} values matched ({len(_ask_years)} years: {', '.join(_ask_years[:4])}{'…' if len(_ask_years)>4 else ''}){_harvest_note}")

            with st.expander("🧠 LLM mapping preview"):
                for sn, sf in filled.items():
                    st.markdown(f"**{sn}**")
                    rows = []
                    for lbl, yv in sf.items():
                        if any(v is not None for v in yv.values()):
                            row = {"Label": lbl}
                            row.update({yr: (f"{v:,.2f}" if v is not None else "—") for yr, v in yv.items()})
                            rows.append(row)
                    if rows:
                        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
                    else:
                        st.caption("No values matched for this sheet.")

            with st.expander("🔬 Year coverage diagnostics (debug)"):
                st.caption("Shows which years have source data, which were sent to LLM, and how many values returned per year.")
                for sn, diag in _sheet_diag.items():
                    st.markdown(f"**{sn}**")
                    dy = diag["data_years"]
                    ay = diag["ask_years"]
                    yc = diag["llm_values_per_year"]
                    col1, col2 = st.columns(2)
                    with col1:
                        st.markdown(f"**Source data years** ({len(dy)}):")
                        st.code(", ".join(dy) if dy else "(none)")
                        st.markdown(f"**Years sent to LLM** ({len(ay)}):")
                        st.code(", ".join(ay) if ay else "(none)")
                    with col2:
                        st.markdown("**LLM values returned per year:**")
                        if yc:
                            yr_rows = [{"Year": yr, "Values": cnt} for yr, cnt in sorted(yc.items(), reverse=True)]
                            st.dataframe(pd.DataFrame(yr_rows), use_container_width=True, hide_index=True)
                        else:
                            st.caption("No values returned.")
                        missing_yrs = [y for y in ay if y not in yc]
                        if missing_yrs:
                            st.warning(f"⚠️ LLM returned 0 values for: {', '.join(missing_yrs)}")
                        # Phase 2 harvest summary
                        _hl = diag.get("harvest_labels", 0)
                        _p2 = diag.get("phase2_fills", 0)
                        if _hl:
                            st.caption(
                                f"🔍 Harvest phase: extracted **{_hl}** raw labels from PDF text "
                                f"→ similarity-matched **{_p2}** additional template slots "
                                f"(fills gaps the main LLM pass missed)"
                            )
                    st.divider()


            # Merge LLM fills into fuzzy-matched filled (LLM wins on conflicts)
            _ai_extra = 0
            for _sn, _sf in _lm_filled.items():
                if _sn not in filled:
                    filled[_sn] = _sf
                else:
                    for _lbl, _yv in _sf.items():
                        if _lbl not in filled[_sn]:
                            filled[_sn][_lbl] = _yv
                        else:
                            for _yr, _v in _yv.items():
                                if filled[_sn][_lbl].get(_yr) is None and _v is not None:
                                    filled[_sn][_lbl][_yr] = _v
                                    _ai_extra += 1
            if _ai_extra:
                st.info(f"🤖 AI Enhance added **{_ai_extra}** extra values on top of fuzzy matches")


        # Sources banner
        sources = []
        if s_count:         sources.append("Screener.in")
        if exc_count:       sources.append("NSE/BSE Exchange")
        if y_count:         sources.append("yfinance")
        if pdf_field_count: sources.append(f"Annual Report ({pdf_label})")
        if pres_text:       sources.append(f"Investor Presentation ({pres_label})")
        if qres_text:
            ann = " [Q4/Annual]" if qres_is_annual else ""
            sources.append(f"Quarterly Result{ann} ({qres_label})")

        # ── Write to template (only if template was uploaded) ──────────────────
        if _has_template:
            progress.progress(92, text="Writing to template…")
            output_bytes, stats = write_to_template(template_bytes, filled, overwrite=overwrite)
            progress.progress(100, text="Done ✓")
            status.success(
                f"✅ Done! **{stats['written']}** cells written · "
                f"{stats['skipped_formula']} formulas preserved · "
                f"{stats['skipped_filled']} existing kept · "
                f"{stats['not_found']} labels unmatched\n\n"
                f"📡 Sources: {' + '.join(sources) if sources else 'none'}"
            )
        else:
            progress.progress(100, text="Done ✓")
            status.success(
                f"✅ Done! Raw Data Excel ready.\n\n"
                f"📡 Sources: {' + '.join(sources) if sources else 'none'}"
            )

        # ── Download buttons ───────────────────────────────────────────────────
        if _has_template:
            _dl_c1, _dl_c2 = st.columns(2)
            with _dl_c1:
                st.download_button(
                    label="⬇️ Raw Data Excel (all fields, 5 sheets)",
                    data=raw_excel_bytes,
                    file_name=f"{nse_symbol}_raw_data.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    help="All extracted data dumped as-is — no mapping. Every label from every source.",
                )
            with _dl_c2:
                st.download_button(
                    label="⬇️ Populated Template",
                    data=output_bytes,
                    file_name=f"{nse_symbol}_populated.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    help="Your template with fuzzy-matched values filled in (+ AI Enhance if enabled).",
                )
        else:
            st.download_button(
                label="⬇️ Download Raw Data Excel (5 sheets)",
                data=raw_excel_bytes,
                file_name=f"{nse_symbol}_raw_data.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
            )

        # ── Write-level diagnostics (template mode only) ──────────────────────
        if _has_template:
            _wb_yr = stats.get("written_by_year", {})
            _fb_yr = stats.get("formula_by_year", {})
            if _wb_yr or _fb_yr:
                with st.expander("📊 Written vs formula-preserved cells by year"):
                    st.caption(
                        "Formula cells are always kept as-is — the app only fills **blank** cells. "
                        "'Formula-preserved' = a matched value was found but the cell already has a formula, so the formula wins."
                    )
                    all_yrs = sorted(set(list(_wb_yr.keys()) + list(_fb_yr.keys())), reverse=True)
                    yr_diag_rows = [
                        {"Year": yr, "Written": _wb_yr.get(yr, 0), "Formula-preserved": _fb_yr.get(yr, 0)}
                        for yr in all_yrs
                    ]
                    st.dataframe(pd.DataFrame(yr_diag_rows), use_container_width=True, hide_index=True)

            audit = stats.get("audit", [])
            if audit:
                st.subheader("📋 Cells Written")
                st.dataframe(pd.DataFrame(audit), use_container_width=True, hide_index=True)

    except Exception as e:
        status.error(f"❌ Pipeline failed: {e}")
        with st.expander("Error details"):
            st.code(traceback.format_exc())
