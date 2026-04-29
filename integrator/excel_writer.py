"""
Excel Writer — non-destructive population of the template.

Rules enforced:
  1. NEVER overwrite a cell that contains a formula.
  2. NEVER overwrite a cell that already has a numeric value.
  3. ONLY write to cells that are empty (value is None or "").
  4. Preserve all styles, formatting, merged cells, and named ranges.
  5. Write a full audit log of every cell touched.
"""

from __future__ import annotations

import copy
import io
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional, Union

import openpyxl
from openpyxl.utils import get_column_letter

from mapper.field_mapper import MappedField, MappingReport
from utils.helpers import get_logger

logger = get_logger(__name__)


@dataclass
class WriteRecord:
    """Audit record for a single cell write."""
    sheet: str
    cell_ref: str            # e.g. "E6"
    row_label: str
    year: str
    statement_type: str
    value: float
    match_method: str
    match_score: float
    extracted_label: str
    confidence_tier: str = "MEDIUM"   # "VERY_HIGH"|"HIGH"|"MEDIUM"|"FLAG"
    web_value: float = None
    web_source: str = ""
    flag_reason: str = ""
    timestamp: str = field(default_factory=lambda: datetime.now().isoformat())


@dataclass
class WriteResult:
    """Summary of a write operation."""
    output_path: str
    cells_written: int
    cells_skipped_formula: int
    cells_skipped_filled: int
    cells_skipped_no_match: int
    cells_review_queued: int = 0       # flagged by web validation → Data Review sheet
    audit_log: list[WriteRecord] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    def to_summary_dict(self) -> dict:
        return {
            "output_path": self.output_path,
            "cells_written": self.cells_written,
            "cells_review_queued": self.cells_review_queued,
            "cells_skipped_formula": self.cells_skipped_formula,
            "cells_skipped_filled": self.cells_skipped_filled,
            "cells_skipped_no_match": self.cells_skipped_no_match,
            "total_warnings": len(self.warnings),
        }


# ── Main writer ────────────────────────────────────────────────────────────────

def write_to_template(
    template_path: Union[str, Path],
    mapping_reports: list[MappingReport],
    output_path: Optional[Union[str, Path]] = None,
    overwrite_existing: bool = False,
    web_validated: bool = False,      # True = use confidence_tier to route writes
    auto_write_tiers: tuple = ("VERY_HIGH", "HIGH", "MEDIUM"),  # tiers that go to main sheet
) -> WriteResult:
    """
    Apply mapping reports to the Excel template and save the output.

    Parameters
    ----------
    template_path : path to the original template
    mapping_reports : list of MappingReport from the field mapper
    output_path : where to save the result (default: template_path with _populated suffix)
    overwrite_existing : if True, overwrite cells that already have numeric values

    Returns
    -------
    WriteResult with audit log and statistics
    """
    template_path = Path(template_path)
    if output_path is None:
        output_path = template_path.parent / (
            template_path.stem + "_populated" + template_path.suffix
        )
    output_path = Path(output_path)

    logger.info(f"Loading template: {template_path}")
    wb = openpyxl.load_workbook(str(template_path))   # opens with formula strings

    audit_log: list[WriteRecord] = []
    warnings: list[str] = []
    stats = {
        "written": 0,
        "skip_formula": 0,
        "skip_filled": 0,
        "skip_no_match": 0,
        "review_queued": 0,
    }

    # Collect all mapped fields across all reports
    all_mapped: list[MappedField] = []
    for report in mapping_reports:
        all_mapped.extend(report.mapped)

    logger.info(
        f"Applying {len(all_mapped)} field mappings to template "
        f"(web_validated={web_validated}, auto_write_tiers={auto_write_tiers})…"
    )

    # Prepare the Data Review sheet (created only if there are flagged fields)
    review_rows: list[dict] = []

    for mf in all_mapped:
        # ── Confidence-tier routing ────────────────────────────────────────────
        # When web_validated=True:
        #   VERY_HIGH / HIGH → write to main template (confirmed by web)
        #   MEDIUM           → write to main template (no web conflict, but unconfirmed)
        #   FLAG             → DO NOT write to main sheet; add to Data Review sheet only
        # When web_validated=False (no web data): write everything as usual.
        if web_validated and mf.confidence_tier == "FLAG":
            review_rows.append({
                "Sheet":           mf.sheet_name,
                "Row Label":       mf.template_label,
                "Year":            mf.year,
                "Statement Type":  mf.statement_type,
                "PDF Value":       round(mf.value, 2),
                "Web Value":       round(mf.web_value, 2) if mf.web_value is not None else "—",
                "Web Source":      mf.web_source or "—",
                "Difference":      (
                    f"{abs(mf.value - mf.web_value) / abs(mf.web_value):.1%}"
                    if mf.web_value else "—"
                ),
                "Flag Reason":     mf.flag_reason or "",
                "Match Method":    mf.match_method,
            })
            stats["review_queued"] += 1
            logger.info(
                f"  REVIEW {mf.sheet_name} [{mf.template_label} | {mf.year}] "
                f"PDF={mf.value} WebValue={mf.web_value} → {mf.flag_reason}"
            )
            continue

        if mf.sheet_name not in wb.sheetnames:
            warnings.append(f"Sheet '{mf.sheet_name}' not found in workbook")
            stats["skip_no_match"] += 1
            continue

        ws = wb[mf.sheet_name]
        cell = ws.cell(mf.row, mf.col)
        cell_ref = f"{get_column_letter(mf.col)}{mf.row}"

        # Guard 1: formula cells
        if isinstance(cell.value, str) and cell.value.startswith("="):
            logger.debug(f"  SKIP formula: {mf.sheet_name}!{cell_ref}")
            stats["skip_formula"] += 1
            continue

        # Guard 2: already has a value
        if cell.value is not None and cell.value != "":
            if not overwrite_existing:
                logger.debug(f"  SKIP filled: {mf.sheet_name}!{cell_ref} = {cell.value!r}")
                stats["skip_filled"] += 1
                continue
            else:
                warnings.append(
                    f"Overwriting existing value in {mf.sheet_name}!{cell_ref}: "
                    f"{cell.value} → {mf.value}"
                )

        # Write the value
        cell.value = round(mf.value, 2) if mf.value is not None else mf.value
        stats["written"] += 1

        record = WriteRecord(
            sheet=mf.sheet_name,
            cell_ref=cell_ref,
            row_label=mf.template_label,
            year=mf.year,
            statement_type=mf.statement_type,
            value=mf.value,
            match_method=mf.match_method,
            match_score=mf.match_score,
            extracted_label=mf.extracted_label,
            confidence_tier=mf.confidence_tier,
            web_value=mf.web_value,
            web_source=mf.web_source,
            flag_reason=mf.flag_reason,
        )
        audit_log.append(record)
        web_tag = (
            f" [web:{mf.web_source}={mf.web_value:.2f}→{mf.confidence_tier}]"
            if mf.web_value is not None else ""
        )
        logger.info(
            f"  WRITE {mf.sheet_name}!{cell_ref} "
            f"[{mf.template_label} | {mf.year} | {mf.statement_type}] "
            f"= {mf.value} (via {mf.match_method}, score={mf.match_score:.2f}){web_tag}"
        )

    # ── Write the Data Review sheet (if any flagged fields) ────────────────────
    if review_rows:
        _write_review_sheet(wb, review_rows)
        logger.info(
            f"  Data Review sheet: {len(review_rows)} flagged field(s) need human review"
        )

    # Save
    wb.save(str(output_path))
    logger.info(
        f"Saved to: {output_path}\n"
        f"  Written: {stats['written']} | "
        f"Review-queued: {stats['review_queued']} | "
        f"Skip-formula: {stats['skip_formula']} | "
        f"Skip-filled: {stats['skip_filled']} | "
        f"Skip-no-match: {stats['skip_no_match']}"
    )

    return WriteResult(
        output_path=str(output_path),
        cells_written=stats["written"],
        cells_skipped_formula=stats["skip_formula"],
        cells_skipped_filled=stats["skip_filled"],
        cells_skipped_no_match=stats["skip_no_match"],
        cells_review_queued=stats["review_queued"],
        audit_log=audit_log,
        warnings=warnings,
    )


def _write_review_sheet(wb, review_rows: list[dict]) -> None:
    """
    Write flagged fields to a 'Data Review' sheet in the workbook.
    This sheet is for human review — values here were NOT written to the
    main template because web data disagreed with the PDF-extracted value
    by more than 25%.

    The analyst reads this sheet, corrects the value if needed, and
    manually pastes it into the appropriate cell in the main sheets.
    """
    SHEET_NAME = "Data Review"

    # Remove existing sheet if present (re-run scenario)
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]

    ws = wb.create_sheet(SHEET_NAME)

    # Header row
    headers = [
        "Sheet", "Row Label", "Year", "Statement Type",
        "PDF Value (INR Cr)", "Web Value (INR Cr)", "Web Source",
        "Difference %", "Flag Reason", "Match Method",
        "Action (Accept / Correct / Skip)",
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(
            fill_type="solid", fgColor="FFC7CE"   # light red header
        )

    # Data rows
    for row_idx, rec in enumerate(review_rows, start=2):
        values = [
            rec.get("Sheet", ""),
            rec.get("Row Label", ""),
            rec.get("Year", ""),
            rec.get("Statement Type", ""),
            rec.get("PDF Value", ""),
            rec.get("Web Value", ""),
            rec.get("Web Source", ""),
            rec.get("Difference", ""),
            rec.get("Flag Reason", ""),
            rec.get("Match Method", ""),
            "",   # analyst fills this in
        ]
        for col_idx, val in enumerate(values, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.fill = openpyxl.styles.PatternFill(
                fill_type="solid", fgColor="FFEB9C"   # light yellow
            )

    # Auto-size columns (approximate)
    col_widths = [15, 40, 8, 16, 18, 18, 12, 12, 60, 14, 30]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_to_bytes(
    template_bytes: bytes,
    mapping_reports: list[MappingReport],
    overwrite_existing: bool = False,
    web_validated: bool = False,
    auto_write_tiers: tuple = ("VERY_HIGH", "HIGH", "MEDIUM"),
) -> tuple[bytes, WriteResult]:
    """
    In-memory version for Streamlit — takes template as bytes, returns output as bytes.
    """
    import tempfile, os

    # Write template to temp file
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_in:
        tmp_in.write(template_bytes)
        tmp_in_path = tmp_in.name

    tmp_out_path = tmp_in_path.replace(".xlsx", "_out.xlsx")

    try:
        result = write_to_template(
            template_path=tmp_in_path,
            mapping_reports=mapping_reports,
            output_path=tmp_out_path,
            overwrite_existing=overwrite_existing,
            web_validated=web_validated,
            auto_write_tiers=auto_write_tiers,
        )
        with open(tmp_out_path, "rb") as f:
            output_bytes = f.read()
        result.output_path = "populated_template.xlsx"
        return output_bytes, result
    finally:
        for p in [tmp_in_path, tmp_out_path]:
            try:
                os.unlink(p)
            except Exception:
                pass


# ── Audit log export ───────────────────────────────────────────────────────────

def audit_log_to_dataframe(audit_log: list[WriteRecord]):
    """Convert audit log to a pandas DataFrame for display."""
    import pandas as pd
    if not audit_log:
        return pd.DataFrame()

    rows = []
    for r in audit_log:
        # Use getattr with defaults so old WriteRecord objects (pre web-validation
        # fields) loaded from session state don't crash.
        web_val = getattr(r, "web_value", None)
        rows.append({
            "Sheet":            r.sheet,
            "Cell":             r.cell_ref,
            "Row Label":        r.row_label,
            "Year":             r.year,
            "Type":             r.statement_type,
            "Value (INR Cr)":   r.value,
            "Web Value":        round(web_val, 2) if web_val is not None else "—",
            "Web Source":       getattr(r, "web_source", "") or "—",
            "Confidence":       getattr(r, "confidence_tier", "MEDIUM"),
            "Extracted Label":  r.extracted_label,
            "Match Method":     r.match_method,
            "Match Score":      f"{r.match_score:.0%}",
            "Flag":             getattr(r, "flag_reason", "") or "",
        })
    return pd.DataFrame(rows)
