"""
Universal Financial Data Extraction & Template Population System
Streamlit Web Application — main entry point

Run:
    streamlit run app.py
"""

from __future__ import annotations

import io
import os
import subprocess
import sys
import time
import traceback
from pathlib import Path

# ── Auto-install missing packages using the SAME Python that is running this
# ── script (fixes venv/interpreter mismatch on Mac). ─────────────────────────
_REQUIRED = {
    "fitz":             "pymupdf>=1.23.0",
    "openpyxl":         "openpyxl>=3.1.2",
    "openai":           "openai>=1.30.0",
    "dotenv":           "python-dotenv>=1.0.0",
    "fuzzywuzzy":       "fuzzywuzzy>=0.18.0",
    "Levenshtein":      "python-Levenshtein>=0.25.0",
    "pandas":           "pandas>=2.2.0",
}

def _ensure_packages():
    missing = []
    for mod, pkg in _REQUIRED.items():
        try:
            __import__(mod)
        except ImportError:
            missing.append(pkg)
    if missing:
        import streamlit as _st
        _st.info(f"Installing missing packages: {missing} — please wait…")
        subprocess.check_call(
            [sys.executable, "-m", "pip", "install", "--quiet"] + missing
        )
        # Force re-import after install
        import importlib
        for mod in list(sys.modules.keys()):
            if any(m.split(".")[0] == mod for m in _REQUIRED):
                importlib.reload(sys.modules[mod])

_ensure_packages()

import pandas as pd
import streamlit as st

# ── Path setup ─────────────────────────────────────────────────────────────────
sys.path.insert(0, str(Path(__file__).parent))

# ── Page config ────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Financial Data Extractor",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Custom CSS ─────────────────────────────────────────────────────────────────
st.markdown("""
<style>
    .main-header {
        font-size: 2.2rem;
        font-weight: 700;
        color: #1E3A5F;
        margin-bottom: 0.2rem;
    }
    .sub-header {
        font-size: 1rem;
        color: #6B7280;
        margin-bottom: 2rem;
    }
    .status-box {
        padding: 0.75rem 1rem;
        border-radius: 0.5rem;
        margin: 0.5rem 0;
    }
    .step-badge {
        background: #1E3A5F;
        color: white;
        padding: 2px 10px;
        border-radius: 12px;
        font-size: 0.8rem;
        margin-right: 8px;
    }
    .metric-card {
        background: #F0F4F8;
        border-left: 4px solid #1E3A5F;
        padding: 0.6rem 1rem;
        border-radius: 0 0.4rem 0.4rem 0;
        margin: 0.3rem 0;
    }
    div[data-testid="stExpander"] details summary p {
        font-weight: 600;
    }
</style>
""", unsafe_allow_html=True)


# ── Session state helpers ──────────────────────────────────────────────────────

_FISCAL_YEARS = [
    "Auto-detect",
    "FY 2026 (Apr 2025 – Mar 2026)",
    "FY 2025 (Apr 2024 – Mar 2025)",
    "FY 2024 (Apr 2023 – Mar 2024)",
    "FY 2023 (Apr 2022 – Mar 2023)",
    "FY 2022 (Apr 2021 – Mar 2022)",
    "FY 2021 (Apr 2020 – Mar 2021)",
]
_YEAR_LABEL_MAP = {
    "FY 2026 (Apr 2025 – Mar 2026)": "F2026",
    "FY 2025 (Apr 2024 – Mar 2025)": "F2025",
    "FY 2024 (Apr 2023 – Mar 2024)": "F2024",
    "FY 2023 (Apr 2022 – Mar 2023)": "F2023",
    "FY 2022 (Apr 2021 – Mar 2022)": "F2022",
    "FY 2021 (Apr 2020 – Mar 2021)": "F2021",
}
_REPORT_TYPES = [
    "Annual Report",
    "Investor Presentation",
    "Quarterly Results",
    "Analyst Presentation",
    "Other",
]
_STMT_TYPES = ["Auto-detect", "Consolidated", "Standalone", "Both"]


def _init_state():
    defaults = {
        "extraction_results": None,
        "mapping_reports": None,
        "write_result": None,
        "output_bytes": None,
        "template_model": None,
        "parsed_docs": None,
        "pdf_metadata": {},   # {filename: {year, report_type, statement_type}}
        "api_key_valid": False,
        "log_messages": [],
        "web_results": None,
        "validation_results": None,
        "company_identifier": None,
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v


def _log(msg: str, level: str = "info"):
    entry = f"[{level.upper()}] {msg}"
    st.session_state.log_messages.append(entry)


# ── Main UI ────────────────────────────────────────────────────────────────────

def main():
    _init_state()

    # ── Header ────────────────────────────────────────────────────────────────
    st.markdown('<div class="main-header">📊 Financial Data Extraction System</div>', unsafe_allow_html=True)
    st.markdown(
        '<div class="sub-header">Universal template population from unstructured financial documents</div>',
        unsafe_allow_html=True,
    )

    # ── Sidebar ────────────────────────────────────────────────────────────────
    with st.sidebar:
        st.markdown("### ⚙️ Configuration")

        api_key = st.text_input(
            "DeepSeek API Key",
            type="password",
            value=os.getenv("DEEPSEEK_API_KEY", ""),
            help="Enter your DeepSeek API key. It won't be stored.",
        )
        if api_key:
            os.environ["DEEPSEEK_API_KEY"] = api_key
            st.session_state.api_key_valid = True
            st.success("✓ API key set")
        else:
            st.warning("⚠ API key required for extraction")

        st.markdown("---")
        st.markdown("### 🔧 Options")

        overwrite_mode = st.checkbox(
            "Overwrite existing values",
            value=False,
            help="If checked, will replace already-filled cells. Default: preserve existing data.",
        )
        use_llm_mapping = st.checkbox(
            "Use LLM for field mapping",
            value=True,
            help="Use DeepSeek for semantic field name matching (recommended).",
        )
        min_confidence = st.slider(
            "Min extraction confidence",
            min_value=0.0, max_value=1.0, value=0.5, step=0.05,
            help="Only include extraction results above this confidence threshold.",
        )

        st.markdown("---")
        st.markdown("### 🌐 Web Validation (optional)")
        st.caption(
            "Provide company identifiers to cross-validate PDF-extracted "
            "values against Yahoo Finance and BSE data."
        )
        nse_sym = st.text_input(
            "NSE Symbol",
            value=st.session_state.get("_nse_sym", ""),
            placeholder="e.g. HDFCBANK",
            help="NSE ticker symbol (preferred — more complete data via Yahoo Finance)",
            key="_nse_sym",
        )
        bse_code = st.text_input(
            "BSE Scrip Code",
            value=st.session_state.get("_bse_code", ""),
            placeholder="e.g. 500180",
            help="BSE scrip code (used for BSE quarterly XBRL data)",
            key="_bse_code",
        )
        if nse_sym or bse_code:
            from web.base import CompanyIdentifier
            st.session_state.company_identifier = CompanyIdentifier(
                name=nse_sym or bse_code,
                nse_symbol=nse_sym.upper().strip() if nse_sym else None,
                bse_code=bse_code.strip() if bse_code else None,
            )
            st.success(f"✓ Company: {st.session_state.company_identifier}")
        else:
            st.session_state.company_identifier = None
            st.info("Web validation skipped — no company identifier provided")

        st.markdown("---")
        st.markdown("### 📋 About")
        st.markdown(
            "Extracts financial data from PDFs and populates Excel templates "
            "using AI-powered field matching.\n\n"
            "**Supported sheets:**\n"
            "- Annual P&L\n"
            "- Annual Balance Sheet\n"
            "- Annual Cash Flow\n"
            "- Quarterly Results"
        )

    # ── Main content — 3-step workflow ─────────────────────────────────────────
    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "📁 1 · Upload Files",
        "🔍 2 · Extract & Preview",
        "✅ 3 · Map & Populate",
        "🌐 Web Validation",
        "📋 Audit Log",
    ])

    # ──────────────────────────────────────────────────────────────────────────
    # TAB 1 — Upload
    # ──────────────────────────────────────────────────────────────────────────
    with tab1:
        st.subheader("Upload Files")
        col1, col2 = st.columns([1, 2])

        with col1:
            st.markdown("#### 📄 Excel Template")
            template_file = st.file_uploader(
                "Upload standardised Excel template (.xlsx)",
                type=["xlsx"],
                key="template_upload",
            )
            if template_file:
                st.success(f"✓ **{template_file.name}** ({template_file.size:,} bytes)")
                _show_template_preview(template_file)

        with col2:
            st.markdown("#### 📑 Financial Documents")
            pdf_files = st.file_uploader(
                "Upload one or more PDFs (annual reports, presentations…)",
                type=["pdf"],
                accept_multiple_files=True,
                key="pdf_upload",
            )

            if pdf_files:
                st.markdown("**Tag each document** so the extractor knows exactly what year it covers:")
                st.markdown("---")
                for pf in pdf_files:
                    with st.container():
                        st.markdown(f"📄 **{pf.name}** &nbsp; <span style='color:#6B7280;font-size:0.85rem'>({pf.size:,} bytes)</span>", unsafe_allow_html=True)
                        c1, c2, c3 = st.columns([2, 2, 2])
                        with c1:
                            fy = st.selectbox(
                                "Fiscal year",
                                _FISCAL_YEARS,
                                index=2,   # default: FY 2025
                                key=f"fy_{pf.name}",
                                help="Which fiscal year does this report cover?",
                            )
                        with c2:
                            rt = st.selectbox(
                                "Report type",
                                _REPORT_TYPES,
                                index=0,
                                key=f"rt_{pf.name}",
                            )
                        with c3:
                            stype = st.selectbox(
                                "Statements",
                                _STMT_TYPES,
                                index=0,
                                key=f"st_{pf.name}",
                                help="Which statements to extract. 'Auto-detect' reads from the PDF.",
                            )
                        st.session_state.pdf_metadata[pf.name] = {
                            "year": _YEAR_LABEL_MAP.get(fy),   # None = auto-detect
                            "report_type": rt,
                            "statement_type": None if stype == "Auto-detect" else stype.lower(),
                        }
                        st.markdown("---")

    # ──────────────────────────────────────────────────────────────────────────
    # TAB 2 — Extract
    # ──────────────────────────────────────────────────────────────────────────
    with tab2:
        st.subheader("Extract Financial Data from PDFs")

        if not template_file or not pdf_files:
            st.info("⬅ Please upload files in the **Upload Files** tab first.")
        elif not st.session_state.api_key_valid:
            st.warning("⬅ Please enter your DeepSeek API key in the sidebar.")
        else:
            col_btn, col_status = st.columns([1, 3])
            with col_btn:
                run_extraction = st.button(
                    "🚀 Run Extraction",
                    type="primary",
                    use_container_width=True,
                )

            if run_extraction:
                _run_extraction(
                    template_file, pdf_files,
                    min_confidence,
                    pdf_metadata=st.session_state.pdf_metadata,
                )

            if st.session_state.extraction_results:
                _show_extraction_preview(st.session_state.extraction_results)

    # ──────────────────────────────────────────────────────────────────────────
    # TAB 3 — Map & Populate
    # ──────────────────────────────────────────────────────────────────────────
    with tab3:
        st.subheader("Map Fields & Populate Template")

        if not st.session_state.extraction_results:
            st.info("⬅ Run extraction first in the **Extract & Preview** tab.")
        else:
            col_btn2, _ = st.columns([1, 3])
            with col_btn2:
                run_mapping = st.button(
                    "🗺️ Map & Write to Template",
                    type="primary",
                    use_container_width=True,
                )

            if run_mapping:
                _run_mapping_and_write(
                    template_file,
                    st.session_state.extraction_results,
                    st.session_state.template_model,
                    overwrite_existing=overwrite_mode,
                    use_llm=use_llm_mapping,
                )

            if st.session_state.write_result:
                _show_write_result()

    # ──────────────────────────────────────────────────────────────────────────
    # TAB 4 — Web Validation
    # ──────────────────────────────────────────────────────────────────────────
    with tab4:
        st.subheader("🌐 Web Cross-Validation")
        st.caption(
            "PDF-extracted values compared against Yahoo Finance (yfinance) "
            "and BSE quarterly data. Values that agree within 10% are marked "
            "HIGH confidence and are auto-write eligible."
        )

        val_results = st.session_state.get("validation_results")
        web_results  = st.session_state.get("web_results")

        if val_results is None and web_results is None:
            if st.session_state.get("company_identifier") is None:
                st.info(
                    "Enter an NSE symbol or BSE code in the sidebar to enable "
                    "web cross-validation."
                )
            else:
                st.info("Run extraction first to see cross-validation results.")
        else:
            # Summary metrics
            if val_results:
                from web.collector import WebCollector
                summary = WebCollector().validation_summary(val_results)
                c1, c2, c3, c4, c5 = st.columns(5)
                c1.metric("Total Fields",  summary["total"])
                c2.metric("✅ Agreed (≤10%)", summary["very_high"] + summary["high"])
                c3.metric("⚠️ Diverged",    summary["medium"])
                c4.metric("🚩 Flagged",     summary["flagged"])
                c5.metric("🔘 No Web Data", summary["no_web"])

                # Full validation table
                rows = []
                for r in val_results:
                    rows.append({
                        "Field":         r.template_field,
                        "Year":          r.year,
                        "PDF Value":     round(r.pdf_value, 2),
                        "Web Value":     round(r.web_value, 2) if r.web_value is not None else "—",
                        "Web Source":    r.web_source or "—",
                        "Diff %":        f"{r.tolerance_pct:.1%}" if r.web_value is not None else "—",
                        "Confidence":    r.confidence_upgrade,
                        "Flag":          r.flag_reason or "",
                    })
                df_val = pd.DataFrame(rows)

                # Colour-code by confidence
                def _color_conf(val):
                    colors = {
                        "VERY_HIGH": "background-color: #d4edda",
                        "HIGH":      "background-color: #d4edda",
                        "MEDIUM":    "background-color: #fff3cd",
                        "FLAG":      "background-color: #f8d7da",
                    }
                    return colors.get(str(val), "")

                filter_conf = st.selectbox(
                    "Filter by confidence",
                    ["All", "VERY_HIGH", "HIGH", "MEDIUM", "FLAG", "No Web Data"],
                )
                if filter_conf == "No Web Data":
                    df_show = df_val[df_val["Web Value"] == "—"]
                elif filter_conf != "All":
                    df_show = df_val[df_val["Confidence"] == filter_conf]
                else:
                    df_show = df_val

                st.dataframe(
                    df_show.style.applymap(_color_conf, subset=["Confidence"]),
                    use_container_width=True,
                    hide_index=True,
                )
                csv_val = df_val.to_csv(index=False)
                st.download_button(
                    "⬇️ Download Validation CSV",
                    data=csv_val,
                    file_name="web_validation.csv",
                    mime="text/csv",
                )

            # Raw web data
            if web_results:
                with st.expander(f"📡 Raw web data ({len(web_results)} data points)"):
                    rows_web = [
                        {
                            "Source":          w.source,
                            "Raw Field":       w.raw_field,
                            "Template Hint":   w.template_field,
                            "Year":            w.year,
                            "Value (INR Cr)":  round(w.value, 2),
                            "Confidence":      w.confidence,
                            "Unit Applied":    w.unit_applied,
                        }
                        for w in web_results
                    ]
                    st.dataframe(pd.DataFrame(rows_web), use_container_width=True, hide_index=True)

    # ──────────────────────────────────────────────────────────────────────────
    # TAB 5 — Audit Log
    # ──────────────────────────────────────────────────────────────────────────
    with tab5:
        st.subheader("Audit Log")
        if st.session_state.write_result:
            from integrator.excel_writer import audit_log_to_dataframe
            df = audit_log_to_dataframe(st.session_state.write_result.audit_log)
            if not df.empty:
                st.dataframe(df, use_container_width=True, hide_index=True)
                csv = df.to_csv(index=False)
                st.download_button(
                    "⬇️ Download Audit CSV",
                    data=csv,
                    file_name="extraction_audit_log.csv",
                    mime="text/csv",
                )
            else:
                st.info("No writes recorded.")

            if st.session_state.write_result.warnings:
                with st.expander(f"⚠️ {len(st.session_state.write_result.warnings)} warnings"):
                    for w in st.session_state.write_result.warnings:
                        st.warning(w)
        else:
            st.info("No audit log yet. Complete the mapping step first.")

        if st.session_state.log_messages:
            with st.expander("🔎 Full pipeline log"):
                st.code("\n".join(st.session_state.log_messages))


# ── Helper UI functions ────────────────────────────────────────────────────────

def _show_template_preview(template_file):
    """Show a compact preview of the template structure."""
    try:
        import openpyxl
        template_file.seek(0)
        wb = openpyxl.load_workbook(io.BytesIO(template_file.read()), data_only=True)
        with st.expander("📊 Template structure preview"):
            cols = st.columns(len(wb.sheetnames))
            for i, sheet_name in enumerate(wb.sheetnames):
                with cols[i]:
                    ws = wb[sheet_name]
                    st.markdown(f"**{sheet_name}**")
                    st.caption(f"{ws.max_row} rows × {ws.max_column} cols")
        template_file.seek(0)
    except Exception:
        pass


def _run_extraction(
    template_file,
    pdf_files,
    min_confidence: float,
    pdf_metadata: dict | None = None,
):
    """Execute the full extraction pipeline with real-time progress."""
    pdf_metadata = pdf_metadata or {}
    progress = st.progress(0, text="Initialising…")
    status = st.empty()
    detail = st.empty()

    try:
        # ── Step 1: Read template ─────────────────────────────────────────────
        status.info("📄 Step 1 / 3 — Reading template structure…")
        progress.progress(5, text="Reading template…")

        from mapper.template_reader import read_template
        import tempfile

        template_file.seek(0)
        template_bytes = template_file.read()

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(template_bytes)
            tmp_path = tmp.name

        template_model = read_template(tmp_path)
        os.unlink(tmp_path)
        st.session_state.template_model = template_model
        _log(f"Template read: {len(template_model.sheets)} sheets")
        detail.success(f"✓ Template: {len(template_model.sheets)} sheets loaded")
        progress.progress(15, text="Template read ✓")

        # ── Step 2: Parse PDFs ────────────────────────────────────────────────
        status.info("📑 Step 2 / 3 — Parsing PDFs…")
        from extractor.pdf_parser import parse_pdf

        parsed_docs = []
        for i, pdf_file in enumerate(pdf_files):
            meta = pdf_metadata.get(pdf_file.name, {})
            year_tag  = meta.get("year")    # e.g. "F2025" or None
            rtype_tag = meta.get("report_type", "")
            stmt_tag  = meta.get("statement_type")  # "consolidated" / "standalone" / None

            tag_str = f"FY={year_tag or 'auto'}, type={rtype_tag or 'auto'}, stmts={stmt_tag or 'auto'}"
            pct = 15 + int(((i + 0.5) / len(pdf_files)) * 25)
            progress.progress(pct, text=f"Parsing {pdf_file.name}…")
            detail.info(f"⏳ Parsing **{pdf_file.name}** [{tag_str}]…")

            pdf_file.seek(0)
            pdf_bytes = pdf_file.read()
            doc = parse_pdf(io.BytesIO(pdf_bytes))
            # Attach user-supplied metadata to the doc for downstream use
            doc._year_override    = year_tag
            doc._stmt_override    = stmt_tag
            doc._report_type      = rtype_tag
            doc._pdf_bytes        = pdf_bytes   # stored for pdfplumber table extraction
            parsed_docs.append(doc)

            n_sec = len(doc.sections)
            _log(f"Parsed '{pdf_file.name}': {n_sec} sections, {doc.total_pages} pages [{tag_str}]")
            detail.success(
                f"✓ **{pdf_file.name}**: {doc.total_pages} pages, "
                f"{n_sec} section(s) detected"
                + (f" · 📅 {year_tag}" if year_tag else "")
            )

        st.session_state.parsed_docs = parsed_docs
        progress.progress(40, text="PDFs parsed ✓")

        # ── Step 3: Extraction ────────────────────────────────────────────────────
        # Primary path: pdfplumber direct table extraction (no LLM for values).
        # pdfplumber reads actual column positions from the PDF, so year
        # assignment is structurally correct. Unit conversion is applied once
        # in Python. LLM is only used for label→template matching (not here).
        #
        # Fallback: if pdfplumber finds 0 values (e.g. scanned/image PDF),
        # we fall back to LLM extraction so you still get something.
        status.info("🔍 Step 3 / 3 — Extracting financial data from tables…")
        from extractor.llm_extractor import (
            extract_from_multiple_docs, extract_from_full_text,
            get_debug_responses, OpenAI, _merge_results,
        )
        from extractor.table_extractor import extract_tables_from_pdf

        live_api_key = os.environ.get("DEEPSEEK_API_KEY", "")
        live_base_url = os.environ.get("DEEPSEEK_BASE_URL", "https://api.deepseek.com")
        client = OpenAI(api_key=live_api_key, base_url=live_base_url) if live_api_key else None

        all_results = []
        extraction_errors: list[str] = []

        progress.progress(50, text="Reading PDF tables…")

        for doc in parsed_docs:
            stmt_type = getattr(doc, "_stmt_override", None) or "consolidated"
            pdf_bytes = getattr(doc, "_pdf_bytes", None)

            if pdf_bytes:
                # ── Primary: direct table extraction ────────────────────────
                detail.info(f"⏳ Extracting tables from {doc.filename}…")
                try:
                    tbl_results = extract_tables_from_pdf(
                        pdf_source=pdf_bytes,
                        parsed_doc=doc,
                        statement_type=stmt_type,
                    )
                except Exception as ex:
                    tbl_results = []
                    extraction_errors.append(f"{doc.filename} (table): {ex}")
                    _log(f"Table extraction failed for {doc.filename}: {ex}", "warning")

                n_tbl = sum(len(r.data) for r in tbl_results)
                _log(f"Table extraction: {doc.filename} → {n_tbl} values")

                if n_tbl > 0:
                    all_results.extend(tbl_results)
                    detail.success(
                        f"📊 {doc.filename}: {n_tbl} values extracted "
                        f"from PDF tables directly"
                    )
                    continue   # table extraction worked — skip LLM for this doc
            else:
                tbl_results = []
                n_tbl = 0

            # ── Fallback: LLM extraction (scanned PDFs, or table extraction failed) ──
            if doc.sections and client:
                detail.info(
                    f"⏳ Table extraction got {n_tbl} values — "
                    f"running LLM fallback for {doc.filename}…"
                )
                try:
                    llm_results, errs = extract_from_multiple_docs(
                        [doc], client=client, max_workers=1
                    )
                    all_results.extend(llm_results)
                    extraction_errors.extend(errs)
                    n_llm = sum(len(r.data) for r in llm_results)
                    detail.info(f"  LLM fallback: {n_llm} additional values")
                except Exception as ex:
                    extraction_errors.append(f"{doc.filename} (llm): {ex}")
            elif doc.sections:
                # No API key — rule-based only
                from extractor.llm_extractor import extract_all_sections
                rb_results = extract_all_sections(doc, client=None, rule_based_threshold=1.0)
                all_results.extend(rb_results)
                detail.info(f"  Rule-based fallback: {sum(len(r.data) for r in rb_results)} values")
            else:
                detail.warning(
                    f"⚠️ No sections and no table data for {doc.filename} — skipped"
                )

        all_results = _merge_results(all_results)
        progress.progress(85, text="Extraction complete…")

        # ── Apply year override: only fix truly wrong years ──────────────────────
        # The rule-based extractor already handles column order and assigns the
        # correct year to each column (current year + prior year).  We must NOT
        # blindly force every result to `override` — that destroys the F2024
        # prior-year column.  We only correct years that are completely outside
        # the expected range (e.g. LLM returns "F2020" for a F2025 document).
        if len(parsed_docs) == 1:
            override = getattr(parsed_docs[0], "_year_override", None)
            if override:
                try:
                    yr_int = int(override[1:])
                    valid_years = {override, f"F{yr_int - 1}"}  # current + prior year
                except (ValueError, IndexError):
                    valid_years = {override}
                for r in all_results:
                    if r.year not in valid_years:
                        r.year = override   # only fix genuinely wrong years

        st.session_state.extraction_results = all_results
        # Store parsed docs for the section-text debug panel below
        st.session_state.parsed_docs = parsed_docs

        total_fields = sum(len(r.data) for r in all_results)
        years = sorted({r.year for r in all_results})
        _log(f"Extraction complete: {len(all_results)} blocks, {total_fields} fields, years={years}")

        # ── Surface the detected unit conversion so the user can verify ──────
        # Scan all result notes for unit_divisor stamps.
        divisors_seen: set[str] = set()
        for r in all_results:
            notes = getattr(r, "notes", "") or ""
            import re as _re
            m = _re.search(r"unit_divisor=([\d.]+)", notes)
            if m:
                divisors_seen.add(m.group(1))
        if divisors_seen:
            _UNIT_NAMES = {
                "1.0":     "INR Crores (no conversion needed)",
                "100.0":   "INR Lakhs ÷ 100 → Crores",
                "10.0":    "INR Millions ÷ 10 → Crores",
                "10000.0": "INR Thousands ÷ 10,000 → Crores",
            }
            unit_strs = [_UNIT_NAMES.get(d, f"÷{d}") for d in sorted(divisors_seen)]
            detail.info(f"📐 Unit detected: {' | '.join(unit_strs)}")

        # ── Web cross-validation (optional — only if company identifier given) ─
        company_id = st.session_state.get("company_identifier")
        if company_id and all_results and years:
            status.info("🌐 Running web cross-validation…")
            progress.progress(90, text="Fetching web data…")
            try:
                from web.collector import WebCollector
                collector = WebCollector()
                web_results = collector.collect(company=company_id, years=years)
                st.session_state.web_results = web_results

                if web_results:
                    # Cross-validate PDF values against web
                    validation = collector.cross_validate(all_results, web_results)
                    st.session_state.validation_results = validation
                    summary = collector.validation_summary(validation)

                    # Fill missing cells from web (primary source for PDF gaps)
                    template_model = st.session_state.get("template_model")
                    if template_model:
                        web_fill = collector.fill_missing_from_web(
                            all_results, web_results, template_model
                        )
                        if web_fill:
                            all_results.extend(web_fill)
                            n_fill = sum(len(r.data) for r in web_fill)
                            _log(f"Web gap-fill: {n_fill} additional field(s) from web source")
                            detail.success(
                                f"🌐 Web: "
                                f"✅ {summary['very_high'] + summary['high']} confirmed · "
                                f"⚠️ {summary['flagged']} flagged · "
                                f"➕ {n_fill} gaps filled from web"
                            )
                        else:
                            detail.success(
                                f"🌐 Web: "
                                f"✅ {summary['very_high'] + summary['high']} confirmed · "
                                f"⚠️ {summary['flagged']} flagged · "
                                f"🔘 {summary['no_web']} no web data"
                            )
                    _log(
                        f"Web validation: {summary['total']} fields | "
                        f"VERY_HIGH={summary['very_high']} HIGH={summary['high']} "
                        f"MEDIUM={summary['medium']} FLAG={summary['flagged']}"
                    )
                    # Update session state with the extended results
                    st.session_state.extraction_results = all_results
                else:
                    detail.info("🌐 Web validation: no data retrieved (check ticker / BSE code)")
                    st.session_state.validation_results = None
            except Exception as web_err:
                _log(f"Web validation failed: {web_err}", level="warning")
                detail.warning(f"⚠️ Web validation skipped: {web_err}")
                st.session_state.web_results = None
                st.session_state.validation_results = None
        else:
            st.session_state.web_results = None
            st.session_state.validation_results = None

        progress.progress(100, text="Done ✓")

        if all_results:
            unit_tag = ""
            if divisors_seen:
                _UNIT_SHORT = {"1.0": "Crores", "100.0": "Lakhs→Cr",
                               "10.0": "Mn→Cr", "10000.0": "000s→Cr"}
                unit_tag = " · 📐 " + " | ".join(
                    _UNIT_SHORT.get(d, f"÷{d}") for d in sorted(divisors_seen)
                )
            status.success(
                f"✅ Done! **{len(all_results)}** result blocks · "
                f"**{total_fields}** fields · years: {', '.join(years)}"
                + unit_tag
                + (" · ⚡ rule-based (no LLM used)" if not live_api_key else "")
            )
            detail.empty()
        else:
            status.warning("⚠️ No data extracted.")

            if extraction_errors:
                st.error("**Extraction errors:**\n" + "\n".join(f"- {e}" for e in extraction_errors))

            debug_responses = get_debug_responses()
            if debug_responses:
                with st.expander("🔎 Raw LLM responses — click to diagnose", expanded=True):
                    st.caption(
                        "DeepSeek returned the text below. "
                        "If you see a JSON array, there may be a parsing issue. "
                        "If you see an error message, check your API key / quota."
                    )
                    for i, entry in enumerate(debug_responses):
                        st.markdown(f"**Response {i+1}** — attempt {entry['attempt']}, {entry['chars']} chars")
                        st.code(entry["snippet"], language="text")
            else:
                st.info(
                    "No LLM response was captured. "
                    "Verify your DeepSeek API key in the sidebar, or the PDF may not contain "
                    "recognisable financial tables."
                )

    except Exception as e:
        tb = traceback.format_exc()
        status.error(f"❌ Extraction failed: {e}")
        with st.expander("Error details"):
            st.code(tb)
        _log(f"EXTRACTION ERROR: {e}\n{tb}", "error")


def _show_extraction_preview(results):
    """Show a structured preview of extracted data."""
    st.markdown("---")
    st.markdown("### 🔍 Extracted Data Preview")

    if not results:
        st.warning("No data extracted. Check your PDFs and API key.")
        return

    # Group by section + statement_type + year
    from collections import defaultdict
    grouped = defaultdict(list)
    for r in results:
        key = f"{r.section} · {r.statement_type.title()} · {r.year}"
        grouped[key].append(r)

    cols = st.columns(min(3, len(grouped)))
    for i, (label, group_results) in enumerate(sorted(grouped.items())):
        col = cols[i % len(cols)]
        with col:
            with st.expander(f"📋 {label}", expanded=i < 3):
                combined_data = {}
                for r in group_results:
                    combined_data.update(r.data)
                if combined_data:
                    df = pd.DataFrame(
                        list(combined_data.items()),
                        columns=["Field", "Value (INR Cr)"]
                    )
                    df["Value (INR Cr)"] = df["Value (INR Cr)"].apply(
                        lambda x: f"{x:,.2f}" if isinstance(x, (int, float)) else x
                    )
                    st.dataframe(df, hide_index=True, use_container_width=True)
                else:
                    st.info("No data")

    # ── Section text debug panel ──────────────────────────────────────────────
    # Shows the raw text the PDF parser found for each section so you can verify
    # the right pages were picked up (and spot note-vs-main-BS confusion, etc.)
    parsed_docs = st.session_state.get("parsed_docs", [])
    if parsed_docs:
        with st.expander("🔬 PDF section text — click to debug extraction", expanded=False):
            st.caption(
                "This shows the raw text extracted from your PDF for each financial section. "
                "If the Balance Sheet text looks like Notes to Accounts (share counts, "
                "detailed schedules) rather than the main statement, the PDF parser found "
                "the wrong pages — share that text here so the scoring can be adjusted."
            )
            for doc in parsed_docs:
                st.markdown(f"**{doc.filename}** — {doc.total_pages} pages, "
                            f"{len(doc.sections)} section(s) detected")
                if not doc.sections:
                    st.warning("No sections detected in this PDF.")
                    continue
                for sec in doc.sections:
                    label = (f"{sec.section_type} · "
                             f"{'Consolidated' if sec.is_consolidated else ''}"
                             f"{'Standalone' if sec.is_standalone else ''} · "
                             f"conf={sec.confidence:.0%} · "
                             f"pages {sec.page_numbers[:5]}"
                             f"{'…' if len(sec.page_numbers) > 5 else ''}")
                    with st.expander(label, expanded=False):
                        # Show first 3000 chars — enough to see if it's the right content
                        preview = sec.text[:3000]
                        st.code(preview, language="text")
                        if len(sec.text) > 3000:
                            st.caption(f"… (showing 3000 / {len(sec.text):,} chars)")


def _run_mapping_and_write(
    template_file,
    extraction_results,
    template_model,
    overwrite_existing: bool,
    use_llm: bool,
):
    """Execute field mapping (with web validation stamping) and write to template."""
    progress = st.progress(0, text="Starting mapping…")
    status = st.empty()

    try:
        # Step 1: Map fields
        status.info("🗺️ Mapping extracted fields to template rows…")
        progress.progress(10, text="Field mapping…")

        from mapper.field_mapper import FieldMapper, stamp_validation, map_direct

        # Detect whether template-guided extraction was used.
        # Template-guided results carry notes="mode=template_guided" and their
        # keys are already template labels, so we skip fuzzy/LLM mapping entirely.
        is_template_guided = any(
            getattr(r, "notes", "") == "mode=template_guided"
            for r in extraction_results
        )

        if is_template_guided:
            status.info("🗺️ Template-guided mode — direct label matching (no fuzzy mapping)…")
            reports = map_direct(extraction_results, template_model)
        else:
            status.info("🗺️ Mapping extracted fields to template rows…")
            mapper = FieldMapper(template_model, use_llm=use_llm)
            reports = mapper.map_results(extraction_results)

        st.session_state.mapping_reports = reports
        progress.progress(50, text="Mapping complete…")

        total_mapped = sum(len(r.mapped) for r in reports)
        total_unmapped = sum(len(r.unmapped) for r in reports)
        mode_tag = " [template-guided]" if is_template_guided else ""
        _log(f"Mapped {total_mapped} fields, {total_unmapped} unmapped{mode_tag}")

        # Step 2: Stamp web validation confidence onto mapped fields
        validation_results = st.session_state.get("validation_results")
        web_validated = False
        if validation_results:
            status.info("🌐 Applying web validation confidence to mapped fields…")
            progress.progress(65, text="Stamping web confidence…")
            stamp_validation(reports, validation_results)
            web_validated = True
            flagged = sum(
                1 for r in reports
                for mf in r.mapped
                if mf.confidence_tier == "FLAG"
            )
            confirmed = sum(
                1 for r in reports
                for mf in r.mapped
                if mf.confidence_tier in ("VERY_HIGH", "HIGH")
            )
            _log(
                f"Web validation stamped: {confirmed} confirmed (HIGH+), "
                f"{flagged} flagged for review"
            )
            status.info(
                f"🌐 Web validation: ✅ {confirmed} confirmed · "
                f"🚩 {flagged} flagged (→ Data Review sheet)"
            )

        # Step 3: Write to template
        status.info("✍️ Writing to template (non-destructive)…")
        progress.progress(75, text="Writing cells…")

        from integrator.excel_writer import write_to_bytes
        template_file.seek(0)
        template_bytes = template_file.read()

        output_bytes, write_result = write_to_bytes(
            template_bytes=template_bytes,
            mapping_reports=reports,
            overwrite_existing=overwrite_existing,
            web_validated=web_validated,
        )

        st.session_state.write_result = write_result
        st.session_state.output_bytes = output_bytes

        progress.progress(100, text="Done ✓")

        review_msg = (
            f" · 🚩 **{write_result.cells_review_queued}** flagged → Data Review sheet"
            if write_result.cells_review_queued else ""
        )
        status.success(
            f"✅ Template populated: "
            f"**{write_result.cells_written}** cells written"
            f"{review_msg} | "
            f"{write_result.cells_skipped_formula} formulas preserved | "
            f"{write_result.cells_skipped_filled} existing values kept"
        )
        _log(
            f"Write complete: {write_result.cells_written} cells written, "
            f"{write_result.cells_review_queued} queued for review"
        )

    except Exception as e:
        tb = traceback.format_exc()
        status.error(f"❌ Mapping/write failed: {e}")
        with st.expander("Error details"):
            st.code(tb)
        _log(f"MAPPING ERROR: {e}\n{tb}", "error")


def _show_write_result():
    """Show the write result with download button."""
    wr = st.session_state.write_result

    st.markdown("---")
    st.markdown("### 📥 Download Populated Template")

    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Cells Written", wr.cells_written)
    with col2:
        st.metric("Formulas Preserved", wr.cells_skipped_formula)
    with col3:
        st.metric("Existing Values Kept", wr.cells_skipped_filled)
    with col4:
        st.metric("Warnings", len(wr.warnings))

    st.markdown("---")
    st.download_button(
        label="⬇️ Download Populated Excel Template",
        data=st.session_state.output_bytes,
        file_name="PPL_Template_Populated.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
    )

    if st.session_state.mapping_reports:
        st.markdown("#### 🗺️ Field Mapping Summary")
        rows = []
        for report in st.session_state.mapping_reports:
            for mf in report.mapped:
                rows.append({
                    "Sheet": mf.sheet_name,
                    "Template Field": mf.template_label,
                    "Extracted As": mf.extracted_label,
                    "Year": mf.year,
                    "Type": mf.statement_type.title(),
                    "Value (INR Cr)": f"{mf.value:,.2f}",
                    "Match": mf.match_method,
                    "Score": f"{mf.match_score:.0%}",
                })
        if rows:
            st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)


if __name__ == "__main__":
    main()
