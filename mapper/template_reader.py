"""
Template Reader — parses the Excel template structure into a queryable model.

Builds a TemplateModel that knows:
  - Which sheets exist
  - For each sheet: which row contains which label
  - For each sheet: which column corresponds to which year/period
  - Which cells already have data (to prevent overwrite)
  - Which cells contain formulas (never overwrite)
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional, Union

import openpyxl
from openpyxl.cell import Cell

from utils.helpers import get_logger, clean_label, normalise_year

logger = get_logger(__name__)


@dataclass
class CellAddress:
    sheet: str
    row: int
    col: int

    @property
    def excel_ref(self) -> str:
        from openpyxl.utils import get_column_letter
        return f"{get_column_letter(self.col)}{self.row}"


@dataclass
class SheetModel:
    """Parsed structure of a single template sheet."""
    name: str
    header_row: int                              # row number of the column-header row
    label_col: int                               # column containing row labels (usually 1)
    year_cols: dict[str, int]                    # year label → column number
    row_index: dict[str, list[int]]              # cleaned label → list of ALL row numbers (handles duplicates across sections)
    formula_cells: set[tuple[int, int]]          # (row, col) of formula cells
    filled_cells: set[tuple[int, int]]           # (row, col) of non-empty non-formula cells
    # For sheets with Consolidated / Standalone sections
    section_offsets: dict[str, int] = field(default_factory=dict)  # "consolidated" | "standalone" → start row


@dataclass
class TemplateModel:
    """Complete parsed model of the Excel template."""
    filepath: str
    sheets: dict[str, SheetModel] = field(default_factory=dict)

    def get_cell_address(
        self,
        sheet_name: str,
        row_label: str,
        year: str,
        statement_type: str = "consolidated",
    ) -> Optional[CellAddress]:
        """
        Return the CellAddress for a given (sheet, row label, year, statement type).
        Returns None if the mapping cannot be resolved.
        """
        sheet = self.sheets.get(sheet_name)
        if not sheet:
            return None

        col = sheet.year_cols.get(year)
        if not col:
            return None

        # Find the correct row, considering consolidated / standalone sections
        row = _resolve_row(sheet, row_label, statement_type)
        if not row:
            return None

        return CellAddress(sheet=sheet_name, row=row, col=col)

    def is_cell_writable(self, sheet_name: str, row: int, col: int) -> bool:
        """Return True if the cell can be safely written (not a formula, not pre-filled)."""
        sheet = self.sheets.get(sheet_name)
        if not sheet:
            return False
        key = (row, col)
        if key in sheet.formula_cells:
            return False
        if key in sheet.filled_cells:
            return False
        return True

    def all_year_columns(self, sheet_name: str) -> list[str]:
        sheet = self.sheets.get(sheet_name)
        return sorted(sheet.year_cols.keys()) if sheet else []

    def all_labels(self, sheet_name: str) -> list[str]:
        sheet = self.sheets.get(sheet_name)
        return list(sheet.row_index.keys()) if sheet else []


# ── Row resolution (handles Consolidated / Standalone sub-sections) ────────────

def _resolve_row(sheet: SheetModel, label: str, statement_type: str) -> Optional[int]:
    """
    Find the row number for a label, respecting consolidated/standalone offsets.
    row_index maps label → list[int] to handle the same label appearing in
    both consolidated and standalone sections.
    """
    cleaned = clean_label(label).lower()

    # Determine section boundaries
    section_start = 1
    section_end = 99999
    if sheet.section_offsets:
        section_key = "standalone" if statement_type == "standalone" else "consolidated"
        section_start = sheet.section_offsets.get(section_key, 1)
        other_starts = [v for k, v in sheet.section_offsets.items() if v > section_start]
        section_end = min(other_starts) if other_starts else 99999

    # Search for label within the correct section
    for raw_label, row_nums in sheet.row_index.items():
        if _label_matches(raw_label, cleaned):
            # Pick the row that falls in the correct section
            for rn in row_nums:
                if section_start <= rn < section_end:
                    return rn
            # Fallback: return the first matching row
            return row_nums[0]

    return None


def _label_matches(template_label: str, query_label: str) -> bool:
    """Flexible label matching: exact after cleaning."""
    tl = clean_label(template_label).lower()
    return tl == query_label


# ── Main parser ────────────────────────────────────────────────────────────────

_ANNUAL_YEAR_PATTERN = re.compile(r"^[Ff]\d{4}$")
_QUARTERLY_PATTERN  = re.compile(r"^[1-4]QF\d{4}$")

# Additional year formats that some templates use — normalised to F20XX
_ALT_YEAR_PATTERNS: list[tuple[re.Pattern, str]] = [
    # "FY2025" / "FY 2025" / "FY25"
    (re.compile(r"^FY\s*(\d{2,4})$", re.IGNORECASE), "fy"),
    # "2024-25" / "2024-2025" / "2024/25"
    (re.compile(r"^(\d{4})[-/](\d{2,4})$"), "range"),
    # Plain "2025" (bare 4-digit year)
    (re.compile(r"^(20\d{2})$"), "bare"),
]

def _normalise_alt_year(val: str) -> Optional[str]:
    """
    Convert alternative year labels to F20XX format.
    Returns None if unrecognised.
    """
    from utils.helpers import normalise_year
    v = val.strip()
    # Already standard
    if _ANNUAL_YEAR_PATTERN.match(v):
        return v
    for pat, kind in _ALT_YEAR_PATTERNS:
        m = pat.match(v)
        if not m:
            continue
        if kind == "fy":
            raw = m.group(1)
            # "25" → 2025, "2025" → 2025
            yr = int(raw) + 2000 if len(raw) == 2 else int(raw)
            return f"F{yr}"
        elif kind == "range":
            first, second = int(m.group(1)), m.group(2)
            # "2024-25" → F2025 (Indian fiscal: ends in the second year)
            end = int(second) if len(second) == 4 else int(first // 100 * 100 + int(second))
            return f"F{end}"
        elif kind == "bare":
            return f"F{m.group(1)}"
    return None

# Rows to skip when building the label index (section headers / meta rows)
_SKIP_LABELS = {
    "consolidated", "standalone", "particulars", "inr crs.",
    "less:", "financial assets:", "assets", "liabilities",
    "equity and liabilities",
    # Note: "non-current assets", "current assets", "non-current liabilities",
    # "current liabilities", "equity" are intentionally NOT here — they are
    # detected as BS subsection boundary markers below.
    "financial liabilities:",
    "income statement - consolidated", "income statement - standalone",
}

# Balance-sheet subsection boundary markers — these rows aren't data rows but
# tell us which sub-section we're currently inside (so we can qualify duplicate
# labels like "Borrowings" that appear in both Non-Current and Current sections).
_BS_SUBSECTION_MARKERS: dict[str, str] = {
    "non-current assets":    "non_current_assets",
    "current assets":        "current_assets",
    "equity":                "equity",
    "non-current liabilities": "non_current_liabilities",
    "current liabilities":   "current_liabilities",
}

# Labels that appear in multiple BS subsections and need a qualifier so the
# mapper can route them to the right row.  Keys are subsection names; values
# map the lower-cased cleaned label → canonical qualified name.
_BS_QUALIFIED_LABELS: dict[str, dict[str, str]] = {
    "non_current_liabilities": {
        "borrowings":        "Borrowings (Non-Current)",
        "lease liabilities": "Lease Liabilities (Non-Current)",
        "provisions":        "Provisions (Non-Current)",
    },
    "current_liabilities": {
        "borrowings":        "Borrowings (Current)",
        "lease liabilities": "Lease Liabilities (Current)",
        "provisions":        "Provisions (Current)",
    },
}


def read_template(filepath: Union[str, Path]) -> TemplateModel:
    """
    Load and parse the Excel template into a TemplateModel.
    """
    path = str(filepath)
    logger.info(f"Reading template: {path}")

    wb = openpyxl.load_workbook(path, data_only=False)   # keep formula strings
    wb_data = openpyxl.load_workbook(path, data_only=True)  # values for filled-cell detection

    model = TemplateModel(filepath=path)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws_data = wb_data[sheet_name]

        sheet_model = _parse_sheet(ws, ws_data, sheet_name)
        model.sheets[sheet_name] = sheet_model
        logger.info(
            f"  Sheet '{sheet_name}': "
            f"{len(sheet_model.year_cols)} year cols, "
            f"{len(sheet_model.row_index)} labelled rows, "
            f"{len(sheet_model.formula_cells)} formula cells, "
            f"{len(sheet_model.filled_cells)} filled cells"
        )

    return model


def _parse_sheet(ws, ws_data, sheet_name: str) -> SheetModel:
    """Parse a single worksheet into a SheetModel."""

    # ── Step 1: Find the header row (contains year labels like F2025) ──────────
    header_row = _find_header_row(ws)
    label_col = 1   # Almost always column A

    # ── Step 2: Build year→column mapping ─────────────────────────────────────
    year_cols: dict[str, int] = {}
    if header_row:
        for col in range(1, ws.max_column + 1):
            val = ws.cell(header_row, col).value
            if val is None:
                continue
            val_str = str(val).strip()
            # Annual year columns: F2022, F2023, F2024, F2025, F2026
            # Keep the FIRST occurrence only — some sheets repeat year headers in a
            # "Common Size" block to the right of the data block, and we must not
            # overwrite the data column with that secondary column.
            if _ANNUAL_YEAR_PATTERN.match(val_str):
                if val_str not in year_cols:
                    year_cols[val_str] = col
            # Quarterly columns: 1QF2025, etc.
            elif _QUARTERLY_PATTERN.match(val_str):
                if val_str not in year_cols:
                    year_cols[val_str] = col
            else:
                # Try alternative formats (FY25, 2024-25, 2025, etc.)
                normalised = _normalise_alt_year(val_str)
                if normalised and normalised not in year_cols:
                    year_cols[normalised] = col

    # ── Step 3: Build label→row mapping ───────────────────────────────────────
    # row_index maps label → list[int] to support duplicate labels across sections
    row_index: dict[str, list[int]] = {}
    section_offsets: dict[str, int] = {}

    # Track the current Balance Sheet subsection (non_current_assets, etc.) so
    # we can qualify duplicate labels like "Borrowings" with (Non-Current)/(Current).
    current_bs_subsection: Optional[str] = None

    # Rows at or before the header row are structural (title, units, column headers)
    # and must never receive extracted data.
    data_start_row = (header_row + 1) if header_row else 1

    for row in range(1, ws.max_row + 1):
        cell_val = ws.cell(row, label_col).value
        if cell_val is None:
            continue
        label = str(cell_val).strip()
        if not label:
            continue

        # Skip every row at or above the header row — these are title / units rows
        if row <= (header_row or 0):
            continue

        label_lower = label.lower().strip()

        # Detect consolidated / standalone section boundaries
        if label_lower == "consolidated":
            section_offsets["consolidated"] = row
            current_bs_subsection = None   # reset BS subsection on new section
            continue
        if label_lower == "standalone":
            section_offsets["standalone"] = row
            current_bs_subsection = None
            continue

        # Detect Balance Sheet sub-section boundary markers
        if label_lower in _BS_SUBSECTION_MARKERS:
            current_bs_subsection = _BS_SUBSECTION_MARKERS[label_lower]
            continue   # boundary marker — no data value, don't add to row_index

        # Skip meta rows
        if label_lower in _SKIP_LABELS:
            continue

        # Build the canonical label for this row
        cleaned = clean_label(label)
        if not cleaned:
            continue

        # For duplicate BS labels (Borrowings, Lease Liabilities, Provisions),
        # qualify the name based on which subsection we're currently inside so
        # the mapper can route Non-Current vs Current correctly.
        if current_bs_subsection and current_bs_subsection in _BS_QUALIFIED_LABELS:
            qualified = _BS_QUALIFIED_LABELS[current_bs_subsection].get(cleaned.lower())
            if qualified:
                cleaned = qualified

        # Add to index — append to list to handle duplicates across sections
        if cleaned not in row_index:
            row_index[cleaned] = []
        row_index[cleaned].append(row)

    # ── Step 4: Identify formula & pre-filled cells ────────────────────────────
    formula_cells: set[tuple[int, int]] = set()
    filled_cells: set[tuple[int, int]] = set()

    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            raw_cell: Cell = ws.cell(row, col)
            data_cell: Cell = ws_data.cell(row, col)

            raw_val = raw_cell.value
            data_val = data_cell.value

            if raw_val is None:
                continue

            if isinstance(raw_val, str) and raw_val.startswith("="):
                formula_cells.add((row, col))
            elif data_val is not None and data_val != "" and col != label_col:
                filled_cells.add((row, col))

    return SheetModel(
        name=sheet_name,
        header_row=header_row or 4,
        label_col=label_col,
        year_cols=year_cols,
        row_index=row_index,
        formula_cells=formula_cells,
        filled_cells=filled_cells,
        section_offsets=section_offsets,
    )


def _find_header_row(ws) -> Optional[int]:
    """
    Scan rows to find the one containing year headers like F2022, F2023, etc.
    Also recognises alternative formats: FY25, 2024-25, 2025, etc.
    Returns the row number (1-based) or None.
    """
    def _is_year_cell(v: str) -> bool:
        return bool(
            _ANNUAL_YEAR_PATTERN.match(v)
            or _QUARTERLY_PATTERN.match(v)
            or _normalise_alt_year(v) is not None
        )

    for row in range(1, min(15, ws.max_row + 1)):
        row_vals = [str(ws.cell(row, c).value or "").strip() for c in range(1, ws.max_column + 1)]
        year_count = sum(1 for v in row_vals if _is_year_cell(v))
        if year_count >= 2:
            return row
    return None
