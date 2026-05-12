"""
pipeline/llm_fill.py
────────────────────
Single-shot LLM-driven template filling. The cleaner replacement for the
deterministic Smart Pipeline.

Flow:
  1. Read template structure (sheets, rows, sections, years)
  2. Extract text from each user-uploaded PDF
  3. Classify each PDF (Standalone / Consolidated / IP / AR / Quarterly)
  4. Build a single big prompt with the manual extraction algorithm
  5. Send to DeepSeek with response_format=json
  6. Parse JSON of fills
  7. Write Excel with confidence-color coding

Skips: source priority, KB canonical matching, fuzzy resolver, multi-source
consensus, tool-calling agent. The LLM does it all.
"""

from __future__ import annotations

import json
import logging
import re
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Optional

logger = logging.getLogger(__name__)


# Cache version stamp — bump this whenever the extraction logic changes so
# old .md sidecars get regenerated on the next run. Without this, the user
# would have to manually delete every cached .md after a fix.
EXTRACTION_VERSION = "v19-pdfplumber-primary-pymupdf-optional"


# ──────────────────────────────────────────────────────────────────────────────
# PDF classification — categorise each input PDF by reading its first 1-2 pages
# ──────────────────────────────────────────────────────────────────────────────

@dataclass
class PDFClassification:
    path:        str
    name:        str
    stmt_type:   str   # "STANDALONE" | "CONSOLIDATED" | "BOTH" | "UNKNOWN"
    period_type: str   # "ANNUAL" | "QUARTERLY" | "INVESTOR_PRESENTATION" | "UNKNOWN"
    title_text:  str
    pages_total: int
    text_preview: str = ""    # first ~500 chars for prompt context
    toc:         list = None  # [(page_num, section_title, scope), ...]


def _build_toc(md_by_page: dict, std_start: Optional[int],
                cons_start: Optional[int], fin_end: Optional[int]) -> list:
    """Build a table-of-contents per PDF. For each page, detect section
    headings (Balance Sheet, P&L, Cash Flow, Notes, etc.) and emit a
    (page_num, section_title, scope) tuple. The LLM uses this TOC to
    navigate the full PDF content."""
    toc = []
    SECTION_PATTERNS = [
        (r"(?:report on the audit of the |independent auditor['']?s? report)\s*(?:.{0,80}(standalone|consolidated))?",
         "Auditor's Report"),
        (r"(standalone|consolidated)?\s*balance sheet", "Balance Sheet"),
        (r"(standalone|consolidated)?\s*statement of (?:profit|financial position)", "Statement of P&L"),
        (r"(standalone|consolidated)?\s*(?:statement of )?cash flow", "Cash Flow"),
        (r"(standalone|consolidated)?\s*statement of changes in equity", "Changes in Equity"),
        (r"notes to (?:the )?(standalone|consolidated)? ?financial statements?", "Notes"),
        (r"property,?\s+plant\s+(?:&|and)\s+equipment", "PPE schedule"),
        (r"investments?\b", "Investments schedule"),
        (r"borrowings?\b", "Borrowings schedule"),
        (r"trade receivables?\b", "Trade Receivables schedule"),
        (r"revenue from operations?\b", "Revenue note"),
        (r"employee benefits?\s+expense", "Employee Benefits note"),
        (r"other expenses?\b", "Other Expenses note"),
        (r"finance costs?\b", "Finance Costs note"),
        (r"tax expense\b|income tax expense\b", "Tax note"),
        (r"segment (?:reporting|information)", "Segment Reporting"),
        (r"contingent liabilities?", "Contingent Liabilities"),
        (r"related party transactions?", "Related Party"),
        (r"management discussion (?:and|&) analysis|md&a|mda\b", "MD&A"),
        (r"director['']?s? report", "Director's Report"),
        (r"corporate governance", "Corporate Governance"),
        (r"notice of (?:the )?annual general", "Notice of AGM"),
        (r"sustainability|brsr", "Sustainability/BRSR"),
        (r"chairman['']?s? (?:message|letter|statement)", "Chairman's Statement"),
    ]
    SECTION_RES = [(re.compile(p, re.I), t) for p, t in SECTION_PATTERNS]

    for pn in sorted(md_by_page.keys()):
        chunk = md_by_page[pn][:1500].lower()
        if not chunk.strip():
            continue
        # Pick the FIRST section pattern that matches the page head
        for rx, title in SECTION_RES:
            m = rx.search(chunk)
            if m:
                # If the pattern captured a scope group, include it
                scope_prefix = ""
                try:
                    g = m.group(1)
                    if g and g.lower() in ("standalone", "consolidated"):
                        scope_prefix = g.capitalize() + " "
                except IndexError:
                    pass
                # Decide overall scope from boundary indices
                if std_start is not None and cons_start is not None:
                    if pn < std_start:
                        ov_scope = "MD&A/Pre"
                    elif std_start <= pn < cons_start:
                        ov_scope = "STANDALONE"
                    elif fin_end and pn >= fin_end:
                        ov_scope = "Post-Financials"
                    else:
                        ov_scope = "CONSOLIDATED"
                else:
                    ov_scope = "—"
                toc.append((pn + 1, f"{scope_prefix}{title}", ov_scope))
                break
    return toc


def _table_to_markdown(table: list[list]) -> str:
    """Format a 2D list (rows × cells) as a GitHub-flavored markdown table."""
    if not table or not table[0]:
        return ""
    # Sanitize cells: replace None with "", strip newlines, collapse whitespace
    def _cell(c) -> str:
        if c is None: return ""
        s = str(c).replace("\n", " ").replace("|", "/").strip()
        return " ".join(s.split())  # collapse multiple spaces

    rows = [[_cell(c) for c in row] for row in table]
    n_cols = max(len(r) for r in rows)
    rows = [r + [""] * (n_cols - len(r)) for r in rows]   # pad short rows
    header, *body = rows
    md = "| " + " | ".join(header) + " |\n"
    md += "|" + "|".join(["---"] * n_cols) + "|\n"
    for row in body:
        md += "| " + " | ".join(row) + " |\n"
    return md


def _pdf_to_markdown(pdf_path: str, force_ocr: bool = False) -> str:
    """
    Convert a PDF to markdown — tables as MD tables, prose as plain text,
    OCR fallback for image-only pages. Result is cached as <pdf_path>.md
    so subsequent runs read the cache directly (skip extraction entirely).
    """
    # Try cache in two locations: next to PDF (preferred) OR in a temp dir
    # if the PDF folder is read-only (common with system-mounted uploads).
    import tempfile, hashlib
    cache_paths = [
        Path(pdf_path).with_suffix(Path(pdf_path).suffix + ".md"),
        Path(tempfile.gettempdir()) / (
            "llm_fill_md_cache_" +
            hashlib.sha1(pdf_path.encode()).hexdigest()[:16] + ".md"
        ),
    ]
    version_stamp = f"<!-- extractor:{EXTRACTION_VERSION} -->"
    for cache in cache_paths:
        if cache.exists() and not force_ocr:
            try:
                cached = cache.read_text(encoding="utf-8")
                if len(cached) > 100 and version_stamp in cached[:200]:
                    logger.info(f"_pdf_to_markdown: cache hit {cache.name} "
                                f"({len(cached)//1000}k chars, {EXTRACTION_VERSION})")
                    return cached
                elif len(cached) > 100:
                    logger.info(f"_pdf_to_markdown: cache stale (old version) "
                                f"{cache.name} — re-extracting")
            except Exception:
                pass

    # ── pymupdf4llm: opt-in only via env var ─────────────────────────────
    # The pymupdf4llm output structure broke the table-row label parser
    # (labels ended up None, values shifted columns). Default OFF until
    # the table-row format is normalized. Set USE_PYMUPDF4LLM=1 to test.
    import os as _os
    if _os.environ.get("USE_PYMUPDF4LLM") == "1" and not force_ocr:
        try:
            import pymupdf4llm
            page_chunks = pymupdf4llm.to_markdown(
                pdf_path, page_chunks=True, write_images=False,
                show_progress=False,
            )
            if page_chunks:
                full_parts = []
                for chunk in page_chunks:
                    pn_meta = chunk.get("metadata", {}).get("page", 0)
                    txt = chunk.get("text", "")
                    if txt.strip():
                        full_parts.append(
                            f"\n## Page {pn_meta + 1}\n"
                            f"<!-- FILE: {Path(pdf_path).name} | "
                            f"PAGE: {pn_meta + 1}/{len(page_chunks)} -->\n"
                            f"{txt}\n"
                        )
                if full_parts:
                    full = (f"<!-- extractor:{EXTRACTION_VERSION} -->\n"
                            + "\n".join(full_parts))
                    for cache in cache_paths:
                        try:
                            cache.write_text(full, encoding="utf-8")
                            break
                        except Exception:
                            continue
                    return full
        except Exception as exc:
            logger.warning(f"pymupdf4llm failed ({exc}); using pdfplumber")

    # ── PRIMARY: pdfplumber + multi-col reconstructor + OCR ──────────────
    import pdfplumber
    parts: list[str] = []
    ocr_pages: list[int] = []
    fitz_pages: list[int] = []   # pages where pdfplumber tables were sparse

    def _is_garbled(text: str) -> bool:
        """
        Detect column-interleaved garbled extraction. Multi-column slide
        layouts read row-by-row produce text where each "row" picks one
        character from each column (e.g. "O b s e r v e d" actually = 4
        columns interleaved). Heuristic: if more than 35% of whitespace-
        split tokens are single characters, treat as garbled and re-OCR.
        """
        if not text or len(text) < 100:
            return False
        words = text.split()
        if len(words) < 30:
            return False
        single_chars = sum(1 for w in words if len(w) == 1 and w.isalpha())
        return single_chars / len(words) > 0.35

    pdf_basename = Path(pdf_path).name
    try:
        with pdfplumber.open(pdf_path) as pdf:
            n = len(pdf.pages)
            for i in range(n):
                page = pdf.pages[i]
                # Explicit per-page header: filename + page number, so the
                # LLM (and pattern matcher) always know which doc + page each
                # value came from. Required when many docs are stitched.
                page_md = (
                    f"\n## Page {i+1}\n"
                    f"<!-- FILE: {pdf_basename} | PAGE: {i+1}/{n} -->\n"
                )
                # Extract tables first
                try:
                    tables = page.extract_tables() or []
                except Exception:
                    tables = []
                # Extract plain text
                txt = page.extract_text() or ""

                # If text is garbled (multi-column interleaving), drop it and
                # mark this page for OCR — OCR respects visual layout.
                if _is_garbled(txt):
                    logger.info(f"_pdf_to_markdown: page {i+1} of {pdf_basename} "
                                f"looks garbled (column-interleaved) — queuing OCR")
                    txt = ""
                    ocr_pages.append(i)

                # Count NON-EMPTY rows AND count distinct columns. A
                # pdfplumber table that has only 1-2 cols on a page with
                # lots of numbers is a sign of column-merging — try multi-
                # column reconstruction before accepting.
                pp_total_rows = 0
                pp_nonempty_rows = 0
                pp_max_cols = 0
                if tables:
                    for t in tables:
                        for row in t:
                            pp_total_rows += 1
                            non_empty = sum(1 for c in row
                                            if (c is not None and str(c).strip()))
                            if non_empty:
                                pp_nonempty_rows += 1
                            if len(row) > pp_max_cols:
                                pp_max_cols = len(row)

                # ── MULTI-COLUMN RECONSTRUCTION ──
                # Only attempt when pdfplumber detected a table (so we know
                # this page actually has tabular content) but collapsed it
                # into ≤2 cols. AND the reconstruction must have at least
                # one column header that looks like a financial period.
                used_reconstruction = False
                if (tables and 1 <= pp_max_cols <= 2 and pp_nonempty_rows >= 4):
                    rebuilt = _reconstruct_multicol_table(page)
                    if rebuilt and len(rebuilt[0][0]) >= 3:
                        headers = rebuilt[0][0]
                        # Require at least 1 detected header to look like
                        # a year/period column (e.g. "March 31, 2024",
                        # "FY2024", "2024", "Quarter", "Year")
                        has_period_header = any(
                            re.search(r"\b(19|20)\d{2}\b|\bF\d{2,4}\b|"
                                      r"\b(March|FY|quarter|year|period|H[12])\b",
                                      str(h), re.I)
                            for h in headers[1:]   # skip "Particulars"
                        )
                        if has_period_header:
                            logger.info(
                                f"_pdf_to_markdown: page {i+1} of {pdf_basename} "
                                f"pdfplumber returned {pp_max_cols}-col table — "
                                f"reconstructed to {len(rebuilt[0][0])} cols, "
                                f"headers: {headers}"
                            )
                            page_md += (
                                f"<!-- TABLE_TYPE: multicol_reconstructed | "
                                f"COLS: {len(rebuilt[0][0])} | "
                                f"HEADERS: {headers} -->\n"
                            )
                            for t in rebuilt:
                                md = _table_to_markdown(t)
                                if md:
                                    page_md += "\n" + md + "\n"
                            used_reconstruction = True
                        else:
                            logger.debug(
                                f"_pdf_to_markdown: page {i+1} reconstruction "
                                f"rejected — no period header in {headers}"
                            )

                if not used_reconstruction:
                    if tables and pp_nonempty_rows >= 3:
                        for t_idx, t in enumerate(tables):
                            md = _table_to_markdown(t)
                            if md:
                                page_md += (
                                    f"<!-- TABLE {t_idx+1}/{len(tables)} | "
                                    f"ROWS: {len(t)} | COLS: {len(t[0]) if t else 0} -->\n"
                                )
                                page_md += "\n" + md + "\n"
                    elif tables:
                        # Sparse / empty-grid table — pdfplumber detected lines but
                        # almost no text in the cells. Try fitz table extraction
                        # before giving up; OCR is a last resort.
                        logger.info(
                            f"_pdf_to_markdown: page {i+1} of {pdf_basename} "
                            f"sparse pdfplumber tables ({pp_nonempty_rows}/{pp_total_rows} "
                            f"non-empty rows) — queuing fitz fallback"
                        )
                        fitz_pages.append(i)

                if txt.strip():
                    page_md += "\n" + txt + "\n"
                else:
                    # Empty text — candidate for OCR
                    if not tables:
                        ocr_pages.append(i)

                parts.append(page_md)
    except Exception as exc:
        logger.warning(f"_pdf_to_markdown: pdfplumber failed for {pdf_path}: {exc}")

    # Fitz table fallback — pages where pdfplumber returned sparse/empty grids
    if fitz_pages:
        logger.info(f"_pdf_to_markdown: fitz-extracting tables for {len(fitz_pages)} "
                    f"sparse pages of {Path(pdf_path).name}")
        for idx in fitz_pages:
            tbls = _fitz_extract_tables(pdf_path, idx)
            if tbls:
                added = 0
                for t in tbls:
                    md = _table_to_markdown(t)
                    if md:
                        if idx < len(parts):
                            parts[idx] += "\n" + md + "\n"
                            added += 1
                if added:
                    logger.info(f"  fitz recovered {added} table(s) on page {idx+1}")
                else:
                    # No tables came back — fall through to OCR
                    ocr_pages.append(idx)
            else:
                ocr_pages.append(idx)

    # OCR pages where pdfplumber yielded nothing
    if ocr_pages:
        logger.info(f"_pdf_to_markdown: OCR-ing {len(ocr_pages)} image-only pages of {Path(pdf_path).name}")
        ocr_results = _ocr_pdf_pages(pdf_path, ocr_pages, max_pages=80)
        for i, txt in ocr_results.items():
            if i < len(parts):
                parts[i] += f"\n[OCR text]\n{txt}\n"

    # Prepend version stamp so future runs know which extractor produced this cache
    full = f"<!-- extractor:{EXTRACTION_VERSION} -->\n" + "\n".join(parts)
    # Try each cache location until one succeeds
    for cache in cache_paths:
        try:
            cache.write_text(full, encoding="utf-8")
            logger.info(f"_pdf_to_markdown: cached {cache.name} ({len(full)//1000}k chars, {EXTRACTION_VERSION})")
            break
        except Exception as exc:
            logger.debug(f"_pdf_to_markdown: cache write to {cache} failed: {exc}")
    return full


def _fitz_extract_tables(pdf_path: str, page_idx: int) -> list[list[list[str]]]:
    """
    Use PyMuPDF's table-detection to extract tables from a single page.
    Returns list of tables (each table is rows × cells of strings). Empty list
    on failure or if no tables found. PyMuPDF's grid heuristic differs from
    pdfplumber's — works better on sparse line-art layouts (e.g. Q1 result
    presentations with light gridlines).
    """
    try:
        import fitz   # PyMuPDF
    except ImportError:
        return []
    out: list[list[list[str]]] = []
    try:
        doc = fitz.open(pdf_path)
        if page_idx < 0 or page_idx >= len(doc):
            doc.close()
            return []
        page = doc[page_idx]
        # find_tables() requires PyMuPDF >= 1.23
        try:
            finder = page.find_tables()
            tables = finder.tables if hasattr(finder, "tables") else list(finder)
        except Exception:
            tables = []
        for tbl in tables:
            try:
                rows = tbl.extract()  # returns rows × cells of strings/None
                # Skip empty / single-row tables
                if not rows or len(rows) < 2:
                    continue
                # Normalise: ensure all cells are strings
                norm = [[("" if c is None else str(c)) for c in r] for r in rows]
                # Skip if all cells empty
                nonempty = sum(1 for r in norm for c in r if c.strip())
                if nonempty < 4:
                    continue
                out.append(norm)
            except Exception as exc:
                logger.debug(f"_fitz_extract_tables: table extract failed: {exc}")
        doc.close()
    except Exception as exc:
        logger.warning(f"_fitz_extract_tables: cannot open {pdf_path} page {page_idx}: {exc}")
    return out


def _reconstruct_multicol_table(page) -> list[list[list[str]]]:
    """
    Reconstruct a multi-column financial table from pdfplumber's word-level
    extraction when extract_tables() merges columns.

    Strategy:
      1. Get all words on the page with bounding boxes.
      2. Group words into rows by y-coordinate (≤4pt vertical tolerance).
      3. For each row, sort by x-coordinate.
      4. Cluster the x-positions of NUMBER words across all rows.
         A "column" is a tight x-cluster (≤25pt internal spread) that
         contains numbers from many rows.
      5. For each row, assign each word to the nearest column cluster.
      6. Emit a markdown-like table with the first non-number word per row
         as the label and one column per cluster.

    Works generically — no company-specific tuning. Triggered when
    pdfplumber returns a table with ≤1 numeric column on a page that
    clearly has more (detected by counting distinct number x-clusters).
    """
    try:
        words = page.extract_words(
            x_tolerance=2, y_tolerance=3, keep_blank_chars=False,
            extra_attrs=["fontname"],
        ) or []
    except Exception:
        return []
    if len(words) < 20:
        return []

    # Step 1: group words into rows by y-midpoint
    rows_by_y: dict[float, list] = {}
    for w in words:
        y_mid = round((w["top"] + w["bottom"]) / 2 / 4) * 4   # 4-pt buckets
        rows_by_y.setdefault(y_mid, []).append(w)
    if len(rows_by_y) < 5:
        return []

    def _is_number(s: str) -> bool:
        s = s.strip().replace(",", "").replace(" ", "")
        if not s:
            return False
        s = s.lstrip("(").rstrip(")").lstrip("-")
        try:
            float(s)
            return True
        except ValueError:
            return False

    # Step 2: collect all number x-positions across all rows
    number_xs: list[float] = []
    for y, ws in rows_by_y.items():
        for w in ws:
            if _is_number(w["text"]):
                number_xs.append((w["x0"] + w["x1"]) / 2)
    if len(number_xs) < 10:
        return []

    # Only consider rows that have ≥2 number tokens — these are real table
    # rows, not prose. Prose contributes scattered numbers that create
    # spurious columns.
    table_rows_yvals = []
    for y, ws in rows_by_y.items():
        nums_in_row = [w for w in ws if _is_number(w["text"])]
        if len(nums_in_row) >= 2:
            table_rows_yvals.append(y)
    if len(table_rows_yvals) < 4:
        return []   # not enough table-like rows

    # Step 3: cluster x-positions of numbers from TABLE rows only
    table_number_xs: list[float] = []
    for y in table_rows_yvals:
        for w in rows_by_y[y]:
            if _is_number(w["text"]):
                table_number_xs.append((w["x0"] + w["x1"]) / 2)

    table_number_xs_sorted = sorted(table_number_xs)
    clusters: list[list[float]] = [[table_number_xs_sorted[0]]]
    for x in table_number_xs_sorted[1:]:
        if x - clusters[-1][-1] > 25:
            clusters.append([x])
        else:
            clusters[-1].append(x)
    # A column must appear in ≥40% of table rows
    n_table_rows = len(table_rows_yvals)
    clusters = [c for c in clusters if len(c) >= max(3, int(n_table_rows * 0.4))]
    if len(clusters) < 2:
        return []   # not a multi-col table

    col_centers = [sum(c) / len(c) for c in clusters]
    col_centers.sort()

    # Try to detect column headers — look at the rows ABOVE the first
    # table row for date-like text (year mentions).
    header_y_max = min(table_rows_yvals)
    header_words: list = []
    for y in sorted(rows_by_y.keys()):
        if y >= header_y_max:
            break
        for w in rows_by_y[y]:
            txt = w["text"].strip()
            if not txt:
                continue
            # Look for year-bearing tokens or column labels
            if (re.search(r"\b(19|20)\d{2}\b", txt)
                    or re.search(r"\bMarch|year|quarter|period|FY|H1|H2", txt, re.I)):
                header_words.append(w)

    # Build per-column header text by clustering header words to col_centers
    headers = ["" for _ in col_centers]
    for w in header_words:
        x_mid = (w["x0"] + w["x1"]) / 2
        best_col = min(range(len(col_centers)),
                       key=lambda i: abs(col_centers[i] - x_mid))
        if abs(col_centers[best_col] - x_mid) <= 40:
            headers[best_col] = (headers[best_col] + " " + w["text"]).strip()

    # Synthesize header — use detected text when available, else Col_N
    final_headers = [(h if h else f"Col_{i+1}") for i, h in enumerate(headers)]

    # Step 4: build table rows
    table: list[list[str]] = []
    for y in sorted(table_rows_yvals):
        row_words = sorted(rows_by_y[y], key=lambda w: w["x0"])
        # Label = words left of the first column center
        label_parts: list[str] = []
        for w in row_words:
            x_mid = (w["x0"] + w["x1"]) / 2
            if x_mid < col_centers[0] - 20 and not _is_number(w["text"]):
                label_parts.append(w["text"])
        label = " ".join(label_parts).strip()

        cells = [""] * len(col_centers)
        for w in row_words:
            if not _is_number(w["text"]):
                continue
            x_mid = (w["x0"] + w["x1"]) / 2
            best_col = min(range(len(col_centers)),
                           key=lambda i: abs(col_centers[i] - x_mid))
            if abs(col_centers[best_col] - x_mid) <= 30:
                if cells[best_col]:
                    cells[best_col] += " " + w["text"]
                else:
                    cells[best_col] = w["text"]

        # Quality gate: need either a label OR ≥2 cells filled
        cells_filled = sum(1 for c in cells if c)
        if not label and cells_filled < 2:
            continue
        # Quality gate 2: at this point we want rows where label exists OR
        # the row has values across most columns (it's clearly tabular)
        if cells_filled == 0:
            continue
        table.append([label] + cells)

    # Final quality check: at least 60% of emitted rows have a label
    rows_with_label = sum(1 for r in table if r[0])
    if not table or rows_with_label < len(table) * 0.5:
        return []

    if len(table) < 3:
        return []

    return [[["Particulars"] + final_headers] + table]


def _ocr_pdf_pages(pdf_path: str, page_indices: list[int],
                    max_pages: int = 40) -> dict[int, str]:
    """
    Run Tesseract OCR on selected pages of a PDF in PARALLEL. Returns
    {page_idx: text}. Used when pdfplumber returns no text or returns
    column-interleaved garbage. Lower DPI (150) and 4-way parallelism
    cut wall time by ~5x compared to serial 200dpi.
    """
    out: dict[int, str] = {}
    try:
        import fitz                    # PyMuPDF
        import pytesseract
        from PIL import Image
        import io as _io
        import concurrent.futures as _cf
    except ImportError as exc:
        logger.warning(f"OCR fallback unavailable: {exc}")
        return out
    try:
        doc = fitz.open(pdf_path)
    except Exception as exc:
        logger.warning(f"OCR: cannot open {pdf_path}: {exc}")
        return out

    def _ocr_one(idx: int) -> tuple[int, str]:
        if idx < 0 or idx >= len(doc):
            return (idx, "")
        try:
            page = doc[idx]
            # 150 DPI is plenty for text recognition (vs 200) and ~2x faster
            pix = page.get_pixmap(dpi=150)
            img = Image.open(_io.BytesIO(pix.tobytes("png")))
            txt = pytesseract.image_to_string(img, config="--psm 6")
            return (idx, txt or "")
        except Exception as exc:
            logger.debug(f"OCR page {idx} failed: {exc}")
            return (idx, "")

    indices = [i for i in page_indices[:max_pages]]
    try:
        # 4 parallel OCR workers — Tesseract releases the GIL during the
        # subprocess call so we get real CPU parallelism.
        with _cf.ThreadPoolExecutor(max_workers=4) as pool:
            for idx, txt in pool.map(_ocr_one, indices):
                if txt and txt.strip():
                    out[idx] = txt
    finally:
        doc.close()
    return out


def _ocr_pdf_pages_serial(pdf_path: str, page_indices: list[int],
                           max_pages: int = 40) -> dict[int, str]:
    """Legacy serial OCR — kept only for reference, not used."""
    out: dict[int, str] = {}
    try:
        import fitz, pytesseract, io as _io
        from PIL import Image
    except ImportError:
        return out
    try:
        doc = fitz.open(pdf_path)
    except Exception:
        return out
    try:
        for idx in page_indices[:max_pages]:
            if idx < 0 or idx >= len(doc):
                continue
            try:
                page = doc[idx]
                pix = page.get_pixmap(dpi=150)
                img = Image.open(_io.BytesIO(pix.tobytes("png")))
                txt = pytesseract.image_to_string(img, config="--psm 6")
                if txt and txt.strip():
                    out[idx] = txt
            except Exception as exc:
                logger.debug(f"OCR page {idx} failed: {exc}")
    finally:
        doc.close()
    return out


def _classify_pdf(pdf_path: str) -> tuple[PDFClassification, str]:
    """Classify a PDF and return (classification, full_relevant_text).

    Uses the markdown sidecar (.md) cache when available — extracts tables
    properly and preserves column alignment (essential for Quarterly result
    PDFs with date-labeled columns). Falls back to OCR for image-only pages.
    """
    import pdfplumber

    p = Path(pdf_path)
    ocr_cache: dict[int, str] = {}     # page_idx → OCR'd text (lazy)

    # ── Generate (or load cached) markdown rendering of the entire PDF ──
    # This includes extracted tables formatted as MD tables. We use this MD
    # as a SUPPLEMENT to the per-page text extraction below — the page-by-page
    # selector still uses raw text for picking relevant pages, but the FINAL
    # text we send to the LLM uses the MD content for its tables.
    md_text = ""
    try:
        md_text = _pdf_to_markdown(str(p))
    except Exception as exc:
        logger.warning(f"_classify_pdf: MD conversion failed for {p.name}: {exc}")

    def _page_text(pdf, idx: int) -> str:
        """Get text for a page; OCR-fallback if pdfplumber returns nothing."""
        t = pdf.pages[idx].extract_text() or ""
        if t.strip():
            return t
        return ocr_cache.get(idx, "")

    with pdfplumber.open(str(p)) as pdf:
        n_pages = len(pdf.pages)
        # Read first 3 pages for title detection — text first
        head_pages_txt: dict[int, str] = {}
        for i in range(min(3, n_pages)):
            head_pages_txt[i] = pdf.pages[i].extract_text() or ""
        head = "\n".join(head_pages_txt.values()).strip()

        # If first 3 pages have very little text (image-only cover), OCR them
        # so title detection works.
        if len(head) < 200:
            logger.info(f"_classify_pdf: {p.name} cover has only {len(head)} chars "
                        f"— running OCR on first 3 pages")
            ocr_cache.update(_ocr_pdf_pages(str(p), list(range(min(3, n_pages)))))
            head = "\n".join(
                ocr_cache.get(i, head_pages_txt.get(i, ""))
                for i in range(min(3, n_pages))
            ).strip()

        # Decide stmt + period
        head_l = head.lower()
        title_l = (p.stem + " " + head[:600]).lower()

        # Match "investor presentation" / "corporate presentation" with any
        # separator (space, hyphen, underscore) — common in download names.
        title_normalised = re.sub(r"[\-_]", " ", title_l)
        if (re.search(r"investor[\s'’]+(presentation|deck|update|brief)",
                      title_normalised)
            or re.search(r"corporate[\s'’]+presentation", title_normalised)
            or "earnings presentation" in title_normalised):
            period_type = "INVESTOR_PRESENTATION"
        elif "quarterly" in title_l or "three months ended" in head_l:
            period_type = "QUARTERLY"
        elif ("annual report" in title_l or "integrated report" in title_l
              or "year ended" in head_l):
            period_type = "ANNUAL"
        else:
            period_type = "UNKNOWN"

        if "consolidated" in title_l and "standalone" not in title_l:
            stmt_type = "CONSOLIDATED"
        elif "standalone" in title_l and "consolidated" not in title_l:
            stmt_type = "STANDALONE"
        elif "consolidated" in title_l and "standalone" in title_l:
            stmt_type = "BOTH"
        elif period_type == "INVESTOR_PRESENTATION":
            stmt_type = "BOTH"   # IPs typically discuss aggregated/consolidated
        elif period_type == "ANNUAL":
            stmt_type = "BOTH"   # ARs have both sections
        else:
            stmt_type = "UNKNOWN"

        # Extract financial-statement pages only — keeps prompt tight.
        # For IPs, take pages with financial keywords. For ARs, focus on the
        # mid-back portion. For quarterly results, the whole doc (usually 2-5 pages).
        relevant_pages: list[int] = []
        if n_pages <= 10:
            relevant_pages = list(range(n_pages))
        elif period_type == "INVESTOR_PRESENTATION":
            # IPs are mostly slides with text — pick pages mentioning revenue/EBITDA/segment.
            # First pass: pdfplumber. If that yields <300 chars total across the
            # whole doc → image-only deck → run OCR on first 60 pages.
            text_chars = 0
            page_texts: dict[int, str] = {}
            for i in range(n_pages):
                t = pdf.pages[i].extract_text() or ""
                page_texts[i] = t
                text_chars += len(t)
            if text_chars < 300 and n_pages > 0:
                logger.info(f"_classify_pdf: IP {p.name} appears image-only "
                            f"({text_chars} chars total) — running OCR on "
                            f"{min(60, n_pages)} pages")
                ocr_cache.update(
                    _ocr_pdf_pages(str(p), list(range(min(60, n_pages))))
                )
                page_texts.update(ocr_cache)

            for i in range(n_pages):
                t = page_texts.get(i, "") or ocr_cache.get(i, "")
                tl = t.lower()
                # Generic financial-content keywords — work for any sector
                if any(k in tl for k in (
                        "revenue", "ebitda", "pat ", "profit", "loss",
                        "segment", "%", "₹", "rs.", "rs ", "crore",
                        "fy2", "growth", "margin", "yoy", "qoq",
                        "balance sheet", "cash flow", "operating",
                        # Banking / NBFC
                        "nim", "casa", "gnpa", "advances", "deposits",
                        # Insurance
                        "vnb", "ape", "persistency",
                        # Retail / FMCG
                        "store", "sssg", "volume", "arpu",
                        # IT
                        "utilization", "attrition",
                        # Auto / Power
                        "units", "capacity", "plf")):
                    relevant_pages.append(i)
                if len(relevant_pages) >= 30:
                    break
        else:  # ANNUAL — find financial statement pages.
            # CRITICAL FIX: build md_by_page FIRST and use it for boundary
            # detection. pdfplumber.extract_text() returns column-interleaved
            # garbage on magazine-layout pages, which makes "standalone
            # balance sheet" appear as "s t a n d a l o n e" (interleaved).
            # The .md cache uses fitz + multi-col reconstructor, which
            # produces clean readable text.
            md_by_page_early: dict[int, str] = {}
            if md_text:
                chunks_e = re.split(r"(?m)^## Page (\d+)\n", md_text)
                for k in range(1, len(chunks_e) - 1, 2):
                    try:
                        pg_num_e = int(chunks_e[k]) - 1
                        md_by_page_early[pg_num_e] = chunks_e[k + 1]
                    except ValueError:
                        continue

            def _page_text_for_boundary(i: int) -> str:
                """Use md cache content if available — it's the clean
                extraction. Fall back to pdfplumber raw text."""
                md_chunk = md_by_page_early.get(i, "")
                if md_chunk.strip():
                    return md_chunk
                try:
                    return pdf.pages[i].extract_text() or ""
                except Exception:
                    return ""

            # Pass A: scan every page, score it, and look for boundary markers
            standalone_pages: list[tuple[int, int]] = []
            consolidated_pages: list[tuple[int, int]] = []
            unscoped_pages: list[tuple[int, int]] = []
            cf_pages: list[tuple[int, int]] = []
            page_texts_for_boundary: dict[int, str] = {}

            sa_boundary_markers = (
                "standalone balance sheet",
                "standalone statement of profit",
                "standalone statement of financial position",
                "standalone cash flow",
                "standalone financial statement",
                "report on the audit of the standalone",
                "audit of the standalone financial",
                "notes to the standalone financial",
                "notes to standalone financial",
            )
            co_boundary_markers = (
                "consolidated balance sheet",
                "consolidated statement of profit",
                "consolidated statement of financial position",
                "consolidated cash flow",
                "consolidated financial statement",
                "report on the audit of the consolidated",
                "audit of the consolidated financial",
                "notes to the consolidated financial",
                "notes to consolidated financial",
            )
            end_markers = (
                "notice of annual general meeting",
                "notice of the annual general meeting",
                "notice to the members",
                "notice to members",
            )
            std_start_idx_a = None
            cons_start_idx_a = None
            fin_end_idx_a = None

            for i in range(n_pages):
                # USE MD CACHE FIRST (clean text), fall back to pdfplumber
                t = _page_text_for_boundary(i)
                tl = t.lower()
                head = tl[:3000]
                page_texts_for_boundary[i] = tl

                if std_start_idx_a is None and any(
                        k in head for k in sa_boundary_markers):
                    std_start_idx_a = i
                if cons_start_idx_a is None and any(
                        k in head for k in co_boundary_markers):
                    cons_start_idx_a = i
                if (cons_start_idx_a is not None and i > cons_start_idx_a
                        and fin_end_idx_a is None
                        and any(k in head for k in end_markers)):
                    fin_end_idx_a = i

                score = 0
                for kw in ("revenue from operations", "balance sheet",
                           "statement of profit", "cash flow",
                           "total assets", "total equity",
                           "ebitda", "profit before tax", "year ended"):
                    if kw in tl:
                        score += 2
                cf_signals = (
                    "cash flow from operating",
                    "cash flow from investing",
                    "cash flow from financing",
                    "net cash flow from operating",
                    "net cash used in investing",
                    "net cash used in financing",
                    "cash generated from operations",
                    "cash and cash equivalents at the end",
                    "depreciation and amortisation expense",
                    "working capital changes",
                    "adjustments for:",
                )
                cf_score = sum(3 for k in cf_signals if k in tl)
                score += cf_score
                if cf_score >= 6:
                    cf_pages.append((score, i))
                if score >= 2:
                    unscoped_pages.append((score, i))   # initial bucket

            logger.info(
                f"_classify_pdf: {p.name} page-selector boundaries: "
                f"std_start={std_start_idx_a}, cons_start={cons_start_idx_a}, "
                f"fin_end={fin_end_idx_a}"
            )

            # Pass B: USE BOUNDARIES to force-include all in-range pages.
            # When std_start..cons_start..fin_end are detected, every page
            # in the Standalone range goes into standalone_pages, every
            # page in the Consolidated range goes into consolidated_pages.
            forced_pages: set[int] = set()
            if std_start_idx_a is not None and cons_start_idx_a is not None:
                # Force pages in Standalone range
                for i in range(std_start_idx_a, cons_start_idx_a):
                    standalone_pages.append((10, i))   # score=10 (high)
                    forced_pages.add(i)
                # Force pages in Consolidated range
                end = fin_end_idx_a if fin_end_idx_a is not None else n_pages
                for i in range(cons_start_idx_a, end):
                    consolidated_pages.append((10, i))
                    forced_pages.add(i)
                logger.info(
                    f"_classify_pdf: {p.name} forced "
                    f"{cons_start_idx_a - std_start_idx_a} STANDALONE pages "
                    f"and {end - cons_start_idx_a} CONSOLIDATED pages "
                    f"into relevant_pages"
                )

            # Remove force-included pages from unscoped bucket
            unscoped_pages = [(s, i) for (s, i) in unscoped_pages
                              if i not in forced_pages]

            # Take all forced + up to 50 from each std/cons bucket (extra
            # cushion) + 15 unscoped + 25 CF.
            standalone_pages.sort(reverse=True)
            consolidated_pages.sort(reverse=True)
            unscoped_pages.sort(reverse=True)
            cf_pages.sort(reverse=True)
            # Take ALL forced standalone+consolidated pages (no cap), plus
            # unscoped + CF. Capping at 80 was dropping Cons pages on big
            # ARs which starved the LLM of F23/F24/F25 context.
            picked = (
                [i for _, i in standalone_pages]
                + [i for _, i in consolidated_pages]
                + [i for _, i in unscoped_pages[:20]]
                + [i for _, i in cf_pages[:30]]
            )
            relevant_pages = sorted(set(picked))
            logger.info(
                f"_classify_pdf: {p.name} relevant_pages selected: "
                f"{len(relevant_pages)} pages "
                f"(std-bucket={len(standalone_pages)}, "
                f"cons-bucket={len(consolidated_pages)}, "
                f"unscoped={len(unscoped_pages)}, cf={len(cf_pages)})"
            )

        # Pre-split the markdown rendering into per-page chunks so we can
        # interleave it with the section-tracking logic below. The MD has
        # "## Page N" markers — split on those.
        md_by_page: dict[int, str] = {}
        if md_text:
            chunks = re.split(r"(?m)^## Page (\d+)\n", md_text)
            # chunks[0] is preamble (empty), then alternating page_num, content
            for k in range(1, len(chunks) - 1, 2):
                try:
                    pg_num = int(chunks[k]) - 1   # 1-based → 0-based
                    md_by_page[pg_num] = chunks[k + 1]
                except ValueError:
                    continue

        # Stateful section tracker — many ARs print the section name once at
        # the top of the section (e.g. "Consolidated Balance Sheet as at March
        # 31, 2024") and subsequent pages just continue the table without
        # restating it. We carry the last-seen section forward so every page
        # gets tagged. Without this hint the LLM picks Standalone (which
        # usually appears first in the AR) and leaves Consolidated blank.
        def _tag_every_table(page_md: str, scope: str) -> str:
            """Inject a [SCOPE=X] marker on its own line immediately before
            every markdown table inside `page_md`. Belt-and-suspenders so even
            multi-table pages can't accidentally leak across scope boundaries."""
            if scope == "UNKNOWN_SCOPE" or not page_md:
                return page_md
            lines = page_md.split("\n")
            out: list[str] = []
            for idx, ln in enumerate(lines):
                stripped = ln.strip()
                # Detect start of a markdown table: |…| followed by |---|…|
                if (stripped.startswith("|") and stripped.endswith("|")
                        and idx + 1 < len(lines)
                        and re.match(r"\|[\s\-:|]+\|", lines[idx + 1].strip())):
                    # Don't double-tag if a SCOPE marker is already on the line
                    # immediately above
                    prev = out[-1].strip() if out else ""
                    if "[SCOPE=" not in prev:
                        out.append(f"[SCOPE={scope}]")
                out.append(ln)
            return "\n".join(out)

        # ── THREE-BUCKET SCOPE DETECTION ──
        # ARs have THREE kinds of pages, not two:
        #   1. STANDALONE financial statement pages (BS / P&L / CF / Notes
        #      to standalone financials) — between Standalone auditor's report
        #      and Consolidated auditor's report
        #   2. CONSOLIDATED financial statement pages — from Consolidated
        #      auditor's report through end of financial-statements section
        #   3. NON-FINANCIAL pages — cover, MD&A, Director's Report, Corporate
        #      Governance, CSR, Sustainability, Notice of AGM, BRSR, etc.
        #      These are read for context (segment data, KPIs) but their
        #      values are NOT tagged as either Standalone or Consolidated.
        #
        # Pass 1: locate the start of Standalone Financial Statements
        #         AND the start of Consolidated Financial Statements
        #         AND the end of Consolidated Financial Statements (if found).
        # Pass 2: tag each page based on its position relative to these
        #         boundaries.
        def _scope_marker_match(text: str, scope_kind: str) -> bool:
            """Return True if text contains a clear scope-section marker."""
            tl = text.lower()
            if scope_kind == "STANDALONE":
                return any(k in tl for k in (
                    "standalone balance sheet",
                    "standalone statement of profit",
                    "standalone statement of financial position",
                    "standalone cash flow",
                    "standalone statement of cash flow",
                    "standalone financial statement",
                    "report on the audit of the standalone",
                    "audit of the standalone financial",
                    "notes to the standalone financial",
                    "notes to standalone financial",
                ))
            elif scope_kind == "CONSOLIDATED":
                return any(k in tl for k in (
                    "consolidated balance sheet",
                    "consolidated statement of profit",
                    "consolidated statement of financial position",
                    "consolidated cash flow",
                    "consolidated statement of cash flow",
                    "consolidated financial statement",
                    "report on the audit of the consolidated",
                    "audit of the consolidated financial",
                    "notes to the consolidated financial",
                    "notes to consolidated financial",
                ))
            return False

        def _end_of_financials(text: str) -> bool:
            """Markers signaling we're past the financial statements section."""
            tl = text.lower()
            return any(k in tl for k in (
                "notice of annual general meeting",
                "notice of the annual general meeting",
                "notice to the members",
                "notice to members",
                "agm notice",
            ))

        # Pass 1: scan in page order, locate scope-section start indices
        std_start_idx = None
        cons_start_idx = None
        fin_end_idx = None
        sorted_pages = sorted(relevant_pages)
        for i in sorted_pages:
            txt = md_by_page.get(i, "") or pdf.pages[i].extract_text() or ""
            if not txt.strip() and i in ocr_cache:
                txt = ocr_cache[i]
            head_text = txt[:3000]
            if not head_text:
                continue
            if std_start_idx is None and _scope_marker_match(head_text, "STANDALONE"):
                std_start_idx = i
            if cons_start_idx is None and _scope_marker_match(head_text, "CONSOLIDATED"):
                cons_start_idx = i
            # Track latest end-of-financials marker AFTER consolidated start
            if cons_start_idx is not None and i > cons_start_idx and _end_of_financials(head_text):
                if fin_end_idx is None:
                    fin_end_idx = i

        logger.info(
            f"_classify_pdf: {p.name} scope boundaries: "
            f"std_start={std_start_idx}, cons_start={cons_start_idx}, "
            f"fin_end={fin_end_idx} (None = no marker found)"
        )

        # Page order: ORIGINAL page order (not reordered). v9 reordering
        # caused regressions — financial pages were getting their year-header
        # context from EARLIER pages in the doc, and reordering broke this.
        # Truncation cap raised to 200K (from 80K) is enough budget.
        full_text_parts = []
        for i in relevant_pages:
            # Prefer markdown rendering (has tables formatted) over raw text
            t = md_by_page.get(i, "")
            if not t.strip():
                t = pdf.pages[i].extract_text() or ""
            # Fall back to OCR'd text if pdfplumber returned nothing for this page
            if not t.strip() and i in ocr_cache:
                t = ocr_cache[i]
            if not t.strip():
                continue

            # THREE-BUCKET assignment based on page position
            if std_start_idx is not None and cons_start_idx is not None:
                # Both sections detected — clean three-bucket split
                if i < std_start_idx:
                    section_hint = "UNKNOWN_SCOPE"      # MD&A / Director's Rpt
                elif std_start_idx <= i < cons_start_idx:
                    section_hint = "STANDALONE"
                elif fin_end_idx is not None and i >= fin_end_idx:
                    section_hint = "UNKNOWN_SCOPE"      # Notice of AGM / appendix
                else:
                    section_hint = "CONSOLIDATED"
            elif std_start_idx is not None and cons_start_idx is None:
                # Only Standalone section detected (standalone-only company)
                section_hint = "STANDALONE" if i >= std_start_idx else "UNKNOWN_SCOPE"
            elif cons_start_idx is not None and std_start_idx is None:
                # Only Consolidated detected — unusual but treat as cons from there
                section_hint = "CONSOLIDATED" if i >= cons_start_idx else "UNKNOWN_SCOPE"
            else:
                # No scope markers found anywhere (e.g. summary IP-like doc)
                section_hint = "UNKNOWN_SCOPE"
            # Inject scope marker before EVERY table on this page
            t_tagged = _tag_every_table(t, section_hint)
            full_text_parts.append(
                f"\n--- {p.name} page {i+1} [SCOPE={section_hint}] ---\n"
                f"{t_tagged}\n"
                f"--- end page {i+1} [SCOPE={section_hint}] ---\n"
            )
        full_text = "\n".join(full_text_parts)

        # Cap total text at 2M chars per PDF. ARs are large (640K+ for the
        # selected pages of a 291-page integrated report) and we MUST NOT
        # truncate Consolidated pages when filling annual templates.
        # 2M is huge for pattern matching (free, deterministic) and only
        # the LLM path needs to worry about context — that's handled by
        # per-call routing, not this single-PDF cap.
        if len(full_text) > 2_000_000:
            full_text = full_text[:2_000_000] + "\n[...truncated...]"

        # Build TOC from the md cache (catches all known section types).
        # Pull boundary indices from prior detection if available.
        try:
            _toc = _build_toc(
                md_by_page,
                std_start=locals().get('std_start_idx_a'),
                cons_start=locals().get('cons_start_idx_a'),
                fin_end=locals().get('fin_end_idx_a'),
            )
        except Exception as _exc:
            logger.debug(f"_classify_pdf: TOC build failed: {_exc}")
            _toc = []

        cls = PDFClassification(
            path=str(p), name=p.name,
            stmt_type=stmt_type, period_type=period_type,
            title_text=head[:200].replace("\n", " "),
            pages_total=n_pages,
            text_preview=head[:500].replace("\n", " "),
            toc=_toc,
        )
    return cls, full_text


# ──────────────────────────────────────────────────────────────────────────────
# Period mapping for Quarterly PDFs
# ──────────────────────────────────────────────────────────────────────────────

def _build_period_mapping(filename: str, head_text: str, full_text: str) -> str:
    """
    Scan the PDF text for column headers like "Three months ended June 30, 2024"
    or "For the year ended March 31, 2024" and return an explicit mapping
    block that the LLM can read at the top of the PDF context.
    """
    import re as _re

    # Pull the header area — first ~8k chars covers the column headers
    # for any quarterly result PDF.
    sample = (head_text + "\n" + full_text[:8000])

    # Find every (month, year) pair in the header area. For each unique
    # date, emit BOTH possible interpretations:
    #   • Quarterly: <month>-end → 1QF / 2QF / 3QF / 4QF<FY>
    #   • Annual (only if month is March): <month>-end → F<FY>
    # The LLM then picks the correct one based on the table column it's
    # reading. This is more robust than trying to disambiguate from
    # surrounding context (which often fails because Indian Q result
    # PDFs lay headers across multiple lines).
    date_re = _re.compile(
        r"(january|february|march|april|may|june|july|august|september|october|november|december)\s+(\d{1,2})\s*,?\s*(\d{4})",
        _re.I,
    )
    qtr_dates: list[tuple[str, str]] = []
    yr_dates: list[tuple[str, str]] = []
    seen_dates: set[tuple[str, str]] = set()
    for m in date_re.finditer(sample):
        mon_word = m.group(1)
        yr_str = m.group(3)
        key = (mon_word.lower(), yr_str)
        if key in seen_dates:
            continue
        seen_dates.add(key)
        # Always emit the quarterly interpretation
        qtr_dates.append((mon_word, yr_str))
        # If it's March, also emit the FY interpretation (Indian fiscal
        # year ends March 31; "Year ended March 31, 2024" → F2024)
        if mon_word.lower().startswith("mar"):
            yr_dates.append((mon_word, yr_str))

    _MONTH_NUM = {"jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
                  "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12}

    def _quarter_label(month: int, calendar_year: int) -> str:
        """
        Indian FY: April–March. Quarter mapping by ENDING month:
          Jun → Q1, Sep → Q2, Dec → Q3, Mar → Q4
        For Q1/Q2/Q3 the FY ending is calendar_year+1.
        For Q4 (Mar) the FY ending is calendar_year.
        """
        if month == 6:  return f"1QF{calendar_year+1}"
        if month == 9:  return f"2QF{calendar_year+1}"
        if month == 12: return f"3QF{calendar_year+1}"
        if month == 3:  return f"4QF{calendar_year}"
        return ""

    def _annual_label(month: int, year: int) -> str:
        # Indian AR: year ended March 31, 2024 → F2024
        if month == 3:
            return f"F{year}"
        # Calendar year ending Dec 31 — fall back to that year
        return f"F{year}"

    seen = set()
    rows: list[str] = []
    # Quarterly columns
    for mon_word, yr_str in qtr_dates:
        mon = _MONTH_NUM.get(mon_word.lower()[:3])
        if not mon:
            continue
        try: yr = int(yr_str)
        except ValueError: continue
        ql = _quarter_label(mon, yr)
        if not ql or ql in seen:
            continue
        seen.add(ql)
        rows.append(f"  - Column for 'Three months ended "
                    f"{mon_word.title()} {yr_str}' → {ql}")
    # Annual / full-year columns
    for mon_word, yr_str in yr_dates:
        mon = _MONTH_NUM.get(mon_word.lower()[:3])
        if not mon:
            continue
        try: yr = int(yr_str)
        except ValueError: continue
        al = _annual_label(mon, yr)
        if al in seen:
            continue
        seen.add(al)
        rows.append(f"  - Column for 'Year ended "
                    f"{mon_word.title()} {yr_str}' → {al}")

    if not rows:
        return ""
    return (
        f"=== PERIOD MAPPING for {filename} ===\n"
        "Use this mapping when filling cells. The PDF labels columns by date;\n"
        "the template asks for Indian-FY quarter / year labels:\n"
        + "\n".join(rows)
        + "\n=== END PERIOD MAPPING ==="
    )


# ──────────────────────────────────────────────────────────────────────────────
# Template structure — flatten to a list the LLM can reason about
# ──────────────────────────────────────────────────────────────────────────────

@dataclass
class TemplateRowSpec:
    sheet:           str
    statement_type:  str          # "Consolidated" / "Standalone" / "" (no section)
    section:         str          # e.g. "ASSETS / Non-Current Assets / Financial Assets:"
    label:           str          # exact template row label
    years_needed:    list[str]    # e.g. ["F2022", "F2023", "F2024", "F2025"]
    is_formula:      bool         # if true, DON'T fill (template auto-computes)


def _read_template_specs(template_path: str) -> list[TemplateRowSpec]:
    """Parse template into a flat list of rows that need filling."""
    from mapper.template_reader import read_template
    tm = read_template(template_path)
    out: list[TemplateRowSpec] = []
    for sheet_name, sm in tm.sheets.items():
        years = list(sm.year_cols.keys())
        if not sm.row_index:
            continue
        for label, row_nums in sm.row_index.items():
            for r in row_nums:
                # Determine statement_type from section_offsets
                stmt = ""
                if "consolidated" in sm.section_offsets and "standalone" in sm.section_offsets:
                    stmt = ("Consolidated"
                            if r < sm.section_offsets["standalone"]
                            else "Standalone")
                elif "consolidated" in sm.section_offsets:
                    stmt = "Consolidated"
                elif "standalone" in sm.section_offsets:
                    stmt = "Standalone"
                # Section path
                section = sm.row_context.get(label, "")
                # Years needed
                years_needed: list[str] = []
                for yr, c in sm.year_cols.items():
                    cell_key = (r, c)
                    if cell_key in sm.formula_cells:
                        continue   # template formula, don't ask LLM
                    if cell_key in sm.filled_cells:
                        continue   # already filled
                    years_needed.append(yr)
                if not years_needed:
                    continue
                is_formula = any(
                    (r, c) in sm.formula_cells for c in sm.year_cols.values()
                )
                out.append(TemplateRowSpec(
                    sheet=sheet_name, statement_type=stmt,
                    section=section, label=label,
                    years_needed=sorted(years_needed),
                    is_formula=is_formula,
                ))
    return out


# ──────────────────────────────────────────────────────────────────────────────
# Prompt building — the heart of this approach
# ──────────────────────────────────────────────────────────────────────────────

_SYSTEM_PROMPT = """You are a senior equity research analyst at a top Indian financial institution. You have 15+ years of experience covering Indian listed companies across every sector — you've worked banks, manufacturing, services, financials, energy, retail, and everything in between. You hold the CFA charter and you've personally built earnings models for hundreds of companies. You know Ind-AS and SEBI filing conventions cold.

Your job: fill the analyst template by reading the source PDFs and filing your conclusions. You're the analyst. You read these statements all day. You know exactly what to look for.

═══════════════════════════════════════════════════════════════════════════
WHAT YOU'RE READING
═══════════════════════════════════════════════════════════════════════════

Annual Reports (AR) — 200-400 page PDFs. Every AR contains BOTH the
Standalone and Consolidated financial statements (P&L, Balance Sheet, Cash
Flow). Our preprocessor has scanned every page and injected scope markers:
  [SCOPE=STANDALONE]   — page is part of the Standalone section
  [SCOPE=CONSOLIDATED] — page is part of the Consolidated section
  [SCOPE=UNKNOWN_SCOPE] — couldn't determine; use other signals
These markers are also injected directly above every markdown table so even
within a single page the routing stays clean. Trust the tags — they're
derived from the AR's own section headings.

NAVIGATING THE PDF — USE THE TABLE OF CONTENTS:
At the top of the input, each PDF has a TOC listing its key pages with
section titles, e.g.:
  page 75  [STANDALONE]    Standalone Balance Sheet
  page 77  [STANDALONE]    Standalone Statement of P&L
  page 113 [CONSOLIDATED]  Consolidated Balance Sheet
  page 114 [CONSOLIDATED]  Consolidated Statement of P&L

For each cell you fill, use the TOC to locate the right page, then read
that page's content (delimited by `--- file.pdf page N [SCOPE=X] ---`).
Schedules (PPE, Borrowings, Notes etc.) are also listed in the TOC — use
them when a cell needs a sub-component value.

Quarterly Result PDFs (QR) — 2-15 pages, filed with SEBI within 45 days of
quarter end. Always have P&L for the quarter + YTD. Q2 and Q4 often also
include a Balance Sheet snapshot. Q4 is special: it has the full-year audited
figures (effectively a mini-AR for that year's P&L). Most QRs report BOTH
Consolidated AND Standalone in adjacent columns — read both.

Investor Presentations (IP) — slide decks (20-60 slides) with operating
metrics, segment-level data, business commentary, and headline financials.
The numbers shown are almost always Consolidated. Use IPs for segment splits
(business-line revenue, geography mix, product/category mix), KPIs, and
ratios that aren't broken out in the AR.

═══════════════════════════════════════════════════════════════════════════
PERIOD MAPPING (INDIAN FY: APRIL–MARCH)
═══════════════════════════════════════════════════════════════════════════

  "Year ended March 31, YYYY"               → F<YYYY>
  "As at March 31, YYYY"                    → F<YYYY>  (Balance Sheet date)
  "Three months ended June 30, YYYY"        → 1QF<YYYY+1>
  "Three months ended September 30, YYYY"   → 2QF<YYYY+1>
  "Three months ended December 31, YYYY"    → 3QF<YYYY+1>
  "Three months ended March 31, YYYY"       → 4QF<YYYY>

Calendar April–March collapses to ONE Indian FY: April 2024 → March 2025 = F2025.
So a column labeled "June 30, 2024" is Q1 of FY25 → 1QF2025.

═══════════════════════════════════════════════════════════════════════════
SCOPE ROUTING (THE ONE RULE THAT MATTERS)
═══════════════════════════════════════════════════════════════════════════

The template tells you, for each row, the section it belongs to:
  statement_type = "Consolidated"  → fill from Consolidated sources only
  statement_type = "Standalone"    → fill from Standalone sources only
  statement_type = ""              → either is acceptable (default: Consolidated)

Inside an AR, just follow the SCOPE tags. Inside a QR, columns are usually
labelled "Consolidated" / "Standalone" or grouped under section headers —
read both and route the values correctly.

═══════════════════════════════════════════════════════════════════════════
YOUR ANALYST TOOLKIT — USE IT, DERIVE FREELY
═══════════════════════════════════════════════════════════════════════════

You are a real analyst. You DERIVE values when the print doesn't match the
template label exactly — compute from components, ratio them, take diffs
across years. The full toolkit:

─── P&L IDENTITIES ───
  COGS = "Cost of materials consumed" + "Purchases of stock-in-trade"
       + "Changes in inventories of finished goods / WIP / stock-in-trade"
       (banks: COGS does not apply; substitute Interest Expense)

  Total Income      = Revenue from operations + Other income
  Gross Profit      = Revenue from operations − COGS
  EBITDA            = Total Income − COGS − Employee − Other Expenses
                    OR = EBIT + D&A  (back-derive when only EBIT printed)
  EBIT              = EBITDA − D&A
  PBT               = EBIT − Finance costs + Other Income + Share of Associates
                       (± Exceptional items)
  PAT               = PBT − Tax expense
  PAT (after MI)    = PAT − Minority Interest (only in Consolidated)

─── BALANCE SHEET IDENTITIES ───
  Total Non-Current Assets = PPE + CWIP + Intangible Assets + Investments
                           + Loans + Other Financial Assets + Deferred Tax
                           + Other Non-Current Assets
  Total Current Assets     = Inventories + Trade Receivables + Cash & Bank
                           + Short-term Investments + Loans + Other Current Assets
  Total Assets             = Non-Current Assets + Current Assets

  Total Equity             = Equity Share Capital + Other Equity (Reserves)
                           (+ Minority Interest in Consolidated)
  Total Non-Current Liab   = Long-term Borrowings + Deferred Tax Liability
                           + Other Non-Current Liabilities + Provisions
  Total Current Liab       = Trade Payables + Short-term Borrowings
                           + Other Current Liabilities + Short-term Provisions
  Total Liabilities        = Non-Current + Current Liabilities
  Balance Sheet identity   : Total Assets = Total Equity + Total Liabilities

─── PER-SHARE & VALUATION ───
  Basic EPS         = (PAT − Preference Dividend) / Weighted-Avg Equity Shares
  Diluted EPS       = (PAT − Pref Div) / Weighted-Avg Diluted Equity Shares
  Book Value/Share  = Total Equity / Outstanding Equity Shares
  Dividend/Share    = Total Dividend Paid / Outstanding Shares
  Dividend Payout % = Total Dividend / PAT × 100
  Retention Ratio   = 1 − Payout Ratio

─── PROFITABILITY RATIOS ───
  Gross Margin %    = Gross Profit / Revenue × 100
  EBITDA Margin %   = EBITDA / Total Income × 100   (or / Revenue)
  EBIT Margin %     = EBIT / Total Income × 100
  PAT Margin %      = PAT / Total Income × 100
  ROE %             = PAT / Average Total Equity × 100
                      Avg = (Opening + Closing) / 2
  ROA %             = PAT / Average Total Assets × 100
  ROCE %            = EBIT / (Avg Total Equity + Avg Long-term Debt) × 100
                      = EBIT / Capital Employed
  RoIC / RoNW       = PAT / Invested Capital × 100  (Indian filings often
                      show "RoNW" = Return on Net Worth; treat as RoE proxy)

─── LEVERAGE & LIQUIDITY ───
  Debt / Equity     = Total Borrowings / Total Equity
  Net Debt          = Total Borrowings − Cash & Bank − Short-term Investments
  Net D/E           = Net Debt / Total Equity
  Interest Coverage = EBIT / Finance Costs
  Current Ratio     = Current Assets / Current Liabilities
  Quick Ratio       = (Current Assets − Inventory) / Current Liabilities

─── WORKING CAPITAL EFFICIENCY ───
  Inventory Days    = (Average Inventory / COGS) × 365
  Debtor Days       = (Average Trade Receivables / Revenue) × 365
  Creditor Days     = (Average Trade Payables / COGS or Purchases) × 365
  Cash Conversion   = Inventory Days + Debtor Days − Creditor Days
  Asset Turnover    = Revenue / Average Total Assets

─── GROWTH METRICS ───
  YoY growth %      = (Current / Prior − 1) × 100
  QoQ growth %      = (Current Q / Prior Q − 1) × 100
  CAGR over N years = (End / Start)^(1/N) − 1
  Constant currency growth — if reported, use printed; else cannot compute

─── SECTOR-SPECIFIC (FILL WHEN TEMPLATE ASKS) ───

Banks / NBFCs:
  NII (Net Interest Income) = Interest Earned − Interest Expended
  NIM %             = NII / Average Interest-Earning Assets × 100
  Cost-to-Income %  = Operating Expenses / (NII + Other Income) × 100
  GNPA %            = Gross NPA / Gross Advances × 100
  NNPA %            = Net NPA / Net Advances × 100
  PCR %             = Provisions / Gross NPA × 100 (provision coverage)
  CAR / CRAR %      = (Tier 1 + Tier 2 Capital) / RWA × 100
  CASA %            = (Current + Savings Deposits) / Total Deposits × 100
  Credit-Deposit %  = Advances / Deposits × 100
  Cost of Funds %   = Interest Expended / Average Interest-Bearing Liabilities
  Yield on Advances = Interest Earned on Advances / Avg Advances × 100
  AUM (NBFC)        = Total Loan Book under management

Insurance:
  APE               = Regular Premium + 10% × Single Premium
  VNB Margin %      = Value of New Business / APE × 100
  EV                = Embedded Value (printed in disclosures)
  Persistency %     = % of policies still in force at month 13/25/37/49/61
  Solvency Ratio    = Available Capital / Required Capital

Retail / FMCG / QSR:
  SSSG %            = Same-Store-Sales Growth (LFL stores YoY)
  Footfall          = Number of customer visits
  Bill Size / ARPU  = Revenue / Customers (or / Bills)
  Volume Growth %   = Volume YoY change
  Realization       = Revenue / Volume (price per unit)
  Store Count       = Operational stores at period end
  Revenue per sqft  = Revenue / Total Retail Area

IT Services:
  Constant Currency Growth — if printed, use; else cannot derive
  Utilization %     = Billable Hours / Total Available Hours × 100
  Attrition %       = (Voluntary Exits / Avg Headcount) × 100 (LTM)
  Realisation / Bill Rate — usually $ per hour
  Offshore Mix %    = Offshore Revenue / Total Revenue × 100

Power / Energy / Utilities:
  PLF % (Plant Load Factor) = Actual Generation / Installed Capacity × 100
  Generation (MUs)  = Million Units of electricity generated
  Realisation / unit = Revenue / Units sold (Rs / kWh)
  Heat Rate / Specific Fuel Consumption — printed in operating data

Telecom:
  ARPU              = Revenue / Average Subscribers
  Subscribers       = Total active connections at period end
  MoU (Minutes of Use) per subscriber
  Data ARPU / Voice ARPU split if printed

Real Estate / Infra / Capital Goods:
  Order Book        = Backlog of confirmed orders, opening + new − executed
  Book-to-Bill      = New Orders / Revenue
  Pre-Sales / Bookings — area or value of units sold (not yet revenue)
  Execution Period  = Order Book / Annual Revenue (years of revenue visibility)

Auto / Auto Ancillary:
  Volumes (units sold) — domestic + exports + commercial vehicle splits
  Realisation        = Revenue / Volumes
  Capacity Util %    = Production / Installed Capacity × 100

Pharma (when applicable):
  R&D / Sales %      = R&D Expenses / Revenue × 100
  US / India / RoW   = geography-split revenue %
  Pipeline / ANDA Filings — printed in IPs only

─── INDIAN AS / IND-AS CONVENTIONS YOU KNOW COLD ───
  • "Depreciation, amortization & impairment" in template = the AR's
    "Depreciation and amortisation expense" (impairment is usually folded in;
     if shown separately, ADD them)
  • "Finance cost" singular vs "Finance costs" plural — same concept
  • "Other Income (Net)" — the AR calls this "Other income"; deduct
    netting items only if explicitly separated (e.g. "Other income, net of
    interest expense" → already netted)
  • "Tax" = "Total tax expense" OR Current + Deferred + Earlier-year adjustments
  • "Share of profit of associates" appears ONLY in Consolidated; Standalone
    P&L doesn't carry this line by construction
  • "Minority Interest" / "Non-Controlling Interest" — only in Consolidated
  • "Exceptional Items" / "Extraordinary Items" — usually one-off impairments,
    restructuring, gain on sale. Keep sign as printed
  • Bracketed numbers like (40.90) are negative → emit as -40.90
  • All numbers go in INR Crores. If the PDF prints Lakhs divide by 100;
    Millions divide by 10; USD divide by ÷ (read FX rate from page footer
    or assume INR 83/USD if not stated). Filings under SEBI rules are
    usually already in Lakhs or Crores — read the unit line above the table
  • Banks report in INR Crores or Lakhs in their P&L; Operating Profit ≠ EBITDA
    for banks (EBITDA concept doesn't apply); use NII-based equivalents

─── CROSS-VALIDATION YOU SHOULD DO AS YOU FILL ───
  • PBT you compute from components should match PBT printed in source ±0.5
  • Total Assets should match Total Equity + Total Liabilities
  • Quarterly values × 4 (very roughly) should match annual — but expect
    seasonality in Q4 (defence, capital goods, building materials, retail
    festive Q3, agri Q4)
  • Net cash change in CF = Closing − Opening cash
  • Sum of segment revenues = Total Revenue (or close, after eliminations)
  • Gross block − Accumulated Depreciation = Net Block of PPE

─── THINGS YOU'D NEVER DO AS A REAL ANALYST ───
You're senior. You wouldn't make these rookie mistakes. Reading these will
keep your output clean — they're not refusal rules, they're craft:

1. ❌ Same value across quarters of one row.
   Revenue/COGS/Employee/Other-Expense/D&A/Finance-Cost/Tax change every
   single quarter in real companies. If your Q1 = Q2 = Q3 = Q4 for the
   same line, you've misread the column headers — go back, re-map the
   "Three months ended..." dates to the right quarter labels. The classic
   trap: a result PDF prints the CURRENT quarter and a COMPARISON quarter
   (same period prior year) and a YTD column — you've matched all three
   columns to the same template year. Re-check.

2. ❌ Consolidated row = Standalone row for the same label/year.
   Indian listed companies report Consolidated and Standalone separately
   for a reason — subsidiaries differ. If your two values match exactly,
   you pulled both from the SAME section of the source and mis-labeled
   one of them. Cross-check the [SCOPE=...] tag on the page you sourced
   each from.

3. ❌ Quarterly value > Annual value for the same line.
   A single quarter cannot exceed the full year. If your number busts
   that, you've got a unit mismatch — Lakhs read as Crores, USD read as
   INR, or you read Q+YTD as Q-only. Convert and re-emit.

4. ❌ Treating a percentage as a rupee value.
   "Segment-X 59%" is a share-of-revenue, not Rs 59 Cr. If the slide shows
   "Revenue ₹9151 Cr — Segment-X 59%", the rupee value of Segment-X is
   9151 × 0.59 = 5399 Cr. Never paste a percentage into a Crores column.

5. ❌ Inventing values when the source doesn't print them.
   When the AR has the Standalone P&L on pages 195-198 and the Consolidated
   P&L on pages 250-253, the AR has BOTH. But if a quarterly result PDF
   only carries Consolidated (skipping Standalone for that filing), you
   simply have no Standalone source — emit nothing for those cells rather
   than copying the Consolidated number across.

6. ❌ Rounding two distinct values to the same rounded number.
   If you compute COGS = 2451.24 and a different label rounds to the
   same printed value, double-check — coincidences exist but they're
   rare. Verify each component independently.

If any of these would happen, you've made a reading error — pause, re-map,
emit the correct value. Real analyst, real reading.

═══════════════════════════════════════════════════════════════════════════
READING FLATTENED PROSE-TABLES — CRITICAL FOR INDIAN ARs
═══════════════════════════════════════════════════════════════════════════

Indian Annual Reports often print Balance Sheet and P&L statements in a
multi-column layout (current period on the left, comparative periods to
the right — typically 2 columns in standard ARs, but 3-10 columns in
integrated reports / 5-year historical summaries). Our PDF extractor
frequently flattens these into PROSE-LIKE LINES where all numeric columns
sit adjacent on a single line.

YOU WILL SEE LINES LIKE:

    (a) Equity Share capital 15 1,193.32 1,185.91
    (b) Other Equity 16 4,068.47 3,937.21
    Total Equity 5,261.79 5,123.12
    Revenue from operations 7,081.55 6,559.10
    Finance Costs 344.18 198.25
    Trade Receivables 9 & 52 909.56 989.22

Or longer multi-period lines (in 5-year historical summaries):

    Revenue from operations 7,081.55 6,559.10 5,388.39 4,602.92 4,109.13
    PAT (186.46) 375.96 191.95 119.46 (49.41)

This pattern is:
    <label>  [optional note number]  <YEAR-1>  <YEAR-2>  [YEAR-3]  ...

The CONTEXT above these lines (a few lines or pages earlier) will tell
you the period columns. Examples of column-header lines you'll see:
    "Particulars  March 31, 2024  March 31, 2023"
    "Particulars  March 31, 2024  March 31, 2023  March 31, 2022  March 31, 2021"
    "As at  31 March 2023  31 March 2022"
    "FY2024  FY2023  FY2022  FY2021  FY2020"
    "Year ended December 31, 2023  Year ended December 31, 2022"

When you encounter a flattened table:
  • Map left-to-right: leftmost number → leftmost year column, etc.
  • Note-reference integers (5, 9 & 52, 15, 16) are not values — skip them
  • Negative values come in brackets: (40.90) → -40.90
  • The same approach works for fiscal years ending March/December/September
    /June/January — read the month from the header

EXAMPLE — given the lines:
    Particulars March 31, 2024 March 31, 2023 March 31, 2022
    Equity Share capital 15 1,250.40 1,193.32 1,185.91

Emit three fills for that row: F2024=1250.40, F2023=1193.32, F2022=1185.91.

DO NOT skip a row just because the data is in prose form rather than a
clean markdown table. The flattened pattern is the NORM for Indian ARs
after PDF extraction. Read every such line, regardless of how many year
columns appear.

═══════════════════════════════════════════════════════════════════════════
SCHEDULES / NOTES / ANNEXURES — THE BREAKDOWN IS LATER IN THE DOC
═══════════════════════════════════════════════════════════════════════════

The headline financial statements (P&L, BS, CF) print ONE-LINE totals like
"Other expenses 1,569" or "Borrowings 4,200" — the BREAKDOWN of those
totals appears in supporting tables that follow. These supporting tables
have many different names depending on the company:

  • "Notes to the financial statements"          (most common in ARs)
  • "Notes forming part of the financial statements"
  • "Schedules forming part of accounts"
  • "Schedule X" (Schedule III of Companies Act format)
  • "Annexures"
  • "Explanatory Notes"
  • "Disclosures"
  • "Additional Information"
  • "Supporting Schedules"
  • Plain numbered notes: "Note 23", "Note 2.5", "(Refer Note 18)"

Headline rows in the AR usually have a Note reference: "Other expenses (Note
23) 1,569". When the template asks for a line item that's part of a
breakdown, FOLLOW the note reference to the schedule pages later in the
PDF, and pull the sub-line you need.

Common schedules you'll need to find:

  Property, Plant & Equipment (PPE) — schedule shows Gross Block, Additions,
    Disposals, Accumulated Depreciation, and Net Block, broken out by asset
    class (Land, Building, Plant & Machinery, Furniture, Vehicles, Computers,
    Office Equipment, Leasehold Improvements). The headline BS shows only
    Net Block; for any asset-class detail, go to this schedule.

  Inventories — breakdown into Raw Materials, Work-in-Progress, Finished
    Goods, Stock-in-Trade, Stores & Spares, Packing Materials, Goods in Transit.

  Trade Receivables — ageing buckets (0-180 days, 180-365, 1-2 years, 2-3,
    >3 years), Secured vs Unsecured, Considered Good / Doubtful.

  Investments — split into Equity (Quoted/Unquoted), Bonds & Debentures,
    Mutual Funds, Subsidiaries, Associates, Joint Ventures. Cost vs Fair
    Value disclosures.

  Borrowings — Long-term + Short-term, Secured vs Unsecured, Currency
    (INR / USD / EUR), Interest Rate range, Maturity schedule, Lender list.

  Revenue from Operations — Sale of Products vs Services vs Royalties vs
    Exports, geography split (Domestic / Exports / Specific countries).

  Other Income — Interest Income, Dividend Income, Net Gain on FX, Net
    Gain on Investments, Government Grants, Liabilities Written Back,
    Miscellaneous Income.

  Employee Benefits Expense — Salaries & Wages, Contribution to PF,
    Gratuity, Leave Encashment, Staff Welfare, ESOP cost, Training.

  Other Expenses — Power & Fuel, Rent, Repairs, Insurance, Rates & Taxes,
    Travelling, Communication, Legal & Professional, Advertising & Promotion,
    Selling Expenses, R&D, Bad Debts, CSR, Donations, Auditor's Remuneration,
    Bank Charges, Loss on Sale of Assets, FX Loss, Misc.

  Finance Costs — Interest on Borrowings, Interest on Lease Liabilities,
    Other Borrowing Costs (processing fees, premium amortization).

  Tax Expense — Current Tax + Deferred Tax + Earlier-Year Adjustments
    + MAT Credit utilized.

  Contingent Liabilities — Disputed tax demands, Guarantees given, Capital
    Commitments. Usually NOT on BS but shown as a note.

  Related Party Transactions — purchases / sales / loans / remuneration
    to KMP and group companies.

  Segment Reporting (Ind-AS 108) — Primary segment (business) and
    Secondary segment (geography) revenue, results, assets. This is GOLD
    for Operating Metrics rows.

When a template row references a sub-component (e.g. "Power & Fuel" under
Other Expenses, or "Goodwill" under Intangibles, or "Domestic Revenue"
under Revenue), GO TO the matching schedule and extract the printed value.
Don't approximate from totals if the breakdown is available.

═══════════════════════════════════════════════════════════════════════════
CASH FLOW — IT'S IN EVERY AR, FIND IT
═══════════════════════════════════════════════════════════════════════════

CF appears in every AR, usually pages 200-280 out of ~300. Three sections:
Operating, Investing, Financing. Then a "Net increase/(decrease) in cash"
total and Opening + Closing cash balances. Common label translations:

  Template "Cash from operating activities" ≡ AR "Net cash flow from operating activities"
  Template "Cash from investing activities" ≡ AR "Net cash used in investing activities"
  Template "Cash from financing activities" ≡ AR "Net cash used in / from financing activities"
  Template "Cash and bank, beginning of year" ≡ AR "Cash and cash equivalents at the beginning of the year"
  Template "Cash and bank, end of year"       ≡ AR "Cash and cash equivalents at the end of the year"
  Template "Net change in cash"               ≡ AR "Net increase/(decrease) in cash and cash equivalents"
  Template "Income tax paid"                  ≡ AR "Direct taxes paid (net of refunds)"

When CF is split into Standalone CF and Consolidated CF in the same AR,
fill BOTH sections of your template.

═══════════════════════════════════════════════════════════════════════════
OPERATING METRICS — SEGMENT-LEVEL DATA AND KPIs
═══════════════════════════════════════════════════════════════════════════

OM rows are usually segment splits (business-line revenue and %, geography
mix, product mix), KPIs (volumes, capacity, store count, employees, AUM,
GNPA%, NIM, ARPU, etc.), or derived ratios (margin %, growth %, mix %).
The richest source is the Investor Presentation slide deck; the AR's
segment-reporting note and management discussion section are good backups.
Read carefully:

  • When an IP shows "Revenue X — Segment-A 60% / Segment-B 30% / Segment-C 10%":
      Total Revenue = X (in the unit printed)
      Segment-A value = X × 0.60  (derive)
      Segment-B value = X × 0.30
      Segment-C value = X × 0.10
    If the same slide deck has a separate slide with the rupee value of
    each segment, use the PRINTED rupee figure (printed > derived).
  • Watch the unit on every slide: IPs sometimes mix INR Crores, INR Lakhs,
    USD millions. Convert to INR Crores before emitting.
  • Different slides have different periods (LTM, FY, YTD, calendar year).
    Match the period label carefully — don't merge LTM with FY.
  • Banking/NBFC metrics (GNPA%, NNPA%, CAR, NIM, RoA, RoE), retail/FMCG
    (volume growth, store count, ARPU), and infrastructure (order book,
    EPC mix) are all valid OM rows — fill them from wherever they're printed.

═══════════════════════════════════════════════════════════════════════════
HOW TO ANSWER
═══════════════════════════════════════════════════════════════════════════

Output JSON. One fill per cell. Schema:
  {
    "sheet": "<exact template sheet name>",
    "label": "<exact template row label, copied verbatim — don't paraphrase>",
    "statement_type": "Consolidated" | "Standalone" | "",
    "year": "<F2025 | 1QF2025 | ...>",
    "value": <number in INR Crores>,
    "confidence": "HIGH" | "MEDIUM" | "LOW"
  }

Confidence guide:
  HIGH   — printed verbatim in the source, exact column match
  MEDIUM — derived via a clean identity (e.g. COGS sum), or value
           confirmed across two sources
  LOW    — single-source, partial match, or unit conversion involved

Skip rows that are formula-computed in the template (Gross Profit, EBITDA,
EBIT, PBT, PAT, % of sales) — Excel calculates those.

Fill confidently. You're the analyst. The source PDFs ARE the data.
"""


def _build_user_message(template_specs: list[TemplateRowSpec],
                         pdfs: list[tuple[PDFClassification, str]]) -> str:
    """Build the user message: PDF metadata + relevant text + template rows."""
    parts: list[str] = []

    parts.append("# AVAILABLE PDFs\n")
    for cls, _ in pdfs:
        parts.append(
            f"\n## {cls.name}\n"
            f"  - Statement type: **{cls.stmt_type}**\n"
            f"  - Period type:   **{cls.period_type}**\n"
            f"  - Total pages:    {cls.pages_total}\n"
            f"  - Title preview:  {cls.title_text[:160]!r}\n"
        )

    # ── PDF CONTENT — REVERTED to simple flat truncation (r22 era) ──
    # Removed: TOC injection, scope-based budget splits, page priority
    # scoring. Those caused cross-contamination by re-ordering content and
    # confusing the LLM about which year/scope a value belongs to. The
    # simple "flat 60K per PDF in page order" gave the best output (r22).
    parts.append("\n\n# PDF CONTENT (financial statement pages only)\n")
    for cls, full_text in pdfs:
        parts.append(f"\n=== {cls.name} ({cls.stmt_type}, {cls.period_type}) ===\n")
        parts.append(full_text[:60_000])
        parts.append("\n")

    parts.append("\n\n# TEMPLATE ROWS TO FILL\n")
    parts.append(
        "Each row below needs values for the listed years. "
        "Use the EXACT label and year strings when emitting JSON.\n\n"
    )
    # Group by sheet for readability
    by_sheet: dict[str, list[TemplateRowSpec]] = {}
    for r in template_specs:
        if r.is_formula:
            continue
        by_sheet.setdefault(r.sheet, []).append(r)
    for sheet, rows in by_sheet.items():
        parts.append(f"\n## Sheet: {sheet}\n")
        for r in rows:
            ctx = f" [in: {r.section}]" if r.section else ""
            stmt = f" [{r.statement_type}]" if r.statement_type else ""
            yrs = ", ".join(r.years_needed)
            parts.append(f"  - {r.label}{stmt}{ctx} → fill: {yrs}\n")

    parts.append("\n\n# OUTPUT\n")
    parts.append(
        "Return a JSON object with key 'fills' containing a JSON array. "
        "Each array element is one cell fill. Fill ONLY what the source PDFs "
        "actually contain — leave the rest empty (do NOT include them in output). "
        "Keep values terse — DO NOT add fields beyond the schema.\n"
        "Schema:\n"
        '{ "fills": [ { "sheet":, "label":, "statement_type":, "year":, '
        '"value":, "confidence": } ] }'
    )

    return "".join(parts)


# ──────────────────────────────────────────────────────────────────────────────
# Excel writing — apply LLM fills with color coding
# ──────────────────────────────────────────────────────────────────────────────

def _write_filled_excel(template_path: str, output_path: str,
                         fills: list[dict]) -> dict:
    """Write the LLM's fills back into the template, color-coded by confidence."""
    import openpyxl
    from openpyxl.styles import PatternFill
    from mapper.template_reader import read_template

    GREEN  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    RED    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    tm = read_template(template_path)
    wb = openpyxl.load_workbook(template_path, data_only=False)

    stats = {"written": 0, "skipped_formula": 0, "skipped_filled": 0,
             "skipped_no_address": 0, "skipped_used_value": 0,
             "skipped_already_written": 0,
             "by_confidence": {},
             "per_source_fills": {},
             "rejection_log": []}
    # CELL LOCK: once any fill lands in (sheet, row, col), no later fill
    # can overwrite it. Pattern fills land first (more accurate), so they
    # win. The LLM only gets to fill GENUINELY EMPTY cells. This prevents
    # cross-section contamination (Standalone value overwriting Consolidated).
    cells_already_written: set[tuple] = set()

    # ── PER-REPORT VALUE POOLS ────────────────────────────────────────────
    # User's design: every PDF (report) gets its own named pool. The master
    # tracker (`used_pool_master`) holds all pool names. Within ONE report,
    # the same value cannot fill two different columns of the same row.
    # ACROSS reports, the same value CAN appear in different cells (legit —
    # e.g. AR-FY23 has Revenue F2023 = 7081, Q4FY23 PDF also has Revenue
    # F2023 = 7081; both should be accepted into the same cell once).
    #
    # The pool name is derived from the fill's "source" field (set by
    # pattern_fill/_extract_prose_pairs to "pattern:filename" or
    # "pattern-prose:filename"). LLM fills don't have a source, so they
    # all go into a single "llm" pool.
    _CONF_RANK = {"HIGH": 0, "MEDIUM": 1, "LOW": 2}
    def _source_rank(fill: dict) -> int:
        """Pattern fills (verbatim from source PDFs) are MORE TRUSTED than
        LLM fills (which can mis-route across scope or year). Pattern wins
        the race for any cell."""
        src = (fill.get("source") or "")
        if src.startswith("pattern"):
            return 0   # pattern fills first
        return 1       # LLM fills second
    fills_sorted = sorted(
        fills,
        key=lambda f: (
            _CONF_RANK.get(f.get("confidence", "MEDIUM"), 1),
            _source_rank(f),
        ),
    )
    used_pool_master: dict[str, dict[tuple, set]] = {}   # pool_name → (sheet, label, stmt) → {values}
    dedup_log: list[str] = []

    def _pool_name(fill: dict) -> str:
        """Derive the pool name from the fill's source field. Each PDF gets
        its own named pool; LLM fills all share the 'llm' pool."""
        src = (fill.get("source") or "llm")
        # Pattern source format: "pattern:<filename>[<scope>]" or "pattern-prose:<filename>[<scope>]"
        m = re.match(r"^pattern(?:-prose)?:([^\[]+)", src)
        if m:
            return m.group(1).strip()
        return "llm"

    def _is_dup_in_row(fill: dict) -> bool:
        v = fill.get("value")
        if not isinstance(v, (int, float)):
            return False
        sheet = fill.get("sheet") or ""
        label = (fill.get("label") or "").strip().lower()
        stmt = (fill.get("statement_type", "") or "").lower() or "consolidated"
        rounded = round(float(v), 2)
        pool_name = _pool_name(fill)
        pool = used_pool_master.setdefault(pool_name, {})
        row_key = (sheet, label, stmt)
        used = pool.setdefault(row_key, set())
        if rounded in used:
            return True
        used.add(rounded)
        return False

    for fill in fills_sorted:
        sheet = fill.get("sheet")
        label = fill.get("label")
        year  = fill.get("year")
        value = fill.get("value")
        conf  = fill.get("confidence", "MEDIUM")
        stmt  = (fill.get("statement_type", "") or "").lower() or "consolidated"
        # LLM frequently appends bracketed metadata to the label that the
        # template doesn't have. Examples observed in the wild:
        #   "Revenue from operations [Standalone]"
        #   "Revenue from operations [Standalone] [in: Consolidated]"
        #   "Revenue from operations [in: Consolidated / Income Statement]"
        # Strip ALL trailing bracketed segments. The statement_type field
        # already carries the section info we need for routing.
        if label:
            label = str(label)
            for _ in range(5):   # repeatedly strip nested suffixes
                stripped = re.sub(
                    r"\s*\[(?:[^\[\]]+)\]\s*$", "",
                    label, flags=re.I,
                ).strip()
                if stripped == label:
                    break
                label = stripped
            fill["label"] = label
        if not sheet or not label or not year or value is None:
            stats["rejection_log"].append({
                "reason": "missing_field",
                "sheet": sheet, "label": label, "year": year,
                "value": value, "stmt": stmt,
            })
            continue

        # USED-VALUE POOL CHECK: if this exact value already filled another
        # column of this same (sheet, row, statement_type), reject this fill.
        # We process highest-confidence first, so HIGH always wins.
        if _is_dup_in_row(fill):
            stats["skipped_used_value"] += 1
            if len(dedup_log) < 30:
                dedup_log.append(
                    f"  USED   {sheet}/{stmt}/{label[:30]} {year}: "
                    f"value {value} already used in another column of this row"
                )
            continue

        addr = tm.get_cell_address(sheet, label, year, statement_type=stmt)
        if not addr:
            # Diagnose WHY get_cell_address failed
            sm_dbg = tm.sheets.get(sheet)
            if sm_dbg is None:
                why = f"sheet '{sheet}' not in template"
            elif year not in sm_dbg.year_cols:
                avail = list(sm_dbg.year_cols.keys())[:8]
                why = f"year '{year}' not in template (available: {avail})"
            else:
                # Year column exists; row resolution must have failed
                from mapper.template_reader import _resolve_row, clean_label
                row = _resolve_row(sm_dbg, label, stmt)
                if row is None:
                    similar = [k for k in sm_dbg.row_index.keys()
                               if clean_label(k).lower()[:15] == clean_label(label).lower()[:15]][:3]
                    why = (f"label '{label}' not found in {sheet}/{stmt} "
                           f"(similar: {similar})")
                else:
                    why = "unknown — addr None but row resolved"
            stats["skipped_no_address"] += 1
            stats["rejection_log"].append({
                "reason": "no_address", "why": why,
                "sheet": sheet, "label": label, "year": year,
                "value": value, "stmt": stmt,
            })
            continue
        sm = tm.sheets.get(sheet)
        if not sm:
            continue
        key = (addr.row, addr.col)
        if key in sm.formula_cells:
            stats["skipped_formula"] += 1
            continue
        if key in sm.filled_cells:
            stats["skipped_filled"] += 1
            continue
        # CELL LOCK — pattern fills land first, lock the cell, LLM can't overwrite
        global_key = (sheet, addr.row, addr.col)
        if global_key in cells_already_written:
            stats["skipped_already_written"] += 1
            continue
        ws = wb[sheet]
        try:
            ws.cell(addr.row, addr.col).value = float(value)
        except (TypeError, ValueError):
            continue
        cells_already_written.add(global_key)
        # Color by confidence
        fill_color = {"HIGH": GREEN, "MEDIUM": YELLOW, "LOW": RED}.get(
            conf.upper(), YELLOW
        )
        ws.cell(addr.row, addr.col).fill = fill_color
        stats["written"] += 1
        stats["by_confidence"][conf] = stats["by_confidence"].get(conf, 0) + 1
        pool_name = _pool_name(fill)
        stats["per_source_fills"][pool_name] = stats["per_source_fills"].get(pool_name, 0) + 1

    if dedup_log:
        stats["dedup_log"] = dedup_log

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return stats


# ──────────────────────────────────────────────────────────────────────────────
# Main entry
# ──────────────────────────────────────────────────────────────────────────────

@dataclass
class LLMFillResult:
    output_path:    str
    cells_written:  int = 0
    by_confidence:  dict[str, int] = field(default_factory=dict)
    pdfs_used:      list[dict] = field(default_factory=list)
    fills:          list[dict] = field(default_factory=list)
    raw_response:   str = ""
    elapsed_sec:    float = 0.0
    warnings:       list[str] = field(default_factory=list)

    def summary(self) -> str:
        bc = self.by_confidence
        return (f"{self.cells_written} cells written "
                f"(HIGH={bc.get('HIGH',0)}, MEDIUM={bc.get('MEDIUM',0)}, "
                f"LOW={bc.get('LOW',0)}) · {len(self.pdfs_used)} PDFs · "
                f"{self.elapsed_sec:.1f}s")


def _auto_fetch_pdfs(symbol: str, years: list[str], save_dir: str,
                     bse_code: Optional[str] = None,
                     name: Optional[str] = None) -> list[str]:
    """
    Pull THREE PDF categories so the LLM has matching sources for every sheet:
      1. Annual Reports        → fills Annual P&L / BS / Cash Flow
      2. Quarterly Result PDFs → fills Quarterly sheet (latest 4 quarters)
      3. Investor Presentations → fills Operating Metrics / segments
    Saves into <save_dir>/ with prefixed names so user can verify.
    """
    from web.base import CompanyIdentifier
    from pipeline.pdf_discovery import discover_all
    from pipeline.pdf_harvest import _download_cached
    from pathlib import Path as _P

    company = CompanyIdentifier(
        name=name or symbol,
        nse_symbol=symbol.upper() if symbol else None,
        bse_code=bse_code,
    )
    save_path = _P(save_dir)
    save_path.mkdir(parents=True, exist_ok=True)
    saved: list[str] = []

    # Final safety net — even if upstream filters miss, the LLM-fill auto-fetch
    # will skip any filename that contains junk markers (audio, transcript,
    # intimation, BM-outcome, AGM notice, etc.). The shared rule lives in
    # web.pdf_fetcher so we get one source of truth.
    try:
        from web.pdf_fetcher import _is_filing_junk
    except Exception:
        def _is_filing_junk(fn: str, link_text: str = "") -> bool:  # noqa
            return False

    # ── COVER-PAGE VERIFICATION ──────────────────────────────────────────
    # Open the PDF in-memory and read the first 1-2 pages to confirm the
    # document type matches what we expected (AR / QR / IP). Filename and
    # announcement-subject heuristics catch most junk; cover-page verification
    # catches the residue (e.g. NSE serves a "intimation about Q1 results"
    # letter when the announcement says "Q1 Financial Results" — same
    # subject, totally different content).
    _AR_COVER_RE = re.compile(
        # Title-page markers
        r"annual\s+report|integrated\s+(annual\s+)?report"
        r"|annual\s+accounts|notice\s+of\s+(the\s+)?(\d+\w*\s+)?annual\s+general"
        r"|board['']?s?\s+report|director['']?s?\s+report"
        # Inside-AR markers (in case page 1 is just a cover image with company name)
        r"|year\s+ended\s+(march|31[\s\-]?march)"
        r"|standalone\s+balance\s+sheet|consolidated\s+balance\s+sheet"
        r"|standalone\s+statement\s+of\s+profit|consolidated\s+statement\s+of\s+profit"
        r"|management\s+discussion\s+and\s+analysis|md&a|mda"
        r"|corporate\s+overview|financial\s+statements?\b"
        r"|chairman['']?s?\s+(message|letter|statement)"
        r"|notice\s+to\s+(the\s+)?(members|shareholders)",
        re.I,
    )
    _QR_COVER_RE = re.compile(
        r"statement\s+of\s+(unaudited|audited|standalone|consolidated).{0,30}financial\s+results?"
        r"|financial\s+results?\s+for\s+the\s+(quarter|three\s+months|half[\s\-]year|year)\s+ended"
        r"|results?\s+for\s+the\s+(quarter|three\s+months|half[\s\-]year|year)\s+ended"
        r"|unaudited\s+financial\s+results?|audited\s+financial\s+results?",
        re.I,
    )
    _IP_COVER_RE = re.compile(
        # Only accept TITLE-page markers — must literally call itself a
        # "presentation" or "deck". Generic financial docs that just
        # mention "₹ Cr" or segment names don't count.
        r"investor[\s'’]+(presentation|deck|update|brief)"
        r"|earnings\s+presentation|results?\s+presentation"
        r"|q[1-4]\s*fy\d{2,4}.{0,30}(presentation|deck)"
        r"|fy\d{2,4}\s+(annual\s+)?(results?\s+)?presentation"
        r"|analyst\s+(day|meet)\s+presentation",
        re.I,
    )
    _NEG_COVER_RE = re.compile(
        r"intimation|outcome\s+of\s+(the\s+)?board\s+meeting|board\s+meeting\s+outcome"
        r"|notice\s+of\s+(meeting|board\s+meeting|postal\s+ballot)"
        r"|transcript|verbatim|recording\s+of"
        r"|remuneration\s+policy|nomination\s+and\s+remuneration"
        r"|whistle.?blower|code\s+of\s+conduct"
        r"|loss\s+of\s+share|duplicate\s+share|appointment\s+of|resignation\s+of",
        re.I,
    )

    def _read_first_pages(data: bytes, n: int = 2) -> str:
        """Extract text from first N pages of a PDF blob. Empty on failure."""
        try:
            import pdfplumber, io
            with pdfplumber.open(io.BytesIO(data)) as pdf:
                texts = []
                for i in range(min(n, len(pdf.pages))):
                    t = pdf.pages[i].extract_text() or ""
                    texts.append(t)
                return "\n".join(texts)
        except Exception as exc:
            logger.debug(f"cover-page read failed: {exc}")
            return ""

    def _verify_cover(data: bytes, kind: str) -> tuple[bool, str]:
        """
        Returns (ok, reason). kind ∈ {"AR","QR","IP"}.
        ARs need a wider read (first ~10 pages) because page 1 is often a
        glossy cover image with just the company name. QR/IPs typically have
        all signal on page 1-2.
        """
        # Negative check on first 2 pages — that's where the "intimation"
        # boilerplate sits in junk filings.
        head_short = _read_first_pages(data, n=2)
        if head_short.strip() and _NEG_COVER_RE.search(head_short[:4000]):
            return False, "cover page is junk filing (intimation/BM-outcome/transcript)"

        # Positive check — read more pages for ARs since title might be on p3+
        n_pages = 10 if kind == "AR" else 3
        head = _read_first_pages(data, n=n_pages)
        if not head.strip():
            return True, "empty extract — kept (image-only PDF likely)"
        sample = head[:25_000] if kind == "AR" else head[:6000]
        if kind == "AR":
            return (bool(_AR_COVER_RE.search(sample)), "no annual-report markers in first 10 pages")
        if kind == "QR":
            return (bool(_QR_COVER_RE.search(sample)), "no quarterly-result markers on cover")
        if kind == "IP":
            return (bool(_IP_COVER_RE.search(sample)), "no investor-presentation markers on cover")
        return True, "unknown kind — kept"

    # ── Sweep stale junk from PRIOR runs ─────────────────────────────────
    # The save folder is cumulative; files from older runs (saved before the
    # current junk filter) keep polluting the visible folder. Delete anything
    # that the current filter would reject so the user only sees real PDFs.
    swept = 0
    for old in save_path.glob("*.pdf"):
        if _is_filing_junk(old.name) or old.name.startswith("user_upload_tmp"):
            try:
                old.unlink()
                swept += 1
            except Exception:
                pass
    if swept:
        logger.info(f"llm_fill auto-fetch: swept {swept} stale junk file(s) from {save_path}")

    def _save(prefix: str, url: str, fn: str, data: bytes) -> Optional[str]:
        fn = (fn or "doc.pdf").replace("/", "_")
        if not fn.lower().endswith(".pdf"):
            fn += ".pdf"
        if _is_filing_junk(fn):
            logger.info(f"llm_fill auto-fetch: skipping junk filename {fn}")
            return None
        out = save_path / f"{prefix}_{fn}"[:200]
        try:
            out.write_bytes(data)
            logger.info(f"llm_fill auto-fetch: saved {out.name} ({len(data)//1024}KB)")
            return str(out)
        except Exception as exc:
            logger.warning(f"llm_fill auto-fetch save failed for {fn}: {exc}")
            return None

    def _try_category(entries: list[dict], kind: str, prefix_fn,
                      take_after_verify: int) -> int:
        """Download candidates, verify cover page, save until target reached."""
        kept = 0
        # Walk further down the list than `take_after_verify` so we can skip
        # verifier-rejected files and still reach the target count.
        for entry in entries[: take_after_verify * 4]:
            if kept >= take_after_verify:
                break
            url = entry.get("url")
            if not url:
                continue
            try:
                data = _download_cached(url)
            except Exception as exc:
                logger.warning(f"{kind} download failed: {exc}")
                continue
            if not data:
                continue
            ok, reason = _verify_cover(data, kind)
            fn = entry.get("filename") or f"{kind.lower()}.pdf"
            if not ok:
                logger.info(f"llm_fill auto-fetch: REJECT {kind} {fn[:60]} "
                            f"— {reason}")
                continue
            saved_path = _save(prefix_fn(entry), url, fn, data)
            if saved_path:
                saved.append(saved_path)
                kept += 1
        return kept

    # ── 1. Annual Reports ────────────────────────────────────────────────
    discovery = discover_all(company, years)
    logger.info(f"llm_fill auto-fetch: {len(discovery.urls)} AR candidates")
    n_ar = _try_category(
        discovery.urls, "AR",
        lambda e: f"AR_{e.get('year') or 'x'}_{e.get('source') or 'auto'}",
        take_after_verify=3,
    )
    logger.info(f"llm_fill auto-fetch: kept {n_ar} ARs after verification")

    # ── 2. Quarterly Result PDFs ─────────────────────────────────────────
    try:
        from web.pdf_fetcher import get_available_quarterly_results
        qrs = get_available_quarterly_results(symbol.upper(), bse_code) or []
    except Exception as exc:
        logger.warning(f"QR discovery failed: {exc}"); qrs = []
    logger.info(f"llm_fill auto-fetch: {len(qrs)} QR candidates")
    n_qr = _try_category(
        qrs, "QR",
        lambda e: f"QR_{e.get('year') or 'x'}_{e.get('quarter') or 'Q'}",
        take_after_verify=4,
    )
    logger.info(f"llm_fill auto-fetch: kept {n_qr} QRs after verification")

    # ── 3. Investor Presentations ────────────────────────────────────────
    try:
        from web.pdf_fetcher import get_available_presentations
        ips = get_available_presentations(symbol.upper(), bse_code) or []
    except Exception as exc:
        logger.warning(f"IP discovery failed: {exc}"); ips = []
    logger.info(f"llm_fill auto-fetch: {len(ips)} IP candidates")
    n_ip = _try_category(
        ips, "IP",
        lambda e: f"IP_{e.get('year') or 'x'}",
        take_after_verify=2,
    )
    logger.info(f"llm_fill auto-fetch: kept {n_ip} IPs after verification")

    logger.info(f"llm_fill auto-fetch: total saved = {len(saved)} PDFs")
    return saved


# ──────────────────────────────────────────────────────────────────────────────
# PATTERN MATCHER — fills "obvious" cells without LLM. Free + fast.
# Strategy: parse markdown tables from cached .md files, match template label
# to table row, map column header date → template year, extract value.
# ──────────────────────────────────────────────────────────────────────────────

_NUM_RE = re.compile(r"-?\(?\s*\d{1,3}(?:[,\s]\d{3})*(?:\.\d+)?\s*\)?")
_MONTH_NAMES = {
    "jan": 3, "feb": 3, "mar": 3, "apr": 1, "may": 1, "jun": 1,
    "jul": 2, "aug": 2, "sep": 2, "oct": 3, "nov": 3, "dec": 3,
}   # month → quarter-end mapping for Indian FY

def _parse_number(cell: str) -> Optional[float]:
    """Parse '1,063.20' / '(40.90)' / '5 ,389.12' → float. Return None if not numeric."""
    if not cell or not cell.strip():
        return None
    s = cell.strip().replace(" ", "").replace(",", "")
    is_neg = "(" in s and ")" in s
    s = s.replace("(", "").replace(")", "")
    s = s.lstrip("-").strip()
    if not s:
        return None
    try:
        v = float(s)
        if is_neg or cell.strip().startswith("-"):
            v = -v
        return v
    except ValueError:
        return None


def _normalise_label(s: str) -> str:
    """Lowercase, strip punctuation/whitespace for fuzzy matching."""
    s = (s or "").lower()
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


# Common CF label aliases — different ARs phrase the same line dozens of ways.
# Each canonical key maps to a list of regex fragments. When a template label
# OR a table row label matches any fragment in a group, it's collapsed to the
# canonical key so they match each other.
_CF_LABEL_ALIASES: dict[str, tuple[str, ...]] = {
    "cf operating":   (
        r"\bnet cash (?:flow )?(?:generated )?from operating",
        r"\bcash (?:flow )?from operating",
        r"\bcash generated from operations",
        r"\boperating activities\b",
    ),
    "cf investing":   (
        r"\bnet cash (?:flow )?(?:used )?in investing",
        r"\bcash (?:flow )?from investing",
        r"\binvesting activities\b",
    ),
    "cf financing":   (
        r"\bnet cash (?:flow )?(?:used )?in financing",
        r"\bcash (?:flow )?from financing",
        r"\bfinancing activities\b",
    ),
    "cf cash beg":    (
        r"\bcash.*beginning(?: of)?(?: the)? (?:year|period)",
        r"\bcash.*at(?: the)? beginning",
        r"\bopening (?:balance of )?cash",
        r"\bcash and bank balances? beginning",
    ),
    "cf cash end":    (
        r"\bcash.*end(?: of)?(?: the)? (?:year|period)",
        r"\bcash.*at(?: the)? end",
        r"\bclosing (?:balance of )?cash",
        r"\bcash and bank balances? end",
    ),
    "cf net change":  (
        r"\bnet (?:increase|decrease|change|change/?\(decrease\)).{0,30}in cash",
        r"\bnet increase\s+decrease.{0,30}in cash",
        r"\bnet change in cash",
    ),
    "cf depreciation": (
        r"\bdepreciation and amortis(?:ation|ation expense)",
        r"\bdepreciation amortisation and impairment",
    ),
    "cf finance cost": (
        r"\bfinance cost(?:s)?\b",
        r"\binterest expense\b",
    ),
    "cf working capital": (
        r"\bworking capital changes",
        r"\bchanges in working capital",
    ),
    "cf tax paid":    (
        r"\bincome tax(?:es)? paid",
        r"\bdirect taxes? paid",
    ),
}

def _cf_canonical(label_norm: str) -> Optional[str]:
    """If `label_norm` (already lowercase) matches a CF alias group, return the
    canonical key. Otherwise None. Used to bridge wording differences between
    template and AR for Cash Flow rows specifically."""
    for canon, pats in _CF_LABEL_ALIASES.items():
        for pat in pats:
            if re.search(pat, label_norm):
                return canon
    return None


# ── UNIVERSAL P&L / BS LABEL ALIASES — generic, sector-neutral ──
# Keyed by canonical concept; each entry has [regex_patterns] for template
# label matching AND [source_patterns] for finding values in the source PDF.
_PL_BS_ALIASES: dict[str, dict[str, tuple[str, ...]]] = {
    "revenue": {
        "template": (r"^revenue from operations?$", r"^total revenue", r"^net revenue", r"^sales\b"),
        "source":   (r"\brevenue from operations?\b", r"\bnet sales\b", r"\btotal revenue\b"),
    },
    "other_income": {
        "template": (r"other income", r"other operating income"),
        "source":   (r"\bother income\b\s*(?:\(net\))?"),
    },
    "cogs_materials": {
        "template": (r"cost of materials consumed", r"cost of raw materials"),
        "source":   (r"\bcost of materials consumed\b", r"\bcost of raw materials"),
    },
    "cogs_purchases": {
        "template": (r"purchases of stock", r"purchases? in trade"),
        "source":   (r"\bpurchases of stock", r"\bpurchase of (?:traded )?goods"),
    },
    "cogs_inv_change": {
        "template": (r"changes in inventor", r"increase\s*\(decrease\) in inventor"),
        "source":   (r"\bchanges in inventor", r"\bdecrease\s*[/\(]\s*increase in inventor"),
    },
    "cogs_total": {
        "template": (r"cost of goods sold", r"^cogs$", r"^less\s*:\s*cogs"),
        # NO direct source — derived as sum of materials + purchases + Δ inventory
        "source":   (),
    },
    "employee": {
        "template": (r"employee benefit", r"staff cost", r"personnel cost"),
        "source":   (r"\bemployee benefits?\s+expense\b", r"\bstaff costs?\b"),
    },
    "depreciation": {
        "template": (r"depreciation", r"d\s*&\s*a", r"d\&a"),
        "source":   (r"\bdepreciation\s+and\s+amortis[ae]tion\b",
                     r"\bdepreciation,?\s+amortis[ae]tion\s+(?:&|and)\s+impairment"),
    },
    "finance_cost": {
        "template": (r"finance costs?", r"interest expense", r"borrowing costs?"),
        "source":   (r"\bfinance costs?\b", r"\binterest expense\b"),
    },
    "other_expenses": {
        "template": (r"other expenses", r"other operating expenses"),
        "source":   (r"\bother expenses\b\s*(?:\(net\))?"),
    },
    "tax_current": {
        "template": (r"current tax",),
        "source":   (r"\bcurrent tax\b",),
    },
    "tax_deferred": {
        "template": (r"deferred tax",),
        "source":   (r"\bdeferred tax\b",),
    },
    "tax_total": {
        "template": (r"^less\s*:?\s*tax\b", r"^tax expense", r"^total tax\b"),
        "source":   (r"\btotal tax expense\b", r"\btax expense\b\s*$"),
    },
    "exceptional": {
        "template": (r"exceptional items?",),
        "source":   (r"\bexceptional items?\b",),
    },
    "associates": {
        "template": (r"share of (?:net )?profit of associates?", r"share of associates?"),
        "source":   (r"\bshare of (?:net )?profit", r"\bshare of associates?\b"),
    },
    "eps_diluted": {
        "template": (r"diluted eps", r"diluted earnings per share"),
        "source":   (r"\bdiluted\b.*?\b(?:eps|earnings per share)\b",
                     r"\bearnings per share.*?diluted\b"),
    },
    "eps_basic": {
        "template": (r"basic eps", r"basic earnings per share"),
        "source":   (r"\bbasic\b.*?\b(?:eps|earnings per share)\b",
                     r"\bearnings per share.*?basic\b"),
    },
    # Balance sheet entries
    "ppe": {
        "template": (r"property,?\s*plant\s*(?:&|and)\s*equipment", r"\bppe\b", r"net block"),
        "source":   (r"\bproperty,?\s*plant\s*(?:&|and)\s*equipment\b"),
    },
    "cwip": {
        "template": (r"capital work[-\s]in[-\s]progress", r"\bcwip\b"),
        "source":   (r"\bcapital work[-\s]in[-\s]progress\b"),
    },
    "goodwill": {
        "template": (r"^goodwill$",),
        "source":   (r"\bgoodwill\b",),
    },
    "intangibles": {
        "template": (r"^intangible assets?\b", r"intangibles$"),
        "source":   (r"\bintangible assets?\b",),
    },
    "rou": {
        "template": (r"right[-\s]of[-\s]use", r"\brou\b", r"rou assets?"),
        "source":   (r"\bright[-\s]of[-\s]use\b",),
    },
    "inventories": {
        "template": (r"^inventories$", r"^stock\b"),
        "source":   (r"\binventories\b",),
    },
    "trade_receivables": {
        "template": (r"trade receivables?", r"debtors?"),
        "source":   (r"\btrade receivables?\b",),
    },
    "cash": {
        "template": (r"cash (?:&|and) cash equivalents?", r"cash (?:&|and) bank"),
        "source":   (r"\bcash (?:and|&) cash equivalents?\b",
                     r"\bcash on hand\b"),
    },
    "equity_share_cap": {
        "template": (r"equity share capital", r"share capital"),
        "source":   (r"\bequity share capital\b", r"\bissued (?:and )?subscribed"),
    },
    "other_equity": {
        "template": (r"other equity", r"reserves (?:&|and) surplus"),
        "source":   (r"\bother equity\b", r"\breserves (?:and|&) surplus\b"),
    },
    "borrowings": {
        "template": (r"^borrowings$",),
        "source":   (r"\bborrowings\b",),
    },
    "trade_payables": {
        "template": (r"trade payables?", r"sundry creditors?"),
        "source":   (r"\btrade payables?\b",),
    },
}


def _pl_bs_canonical_template(label_norm: str) -> Optional[str]:
    """Match a TEMPLATE label to a canonical concept."""
    for canon, patterns in _PL_BS_ALIASES.items():
        for pat in patterns.get("template", ()):
            if re.search(pat, label_norm, re.I):
                return canon
    return None


def _pl_bs_canonical_source(line_norm: str) -> Optional[str]:
    """Match a SOURCE line to a canonical concept."""
    for canon, patterns in _PL_BS_ALIASES.items():
        for pat in patterns.get("source", ()):
            if re.search(pat, line_norm, re.I):
                return canon
    return None


def _derive_aggregated_values(pairs: list) -> list:
    """Derive computed values from emitted pairs.
    pairs is list of (label, year, scope, value, kind).
    Derivations applied:
      COGS = materials + purchases + Δ inventory
      Tax  = current + deferred
    """
    # Group existing pairs by (canon, year, scope)
    by_key = {}
    for label, year, scope, val, kind in pairs:
        ln = (label or "").lower()
        canon = _pl_bs_canonical_source(ln)
        if canon:
            by_key.setdefault((canon, year, scope), []).append(val)

    derived = []
    for (year, scope) in set((y, s) for c, y, s in by_key.keys()):
        # COGS = materials + purchases + inv_change
        mat = by_key.get(("cogs_materials", year, scope), [])
        pur = by_key.get(("cogs_purchases", year, scope), [])
        inv = by_key.get(("cogs_inv_change", year, scope), [])
        if mat and pur and inv:
            cogs_val = mat[0] + pur[0] + inv[0]
            derived.append(("Less: Cost of goods sold", year, scope, cogs_val, "derived_cogs"))
        # Tax = current + deferred (if both present)
        cur = by_key.get(("tax_current", year, scope), [])
        defrd = by_key.get(("tax_deferred", year, scope), [])
        if cur and defrd:
            tax_val = cur[0] + defrd[0]
            derived.append(("Less: Tax", year, scope, tax_val, "derived_tax"))
    return derived


def _column_to_year_label(header: str) -> Optional[str]:
    """
    Map a markdown table column header to a template year/quarter label.
      "Three months ended June 30, 2024" → "1QF2025"
      "Year ended March 31, 2024"        → "F2024"
      "March 31, 2025"                   → "F2025"  (annual context)
      "F2025"                            → "F2025"
    """
    if not header:
        return None
    h = header.lower()
    # Already in template format
    m = re.match(r"^\s*(f\d{4}|[1-4]qf\d{4})\s*$", h)
    if m:
        return m.group(1).upper().replace("F", "F")
    # Date-formatted columns
    m = re.search(
        r"(january|february|march|april|may|june|july|august|september|october|november|december)\s+\d{1,2}\s*,?\s*(\d{4})",
        h,
    )
    if not m:
        return None
    mon_word, yr_str = m.group(1), m.group(2)
    mon = list(_MONTH_NAMES.keys()).index(mon_word[:3])   # 0..11
    is_quarter = "three months" in h or "quarter ended" in h
    is_annual  = "year ended" in h or "for the year" in h
    yr = int(yr_str)
    if is_annual:
        # Year ended March 31, 2024 → F2024
        return f"F{yr}"
    if is_quarter:
        # Indian FY: Mar=Q4 of yr, Jun=Q1 of yr+1, Sep=Q2 of yr+1, Dec=Q3 of yr+1
        if mon_word.startswith("mar"):  return f"4QF{yr}"
        if mon_word.startswith("jun"):  return f"1QF{yr+1}"
        if mon_word.startswith("sep"):  return f"2QF{yr+1}"
        if mon_word.startswith("dec"):  return f"3QF{yr+1}"
    # Bare date column — could be quarter-end or year-end. Default to annual.
    if mon_word.startswith("mar"):
        return f"F{yr}"
    return None


def _parse_md_tables(md_text: str
                      ) -> list[tuple[list[str], list[list[str]], str]]:
    """
    Parse markdown tables out of an MD blob.
    Returns list of (header_row, body_rows, scope) tuples where scope is
    "STANDALONE" / "CONSOLIDATED" / "UNKNOWN" — based on the most recent
    section heading that appeared above the table. The classifier emits
    "[SCOPE=...]" markers on every page boundary, and many ARs print
    section names like "Consolidated Balance Sheet" in heading rows.
    """
    tables: list[tuple[list[str], list[list[str]], str]] = []
    lines = md_text.split("\n")
    current_scope = "UNKNOWN"
    i = 0
    while i < len(lines):
        line = lines[i].strip()

        # Track section context lines that come BEFORE the table
        ll = line.lower()
        # Scope markers from classifier: "--- file.pdf page 12 [SCOPE=CONSOLIDATED] ---"
        m = re.search(r"\[SCOPE=([A-Z_]+)\]", line)
        if m:
            v = m.group(1)
            if v in ("STANDALONE", "CONSOLIDATED"):
                current_scope = v
            elif v in ("UNKNOWN_SCOPE", "UNKNOWN"):
                current_scope = "UNKNOWN"
        # Section headings printed in the document itself
        elif (("consolidated balance sheet" in ll
               or "consolidated statement of profit" in ll
               or "consolidated statement of financial position" in ll
               or "consolidated cash flow" in ll
               or "consolidated financial statement" in ll)):
            current_scope = "CONSOLIDATED"
        elif (("standalone balance sheet" in ll
               or "standalone statement of profit" in ll
               or "standalone statement of financial position" in ll
               or "standalone cash flow" in ll
               or "standalone financial statement" in ll)):
            current_scope = "STANDALONE"

        # Markdown table starts with |...| then |---|---|
        if (line.startswith("|") and line.endswith("|") and
                i + 1 < len(lines) and re.match(r"\|[\s\-:|]+\|", lines[i+1].strip())):
            header = [c.strip() for c in line.strip("|").split("|")]
            body: list[list[str]] = []
            j = i + 2
            while j < len(lines) and lines[j].strip().startswith("|"):
                row = [c.strip() for c in lines[j].strip().strip("|").split("|")]
                if len(row) > 0:
                    body.append(row)
                j += 1
            if header and body:
                tables.append((header, body, current_scope))
            i = j
            continue
        i += 1
    return tables


def _extract_prose_pairs(md_text: str, cls: "PDFClassification"
                          ) -> list[tuple[str, str, str, float, str]]:
    """
    Scan markdown text for flattened multi-column financial lines:

        <label>  [optional note ref]  <YEAR-1>  <YEAR-2>  [YEAR-3] ...

    Works for ANY number of comparative columns (typically 2 in Indian
    annual reports, but some integrated reports / 5-year historical
    tables show 3-10 years side by side). Works for ANY fiscal year-end
    (March 31 / December 31 / September 30 / June 30 / January 31 — any
    month-end pattern).

    Year detection: scans every line for date-headers like
        "March 31, 2024"
        "31 March 2024"
        "December 31, 2023"
        "31st March, 2023"
        "FY2024"
        "Year ended March 31, 2024"
    and stores the ordered list of years detected. Subsequent data lines
    map their N numbers to the N most-recent year columns left-to-right.

    Returns list of (label, year_label, scope, value, kind) tuples.
    """
    out: list[tuple[str, str, str, float, str]] = []
    if not md_text:
        return out

    lines = md_text.split("\n")
    # current_years is the LEFT-to-RIGHT ordered list of year labels for the
    # current section. The most-recent column-header line sets this.
    current_years: list[str] = []
    current_scope = "UNKNOWN"

    NUM = r"-?\(?\s*\d{1,3}(?:,\d{3})*(?:\.\d+)?\s*\)?"
    # Match: label + ≥2 trailing numbers. Use a regex that captures all numbers.
    LINE_RE = re.compile(
        rf"^\s*([A-Za-z][A-Za-z0-9 ,&/\-()'\"]{{3,100}}?)\s+"
        rf"(?:(\d{{1,3}}(?:\s*&\s*\d{{1,3}})?)\s+)?"
        rf"((?:{NUM}\s+){{1,9}}{NUM})\s*$"
    )

    # Month names → indexes (any month-end date might appear as period header)
    MONTH_RE = (r"(?:January|February|March|April|May|June|"
                r"July|August|September|October|November|December|"
                r"Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)")

    # Date patterns for column-header detection (all return year)
    DATE_PATTERNS = [
        # "March 31, 2023" or "31 March 2023" or "31st March, 2023"
        rf"{MONTH_RE}\s+\d{{1,2}},?\s*(\d{{4}})",
        rf"\d{{1,2}}(?:st|nd|rd|th)?\s+{MONTH_RE},?\s*(\d{{4}})",
        # "FY2024" / "FY 24" / "F2024"
        r"\bF[Y]?\s*(\d{4})\b",
        r"\bF[Y]?\s*['‘’]?(\d{2})\b",   # FY'24 / FY24 / FY 24
        # Bare "Year ended <date>"
        rf"Year\s+ended\s+{MONTH_RE}\s+\d{{1,2}},?\s*(\d{{4}})",
        # "As at <date>"
        rf"As\s+at\s+{MONTH_RE}\s+\d{{1,2}},?\s*(\d{{4}})",
    ]
    DATE_REGEXES = [re.compile(p, re.I) for p in DATE_PATTERNS]

    def _parse_num(s: str) -> Optional[float]:
        s = s.strip().replace(",", "").replace(" ", "")
        if not s:
            return None
        neg = "(" in s and ")" in s
        s = s.replace("(", "").replace(")", "")
        try:
            v = float(s)
            return -v if neg else v
        except ValueError:
            return None

    def _is_year_header_line(line: str) -> list[str]:
        """If line is a column-header line, return ordered list of F<YYYY>
        labels for each column detected. Else return []."""
        found: list[tuple[int, int]] = []   # (start_pos, year)
        for rx in DATE_REGEXES:
            for m in rx.finditer(line):
                try:
                    yr_str = m.group(1)
                    yr = int(yr_str)
                    # Handle 2-digit years (e.g. "FY24" → 2024)
                    if yr < 100:
                        yr = 2000 + yr if yr < 50 else 1900 + yr
                    if 1990 <= yr <= 2050:
                        found.append((m.start(), yr))
                except (ValueError, IndexError):
                    pass
        if not found:
            return []
        # Deduplicate years by position (within 30 chars = same column)
        found.sort()
        result: list[str] = []
        last_pos = -100
        last_yr = -1
        for pos, yr in found:
            if yr == last_yr and pos - last_pos < 30:
                continue
            result.append(f"F{yr}")
            last_pos = pos
            last_yr = yr
        return result

    for line in lines:
        ll = line.strip()
        if ll.startswith("|") or ll.startswith("---") or not ll:
            continue
        if len(ll) > 300:
            continue

        # Track scope
        sm = re.search(r"\[SCOPE=([A-Z_]+)\]", ll)
        if sm:
            v = sm.group(1)
            if v in ("STANDALONE", "CONSOLIDATED"):
                current_scope = v
            elif v in ("UNKNOWN_SCOPE", "UNKNOWN"):
                current_scope = "UNKNOWN"

        # Detect column headers (multi-year lines) — update current_years
        # only if line has at least 2 year mentions (else single year could
        # be a date in body text)
        header_years = _is_year_header_line(ll)
        if len(header_years) >= 2:
            current_years = header_years
        elif len(header_years) == 1 and not current_years:
            # Bootstrap with single year — assume prior column = yr - 1
            try:
                yr = int(header_years[0][1:])
                current_years = [f"F{yr}", f"F{yr-1}"]
            except ValueError:
                pass

        if not current_years:
            continue

        m = LINE_RE.match(ll)
        if not m:
            continue
        label = m.group(1).strip().rstrip("- ").rstrip(":")
        label = re.sub(r"^\s*\(?\s*[a-z0-9]+\s*[\)\.]\s*", "", label, flags=re.I).strip()
        if len(label) < 4:
            continue

        # Parse all trailing numbers
        nums_text = m.group(3)
        num_strings = re.findall(NUM, nums_text)
        values = [_parse_num(n) for n in num_strings]
        values = [v for v in values if v is not None]
        if len(values) < 2:
            continue

        # Skip year-like numbers (all are 1990-2050 with no decimals)
        if all(1990 < abs(v) < 2100 and v == int(v) for v in values):
            continue
        # Skip if all values are tiny (likely percentages)
        if all(abs(v) < 1 for v in values):
            continue

        # Assign values to years LEFT-to-RIGHT. If we have N values and M
        # years, take min(N, M). The leftmost number maps to the leftmost
        # year in the header.
        n_to_assign = min(len(values), len(current_years))
        for i in range(n_to_assign):
            yr_lbl = current_years[i]
            val = values[i]
            kind = "current" if i == 0 else ("prior" if i == 1 else f"col{i+1}")
            out.append((label, yr_lbl, current_scope, val, kind))

    return out


def _pattern_fill(template_specs: list[TemplateRowSpec],
                   pdf_data: list[tuple[PDFClassification, str]]) -> list[dict]:
    """
    Run pattern matching on cached MD tables. Returns list of fill dicts in
    the same shape as LLM fills. PRECISION-first: only emit a fill when:
      • Template label matches a table row (exact normalised or 90% substring)
      • Column header maps unambiguously to a template year
      • Value parses cleanly as a number
      • PDF stmt_type AND per-table SCOPE marker match the spec.statement_type
        (a Standalone PDF cannot fill a Consolidated row, and vice versa)
    """
    fills: list[dict] = []
    seen_keys: set[tuple] = set()    # (sheet, label, year, stmt) → already-emitted
    skipped_scope = 0                # diagnostic counter

    # Index template specs by sheet
    by_sheet: dict[str, list[TemplateRowSpec]] = {}
    for s in template_specs:
        if s.is_formula:
            continue
        by_sheet.setdefault(s.sheet, []).append(s)

    def _scopes_compatible(spec_stmt: str, table_scope: str,
                           pdf_stmt: str) -> bool:
        """
        Decide whether a pattern fill can land. Lenient: only blocks when
        the table scope DEFINITIVELY mismatches the spec. Anything ambiguous
        is allowed — let the LLM and downstream readers sort it out.
        """
        ss = (spec_stmt or "").lower()
        ts = (table_scope or "").upper()
        ps = (pdf_stmt or "").upper()

        if not ss:
            return True   # spec has no preference

        # Block only when both signals say "wrong scope"
        if ts == "STANDALONE" and ss == "consolidated":
            return False
        if ts == "CONSOLIDATED" and ss == "standalone":
            return False
        # PDF-level conflict (and table has no scope marker)
        if ts not in ("CONSOLIDATED", "STANDALONE"):
            if ps == "STANDALONE" and ss == "consolidated":
                return False
            if ps == "CONSOLIDATED" and ss == "standalone":
                return False
        return True   # accept anything not definitively mismatched

    for cls, md_text in pdf_data:
        tables = _parse_md_tables(md_text)
        if not tables:
            continue
        scopes_seen = {t[2] for t in tables}
        logger.info(
            f"_pattern_fill: {cls.name} [stmt={cls.stmt_type}] "
            f"has {len(tables)} MD tables across scopes {sorted(scopes_seen)}"
        )

        for header, body, table_scope in tables:
            # Map each column index → template year label
            col_year_map: dict[int, str] = {}
            for ci, hdr in enumerate(header):
                yl = _column_to_year_label(hdr)
                if yl:
                    col_year_map[ci] = yl
            if not col_year_map:
                continue   # table has no year columns we recognize

            # For each table row, try to match the first cell to a template label
            for row in body:
                if not row or not row[0].strip():
                    continue
                row_label_raw = row[0]
                row_label_norm = _normalise_label(row_label_raw)
                if not row_label_norm or len(row_label_norm) < 4:
                    continue

                # Check every sheet's specs for a match
                for sheet_name, specs in by_sheet.items():
                    for spec in specs:
                        # ── Scope gate: BLOCK cross-section contamination ──
                        if not _scopes_compatible(
                                spec.statement_type, table_scope, cls.stmt_type):
                            skipped_scope += 1
                            continue

                        spec_norm = _normalise_label(spec.label)
                        if not spec_norm:
                            continue
                        # CF-specific bridge: when BOTH labels map to the same
                        # CF canonical, accept the match even if surface words
                        # differ. Skips length sanity for CF rows.
                        is_cf_match = False
                        sheet_lc = sheet_name.lower()
                        if ("cash flow" in sheet_lc or " cf" in sheet_lc
                                or sheet_lc.endswith("cf")
                                or "cashflow" in sheet_lc):
                            spec_canon = _cf_canonical(spec_norm)
                            row_canon = _cf_canonical(row_label_norm)
                            if spec_canon and spec_canon == row_canon:
                                is_cf_match = True
                        # Match: exact normalised OR one fully contains the other
                        if not is_cf_match and not (
                                spec_norm == row_label_norm
                                or spec_norm in row_label_norm
                                or row_label_norm in spec_norm):
                            continue
                        # Length sanity — avoid "Tax" matching "Tax expense excluding deferred"
                        # CF canonical match bypasses this (we trust the regex aliases).
                        if not is_cf_match and abs(len(spec_norm) - len(row_label_norm)) > 25:
                            continue
                        # Extract values for each year column
                        for ci, yr in col_year_map.items():
                            if yr not in spec.years_needed:
                                continue
                            if ci >= len(row):
                                continue
                            v = _parse_number(row[ci])
                            if v is None:
                                continue
                            stmt = (spec.statement_type or "").lower() or "consolidated"
                            key = (spec.sheet, spec.label, yr, stmt)
                            if key in seen_keys:
                                continue
                            seen_keys.add(key)
                            # Effective scope for citation tracking
                            eff = (table_scope if table_scope in
                                   ("CONSOLIDATED", "STANDALONE")
                                   else cls.stmt_type)
                            fills.append({
                                "sheet": spec.sheet,
                                "label": spec.label,
                                "statement_type": spec.statement_type or "",
                                "year": yr,
                                "value": v,
                                "confidence": "HIGH",
                                "source": f"pattern:{cls.name}[{eff}]",
                            })
    pre_prose = len(fills)

    # ── PROSE PAIR EXTRACTOR ──
    # In addition to markdown tables, scan the raw md text for flattened
    # two-column lines like "Equity Share capital 15 1,193.32 1,185.91".
    # This catches the Indian-AR pattern where pdfplumber merged the
    # comparative-year column into prose. Same scope-matching rules apply.
    for cls, md_text in pdf_data:
        prose_pairs = _extract_prose_pairs(md_text, cls)
        # Add derived values (COGS = sum of components, Tax = current+deferred)
        # so the pattern matcher can fill rows that don't have a direct
        # printed line in the source (which is normal for Indian ARs that
        # report components rather than totals).
        prose_pairs = prose_pairs + _derive_aggregated_values(prose_pairs)
        if not prose_pairs:
            continue
        logger.info(
            f"_pattern_fill: {cls.name} → {len(prose_pairs)} prose-line "
            f"label-value pairs detected"
        )
        for label_raw, yr, scope, v, kind in prose_pairs:
            label_norm = _normalise_label(label_raw)
            if not label_norm or len(label_norm) < 4:
                continue
            # Find a matching template spec
            for sheet_name, specs in by_sheet.items():
                for spec in specs:
                    if yr not in spec.years_needed:
                        continue
                    if not _scopes_compatible(
                            spec.statement_type, scope, cls.stmt_type):
                        skipped_scope += 1
                        continue
                    spec_norm = _normalise_label(spec.label)
                    if not spec_norm:
                        continue
                    if not (spec_norm == label_norm
                            or spec_norm in label_norm
                            or label_norm in spec_norm):
                        continue
                    if abs(len(spec_norm) - len(label_norm)) > 25:
                        continue
                    stmt_key = (spec.statement_type or "").lower() or "consolidated"
                    key = (spec.sheet, spec.label, yr, stmt_key)
                    if key in seen_keys:
                        continue
                    seen_keys.add(key)
                    eff = scope if scope in ("CONSOLIDATED", "STANDALONE") else cls.stmt_type
                    fills.append({
                        "sheet": spec.sheet,
                        "label": spec.label,
                        "statement_type": spec.statement_type or "",
                        "year": yr,
                        "value": v,
                        "confidence": "HIGH",
                        "source": f"pattern-prose:{cls.name}[{eff}]",
                    })
    prose_added = len(fills) - pre_prose

    logger.info(
        f"_pattern_fill: emitted {len(fills)} cells "
        f"({pre_prose} from MD tables + {prose_added} from prose pairs, "
        f"skipped {skipped_scope} due to scope mismatch)"
    )
    return fills


def llm_fill_template(
    template_path: str,
    pdf_paths:     Optional[list[str]] = None,
    output_path:   str = "",
    api_key:       str = "",
    model:         str = "deepseek-chat",
    base_url:      str = "https://api.deepseek.com",
    max_tokens:    int = 8192,   # DeepSeek's hard ceiling — see salvage logic below
    auto_fetch:    bool = False,
    nse_symbol:    Optional[str] = None,
    bse_code:      Optional[str] = None,
    company_name:  Optional[str] = None,
    years:         Optional[list[str]] = None,
    skip_pattern_match: bool = False,   # if True, send everything to LLM (no pre-pass)
) -> LLMFillResult:
    """
    Fill the template using a single LLM call.

    Either provide `pdf_paths` (user-uploaded files) OR set `auto_fetch=True`
    with `nse_symbol` so the function discovers + downloads the right PDFs
    via the pdf_discovery layers. Auto-fetched files persist to
    output/downloaded_pdfs/<SYMBOL>/ for inspection.
    """
    t0 = time.time()
    res = LLMFillResult(output_path=output_path)

    # 1. Read template structure
    template_specs = _read_template_specs(template_path)
    logger.info(f"llm_fill: template has {len(template_specs)} fillable rows")

    # 1.5 — Auto-fetch PDFs if requested
    pdf_paths = list(pdf_paths or [])
    if auto_fetch and nse_symbol:
        from datetime import date as _date
        yrs = years or [f"F{_date.today().year - i}" for i in range(0, 4)]
        save_root = (Path(__file__).parent.parent / "output"
                     / "downloaded_pdfs" / nse_symbol.upper())
        try:
            fetched = _auto_fetch_pdfs(
                nse_symbol, yrs, str(save_root),
                bse_code=bse_code, name=company_name,
            )
            pdf_paths.extend(fetched)
            res.warnings.append(
                f"Auto-fetched {len(fetched)} PDFs into {save_root}"
            )
        except Exception as exc:
            res.warnings.append(f"Auto-fetch failed: {exc}")

    # 2. Classify and extract text from each PDF
    # ── PRE-FILTER: dedupe by content hash + drop junk filings ──────────
    # Folder accumulates duplicates (same AR saved as AR_*.pdf and F*_.pdf
    # by different code paths) and junk leftovers from prior runs (audio,
    # transcripts, intimations, BMOutcomes, remuneration policies). 15 raw
    # PDFs collapse to ~4 unique real docs after this step. Without it, the
    # LLM prompt blows past DeepSeek's 64k context window and older ARs get
    # silently dropped (which is why F2022/F2023 BS rows stayed empty).
    import hashlib as _hl
    try:
        from web.pdf_fetcher import _is_filing_junk
    except Exception:
        def _is_filing_junk(fn: str, link_text: str = "") -> bool:  # noqa
            return False

    # Composite-key dedupe: (file_size, head_hash, tail_hash). Hashing only
    # the first 1MB previously killed legitimate ARs whose cover/boilerplate
    # was the same as another year's AR. Now we also hash the LAST 256KB AND
    # include file size — only files identical end-to-end collide.
    seen_keys: set[tuple] = set()
    deduped_paths: list[str] = []
    skipped_dup = skipped_junk = 0
    dedupe_log: list[str] = []   # surfaced to user so they can audit
    for pp in pdf_paths:
        try:
            fn = Path(pp).name
            if _is_filing_junk(fn):
                skipped_junk += 1
                dedupe_log.append(f"  JUNK   {fn}")
                logger.info(f"llm_fill: skip junk {fn}")
                continue
            import os as _os
            size = _os.path.getsize(pp)
            with open(pp, "rb") as _fh:
                head = _fh.read(1_048_576)            # first 1MB
                if size > 1_048_576 + 262_144:
                    _fh.seek(-262_144, 2)             # last 256KB
                    tail = _fh.read(262_144)
                else:
                    tail = b""
            h_head = _hl.sha1(head).hexdigest()
            h_tail = _hl.sha1(tail).hexdigest()
            key = (size, h_head, h_tail)
            if key in seen_keys:
                skipped_dup += 1
                dedupe_log.append(
                    f"  DUP    {fn} (size={size}, head={h_head[:6]}, tail={h_tail[:6]})"
                )
                logger.info(f"llm_fill: skip duplicate {fn} "
                            f"(size={size}, head={h_head[:8]}, tail={h_tail[:8]})")
                continue
            seen_keys.add(key)
            deduped_paths.append(pp)
            dedupe_log.append(
                f"  KEEP   {fn} (size={size}, head={h_head[:6]}, tail={h_tail[:6]})"
            )
        except Exception as exc:
            logger.warning(f"llm_fill dedupe: {pp} → {exc}")
            deduped_paths.append(pp)   # don't lose a PDF because hashing crashed
    if skipped_dup or skipped_junk:
        res.warnings.append(
            f"Pre-filter: kept {len(deduped_paths)} unique real PDFs "
            f"(dropped {skipped_dup} duplicates, {skipped_junk} junk filings)"
        )
    # Always surface the dedupe log so user can spot mis-kills like
    # "F2022_AR and F2023_AR were treated as duplicates" — copy/paste shows
    # the head/tail hashes so they can be compared at a glance.
    if dedupe_log:
        res.warnings.append("Dedupe log:\n" + "\n".join(dedupe_log[:40]))
    pdf_paths = deduped_paths

    pdf_data: list[tuple[PDFClassification, str]] = []
    for pp in pdf_paths:
        try:
            cls, text = _classify_pdf(pp)
            pdf_data.append((cls, text))
            res.pdfs_used.append({
                "name": cls.name, "stmt_type": cls.stmt_type,
                "period_type": cls.period_type, "pages": cls.pages_total,
                "text_chars": len(text),
                "path": pp,
            })
            logger.info(f"llm_fill: classified {cls.name} → "
                        f"{cls.stmt_type}/{cls.period_type}, {len(text)} chars")
        except Exception as exc:
            res.warnings.append(f"PDF classify failed for {pp}: {exc}")

    if not pdf_data:
        res.warnings.append("No usable PDFs — nothing to fill.")
        res.elapsed_sec = time.time() - t0
        return res

    # ── 2.5: PATTERN MATCH PRE-PASS — fills "obvious" cells for free ─────
    # Walks the cached MD tables and extracts cells where label matches
    # exactly. Then we remove those (sheet, year, label, stmt) combos from
    # template_specs so the LLM doesn't redo work. Cuts API calls by 50-70%.
    # Skipped when skip_pattern_match=True (LLM-only mode).
    if not skip_pattern_match:
        pattern_fills = _pattern_fill(template_specs, pdf_data)
        res.fills.extend(pattern_fills)
        if pattern_fills:
            res.warnings.append(
                f"Pattern-match pre-pass filled {len(pattern_fills)} cells (free, no LLM)"
            )
        pattern_filled_keys = {
            (f.get("sheet"), _normalise_label(f.get("label", "")),
             f.get("year"), (f.get("statement_type") or "").lower())
            for f in pattern_fills
        }

        # Subtract already-filled cells from each spec's years_needed
        remaining_specs: list[TemplateRowSpec] = []
        for s in template_specs:
            if s.is_formula:
                remaining_specs.append(s)
                continue
            norm = _normalise_label(s.label)
            stmt = (s.statement_type or "").lower() or "consolidated"
            new_years = [
                yr for yr in s.years_needed
                if (s.sheet, norm, yr, stmt) not in pattern_filled_keys
            ]
            if new_years:
                remaining_specs.append(TemplateRowSpec(
                    sheet=s.sheet, statement_type=s.statement_type,
                    section=s.section, label=s.label,
                    years_needed=new_years, is_formula=s.is_formula,
                ))
        template_specs = remaining_specs
        logger.info(f"llm_fill: after pattern pre-pass, {len(template_specs)} specs "
                    f"remain for LLM ({len(pattern_fills)} filled by pattern)")
    else:
        logger.info("llm_fill: skip_pattern_match=True → sending all specs to LLM")

    # 3. Build prompts — chunk by sheet so we never blow past max_tokens.
    # Each sheet gets its own LLM call; fills are merged at the end. The PDF
    # context is the same in every call (cheap on input tokens, trivially
    # cached by DeepSeek) but the OUTPUT for one sheet always fits.
    by_sheet: dict[str, list[TemplateRowSpec]] = {}
    for r in template_specs:
        if r.is_formula:
            continue
        by_sheet.setdefault(r.sheet, []).append(r)
    sheet_order = list(by_sheet.keys())
    logger.info(f"llm_fill: chunking into {len(sheet_order)} per-sheet calls "
                f"(sheets: {sheet_order})")

    # 4. Call DeepSeek — one call per sheet
    try:
        from openai import OpenAI
    except ImportError:
        res.warnings.append("openai package not installed")
        return res

    client = OpenAI(api_key=api_key, base_url=base_url)

    # ── Token sanity check — one tiny call. If auth fails, abort the whole
    # run before chunks waste minutes hitting the same 401 each.
    try:
        client.chat.completions.create(
            model=model,
            messages=[{"role": "user", "content": "ping"}],
            max_tokens=4, temperature=0,
        )
        logger.info("llm_fill: token sanity check passed")
    except Exception as exc:
        err_str = str(exc)
        if "401" in err_str or "auth" in err_str.lower() or "invalid" in err_str.lower():
            res.warnings.append(
                f"❌ Auth failed (401). Token rejected by {base_url}. "
                f"Re-copy the token, ensure no whitespace, and verify it hasn't "
                f"expired (AI Pipe tokens last 7 days). Error: {err_str[:200]}"
            )
            res.elapsed_sec = time.time() - t0
            return res
        # Non-auth error (e.g. 429 from quota) — let chunks handle it normally
        logger.warning(f"llm_fill: sanity-check non-auth error, continuing: {exc}")

    fills: list = []
    raw_responses: list[str] = []

    def _try_parse(text: str) -> list:
        try:
            obj = json.loads(text)
        except json.JSONDecodeError:
            return []
        if isinstance(obj, dict):
            f = obj.get("fills", [])
            return f if isinstance(f, list) else []
        if isinstance(obj, list):
            return obj
        return []

    def _salvage(raw: str) -> list:
        """Try multiple recovery strategies on a (possibly truncated) JSON blob."""
        out = _try_parse(raw)
        if out:
            return out
        cleaned = re.sub(r"^```(?:json)?\s*", "", raw.strip(), flags=re.I)
        cleaned = re.sub(r"\s*```\s*$", "", cleaned)
        out = _try_parse(cleaned)
        if out:
            return out
        # Brace-balanced scan
        try:
            start = raw.index("{")
            depth = 0; end = -1
            for i, ch in enumerate(raw[start:], start):
                if ch == "{": depth += 1
                elif ch == "}":
                    depth -= 1
                    if depth == 0: end = i + 1; break
            if end > start:
                out = _try_parse(raw[start:end])
                if out: return out
        except ValueError:
            pass
        # Truncation salvage — close the array at the last complete fill
        arr_start = raw.find('"fills"')
        if arr_start != -1:
            arr_open = raw.find('[', arr_start)
            if arr_open != -1:
                depth = 0; last_end = -1; in_str = False; escape = False
                for i in range(arr_open + 1, len(raw)):
                    ch = raw[i]
                    if escape: escape = False; continue
                    if ch == "\\": escape = True; continue
                    if ch == '"': in_str = not in_str; continue
                    if in_str: continue
                    if ch == "{": depth += 1
                    elif ch == "}":
                        depth -= 1
                        if depth == 0: last_end = i
                if last_end != -1:
                    return _try_parse(raw[arr_open:last_end + 1] + "]")
        return []

    # ── Per-sheet PDF routing ─────────────────────────────────────────────
    # Sending all 15 PDFs into every call wastes input tokens and starves
    # the output budget. Route by sheet purpose:
    #   Annual P&L / BS / CF       → ANNUAL Reports (and IP for context)
    #   Quarterly                   → QUARTERLY result PDFs (+ IP for context)
    #   Operating Metrics           → IP only (segment splits/operating data)
    # Always include matching statement_type or BOTH types when uncertain.
    def _route_pdfs(sheet_name: str) -> list[tuple[PDFClassification, str]]:
        sn = sheet_name.lower()
        out_pdfs: list[tuple[PDFClassification, str]] = []
        if "quarterly" in sn:
            for cls, txt in pdf_data:
                if cls.period_type == "QUARTERLY":
                    out_pdfs.append((cls, txt))
            # IPs sometimes carry quarterly segment data
            for cls, txt in pdf_data:
                if cls.period_type == "INVESTOR_PRESENTATION":
                    out_pdfs.append((cls, txt))
        elif "operating" in sn or "metric" in sn or "segment" in sn:
            # OM — primary source is IP (richest segment data), but ARs
            # often have segment notes too. Send both; the analyst will pick
            # the right printed values.
            for cls, txt in pdf_data:
                if cls.period_type == "INVESTOR_PRESENTATION":
                    out_pdfs.append((cls, txt))
            for cls, txt in pdf_data:
                if cls.period_type == "ANNUAL":
                    out_pdfs.append((cls, txt))
        else:    # Annual P&L / BS / CF / unclassified
            for cls, txt in pdf_data:
                if cls.period_type == "ANNUAL":
                    out_pdfs.append((cls, txt))
            # Q4 quarterly result has full-year figures — still useful for annuals
            for cls, txt in pdf_data:
                if cls.period_type == "QUARTERLY":
                    out_pdfs.append((cls, txt))
        return out_pdfs or list(pdf_data)

    def _call_one(sheet_label: str, specs_subset: list[TemplateRowSpec],
                  routed_pdfs: Optional[list] = None) -> list:
        """Make one LLM call for the given specs subset; return fills (possibly empty)."""
        if not specs_subset:
            return []
        pdfs_to_send = routed_pdfs if routed_pdfs is not None else pdf_data
        # Skip the call entirely if there are no PDFs to read from.
        # Saves API cost AND prevents LLM from inventing values out of nothing.
        if not pdfs_to_send:
            logger.info(f"llm_fill[{sheet_label}]: 0 routed PDFs → skip call")
            return []
        user_msg = _build_user_message(specs_subset, pdfs_to_send)
        msg_size = len(_SYSTEM_PROMPT) + len(user_msg)
        logger.info(f"llm_fill[{sheet_label}]: prompt ≈ {msg_size//1000}k chars, "
                    f"{len(specs_subset)} rows, {len(pdfs_to_send)} PDFs")
        # ── 429 retry-with-backoff ──────────────────────────────────────
        # Free tiers throttle aggressively. Distinguish two flavors:
        #   • Per-MINUTE limit hit → wait the suggested delay, retry
        #   • Per-DAY limit hit ("limit: 0", "PerDayPer...") → ABORT immediately,
        #     no retry will succeed today. Set a flag so other in-flight
        #     chunks also abort instead of waiting in useless 60s loops.
        if getattr(_call_one, "_quota_dead_for_day", False):
            return []   # another chunk already detected daily exhaustion

        resp = None
        last_err = None
        for attempt in range(4):
            try:
                resp = client.chat.completions.create(
                    model=model,
                    messages=[
                        {"role": "system", "content": _SYSTEM_PROMPT},
                        {"role": "user",   "content": user_msg},
                    ],
                    response_format={"type": "json_object"},
                    temperature=0,
                    max_tokens=max_tokens,
                )
                break
            except Exception as exc:
                last_err = exc
                err_str = str(exc)
                # Daily quota dead → mark global flag, abort everyone
                if ("PerDay" in err_str or "limit: 0" in err_str
                        or "RequestsPerDay" in err_str):
                    _call_one._quota_dead_for_day = True
                    logger.warning(f"llm_fill[{sheet_label}]: DAILY quota "
                                   f"exhausted — aborting all remaining "
                                   f"chunks (no retry will help today)")
                    return []
                # Non-rate-limit error — don't retry
                if "429" not in err_str and "rate" not in err_str.lower():
                    break
                # Per-minute rate limit — sleep then retry
                wait_s = 30
                m = re.search(r"retry.{0,5}in\s+(\d+(?:\.\d+)?)\s*s", err_str, re.I)
                if m:
                    wait_s = min(int(float(m.group(1))) + 2, 60)
                m2 = re.search(r"retryDelay['\"]?\s*[:=]\s*['\"]?(\d+)", err_str)
                if m2:
                    wait_s = min(int(m2.group(1)) + 2, 60)
                logger.info(f"llm_fill[{sheet_label}]: 429 hit, waiting "
                            f"{wait_s}s (attempt {attempt+1}/4)")
                time.sleep(wait_s)

        if resp is None:
            res.warnings.append(f"LLM call failed for '{sheet_label}': {last_err}")
            return []
        raw = resp.choices[0].message.content or ""
        finish = resp.choices[0].finish_reason if resp.choices else "?"
        raw_responses.append(f"=== {sheet_label} ({len(raw)} chars, finish={finish}) ===\n{raw[:1500]}")

        out = _salvage(raw)

        # Distinguish "LLM correctly returned empty fills" from "couldn't parse".
        # An empty {"fills": []} is a valid response — the LLM read the PDFs
        # and decided this chunk has no data. Don't surface it as a warning.
        is_legit_empty = (
            finish == "stop"
            and re.search(r'"fills"\s*:\s*\[\s*\]', raw) is not None
        )
        if not out:
            if is_legit_empty:
                logger.info(f"llm_fill[{sheet_label}]: 0 fills "
                            f"(LLM found nothing for this chunk)")
            else:
                res.warnings.append(
                    f"'{sheet_label}': could not parse response "
                    f"({len(raw)} chars, finish={finish}). First 200: {raw[:200]!r}"
                )
        else:
            logger.info(f"llm_fill[{sheet_label}]: {len(out)} fills "
                        f"(finish={finish}, {len(raw)} chars)")
            if finish == "length":
                res.warnings.append(
                    f"'{sheet_label}': output truncated; salvaged {len(out)} complete fills."
                )
        return out

    # ── PER-PDF SEQUENTIAL LLM PROCESSING ─────────────────────────────────
    # User-designed architecture (v15+): process ONE PDF at a time. Each
    # LLM call sees:
    #   - Just that PDF's content (~60K chars)
    #   - Only the template cells STILL EMPTY (not what pattern matcher or
    #     prior PDFs already filled)
    # Cells lock progressively. No cross-PDF mixing in any single prompt.
    # Smaller context per call → cheaper, more accurate.
    #
    # PDF priority order (most-authoritative first):
    #   1. ANNUAL reports — most recent first (FY24-25 → FY23-24 → FY22-23)
    #   2. QUARTERLY results — Q4 first (full-year audited), then Q3..Q1
    #   3. INVESTOR_PRESENTATION — fills Operating Metrics last
    def _pdf_priority(item):
        cls, _ = item
        period = (cls.period_type or "").upper()
        # Extract year hint from filename/title
        m = re.search(r"(\d{4})", cls.name + " " + cls.title_text)
        yr = int(m.group(1)) if m else 0
        # Q4 boost
        is_q4 = "q4" in cls.name.lower() or "4qf" in cls.name.lower()
        if period == "ANNUAL":
            return (0, -yr)
        if period == "QUARTERLY":
            return (1, 0 if is_q4 else 1, -yr)
        if period == "INVESTOR_PRESENTATION":
            return (2, -yr)
        return (3, -yr)
    pdf_data_sorted = sorted(pdf_data, key=_pdf_priority)
    logger.info(f"llm_fill: per-PDF processing order: "
                f"{[c.name for c, _ in pdf_data_sorted]}")

    # Track which (sheet, label, year, stmt) cells are already filled
    def _fill_key(f):
        return (f.get("sheet"), (f.get("label") or "").lower().strip(),
                f.get("year"),
                (f.get("statement_type") or "").lower().strip())
    already_filled: set = {_fill_key(f) for f in fills}
    logger.info(f"llm_fill: starting per-PDF pass, "
                f"{len(already_filled)} cells already filled by pattern")

    # Build all empty-cell specs ONCE (before parallel calls). Each PDF
    # call sees the same starting state. Writer cell-lock dedupes results.
    remaining_specs_initial: list[TemplateRowSpec] = []
    for spec in template_specs:
        if spec.is_formula:
            continue
        stmt = (spec.statement_type or "").lower().strip()
        label_norm = (spec.label or "").lower().strip()
        unfilled_years = [
            yr for yr in spec.years_needed
            if (spec.sheet, label_norm, yr, stmt) not in already_filled
        ]
        if unfilled_years:
            remaining_specs_initial.append(TemplateRowSpec(
                sheet=spec.sheet, statement_type=spec.statement_type,
                section=spec.section, label=spec.label,
                years_needed=unfilled_years, is_formula=False,
            ))

    if not remaining_specs_initial:
        logger.info("llm_fill: all cells filled by pattern — no LLM calls needed")
    else:
        n_cells_remaining = sum(len(s.years_needed) for s in remaining_specs_initial)
        logger.info(f"llm_fill: {n_cells_remaining} empty cells across "
                    f"{len(pdf_data_sorted)} PDFs (parallel)")

        # Parallel per-PDF calls. Each gets the SAME empty-cells target.
        # Cell-lock in writer ensures pattern fills win and per-PDF fills
        # dedupe at write time. Concurrency 4-5 keeps total runtime ~30-60s.
        import concurrent.futures as _cf
        is_gemini = "gemini" in (model or "").lower() or "googleapis" in (base_url or "").lower()
        parallelism = 2 if is_gemini else 5
        logger.info(f"llm_fill: per-PDF parallelism={parallelism}")

        def _run_one_pdf(idx: int) -> list:
            cls_i, full_text_i = pdf_data_sorted[idx]
            label_i = f"PDF {idx+1}/{len(pdf_data_sorted)}: {cls_i.name}"
            try:
                pdf_fills = _call_one(label_i, remaining_specs_initial,
                                      [(cls_i, full_text_i)])
                # Tag source so writer per-source pool works
                for f in pdf_fills:
                    f["source"] = f.get("source", f"llm:{cls_i.name}")
                return pdf_fills
            except Exception as exc:
                res.warnings.append(f"{label_i} raised: {exc}")
                return []

        with _cf.ThreadPoolExecutor(max_workers=parallelism) as pool:
            futures = {pool.submit(_run_one_pdf, i): i
                       for i in range(len(pdf_data_sorted))}
            for fut in _cf.as_completed(futures):
                pdf_fills = fut.result()
                fills.extend(pdf_fills)

    res.raw_response = "\n\n".join(raw_responses)[:8000]
    res.fills = fills
    logger.info(f"llm_fill: TOTAL {len(fills)} fills across {len(sheet_order)} sheets")

    if not fills:
        res.warnings.append(
            "All sheet calls failed. Check the warnings above and PDF inputs."
        )

    # 6. Write Excel
    stats = _write_filled_excel(template_path, output_path, fills)
    res.cells_written = stats["written"]
    res.by_confidence = stats["by_confidence"]
    if stats.get("skipped_already_written", 0):
        res.warnings.append(
            f"Cell-lock prevented {stats['skipped_already_written']} overwrites "
            f"(pattern fills land first; LLM can't overwrite them with "
            f"cross-section or wrong-year values)"
        )
    if stats.get("skipped_used_value", 0):
        res.warnings.append(
            f"Per-report pool rejected {stats['skipped_used_value']} duplicates "
            f"(same value re-used in another column of the same row, "
            f"within one source)"
        )
        if stats.get("dedup_log"):
            res.warnings.append(
                "Dedup samples:\n" + "\n".join(stats["dedup_log"][:10])
            )

    # Per-source contribution report — shows which PDFs are silent
    if stats.get("per_source_fills"):
        per_src = stats["per_source_fills"]
        sorted_src = sorted(per_src.items(), key=lambda kv: -kv[1])
        lines = [f"  {src}: {n} fills" for src, n in sorted_src]
        res.warnings.append(
            f"Per-report pool contributions ({len(per_src)} sources):\n"
            + "\n".join(lines)
        )
        # Flag sources that contributed zero
        all_pdf_names = {pp.get("name", "") for pp in res.pdfs_used}
        contributing = {src for src in per_src if src != "llm"}
        silent = all_pdf_names - contributing
        if silent:
            res.warnings.append(
                f"⚠ Silent PDFs (no pattern fills): "
                + ", ".join(sorted(silent)[:8])
                + " — LLM may have filled from these, check raw_response"
            )
    if stats["skipped_no_address"]:
        res.warnings.append(
            f"{stats['skipped_no_address']} fills couldn't be placed "
            f"(label/year mismatch with template)"
        )
        # Surface the first 8 rejections in detail so user can see *why*
        # without having to re-run. Group by reason to keep it short.
        rej_log = stats.get("rejection_log", [])
        seen_reasons = {}
        for r in rej_log:
            why = r.get("why", r.get("reason", "?"))
            if why not in seen_reasons:
                seen_reasons[why] = []
            if len(seen_reasons[why]) < 3:
                seen_reasons[why].append(
                    f"  [{r.get('sheet')}/{r.get('stmt')}/{r.get('year')}] "
                    f"label={r.get('label')!r} value={r.get('value')}"
                )
        rej_summary = []
        for why, items in list(seen_reasons.items())[:6]:
            rej_summary.append(f"• {why}")
            rej_summary.extend(items)
        if rej_summary:
            res.warnings.append("Rejection details:\n" + "\n".join(rej_summary))

    # 6.5 — Post-write hallucination guards. These do NOT tell the LLM
    # what to do — they run AFTER the LLM has written, and remove only
    # the three failure modes we keep seeing empirically:
    #   1. Same value duplicated across ≥3 quarters of one row (LLM pattern-fill)
    #   2. Standalone cell == Consolidated cell for same row/year (Indian
    #      listed companies always differ — back-fill error)
    #   3. Quarterly cell > 1.5× annual (units / magnitude error)
    # Disable only for diagnostics via DISABLE_POST_WRITE_GUARDS=1.
    import os as _os
    if _os.environ.get("DISABLE_POST_WRITE_GUARDS") != "1":
        try:
            g_stats = _post_write_guards(output_path, template_path)
            total_blanked = (g_stats["duplicate_quarterly"]
                             + g_stats["cons_eq_standalone"]
                             + g_stats["magnitude_too_high"])
            if total_blanked:
                res.warnings.append(
                    f"Cleaned {total_blanked} hallucinated cells "
                    f"(dup-quarter: {g_stats['duplicate_quarterly']}, "
                    f"cons=stand: {g_stats['cons_eq_standalone']}, "
                    f"magnitude: {g_stats['magnitude_too_high']})"
                )
        except Exception as exc:
            res.warnings.append(f"Post-write guards crashed: {exc}")

    # 7. Screener fallback for sheets that came out empty.
    # Per user: "for empty slide or sheet put the data from screener". This
    # runs AFTER pattern + LLM and inspects the output Excel. Any sheet with
    # < 8 cells filled gets a "From Screener.in" block appended at the bottom
    # (Cash Flow is the typical offender — templates often lack CF labels and
    # nothing else can fill it).
    if nse_symbol:
        try:
            sf_stats = _screener_fallback(
                output_path=output_path,
                template_path=template_path,
                nse_symbol=nse_symbol,
                bse_code=bse_code,
            )
            if sf_stats["sheets_filled"]:
                res.warnings.append(
                    "Screener fallback added "
                    f"{sf_stats['cells_added']} cells across "
                    f"{len(sf_stats['sheets_filled'])} empty sheet(s): "
                    + ", ".join(
                        f"{s}({src}:{n})" for s, src, n, _ in sf_stats["sheets_filled"]
                    )
                )
            for err in sf_stats.get("errors", []):
                res.warnings.append(f"Screener fallback: {err}")
        except Exception as exc:
            res.warnings.append(f"Screener fallback crashed: {exc}")

    res.elapsed_sec = time.time() - t0
    return res


# ──────────────────────────────────────────────────────────────────────────────
# Post-write hallucination guards — deterministic safety net for the LLM.
# Catch the three failure modes we keep seeing:
#   1. Same value repeated across ≥3 quarters (Quarterly Revenue 1099.96 ×3)
#   2. Standalone cell == Consolidated cell for same label/year (must differ)
#   3. Quarterly cell magnitude wildly above annual (units error like 37200)
# ──────────────────────────────────────────────────────────────────────────────

def _post_write_guards(output_path: str, template_path: str) -> dict:
    """
    Run sanity checks on the just-written Excel and BLANK any cell that's
    obviously wrong. Returns stats so we can report to the user.

    Conservative: only blanks values we're highly confident are hallucinations.
    Empty cells are user-recoverable; wrong cells get pasted into truth tables.
    """
    stats = {
        "duplicate_quarterly":  0,    # rule 1
        "cons_eq_standalone":   0,    # rule 2
        "magnitude_too_high":   0,    # rule 3
        "details":              [],
    }
    try:
        import openpyxl
        from mapper.template_reader import read_template
    except ImportError:
        return stats

    try:
        tm = read_template(template_path)
        wb = openpyxl.load_workbook(output_path, data_only=False)
    except Exception as exc:
        stats["details"].append(f"guard reopen failed: {exc}")
        return stats

    def _is_qtr_col(label: str) -> bool:
        return bool(re.match(r"^[1-4]QF\d{4}$", (label or "").strip()))

    # ── Rule 1: Quarterly — same value appearing in ≥3 of the row's quarter cells.
    # Indian listed companies don't post identical Revenue/COGS/expenses
    # quarter-after-quarter. Three or more identical = hallucinated.
    for sheet_name, sm in tm.sheets.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        qtr_cols = {yr: c for yr, c in (sm.year_cols or {}).items() if _is_qtr_col(yr)}
        if len(qtr_cols) < 3:
            continue   # Annual sheets — skip rule 1
        for label, row_nums in sm.row_index.items():
            for r in row_nums:
                # Collect (year, column, value) for this row's quarter cells
                cells = []
                for yr, c in qtr_cols.items():
                    v = ws.cell(row=r, column=c).value
                    if isinstance(v, (int, float)) and v != 0:
                        cells.append((yr, c, v))
                if len(cells) < 3:
                    continue
                # NOTE: the writer's "used-value pool" already prevents the
                # same value from being assigned to multiple columns of the
                # same row at write time. This per-row post-scan is now
                # redundant — kept as a safety net only for the rare case
                # where the writer was bypassed (e.g. pattern fills + LLM
                # fills both targeting the same row from different sources).
                # Threshold raised to ≥4 so it only catches gross overlap.
                from collections import Counter
                value_counts = Counter(round(v, 2) for _, _, v in cells)
                for val, count in value_counts.items():
                    if count >= 4:
                        blanked = 0
                        for (yr, c, v) in cells:
                            if round(v, 2) == val:
                                ws.cell(row=r, column=c).value = None
                                blanked += 1
                        stats["duplicate_quarterly"] += blanked
                        stats["details"].append(
                            f"  Q-DUP  {sheet_name!r} row{r} ({label[:40]!r}): "
                            f"value {val} appeared {count}× → blanked {blanked} cells"
                        )

    # ── Rule 2: Annual — Standalone cell == Consolidated cell for same label/year.
    # By definition they must differ (subsidiaries differ). When identical,
    # the LLM almost certainly back-filled Standalone from Consolidated.
    # We blank the STANDALONE side, since Consolidated is usually more reliable.
    for sheet_name, sm in tm.sheets.items():
        if sheet_name not in wb.sheetnames:
            continue
        section_offsets = getattr(sm, "section_offsets", {}) or {}
        if "consolidated" not in section_offsets or "standalone" not in section_offsets:
            continue   # Sheet doesn't split into both sections — skip
        ws = wb[sheet_name]
        cons_start = section_offsets["consolidated"]
        stand_start = section_offsets["standalone"]
        # For each label in the sheet, find the cons-row and standalone-row
        for label, row_nums in sm.row_index.items():
            cons_rows = [r for r in row_nums if r < stand_start]
            stand_rows = [r for r in row_nums if r >= stand_start]
            if not (cons_rows and stand_rows):
                continue
            cr, sr = cons_rows[0], stand_rows[0]
            for yr, c in (sm.year_cols or {}).items():
                vc = ws.cell(row=cr, column=c).value
                vs = ws.cell(row=sr, column=c).value
                if (isinstance(vc, (int, float)) and isinstance(vs, (int, float))
                        and vc != 0 and vs != 0
                        and abs(vc - vs) < 0.01):
                    ws.cell(row=sr, column=c).value = None
                    stats["cons_eq_standalone"] += 1
                    stats["details"].append(
                        f"  S=C    {sheet_name!r} {label[:40]!r} {yr}: "
                        f"standalone={vs} == consolidated={vc} → blanked standalone"
                    )

    # ── Rule 3: Quarterly cell magnitude check vs annual sibling.
    # If we have an Annual sheet with F<YYYY> and a Quarterly sheet with
    # NQF<YYYY> for the SAME label, no single quarter should exceed annual
    # × 1.5 (some single quarters legitimately concentrate — e.g. one-time
    # finance cost spikes). We compare against the MAX of Cons/Standalone
    # annual values so we don't false-positive on Quarterly Consolidated
    # values when the Standalone annual is the only one recorded.
    annual_lookup: dict[tuple, list[float]] = {}   # (label_normalised, year) → [values]
    for sheet_name, sm in tm.sheets.items():
        if sheet_name not in wb.sheetnames:
            continue
        annual_cols = {yr: c for yr, c in (sm.year_cols or {}).items()
                       if re.match(r"^F\d{4}$", yr)}
        if not annual_cols:
            continue
        ws = wb[sheet_name]
        for label, row_nums in sm.row_index.items():
            ln = _normalise_label(label)
            if not ln:
                continue
            for r in row_nums:
                for yr, c in annual_cols.items():
                    v = ws.cell(row=r, column=c).value
                    if isinstance(v, (int, float)) and v != 0:
                        annual_lookup.setdefault((ln, yr), []).append(abs(v))
    # Now scan Quarterly sheets and compare to the MAX of all annual values
    # for that (label, year) — covers both Cons and Standalone in one shot.
    for sheet_name, sm in tm.sheets.items():
        if sheet_name not in wb.sheetnames:
            continue
        qtr_cols = {yr: c for yr, c in (sm.year_cols or {}).items() if _is_qtr_col(yr)}
        if not qtr_cols:
            continue
        ws = wb[sheet_name]
        for label, row_nums in sm.row_index.items():
            ln = _normalise_label(label)
            if not ln:
                continue
            for r in row_nums:
                for yr, c in qtr_cols.items():
                    v = ws.cell(row=r, column=c).value
                    if not isinstance(v, (int, float)) or v == 0:
                        continue
                    m = re.match(r"^[1-4]QF(\d{4})$", yr)
                    if not m:
                        continue
                    annual_yr = f"F{m.group(1)}"
                    annual_vals = annual_lookup.get((ln, annual_yr), [])
                    if not annual_vals:
                        continue
                    annual_max = max(annual_vals)
                    if abs(v) > annual_max * 1.5:
                        ws.cell(row=r, column=c).value = None
                        stats["magnitude_too_high"] += 1
                        stats["details"].append(
                            f"  MAGN   {sheet_name!r} {label[:40]!r} {yr}: "
                            f"|{v}| > 1.5×max-annual({annual_max}) "
                            f"({annual_yr}) → blanked"
                        )

    try:
        wb.save(output_path)
    except Exception as exc:
        stats["details"].append(f"guard save failed: {exc}")
    return stats


# ──────────────────────────────────────────────────────────────────────────────
# Screener fallback — fills sheets that came out empty after pattern + LLM.
# User insisted: "for empty slide or sheet put the data from screener".
# ──────────────────────────────────────────────────────────────────────────────

# Map template sheet name keywords → Screener section name (returned by
# fetch_screener_raw). Order matters: more-specific keys first.
_SCREENER_SHEET_MAP: tuple[tuple[str, str], ...] = (
    # Cash flow
    ("cash flow",            "Cash Flow"),
    ("cashflow",             "Cash Flow"),
    ("statement of cash",    "Cash Flow"),
    ("cash flows",           "Cash Flow"),
    # Balance sheet
    ("balance sheet",        "Balance Sheet"),
    ("statement of financial position", "Balance Sheet"),
    ("financial position",   "Balance Sheet"),
    ("balance",              "Balance Sheet"),
    # Operating metrics / ratios
    ("operating metric",     "Operating Metrics"),
    ("operating ratio",      "Operating Metrics"),
    ("segment",              "Operating Metrics"),
    ("kpi",                  "Operating Metrics"),
    ("ratio",                "Operating Metrics"),
    ("dupont",               "Operating Metrics"),
    # Quarterly
    ("quarterly",            "Quarterly"),
    ("quarter",              "Quarterly"),
    ("interim",              "Quarterly"),
    ("qoq",                  "Quarterly"),
    # P&L
    ("p & l",                "P&L"),
    ("p&l",                  "P&L"),
    ("p and l",              "P&L"),
    ("profit and loss",      "P&L"),
    ("profit & loss",        "P&L"),
    ("profitability",        "P&L"),
    ("income statement",     "P&L"),
    ("statement of profit",  "P&L"),
    ("statement of income",  "P&L"),
    ("p l",                  "P&L"),
    ("profit",               "P&L"),
    ("income",               "P&L"),
    ("revenue",              "P&L"),
    ("earnings",             "P&L"),
)


def _screener_section_for_sheet(sheet_name: str) -> Optional[str]:
    """Decide which Screener section (P&L / Balance Sheet / Cash Flow / etc.)
    matches a given template sheet name. Returns None when there is no
    reasonable mapping — caller then dumps the FULL Screener export."""
    sn = (sheet_name or "").lower().strip()
    if not sn:
        return None
    # Direct keyword match (specific keys win because they're first)
    for key, screener_section in _SCREENER_SHEET_MAP:
        if key in sn:
            return screener_section
    # Tokenised fallback for short codes like "CF" / "BS" / "PL" / "OM"
    tokens = re.split(r"[\s\-_/.,()]+", sn)
    if "cf" in tokens:
        return "Cash Flow"
    if "bs" in tokens:
        return "Balance Sheet"
    if "pl" in tokens or "pnl" in tokens or "p&l" in tokens:
        return "P&L"
    if "om" in tokens or "ratios" in tokens:
        return "Operating Metrics"
    if "q" in tokens or any(re.fullmatch(r"q[1-4]", t) for t in tokens):
        return "Quarterly"
    return None


def _count_filled_year_cells(ws, year_cols: dict, row_index: dict) -> int:
    """Count cells in a worksheet that already have a numeric value at the
    intersection of a known template row and a known year column."""
    n = 0
    for label, row_nums in row_index.items():
        for r in row_nums:
            for yr, c in year_cols.items():
                v = ws.cell(row=r, column=c).value
                if isinstance(v, (int, float)) and v != 0:
                    n += 1
                elif isinstance(v, str) and v.strip():
                    # Allow non-numeric strings (e.g. "NA") as evidence of fill
                    try:
                        float(v.replace(",", "").strip())
                        n += 1
                    except ValueError:
                        pass
    return n


def _detect_sheet_scope(sheet_name: str, template_sheet_meta=None) -> str:
    """
    Decide whether a template sheet's empty data should be filled from
    CONSOLIDATED Screener, STANDALONE Screener, or BOTH.

    Returns: "CONSOLIDATED" | "STANDALONE" | "BOTH"

    Order of evidence:
      1. Explicit keyword in sheet name ("Standalone P&L" / "P&L (Cons)")
      2. Template's section_offsets shows BOTH sections inside the sheet → BOTH
      3. Default → CONSOLIDATED (Indian listed companies usually report on
         consolidated; truth files prefer this when no other hint)
    """
    sn = (sheet_name or "").lower()
    # Standalone wins explicit mention because it's the "less default" option
    if "standalone" in sn or " sa " in f" {sn} " or sn.endswith(" sa"):
        return "STANDALONE"
    if "consolidated" in sn or "(cons)" in sn or " consol" in sn:
        return "CONSOLIDATED"
    # Inspect template meta — if the sheet's body has BOTH sections marked,
    # the template wants both blocks populated.
    try:
        if template_sheet_meta is not None:
            so = getattr(template_sheet_meta, "section_offsets", {}) or {}
            if "consolidated" in so and "standalone" in so:
                return "BOTH"
            if "standalone" in so:
                return "STANDALONE"
            if "consolidated" in so:
                return "CONSOLIDATED"
    except Exception:
        pass
    return "CONSOLIDATED"


def _screener_fallback(output_path: str, template_path: str,
                        nse_symbol: Optional[str], bse_code: Optional[str],
                        empty_threshold: int = 2) -> dict:
    """
    Post-fill step. After pattern + LLM ran, walk through every sheet in the
    output Excel. For sheets with fewer than `empty_threshold` filled cells
    (default: 2 — basically empty sheets only), fetch Screener.in data and
    dump the matching section into a fresh block at the bottom of that sheet
    (clearly labelled "From Screener.in"). The strict default avoids
    polluting partially-filled sheets — user said "empty sheet" → empty.

    Scope routing: fetches BOTH consolidated and standalone Screener data
    upfront. Per-sheet, decides which to write based on sheet name (e.g.
    "Standalone P&L" → standalone) or template structure (sheet with both
    sections → write both blocks). Sheets without scope hints default to
    CONSOLIDATED, the standard for Indian listed companies.

    Returns stats dict: { sheets_filled, sheets_skipped, cells_added, errors }
    """
    stats = {"sheets_filled": [], "sheets_skipped": [], "cells_added": 0,
             "errors": []}
    if not nse_symbol:
        stats["errors"].append("Screener fallback skipped: no NSE symbol")
        return stats

    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from mapper.template_reader import read_template
    except ImportError as exc:
        stats["errors"].append(f"openpyxl/template_reader missing: {exc}")
        return stats

    # Fetch BOTH consolidated AND standalone Screener data upfront so we can
    # route per-sheet. Indian listed companies report both versions — they're
    # genuinely different numbers (subsidiaries differ in consolidated).
    # Try multiple import paths because web_only_app runs as a Streamlit
    # script (not always importable as a module when called from within
    # another script execution context).
    screener_cons: dict = {}
    screener_stand: dict = {}
    fetch_fn = None
    import_errors: list[str] = []
    for modpath in ("web.screener_raw", "web_only_app", "screener_raw"):
        try:
            import importlib
            m = importlib.import_module(modpath)
            fn = getattr(m, "fetch_screener_raw", None)
            if fn is not None:
                fetch_fn = fn
                logger.info(f"_screener_fallback: using fetch_screener_raw "
                            f"from '{modpath}'")
                break
        except Exception as exc:
            import_errors.append(f"{modpath}: {exc}")

    if fetch_fn is None:
        stats["errors"].append(
            "fetch_screener_raw not importable — tried: "
            + "; ".join(import_errors)
        )
        return stats

    try:
        screener_cons  = fetch_fn(nse_symbol, consolidated=True)  or {}
        screener_stand = fetch_fn(nse_symbol, consolidated=False) or {}
    except Exception as exc:
        stats["errors"].append(f"Screener fetch failed: {exc}")
        return stats

    if not screener_cons and not screener_stand:
        stats["errors"].append(
            f"Screener returned no data for {nse_symbol} "
            f"(company may not be on screener.in or network blocked)"
        )
        return stats

    logger.info(
        f"_screener_fallback: got Screener for {nse_symbol} — "
        f"consolidated sections {list(screener_cons.keys())}, "
        f"standalone sections {list(screener_stand.keys())}"
    )

    try:
        tm = read_template(template_path)
        wb = openpyxl.load_workbook(output_path, data_only=False)
    except Exception as exc:
        stats["errors"].append(f"Reopen output for fallback failed: {exc}")
        return stats

    HEADER_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1",
                              fill_type="solid")
    HEADER_FONT = Font(bold=True, color="1F4E78")
    SCREENER_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC",
                                 fill_type="solid")

    for sheet_name, sm in tm.sheets.items():
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else None
        if ws is None:
            stats["sheets_skipped"].append((sheet_name, "no such ws"))
            continue

        # Skip non-financial sheets (Manufacturing Base, Guidance, etc.)
        # If the template didn't detect ANY year columns for this sheet, it's
        # not a numeric/financial sheet — Screener data would be irrelevant.
        year_cols = getattr(sm, "year_cols", {}) or {}
        if not year_cols:
            stats["sheets_skipped"].append(
                (sheet_name, "no year columns — non-financial sheet")
            )
            continue

        # How full is this sheet already?
        filled = _count_filled_year_cells(ws, sm.year_cols, sm.row_index)

        if filled >= empty_threshold:
            stats["sheets_skipped"].append(
                (sheet_name, f"already has {filled} cells")
            )
            continue

        # Decide which Screener section(s) AND which scope(s) to dump.
        scope = _detect_sheet_scope(sheet_name, sm)   # CONSOLIDATED/STANDALONE/BOTH
        screener_section = _screener_section_for_sheet(sheet_name)

        if screener_section is not None:
            sections_to_dump: list[str] = [screener_section]
        else:
            # Empty sheet with unrecognised name — dump every section we have.
            # Use the scope-appropriate source's section list.
            primary_src = (screener_stand if scope == "STANDALONE"
                           else screener_cons or screener_stand)
            sections_to_dump = list(primary_src.keys())
            if not sections_to_dump:
                stats["sheets_skipped"].append(
                    (sheet_name, "no screener data available")
                )
                continue
            logger.info(
                f"_screener_fallback: sheet '{sheet_name}' has no name match "
                f"AND is empty → dumping ALL Screener sections "
                f"({sections_to_dump}) at scope {scope}"
            )

        # Build the (source_label, source_dict) list based on detected scope
        if scope == "BOTH":
            sources_to_write = [
                ("Consolidated", screener_cons),
                ("Standalone",   screener_stand),
            ]
        elif scope == "STANDALONE":
            sources_to_write = [("Standalone", screener_stand or screener_cons)]
        else:   # CONSOLIDATED (default)
            sources_to_write = [("Consolidated", screener_cons or screener_stand)]

        # Find first empty row to start writing — leave 2 blank rows separator
        last_row = ws.max_row or 1
        while last_row > 1 and all(
            ws.cell(row=last_row, column=c).value in (None, "")
            for c in range(1, (ws.max_column or 1) + 1)
        ):
            last_row -= 1
        cur_row = last_row + 3 if last_row > 1 else 1

        added_total = 0
        blocks_written: list[str] = []   # human-readable summary tokens
        for scope_label, src in sources_to_write:
            if not src:
                continue
            for section in sections_to_dump:
                block = src.get(section)
                if not block:
                    continue
                years = block.get("years") or []
                data = block.get("data") or {}
                if not years or not data:
                    continue

                # Title row — includes scope so user can tell them apart
                title = (
                    f"From Screener.in — {section} [{scope_label}] "
                    f"({nse_symbol})"
                )
                ws.cell(row=cur_row, column=1, value=title)
                ws.cell(row=cur_row, column=1).font = HEADER_FONT
                ws.cell(row=cur_row, column=1).fill = HEADER_FILL
                cur_row += 1

                # Header row: "Particulars" + each year
                ws.cell(row=cur_row, column=1, value="Particulars").font = HEADER_FONT
                ws.cell(row=cur_row, column=1).fill = HEADER_FILL
                for ci, yr in enumerate(years, start=2):
                    ws.cell(row=cur_row, column=ci, value=yr).font = HEADER_FONT
                    ws.cell(row=cur_row, column=ci).fill = HEADER_FILL
                cur_row += 1

                # Data rows
                section_added = 0
                for label, year_vals in data.items():
                    ws.cell(row=cur_row, column=1, value=label)
                    for ci, yr in enumerate(years, start=2):
                        v = year_vals.get(yr)
                        if v is not None:
                            ws.cell(row=cur_row, column=ci, value=v).fill = SCREENER_FILL
                            section_added += 1
                    cur_row += 1
                cur_row += 2   # 2-row gap before next block
                added_total += section_added
                blocks_written.append(f"{section}[{scope_label}]")

        if added_total == 0:
            stats["sheets_skipped"].append(
                (sheet_name, "no usable screener data for this sheet")
            )
            continue

        match_label = (
            blocks_written[0] if len(blocks_written) == 1
            else f"{len(blocks_written)}blocks"
        )
        stats["sheets_filled"].append(
            (sheet_name, match_label, added_total, filled)
        )
        stats["cells_added"] += added_total
        logger.info(
            f"_screener_fallback: sheet '{sheet_name}' (scope={scope}) had "
            f"{filled} filled cells; appended {len(blocks_written)} block(s): "
            f"{blocks_written} → {added_total} cells"
        )

    try:
        wb.save(output_path)
    except Exception as exc:
        stats["errors"].append(f"Could not save fallback to {output_path}: {exc}")
        return stats

    return stats


__all__ = ["llm_fill_template", "LLMFillResult"]
