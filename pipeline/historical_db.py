"""
pipeline/historical_db.py
─────────────────────────
Per-company historical values database.

Stored as JSON files at  db/historical/<SYMBOL>.json:

    {
      "symbol":       "PPLPHARMA",
      "name":         "Piramal Pharma Limited",
      "sector":       "pharma",
      "last_updated": "2026-05-03T15:42:11",
      "values": {
        "revenue_from_operations": {"F2024": 8171.16, "F2023": 7081.55, ...},
        "profit_after_tax":        {"F2024": 80.6,    "F2023": -179.5, ...},
        ...
      }
    }

Keys are KB canonicals (so different runs / different label phrasings
collapse to the same row). Values are floats in INR Crores.

Used for:
  • Validation: flag any new fill that swings >5× from prior year
  • Bootstrapping: seed from a user-supplied "known-good" template
  • Cross-run continuity: each successful run augments the DB so future
    runs of the same company are validated against more years.
"""

from __future__ import annotations

import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)

_DB_ROOT = Path(__file__).parent.parent / "db" / "historical"


def _path_for(symbol: str) -> Path:
    return _DB_ROOT / f"{symbol.upper()}.json"


def _ensure_dir() -> None:
    _DB_ROOT.mkdir(parents=True, exist_ok=True)


def load_historical(symbol: str) -> dict:
    """Load the JSON record for a symbol, or empty dict if none."""
    if not symbol:
        return {}
    p = _path_for(symbol)
    if not p.exists():
        return {}
    try:
        with open(p) as f:
            return json.load(f)
    except Exception as exc:
        logger.warning(f"historical_db: read failed for {p.name}: {exc}")
        return {}


def get_value(symbol: str, canonical: str, year: str) -> Optional[float]:
    """Return historical value for (symbol, canonical, year), or None."""
    rec = load_historical(symbol)
    v = rec.get("values", {}).get(canonical, {}).get(year)
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def get_prior_year_value(symbol: str, canonical: str,
                          year: str) -> Optional[tuple[str, float]]:
    """Return (prior_year_label, value) for the closest preceding year, or None."""
    rec = load_historical(symbol)
    canon_vals = rec.get("values", {}).get(canonical, {})
    if not canon_vals:
        return None
    try:
        target = int(year.lstrip("Ff"))
    except ValueError:
        return None
    candidates = []
    for yr_lbl, val in canon_vals.items():
        try:
            yr_int = int(yr_lbl.lstrip("Ff"))
            if yr_int < target and val is not None:
                candidates.append((yr_int, yr_lbl, float(val)))
        except (ValueError, TypeError):
            continue
    if not candidates:
        return None
    candidates.sort(reverse=True)   # closest preceding year first
    return candidates[0][1], candidates[0][2]


# ── Writing ──────────────────────────────────────────────────────────────────

def save_run(symbol: str,
             values_by_canonical: dict[str, dict[str, float]],
             sector: str = "general",
             name: str = "",
             only_high_confidence: bool = True,
             confidence_meta: Optional[dict] = None) -> int:
    """
    Persist a successful run's values into the historical DB.

    `values_by_canonical` : {canonical_key: {year: value}}
    `confidence_meta`     : {(canonical, year): {"confidence": "HIGH"/...}}
                            If supplied AND only_high_confidence=True, we
                            only persist HIGH-confidence cells (prevents bad
                            extractions from poisoning the DB).
    """
    if not symbol or not values_by_canonical:
        return 0
    _ensure_dir()
    p = _path_for(symbol)

    # Load existing record (we MERGE — don't overwrite years we already have)
    existing = load_historical(symbol)
    existing.setdefault("symbol", symbol.upper())
    existing.setdefault("name", name or symbol)
    existing.setdefault("sector", sector)
    existing["last_updated"] = datetime.now().isoformat(timespec="seconds")
    existing.setdefault("values", {})

    n_new = 0
    for canon, yr_vals in values_by_canonical.items():
        if not canon:
            continue
        existing["values"].setdefault(canon, {})
        for yr, v in yr_vals.items():
            try:
                fv = float(v) if v is not None else None
            except (TypeError, ValueError):
                continue
            if fv is None:
                continue
            # Optional gate: only persist if we have HIGH confidence
            if only_high_confidence and confidence_meta is not None:
                cm = confidence_meta.get((canon, yr), {})
                if cm.get("confidence") != "HIGH":
                    continue
            # Don't overwrite existing values with different ones — preserve
            # the originally-saved truth. New values for new years get added.
            if yr in existing["values"][canon]:
                continue
            existing["values"][canon][yr] = fv
            n_new += 1

    try:
        with open(p, "w") as f:
            json.dump(existing, f, indent=2, sort_keys=True)
        logger.info(f"historical_db: saved {n_new} new (canon, year) "
                    f"values for {symbol} → {p}")
    except Exception as exc:
        logger.warning(f"historical_db: save failed: {exc}")
    return n_new


# ── Seeding from a filled template ──────────────────────────────────────────

def seed_from_template(symbol: str, template_path: str,
                       sector: str = "general", name: str = "") -> int:
    """
    Read a filled template Excel (e.g. a 'known-good' truth file) and store
    every numeric cell whose row label resolves to a KB canonical.

    Returns the number of (canonical, year) cells persisted.
    """
    import openpyxl
    from pipeline.kb_loader import kb_entry_for, load_merged_kb

    p = Path(template_path)
    if not p.exists():
        logger.warning(f"historical_db: template not found: {template_path}")
        return 0

    wb = openpyxl.load_workbook(str(p), data_only=True)
    values_by_canonical: dict[str, dict[str, float]] = {}
    kb_dict = load_merged_kb(sector)

    import re
    YEAR_RE = re.compile(r"^[Ff]\d{4}$")

    for sn in wb.sheetnames:
        ws = wb[sn]
        # Find header row and year columns
        header_row = None
        year_cols = {}
        for r in range(1, min(15, ws.max_row + 1)):
            cnt = 0
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if v and YEAR_RE.match(str(v).strip()):
                    cnt += 1
            if cnt >= 2 and header_row is None:
                header_row = r
                for c in range(1, ws.max_column + 1):
                    v = ws.cell(r, c).value
                    if v and YEAR_RE.match(str(v).strip()):
                        year_cols[str(v).strip()] = c
                break
        if not header_row or not year_cols:
            continue

        for r in range(header_row + 1, ws.max_row + 1):
            label = ws.cell(r, 1).value
            if not label:
                continue
            label = str(label).strip()
            if not label:
                continue
            entry = kb_entry_for(label, sector)
            if not entry:
                continue
            canon = None
            for ck, info in kb_dict.items():
                if info is entry:
                    canon = ck; break
            if not canon:
                continue
            for yr, col in year_cols.items():
                v = ws.cell(r, col).value
                if isinstance(v, (int, float)) and v not in (0, None):
                    values_by_canonical.setdefault(canon, {}).setdefault(yr, float(v))

    n_saved = save_run(symbol, values_by_canonical, sector=sector, name=name,
                       only_high_confidence=False)
    return n_saved


__all__ = [
    "load_historical", "save_run", "seed_from_template",
    "get_value", "get_prior_year_value",
]
