"""
pipeline/orchestrator.py
────────────────────────
The top-level "Smart Pipeline" that turns a company name + a blank Excel
template into a fully populated workbook, sector-aware and validated.

Sequence
────────
 1. Read template structure (any layout — driven by mapper.template_reader).
 2. Detect sector for the company (banking / NBFC / insurance / general / …)
    so the right KB overlay and validation rules are loaded.
 3. Build a per-sheet StatementPlan (consolidated, standalone, or both)
    based on what the template expects.
 4. Source pull — XBRL → Screener → Tickertape → yfinance → optional PDF.
    All cached on disk so reruns don't hammer the sources.
 5. Build a unified pool: {label: {year: value}} with provenance per cell.
 6. For each sheet × statement_type:
      a. KB-aware label mapping (uses pipeline.mapping.llm_map_with_kb if a
         client is provided; falls back to deterministic alias matching).
      b. Derivation pass — fills computed rows via accounting identities and
         KB formulas (margins, sub-totals, NII, etc.).
      c. Validation gate (sector-aware).
      d. Apply safe auto-fixes from the validation report.
 7. Final write to the workbook via integrator.excel_writer (non-destructive).
 8. Returns OrchestratorResult with audit, validation, and source provenance.

This module is fully decoupled from Streamlit. It can be called from a CLI,
a test, or wired into web_only_app.py.
"""

from __future__ import annotations

import logging
import re
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Callable, Optional

from web.base import CompanyIdentifier, WebDataResult

from pipeline import cache as cache_mod
from pipeline.kb_loader     import (
    kb_entry_for, kb_context_for, load_merged_kb,
)
from pipeline.sector_router import detect_sector, is_financial_services
from pipeline.stmt_type_resolver import resolve_for_template, StatementPlan
from pipeline.validation_gate    import (
    validate_sheet, apply_auto_fixes, ValidationReport,
)

logger = logging.getLogger(__name__)


# ── Result classes ──────────────────────────────────────────────────────────

@dataclass
class CellProvenance:
    """Where a single (sheet, label, year) value came from."""
    sheet:        str
    label:        str
    year:         str
    value:        float
    source:       str             # "xbrl" | "screener" | "yfinance" | "tickertape" | "derived" | "auto_fix"
    confidence:   str             # "VERY_HIGH" | "HIGH" | "MEDIUM" | "LOW"
    extracted_as: Optional[str] = None  # the source-side label that mapped here
    notes:        str = ""


@dataclass
class OrchestratorResult:
    company:      CompanyIdentifier
    sector:       str
    sector_conf:  float
    template_path: str
    output_path:  str
    years:        list[str]
    pool_size:    int = 0
    cells_written: int = 0
    cells_review_queued: int = 0
    cells_skipped_formula: int = 0
    cells_skipped_filled: int = 0
    auto_fixes_applied: int = 0
    sheets_processed: list[str] = field(default_factory=list)
    provenance:   list[CellProvenance] = field(default_factory=list)
    validation_reports: dict[str, ValidationReport] = field(default_factory=dict)
    warnings:     list[str] = field(default_factory=list)
    elapsed_sec:  float = 0.0
    source_status: list[dict] = field(default_factory=list)  # cached/fresh/empty per source
    agent_activity: list[dict] = field(default_factory=list)  # per sheet × stmt
    agent_dropped: list[dict] = field(default_factory=list)   # fills the agent submitted that got dropped
    consensus_meta: dict[tuple[str, str], dict] = field(default_factory=dict)  # per-cell confidence (HIGH/MEDIUM/FLAGGED) from multi-source agreement
    confidence_counts: dict[str, int] = field(default_factory=dict)  # tally of how many cells got each confidence tier
    validation_flagged: set = field(default_factory=set)  # (sheet, label, year) cells where retry couldn't fix the issue

    def summary(self) -> str:
        cached = sum(1 for s in self.source_status if s.get("cached"))
        live   = sum(1 for s in self.source_status if not s.get("cached"))
        agent_runs   = len(self.agent_activity)
        agent_fills  = sum(a.get("fills_landed", 0) for a in self.agent_activity)
        agent_calls  = sum(a.get("tool_calls", 0) for a in self.agent_activity)
        agent_str = (f" · agent: {agent_runs} runs, {agent_calls} tool calls, "
                     f"{agent_fills} fills" if agent_runs else "")
        return (
            f"{self.company.name} [{self.sector}] · "
            f"{len(self.sheets_processed)} sheets · "
            f"{self.cells_written} cells written, "
            f"{self.auto_fixes_applied} auto-fixes, "
            f"{self.cells_review_queued} flagged · "
            f"pool={self.pool_size} · sources={live} live + {cached} cached"
            f"{agent_str} · {self.elapsed_sec:.1f}s"
        )


# ── Source pull layer ───────────────────────────────────────────────────────

@dataclass
class SourceFetchResult:
    source:  str
    items:   list[WebDataResult]
    error:   Optional[str] = None
    cached:  bool = False


# Module-level switches set by run_pipeline before fetch helpers run.
_CACHE_BYPASS = False


def _sanity_check_source(items: list, source_name: str,
                         min_items: int = 3) -> tuple[bool, str]:
    """
    Generic sanity-check for a connector's output. Company-agnostic.
    Returns (ok, reason_if_bad).

    Rules:
      • At least `min_items` items
      • Sum of absolute values > 0 (no all-zero responses)
      • At least one numeric value > 1 (filters noise like 0.0001)
      • Has at least one label with > 4 alphabetic chars (filters pure-noise)
    """
    if not items:
        return False, "empty"
    if len(items) < min_items:
        return False, f"only {len(items)} item(s) (min {min_items})"
    total_abs = 0.0
    real_label_count = 0
    big_value_count = 0
    for it in items:
        try:
            v = float(getattr(it, "value", None) or 0)
        except (TypeError, ValueError):
            v = 0
        total_abs += abs(v)
        if abs(v) > 1.0:
            big_value_count += 1
        lbl = (getattr(it, "raw_field", None) or "").strip()
        if sum(1 for c in lbl if c.isalpha()) > 4:
            real_label_count += 1
    if total_abs == 0:
        return False, "all values are zero"
    if big_value_count == 0:
        return False, "no values > 1 (looks like noise)"
    if real_label_count == 0:
        return False, "no labels with real text"
    return True, f"ok ({len(items)} items, sum_abs={total_abs:,.0f})"


def _cached_get(key):
    if _CACHE_BYPASS:
        return None
    return cache_mod.get(key)


def _cached_put(key, value, min_items: int = 1):
    """Only cache results that pass a basic quality bar."""
    try:
        n = len(value) if hasattr(value, "__len__") else (1 if value else 0)
    except Exception:
        n = 0
    if n < min_items:
        return False
    cache_mod.put(key, value)
    return True


def _try_xbrl(company: CompanyIdentifier, years: list[str], stmt: str) -> SourceFetchResult:
    key = ("xbrl", company.nse_symbol or company.bse_code, tuple(sorted(years)), stmt)
    cached = _cached_get(key)
    if cached is not None:
        return SourceFetchResult("xbrl", cached, cached=True)
    try:
        from web.xbrl_extractor import extract_from_xbrl
        results = extract_from_xbrl(
            symbol=company.nse_symbol or "",
            years=years,
            bse_code=company.bse_code,
            statement_type=stmt,
        )
        # XBRL returns ExtractionResult objects; convert to WebDataResult-like
        wds: list[WebDataResult] = []
        for r in results:
            for label, val in (r.data or {}).items():
                if val is None:
                    continue
                try:
                    fv = float(val)
                except (TypeError, ValueError):
                    continue
                wds.append(WebDataResult(
                    source="xbrl",
                    raw_field=label,
                    template_field="",
                    value=fv,
                    year=r.year or "",
                    confidence="HIGH",
                    unit_applied="INR Crores (XBRL)",
                ))
        _cached_put(key, wds, min_items=3)
        return SourceFetchResult("xbrl", wds)
    except Exception as exc:
        logger.warning(f"xbrl fetch failed: {exc}")
        return SourceFetchResult("xbrl", [], error=str(exc))


def _fetch_with_sanity(source_name: str, key: tuple,
                       fetcher, min_items: int = 3) -> SourceFetchResult:
    """Common helper: cache lookup → fetch → sanity check → conditional cache."""
    cached = _cached_get(key)
    if cached is not None:
        return SourceFetchResult(source_name, cached, cached=True)
    try:
        items = fetcher()
    except Exception as exc:
        logger.warning(f"{source_name} fetch failed: {exc}")
        return SourceFetchResult(source_name, [], error=str(exc))
    ok, reason = _sanity_check_source(items, source_name, min_items=min_items)
    if not ok:
        logger.info(f"{source_name}: sanity-check FAILED ({reason}) — not cached")
        return SourceFetchResult(source_name, items if items else [],
                                 error=f"sanity_failed: {reason}")
    _cached_put(key, items, min_items=min_items)
    return SourceFetchResult(source_name, items)


def _try_screener(company: CompanyIdentifier, years: list[str]) -> SourceFetchResult:
    from web.screener_connector import ScreenerConnector
    key = ("screener", company.nse_symbol or company.bse_code, tuple(sorted(years)))
    return _fetch_with_sanity("screener", key,
                              lambda: ScreenerConnector().fetch(company, years))


def _try_tickertape(company: CompanyIdentifier, years: list[str]) -> SourceFetchResult:
    from web.tickertape_connector import TickertapeConnector
    key = ("tickertape", company.nse_symbol or company.bse_code, tuple(sorted(years)))
    return _fetch_with_sanity("tickertape", key,
                              lambda: TickertapeConnector().fetch(company, years))


def _try_yfinance(company: CompanyIdentifier, years: list[str]) -> SourceFetchResult:
    from web.yfinance_connector import YFinanceConnector
    key = ("yfinance", company.yfinance_ticker, tuple(sorted(years)))
    return _fetch_with_sanity("yfinance", key,
                              lambda: YFinanceConnector().fetch(company, years))


# Source priority (later sources never overwrite earlier on the same year)
#
# AUDIT 2026-05-11 — Priorities reordered.
# Scoring against PPL analyst workbook showed raw_excel achieved 11%
# accuracy while the LLM-enhanced pipeline ("smart") achieved 1%.
# raw_excel had exact matches on FY Revenue (6559 vs GT 6559.1) where
# the smart pipeline hallucinated (5775). Conclusion: the LLM/PDF tier
# was systematically OVERWRITING correct values from the deterministic
# layer. raw_excel is now treated as near-ground-truth (priority 95);
# only XBRL outranks it.
#
# The new "derived_agg" source comes from pipeline.aggregation_layer
# (deterministic post-pool aggregation: GP=Rev-COGS, EBITDA=GP-Emp-Oth,
# Total Assets = sum of components, etc.). It runs after _build_pool
# and ranks just below raw_excel — formulaic totals are trustworthy
# given trustworthy leaves.
_SOURCE_PRIORITY = {
    "xbrl":           100,  # exchange-filed structured XBRL — direct source
    "raw_excel":       95,  # PROMOTED from 25 (audit 2026-05-11).
    "derived_agg":     93,  # NEW — deterministic post-pool aggregation
    "pdf":             90,  # annual report PDF tables
    "screener":        80,  # scrapes PDF data into a clean schema
    "stockanalysis":   75,
    "tickertape":      70,
    "yfinance":        40,  # reformulates statements; can pollute pool
    "derived":         50,  # legacy derivation_agent (rule-based)
    "presentation":    35,
    "web":             10,
    "auto_fix":         5,
}


# ── yfinance noise filter ────────────────────────────────────────────────────
# yfinance returns ~700 items per company including weird normalisation
# fields and signed cash-flow items that DON'T match Indian-AR conventions.
# Filter aggressively at intake.

_YFINANCE_NOISE_LABELS = re.compile(
    r"^(?:tax effect of|tax rate for|normalized\s|reconciled\s|"
    r"diluted average|basic average|special income|"
    r"net non operating interest|otherunder preferred|"
    r"(?:repurchase|issuance) of capital|"
    r"(?:proceeds|repayment) of (?:debt|long term debt))",
    re.IGNORECASE,
)


def _normalise_yfinance_item(label: str, value: float) -> tuple[float, bool]:
    """
    yfinance reports cash-flow-style items as NEGATIVE (it's a sources-and-uses
    view). Indian-AR convention is positive magnitudes. Flip sign for known
    cash-outflow items so they don't pollute P&L rows when matched.
    Returns (corrected_value, keep_in_pool).
    """
    if _YFINANCE_NOISE_LABELS.match(label or ""):
        return value, False   # drop entirely
    lc = (label or "").lower()
    # Items that yfinance signs negative but Indian P&L treats as positive
    flip_to_positive = (
        "interest expense", "tax provision", "income tax",
        "depreciation amortization", "depreciation amortization depletion",
        "reconciled depreciation",
    )
    if value < 0 and any(k in lc for k in flip_to_positive):
        return -value, True
    return value, True


# ── Garbage-label detection ──────────────────────────────────────────────────
# PDF parsers sometimes emit fragmented text like "( iii ) O t h e r F in a n c i".
# These shouldn't enter the pool because they'll randomly match template labels
# via fuzzy-overlap and inject wrong values.

_AUDITOR_INTRO_RE = re.compile(
    r"\bfor\s+[A-Z][\w&\.\-]*(?:\s+[A-Z][\w&\.\-]*){0,5}\s*(?:LLP|& Co|& Associates|"
    r"Chartered Accountants)\b", re.IGNORECASE,
)
_DIRECTORSHIP_RE = re.compile(
    r"\bDIN\s*[:#]?\s*\d{6,}", re.IGNORECASE,
)
_SIGNATORY_ROLE_RE = re.compile(
    r"\b(?:chairman|managing director|whole.?time director|chief financial officer|"
    r"chief executive officer|company secretary|partner|membership(?:\s+no)?|"
    r"registration(?:\s+no)?)\b",
    re.IGNORECASE,
)
_PLACE_DATE_RE = re.compile(
    r"^(?:place|date)\s*[:#]", re.IGNORECASE,
)


def _is_garbage_label(label: str) -> bool:
    """
    Heuristic — flag labels that look like PDF text-extraction noise OR
    auditor-report / signatory boilerplate. Company-agnostic: no specific
    company name, person name, or audit-firm name is hardcoded.
    """
    if not label:
        return True
    s = label.strip()
    if len(s) < 2 or len(s) > 200:
        return True
    words = re.split(r"\s+", s)
    if not words:
        return True
    # Single-character word ratio — PDF fragmentation produces many of these
    short_words = sum(1 for w in words if len(re.sub(r"[^a-zA-Z0-9]", "", w)) <= 1)
    if len(words) >= 4 and short_words / len(words) > 0.45:
        return True
    # Very low letter content (mostly punctuation/digits)
    alpha = sum(1 for c in s if c.isalpha())
    if alpha < 3:
        return True
    # Boilerplate from audit reports / signatory blocks — generic patterns
    if (_AUDITOR_INTRO_RE.search(s)
            or _DIRECTORSHIP_RE.search(s)
            or _PLACE_DATE_RE.search(s)
            or _SIGNATORY_ROLE_RE.search(s)):
        return True
    # Standard auditor-report opening boilerplate (any audit firm)
    sl = s.lower()
    if "in terms of our report" in sl or "in terms of our attached report" in sl:
        return True
    return False


def _consensus_pick(items: list[tuple[float, str, str]]
                    ) -> tuple[float, str, int, list[str]]:
    """
    Given a list of (value, source, original_label) for the SAME canonical+year,
    cluster the values into agreement-groups (within 5% of each other), then
    pick the cluster with the most distinct sources. Returns
        (chosen_value, confidence_tier, n_unique_sources, source_list)

    Confidence tiers:
       HIGH    : ≥ 2 distinct sources agree (within 5%)
       MEDIUM  : single-source value
       FLAGGED : multiple sources but they disagree → use highest-priority
    """
    if not items:
        return 0.0, "LOW", 0, []
    if len(items) == 1:
        return items[0][0], "MEDIUM", 1, [items[0][1]]

    # Cluster values that agree within 5% relative tolerance
    clusters: list[dict] = []
    for v, src, lbl in items:
        placed = False
        for c in clusters:
            denom = max(abs(c["median"]), abs(v), 1.0)
            if abs(v - c["median"]) / denom < 0.05:
                c["items"].append((v, src, lbl))
                vals = sorted(x[0] for x in c["items"])
                c["median"] = vals[len(vals) // 2]
                placed = True
                break
        if not placed:
            clusters.append({"median": v, "items": [(v, src, lbl)]})

    def _cluster_score(c: dict) -> tuple[int, int]:
        srcs = {x[1] for x in c["items"]}
        max_prio = max((_SOURCE_PRIORITY.get(s, 0) for s in srcs), default=0)
        return (len(srcs), max_prio)

    best = max(clusters, key=_cluster_score)
    val = best["median"]
    n_unique = len({x[1] for x in best["items"]})
    src_list = sorted({x[1] for x in best["items"]})

    if len(clusters) > 1 and n_unique == 1:
        # Multiple disagreeing sources, single best — flagged for review
        conf = "FLAGGED"
    elif n_unique >= 2:
        conf = "HIGH"
    else:
        conf = "MEDIUM"
    return val, conf, n_unique, src_list


def _build_pool(
    fetched: list[SourceFetchResult],
    sector: str = "general",
) -> tuple[
    dict[str, dict[str, float]],                 # pool: label → year → value
    dict[tuple[str, str], list[CellProvenance]], # provenance: (label, year) → list
    dict[tuple[str, str], dict],                 # consensus: (label_lower, year) → meta
]:
    """
    Aggregate WebDataResult objects across sources into a unified pool using
    multi-source CONSENSUS. For each (KB-canonical-or-label, year), collects
    every source's value, clusters by 5% agreement, and picks the cluster
    with the most distinct sources. Tags every chosen value with confidence
    (HIGH/MEDIUM/FLAGGED) so downstream can prefer agreed values.
    """
    # Build full pool keeping EVERY source label as a separate pool entry
    # (don't collapse). Source-priority winner determines value; consensus
    # is computed SEPARATELY by canonical-key for confidence tagging.
    pool: dict[str, dict[str, float]] = {}
    provenance: dict[tuple[str, str], list[CellProvenance]] = {}
    chosen_priority: dict[tuple[str, str], int] = {}
    # Track all source values per canonical+year for separate consensus check
    canonical_obs: dict[tuple[str, str], list[tuple[float, str]]] = {}
    n_garbage = 0
    n_yf_filtered = 0

    for fr in fetched:
        for w in fr.items:
            label = (w.raw_field or "").strip()
            year  = (w.year or "").strip()
            if not label or not year:
                continue
            if _is_garbage_label(label):
                n_garbage += 1
                continue
            try:
                v = float(w.value)
            except (TypeError, ValueError):
                continue
            if w.source == "yfinance":
                v, keep = _normalise_yfinance_item(label, v)
                if not keep:
                    n_yf_filtered += 1
                    continue
            llower = label.lower()
            _is_loss_eligible = any(k in llower for k in (
                "profit", "earnings", "income before tax", "income after tax",
                "comprehensive income", "exceptional", "gain", "loss",
            ))
            if v < 0 and abs(v) > 0.5 and not _is_loss_eligible:
                v = abs(v)
            key = (llower, year)
            prio = _SOURCE_PRIORITY.get(w.source, 0)

            provenance.setdefault(key, []).append(CellProvenance(
                sheet="", label=label, year=year, value=v,
                source=w.source, confidence=w.confidence or "MEDIUM",
                extracted_as=label, notes=w.unit_applied or "",
            ))

            # Pool: source-priority winner per (label_lower, year)
            if prio > chosen_priority.get(key, -1):
                pool.setdefault(label, {})[year] = v
                chosen_priority[key] = prio

            # Track for canonical-level consensus (DOESN'T collapse pool)
            canon = _kb_canonical_for_pool_label(label, sector)
            if canon:
                canonical_obs.setdefault((canon, year), []).append((v, w.source))

    # Compute consensus tags per (canonical, year) — NOT used to pick pool
    # values, only to tag confidence of the chosen value
    consensus_meta: dict[tuple[str, str], dict] = {}
    for (canon, year), obs in canonical_obs.items():
        _, conf, n_src, src_list = _consensus_pick(
            [(v, s, "") for v, s in obs]
        )
        consensus_meta[(canon, year)] = {
            "confidence": conf, "n_sources": n_src, "sources": src_list,
        }

    if n_garbage:
        logger.info(f"_build_pool: filtered {n_garbage} garbage labels")
    if n_yf_filtered:
        logger.info(f"_build_pool: filtered {n_yf_filtered} yfinance noise items")

    high = sum(1 for c in consensus_meta.values() if c["confidence"] == "HIGH")
    med  = sum(1 for c in consensus_meta.values() if c["confidence"] == "MEDIUM")
    flg  = sum(1 for c in consensus_meta.values() if c["confidence"] == "FLAGGED")
    logger.info(
        f"_build_pool: {len(pool)} pool entries; consensus tags: "
        f"{high} HIGH, {med} MEDIUM, {flg} FLAGGED"
    )
    return pool, provenance, consensus_meta


def _canonical_dedup_pool(
    pool: dict[str, dict[str, float]],
    sector: str,
) -> dict[str, dict[str, float]]:
    """
    Collapse pool entries that resolve to the same KB canonical concept.

    Example: 'Net Profit', 'PAT', 'Profit for the year', 'Net Profit for the
    Year' all map to canonical 'profit_after_tax' → keep ONE entry per (canonical,
    year), preferring the highest-magnitude value (sub-items are usually smaller
    than the main statement value).

    Pool entries with NO canonical mapping are kept as-is.
    """
    grouped: dict[str, dict[str, dict[str, float]]] = {}
    no_canonical: dict[str, dict[str, float]] = {}

    for label, yr_vals in pool.items():
        ck = _kb_canonical_for_pool_label(label, sector)
        if ck is None:
            no_canonical[label] = yr_vals
            continue
        bucket = grouped.setdefault(ck, {})
        for yr, v in yr_vals.items():
            existing = bucket.get(yr)
            # Prefer the larger absolute value — sub-items are typically smaller
            if existing is None or abs(v) > abs(existing):
                bucket[yr] = v

    # Build the deduped pool
    out: dict[str, dict[str, float]] = {}
    out.update(no_canonical)
    for ck, yr_vals in grouped.items():
        # Use the canonical name humanised as the pool key
        nice = ck.replace("_", " ").title()
        out[nice] = yr_vals

    if len(pool) != len(out):
        logger.info(
            f"_canonical_dedup_pool: {len(pool)} → {len(out)} unique entries "
            f"(merged {len(pool) - len(out)} aliases)"
        )
    return out


# ── KB-aware mapping (deterministic + optional LLM) ─────────────────────────

from functools import lru_cache

@lru_cache(maxsize=20000)
def _kb_canonical_for_pool_label(label: str, sector: str) -> Optional[str]:
    """Return canonical KB key for a pool label, or None. Cached — called
    O(templates × pool) times so caching is essential for performance."""
    entry = kb_entry_for(label, sector)
    if not entry:
        return None
    kb = load_merged_kb(sector)
    for ck, info in kb.items():
        if info is entry:
            return ck
    return None


# Operator prefixes / suffixes that templates use for arithmetic readability.
# These don't carry semantic meaning and block alias matching, so we strip them
# before comparing template labels to pool labels.
_PREFIX_PATTERNS = re.compile(
    r"^\s*(?:less\s*:|add\s*:|plus\s*:|minus\s*:)\s*",
    re.IGNORECASE,
)
_NOISE_PATTERNS = [
    re.compile(r"\(net\)\s*$",    re.IGNORECASE),
    re.compile(r"\(gross\)\s*$",  re.IGNORECASE),
    re.compile(r"\(net of tax\)\s*$", re.IGNORECASE),
    re.compile(r"\(consolidated\)\s*$", re.IGNORECASE),
    re.compile(r"\(standalone\)\s*$", re.IGNORECASE),
    re.compile(r"\(in\s*inr.*\)\s*$", re.IGNORECASE),
    re.compile(r"&\s*impairment\b", re.IGNORECASE),
    re.compile(r"\bnet\s+of\s+\w+\b", re.IGNORECASE),
]
# Token synonyms used to enrich word-overlap matching.
_TOKEN_SYNONYMS: dict[str, set[str]] = {
    # Spelling variants — safe 1:1 substitutions
    "amortisation": {"amortization", "amortisation"},
    "amortization": {"amortization", "amortisation"},
    "depreciation": {"depreciation", "depn"},
    # Context-specific financial synonyms — kept narrow to avoid false positives.
    # DO NOT add broad supersets like profit→{income,earnings} because that inflates
    # overlap for labels like "Add: Share of net profit of associates" and causes
    # them to wrongly match "Net Profit" (canonical profit_after_tax).
    "tax":          {"tax", "taxation"},
    "interest":     {"interest", "finance"},
    "finance":      {"finance", "interest"},
    "ppe":          {"ppe", "property", "plant", "equipment"},
    "fixed":        {"fixed", "ppe"},
    "receivables":  {"receivables", "debtors"},
    "payables":     {"payables", "creditors"},
    "borrowings":   {"borrowings", "debt", "loans"},
    "cash":         {"cash", "bank"},
    "employee":     {"employee", "staff", "personnel"},
}


def _strip_operator(label: str) -> str:
    """Remove arithmetic-readability prefixes like 'Less:' and noise suffixes."""
    s = _PREFIX_PATTERNS.sub("", label or "")
    for pat in _NOISE_PATTERNS:
        s = pat.sub("", s)
    return s.strip()


def _normalise(s: str) -> str:
    s = _strip_operator(s or "")
    return re.sub(r"[^a-z0-9]+", " ", s.lower()).strip()


def _expand_tokens(words: set[str]) -> set[str]:
    """Add known synonyms so 'depreciation amortization' matches 'D&A'."""
    out = set(words)
    for w in list(words):
        if w in _TOKEN_SYNONYMS:
            out |= _TOKEN_SYNONYMS[w]
    return out


def _deterministic_match(
    template_label: str,
    pool_keys: list[str],
    sector: str,
    pool: Optional[dict[str, dict[str, float]]] = None,
    sibling_magnitude: Optional[float] = None,
) -> Optional[str]:
    """
    Try exact KB-canonical match first, then exact string match, then Jaccard.

    Critical design note (BUG FIX):
      If the template label has a KB canonical but NO pool label maps to the
      same canonical, we MUST fall through to string/Jaccard matching rather
      than returning None immediately.  The old code returned None here, which
      caused every well-known financial label (Revenue, EBITDA, PAT, …) to fail
      completely unless an exact canonical match was found in the pool — even
      when a pool label with very similar text existed.  This was the primary
      cause of the 9% fill rate.

    When MULTIPLE pool labels resolve to the same canonical (e.g. several
    "Finance Cost" entries from different sections of a fragmented PDF),
    we pick the candidate whose magnitude is most plausible — biggest
    absolute value if no `sibling_magnitude` hint is given.
    """
    if not pool_keys:
        return None

    def _max_abs(pk: str) -> float:
        if not pool:
            return 0.0
        return max((abs(float(v)) for v in pool.get(pk, {}).values()
                    if v is not None), default=0.0)

    # 1. Both sides resolve to the same KB canonical (most reliable).
    #    When multiple pool items map to same canonical, pick by magnitude —
    #    UNLESS the template label is itself a sub-item (e.g. "of Subsidiary X")
    #    in which case the smaller value is correct.
    tmpl_canon = _kb_canonical_for_pool_label(template_label, sector)
    if tmpl_canon:
        same_canon = [pk for pk in pool_keys
                      if _kb_canonical_for_pool_label(pk, sector) == tmpl_canon]
        if same_canon:
            # Found at least one pool label with the same KB canonical.
            if len(same_canon) == 1 or pool is None:
                return same_canon[0]

            # Multiple pool items share the canonical — pick by magnitude.
            # Step A — magnitude prior: pick the largest-magnitude candidate
            # (the main statement row, not a footnote sliver).
            # Step B — outlier guard: if any candidate is >5× the median, that
            # candidate is likely a TOTAL row pretending to match a component
            # row (e.g. Total Equity matching "Equity Share Capital"). In that
            # case, drop the outlier and re-apply Step A.
            def _max_abs(pk: str) -> float:
                return max((abs(float(v)) for v in pool.get(pk, {}).values()
                            if v is not None), default=0.0)

            magnitudes = [(_max_abs(pk), pk) for pk in same_canon]
            magnitudes.sort(reverse=True)
            if len(magnitudes) >= 3:
                median_mag = magnitudes[len(magnitudes) // 2][0]
                if median_mag > 0 and magnitudes[0][0] > 5 * median_mag:
                    return magnitudes[1][1]
            return magnitudes[0][1]
        # else: tmpl_canon was found but NO pool label matches it.
        # DO NOT return None here — fall through to string/Jaccard matching
        # so labels with KB entries still get a chance to match by text.

    # 2. Exact normalised string match (catches case/punctuation variants)
    tmpl_n = _normalise(template_label)
    for pk in pool_keys:
        if _normalise(pk) == tmpl_n:
            return pk

    # 3. Word-overlap match with synonym expansion (>= 55%).
    #    Collect ALL plausible candidates first; pick by combined score.
    #    The canonical guard on pool labels: skip pool labels that already
    #    have a DIFFERENT canonical (they belong to a different concept).
    tmpl_words = _expand_tokens(set(tmpl_n.split()))
    if not tmpl_words:
        return None
    candidates: list[tuple[float, str]] = []   # (score, pool_key)
    for pk in pool_keys:
        pk_canon = _kb_canonical_for_pool_label(pk, sector)
        # Skip pool labels locked to a DIFFERENT canonical than the template.
        # If tmpl_canon is None we allow any pool label through, BUT we apply a
        # stricter score threshold for pool labels that DO have a canonical
        # (they're reserved for the canonical-path match in tier 1/2 of the sort;
        # letting them be stolen by vague non-canonical template labels causes
        # wrong matches like "Share of net profit of associates" → "Net Profit").
        if pk_canon is not None and tmpl_canon is not None and pk_canon != tmpl_canon:
            continue
        pk_words = _expand_tokens(set(_normalise(pk).split()))
        if not pk_words:
            continue
        intersection = len(tmpl_words & pk_words)
        union = max(len(tmpl_words | pk_words), 1)
        jaccard = intersection / union
        if min(len(tmpl_words), len(pk_words)) >= 3:
            contain = intersection / min(len(tmpl_words), len(pk_words))
            score = max(jaccard, 0.7 * contain)
        else:
            score = jaccard
        # Raise bar to 0.70 when the pool label has a KB canonical but the
        # template label doesn't — that pool label "belongs" to a specific
        # concept and should only match if the overlap is very high.
        threshold = 0.70 if (pk_canon is not None and tmpl_canon is None) else 0.55
        if score >= threshold:
            candidates.append((score, pk))
    if not candidates:
        return None
    candidates.sort(reverse=True)
    return candidates[0][1]


# ── Magnitude / identity guards ─────────────────────────────────────────────
# When a candidate pool value is wildly inconsistent with what we already know,
# refuse the match. This catches the classic "Other Expenses → matched to
# Total Expenses" case where the candidate value is ~3× the truth.

def _violates_magnitude(
    template_label: str,
    candidate_values: dict[str, float],     # {year: value}
    sheet_filled: dict[str, dict[str, float]],
    sector: str,
) -> Optional[str]:
    """
    Returns a reason string if the candidate values violate sanity rules
    given what's already filled in this sheet, else None.
    """
    tlbl_norm = _normalise(template_label)
    is_total_label = any(t in tlbl_norm for t in ("total", "sum"))

    # Rule 1: For sub-component rows (NOT total/sum), value must not exceed
    # the value of any already-filled "total" row in the same year, in the
    # same sub-context.
    if not is_total_label:
        for filled_lbl, yr_vals in sheet_filled.items():
            fnorm = _normalise(filled_lbl)
            if not any(t in fnorm for t in ("total", "sum")):
                continue
            # Crude context check — share a non-trivial keyword (e.g. "expenses",
            # "current", "non current", "assets", "liabilities")
            common = set(tlbl_norm.split()) & set(fnorm.split())
            common -= {"total", "sum", "of", "the"}
            if not common:
                continue
            for yr, v in candidate_values.items():
                tot = yr_vals.get(yr)
                if tot is None or tot <= 0:
                    continue
                if v > tot * 1.05:   # 5% slack for rounding
                    return (f"value {v:,.1f} > matching total "
                            f"'{filled_lbl}' = {tot:,.1f} for {yr}")

    # Rule 2: YoY swing > 10x (likely scale error from a different unit)
    yrs_sorted = sorted(candidate_values.keys())
    for i in range(1, len(yrs_sorted)):
        prev = candidate_values.get(yrs_sorted[i - 1])
        curr = candidate_values.get(yrs_sorted[i])
        if prev is None or curr is None:
            continue
        if abs(prev) < 1.0:
            continue
        ratio = abs(curr) / max(abs(prev), 1e-6)
        if ratio > 10.0 or ratio < 0.1:
            return (f"YoY swing {ratio:.1f}x for {yrs_sorted[i-1]}→"
                    f"{yrs_sorted[i]} (likely scale error)")

    return None


def _all_candidates(
    template_label: str,
    pool: dict[str, dict[str, float]],
    sector: str,
    used: Optional[set[str]] = None,
) -> list[str]:
    """
    Return ALL pool keys that could match this template label, ranked best-first.
    Used by the validation-retry gate to try alternative candidates when the
    first pick fails an accounting identity.
    """
    used = used or set()
    avail = [pk for pk in pool.keys() if pk not in used]
    if not avail:
        return []

    # Layer 1: same KB canonical
    tmpl_canon = _kb_canonical_for_pool_label(template_label, sector)
    canonical_matches = []
    if tmpl_canon:
        canonical_matches = [
            pk for pk in avail
            if _kb_canonical_for_pool_label(pk, sector) == tmpl_canon
        ]

    # Layer 2: token overlap
    from difflib import SequenceMatcher as _SM
    tlbl_n = _normalise(template_label)
    tlbl_words = _expand_tokens(set(tlbl_n.split()))
    fuzzy_matches: list[tuple[float, str]] = []
    for pk in avail:
        if pk in canonical_matches:
            continue
        pk_canon = _kb_canonical_for_pool_label(pk, sector)
        if pk_canon and pk_canon != tmpl_canon:
            continue
        pk_words = _expand_tokens(set(_normalise(pk).split()))
        if not pk_words:
            continue
        intersection = len(tlbl_words & pk_words)
        union = max(len(tlbl_words | pk_words), 1)
        jaccard = intersection / union
        if jaccard >= 0.30:
            fuzzy_matches.append((jaccard, pk))
    fuzzy_matches.sort(reverse=True)
    return canonical_matches + [pk for _, pk in fuzzy_matches]


def _map_template_to_pool(
    template_labels: list[str],
    pool: dict[str, dict[str, float]],
    sector: str,
    sheet_name: str,
    years: list[str],
    client: Any = None,
    model: str = "",
) -> dict[str, str]:
    """
    For each template label, find the best matching pool label.
    Returns {template_label: pool_label} (only successful matches).
    Two-phase: first match high-confidence "total" rows, then sub-components
    (so magnitude guards have totals to check against).
    """
    pool_keys = list(pool.keys())
    used: set[str] = set()
    matches: dict[str, str] = {}

    # Track what's been matched so far, with values, for magnitude guard
    matched_values: dict[str, dict[str, float]] = {}

    # Sort priority (lowest number = matched first):
    #   0 — has KB canonical AND is a total row  (anchors for magnitude guard)
    #   1 — has KB canonical, not a total        (main line items — matched before vague labels)
    #   2 — no KB canonical, is a total
    #   3 — no KB canonical, not a total         (descriptive / supplemental rows last)
    #
    # This prevents vague labels like "Add: Share of net profit of associates"
    # (no KB canonical) from stealing pool entries like "Net Profit" (canonical
    # profit_after_tax) that should go to "PAT" (canonical profit_after_tax).
    def _sort_key(lbl: str) -> tuple:
        n = _normalise(lbl)
        is_total = any(t in n for t in ("total", "sum"))
        has_canon = _kb_canonical_for_pool_label(lbl, sector) is not None
        tier = (0 if (has_canon and is_total)
                else 1 if has_canon
                else 2 if is_total
                else 3)
        return (tier, lbl)
    sorted_labels = sorted(template_labels, key=_sort_key)

    # 1. Deterministic pass with magnitude guard
    unresolved: list[str] = []
    for tlbl in sorted_labels:
        cand = _deterministic_match(
            tlbl, [k for k in pool_keys if k not in used],
            sector, pool=pool,
        )
        if cand:
            cand_vals = pool.get(cand, {})
            violation = _violates_magnitude(tlbl, cand_vals, matched_values, sector)
            if violation:
                logger.debug(
                    f"map: rejected '{tlbl}' → '{cand}' on '{sheet_name}': "
                    f"{violation}"
                )
                unresolved.append(tlbl)
                continue
            matches[tlbl] = cand
            used.add(cand)
            matched_values[tlbl] = cand_vals
        else:
            unresolved.append(tlbl)

    # 2. LLM-aware pass for the leftovers
    if unresolved and client is not None:
        try:
            from pipeline.mapping import llm_map_with_kb
            lbl_map = llm_map_with_kb(
                sheet_name=sheet_name,
                unresolved_labels=unresolved,
                pool_labels=[k for k in pool_keys if k not in used],
                pool=pool,
                years=years,
                sector=sector,
                client=client,
                model=model or "deepseek-chat",
            )
            for tlbl, plbl in lbl_map.items():
                if plbl and plbl in pool and plbl not in used:
                    matches[tlbl] = plbl
                    used.add(plbl)
        except Exception as exc:
            logger.warning(f"LLM mapping pass failed for sheet '{sheet_name}': {exc}")

    return matches


# ── Derivation: fill computed rows using KB formulas + identities ────────────

def _try_derive(
    template_label: str,
    filled_so_far: dict[str, dict[str, float]],
    years: list[str],
    sector: str,
) -> dict[str, float]:
    """
    Use KB formulas + general identities to compute (label, year) values from
    what's already filled. Returns {year: value} for years we could derive.
    """
    out: dict[str, float] = {}
    entry = kb_entry_for(template_label, sector)
    if not entry:
        return out
    formula = entry.get("formula")
    if not formula:
        return out

    # Helper: lookup any of a list of possible labels in filled
    def _lookup(label_candidates: list[str], yr: str) -> Optional[float]:
        for lbl in label_candidates:
            if lbl in filled_so_far and filled_so_far[lbl].get(yr) is not None:
                try:
                    return float(filled_so_far[lbl][yr])
                except (TypeError, ValueError):
                    pass
            # Try canonical resolution
            kb = load_merged_kb(sector)
            target_entry = kb_entry_for(lbl, sector)
            if not target_entry:
                continue
            for cand_lbl, yr_vals in filled_so_far.items():
                if kb_entry_for(cand_lbl, sector) is target_entry:
                    v = yr_vals.get(yr)
                    if v is not None:
                        try:
                            return float(v)
                        except (TypeError, ValueError):
                            pass
        return None

    # Formula can be:
    #   • String like "interest_earned - interest_expended" / "Revenue + Other Income"
    #   • List ["IS_REV_OPS", "IS_OTHER_INC"] — interpret as positive sum
    if isinstance(formula, list):
        for yr in years:
            total = 0.0
            ok = True
            for term in formula:
                v = _lookup([term, term.replace("_", " ")], yr)
                if v is None:
                    ok = False
                    break
                total += v
            if ok:
                out[yr] = round(total, 4)
        return out

    if isinstance(formula, str):
        # Tokenise into (sign, label) pairs
        f = formula.replace("×", "*").replace("÷", "/")
        tokens = re.split(r"\s*([+\-*/()])\s*", f)
        # Substitute label-like tokens with looked-up numbers
        for yr in years:
            substituted: list[str] = []
            ok = True
            for t in tokens:
                t = t.strip()
                if not t:
                    continue
                if t in {"+", "-", "*", "/", "(", ")"}:
                    substituted.append(t)
                    continue
                # Numeric literal
                try:
                    float(t)
                    substituted.append(t)
                    continue
                except ValueError:
                    pass
                # Treat as label
                v = _lookup([t, t.replace("_", " ")], yr)
                if v is None:
                    ok = False
                    break
                substituted.append(f"({v})")
            if not ok:
                continue
            try:
                expr = " ".join(substituted)
                # Safe eval — only digits, decimals, and operators present
                if re.fullmatch(r"[\d\s\.\(\)\+\-\*\/]*", expr):
                    val = eval(expr)   # noqa: S307 - sandboxed by regex
                    out[yr] = round(float(val), 4)
            except Exception:
                continue
    return out


def _derive_pass(
    template_labels: list[str],
    filled_so_far: dict[str, dict[str, float]],
    years: list[str],
    sector: str,
    max_iter: int = 3,
) -> int:
    """
    Iteratively derive values until no more progress. Useful when one
    derivation enables another (margin → ratio → …).
    Returns count of cells filled this way.
    """
    total_added = 0
    for it in range(max_iter):
        added_this_iter = 0
        for tlbl in template_labels:
            current = filled_so_far.get(tlbl, {})
            missing_years = [y for y in years if current.get(y) is None]
            if not missing_years:
                continue
            derived = _try_derive(tlbl, filled_so_far, missing_years, sector)
            for yr, v in derived.items():
                filled_so_far.setdefault(tlbl, {})[yr] = v
                added_this_iter += 1
        total_added += added_this_iter
        if added_this_iter == 0:
            break
    return total_added


# ── Quarterly de-duplication helper ──────────────────────────────────────────

def _filter_year_columns(years: list[str], year_cols: dict[str, int]) -> list[str]:
    """Restrict the year list to those present in this sheet's columns."""
    return [y for y in years if y in year_cols]


# ── Main orchestrator ───────────────────────────────────────────────────────

def _validate_with_retry(
    sheet_name: str,
    sheet_filled: dict[str, dict[str, Optional[float]]],
    template_labels: list[str],
    years: list[str],
    pool: dict[str, dict[str, float]],
    sector: str,
    matches: dict[str, str],
    max_retries: int = 3,
) -> tuple[set[tuple[str, str]], int]:
    """
    Run sector-aware accounting validation. For each violation, try alternative
    pool candidates for the suspect cell to see if any makes the identity pass.

    Returns (flagged_cells, fixes_applied)
       flagged_cells: set of (label, year) tuples that couldn't be reconciled
       fixes_applied: count of cells whose value was changed to fix validation
    """
    flagged: set[tuple[str, str]] = set()
    fixes_applied = 0

    for retry in range(max_retries):
        report = validate_sheet(sheet_name, sheet_filled, years, sector=sector)
        if not report.issues:
            break

        progress_this_round = 0
        for issue in report.issues:
            if issue.severity not in ("high", "medium"):
                continue
            suspect_lbl = issue.label
            suspect_yr  = issue.year
            if (suspect_lbl, suspect_yr) in flagged:
                continue
            # Find the actual template label that maps to suspect (canon-aware)
            target_lbl = None
            for tlbl in sheet_filled.keys():
                if (_normalise(tlbl) == _normalise(suspect_lbl)
                        or _kb_canonical_for_pool_label(tlbl, sector)
                        == _kb_canonical_for_pool_label(suspect_lbl, sector)):
                    target_lbl = tlbl
                    break
            if not target_lbl:
                flagged.add((suspect_lbl, suspect_yr))
                continue
            # Get the alternative candidates we haven't tried yet
            already_tried = {matches.get(target_lbl)} if matches.get(target_lbl) else set()
            already_tried |= {pk for pk in pool.keys()
                              if pk == matches.get(target_lbl)}
            candidates = _all_candidates(target_lbl, pool, sector,
                                          used=already_tried)
            if not candidates:
                flagged.add((target_lbl, suspect_yr))
                continue
            # Try each alternative — pick the one whose value would BEST fix
            # the violation. Use issue.suggested_fix if available; else just
            # pick the first alternative whose magnitude is in plausible range.
            chosen_alt = None
            for cand in candidates[:5]:
                cand_val = pool.get(cand, {}).get(suspect_yr)
                if cand_val is None:
                    continue
                # Sanity: not the same value (no point retrying same)
                cur_val = sheet_filled.get(target_lbl, {}).get(suspect_yr)
                if cur_val is not None and abs(cand_val - cur_val) < 0.5:
                    continue
                # Try this candidate
                old_val = sheet_filled.get(target_lbl, {}).get(suspect_yr)
                sheet_filled.setdefault(target_lbl, {})[suspect_yr] = float(cand_val)
                # Quick check: does this specific issue go away?
                new_report = validate_sheet(
                    sheet_name, sheet_filled, [suspect_yr], sector=sector,
                )
                still_failing = any(
                    i.rule == issue.rule and i.label == suspect_lbl
                    and i.year == suspect_yr
                    for i in new_report.issues
                )
                if not still_failing:
                    chosen_alt = cand
                    matches[target_lbl] = cand
                    progress_this_round += 1
                    fixes_applied += 1
                    logger.info(
                        f"validation retry: '{target_lbl}' {suspect_yr} "
                        f"{old_val} → {cand_val} (issue '{issue.rule}' resolved)"
                    )
                    break
                else:
                    # Restore old value
                    if old_val is None:
                        sheet_filled.setdefault(target_lbl, {})[suspect_yr] = None
                    else:
                        sheet_filled[target_lbl][suspect_yr] = old_val
            if chosen_alt is None:
                flagged.add((target_lbl, suspect_yr))

        if progress_this_round == 0:
            break

    return flagged, fixes_applied


def run_pipeline(
    company: CompanyIdentifier,
    template_path: str,
    output_path: str,
    years: list[str],
    statement_type: Optional[str] = None,    # "consolidated" | "standalone" | "both" | None=auto
    explicit_sector: Optional[str] = None,
    client: Any = None,
    model: str = "deepseek-chat",
    progress_cb: Optional[Callable[[str, float], None]] = None,
    overwrite_existing: bool = False,
    quarters: Optional[list[str]] = None,    # e.g. ["1QF2024",…]; if None, derive from template
    use_agent_gapfill: bool = False,         # run tool-calling agent for empty cells
    agent_min_empty_pct: float = 0.20,       # only invoke agent if >=20% empty
    pdf_text: str = "",                       # explicit PDF text for agent search_pdf_pages
    use_pdf_harvest: bool = False,           # also fetch + parse company's annual report PDFs
    user_pdf_paths: Optional[list[str]] = None,  # local PDF paths the user uploaded
    raw_excel_path: Optional[str] = None,    # *_raw_data.xlsx to feed as a high-priority source
    bypass_cache: bool = False,              # force re-fetch from sources (ignore disk cache)
    clear_cache: bool = False,               # wipe entire cache before starting
    use_yfinance: bool = False,              # yfinance reformulates with weird signs; off by default
    save_pdfs_to: Optional[str] = None,      # if set, downloaded PDFs go here with readable names
    pdf_discover_only: bool = False,         # download PDFs but skip extraction
) -> OrchestratorResult:
    """
    End-to-end pipeline. See module docstring for the sequence.
    """
    t0 = time.time()
    progress_cb = progress_cb or (lambda *_: None)

    # ── 0. Cache controls (set BEFORE any source pull) ──────────────────────
    global _CACHE_BYPASS
    _CACHE_BYPASS = bool(bypass_cache)
    if clear_cache:
        n_cleared = cache_mod.invalidate()
        logger.info(f"orchestrator: cleared {n_cleared} cache entries before run")

    # ── 1. Read template ─────────────────────────────────────────────────────
    progress_cb("Reading template…", 0.02)
    from mapper.template_reader import read_template
    template_model = read_template(template_path)

    result = OrchestratorResult(
        company=company,
        sector="general",
        sector_conf=0.5,
        template_path=template_path,
        output_path=output_path,
        years=list(years),
    )

    # ── 2. Detect sector ─────────────────────────────────────────────────────
    progress_cb("Detecting sector…", 0.05)
    det = detect_sector(
        nse_symbol=company.nse_symbol,
        name=company.name,
        explicit_sector=explicit_sector,
    )
    result.sector = det.sector
    result.sector_conf = det.confidence
    logger.info(f"orchestrator: sector → {det}")

    # ── 3. Statement plan per sheet ──────────────────────────────────────────
    plans = resolve_for_template(template_model, user_override=statement_type)
    needed_stmt_types = sorted({s for plan in plans.values() for s in plan.fetch})

    # ── 4. Pull sources (cached) ─────────────────────────────────────────────
    progress_cb("Fetching XBRL…", 0.10)
    source_results: list[SourceFetchResult] = []
    for stmt in needed_stmt_types:
        fr = _try_xbrl(company, years, stmt)
        source_results.append(fr)
        result.source_status.append({
            "source": f"xbrl[{stmt}]", "items": len(fr.items),
            "cached": fr.cached, "error": fr.error,
        })
        logger.info(f"orchestrator: xbrl[{stmt}] → {len(fr.items)} items "
                    f"({'cached' if fr.cached else 'live'})")

    progress_cb("Fetching Screener…", 0.20)
    fr = _try_screener(company, years); source_results.append(fr)
    result.source_status.append({"source": "screener", "items": len(fr.items),
                                 "cached": fr.cached, "error": fr.error})

    progress_cb("Fetching Tickertape…", 0.30)
    fr = _try_tickertape(company, years); source_results.append(fr)
    result.source_status.append({"source": "tickertape", "items": len(fr.items),
                                 "cached": fr.cached, "error": fr.error})

    if use_yfinance:
        progress_cb("Fetching Yahoo Finance…", 0.40)
        fr = _try_yfinance(company, years); source_results.append(fr)
        result.source_status.append({"source": "yfinance", "items": len(fr.items),
                                     "cached": fr.cached, "error": fr.error})
    else:
        result.source_status.append({
            "source": "yfinance", "items": 0, "cached": False,
            "error": "DISABLED by default — yfinance reformulates with non-Indian-AR "
                     "sign convention; enable only if screener+PDF coverage is "
                     "insufficient",
        })

    # ── 4a. Raw Excel source (low priority — user-supplied hint, not truth) ─
    if raw_excel_path:
        try:
            from pipeline.raw_excel_source import load_raw_excel
            raw_items = load_raw_excel(raw_excel_path, trust="fallback")
            if raw_items:
                source_results.append(SourceFetchResult("raw_excel", raw_items))
                result.source_status.append({
                    "source": "raw_excel", "items": len(raw_items),
                    "cached": False, "error": None,
                })
                logger.info(f"orchestrator: raw_excel → {len(raw_items)} items "
                            f"(priority demoted below pdf — treated as hint)")
        except Exception as exc:
            logger.warning(f"orchestrator: raw_excel load failed: {exc}")
            result.warnings.append(f"raw_excel load failed: {exc}")

    # ── 4b. PDF harvest (annual reports → tables + agent text) ───────────────
    harvested_pdf_text = pdf_text or ""
    # User-supplied PDFs ALWAYS get processed if provided, even when
    # use_pdf_harvest=False (which only disables auto-discovery).
    should_run_pdf = use_pdf_harvest or bool(user_pdf_paths)
    if should_run_pdf:
        progress_cb("Harvesting annual-report PDFs…", 0.45)
        try:
            from pipeline.pdf_harvest import harvest_pdfs
            pdf_stmt = needed_stmt_types[0] if needed_stmt_types else "consolidated"
            # Default save_pdfs_to: output/downloaded_pdfs/<SYMBOL>/
            _save_target = save_pdfs_to
            if _save_target is None:
                _proj_root = str(Path(__file__).parent.parent)
                _sym = (company.nse_symbol or company.name or "company").upper()
                _save_target = str(Path(_proj_root) / "output"
                                    / "downloaded_pdfs" / _sym)
            pdf_res = harvest_pdfs(
                company=company,
                years=years,
                statement_type=pdf_stmt,
                user_pdf_paths=user_pdf_paths,
                save_dir=_save_target,
                discover_only=pdf_discover_only,
            )
            if pdf_res.pool_items:
                source_results.append(SourceFetchResult("pdf", pdf_res.pool_items))
            if pdf_res.pdf_text:
                # Concatenate any caller-supplied text with harvested
                harvested_pdf_text = (harvested_pdf_text + "\n" + pdf_res.pdf_text).strip()
            for w in pdf_res.errors:
                result.warnings.append(f"PDF harvest: {w}")
            # Surface PDF harvest in source_status panel — per-layer detail.
            # Include saved file paths so the UI can show clickable links.
            saved_files = [u for u in pdf_res.pdfs_used if u.get("saved_to")]
            result.source_status.append({
                "source": f"pdf_harvest ({len(pdf_res.pdfs_used)} files)",
                "items":  len(pdf_res.pool_items),
                "cached": False,
                "error":  None if pdf_res.pdfs_used else (
                    "no PDFs discovered — see per-layer detail below"
                ),
                "saved_files": [
                    {
                        "name":   Path(u["saved_to"]).name,
                        "path":   u["saved_to"],
                        "year":   u.get("year") or "?",
                        "source": u.get("source") or "?",
                        "size_mb": u.get("size_mb") or 0,
                        "cells":  u.get("structured_cells", 0),
                        "text_chars": u.get("text_chars", 0),
                        "url":    u.get("url", ""),
                    }
                    for u in saved_files
                ],
            })
            # One row per discovery layer (NSE / BSE / Screener / IR site / search)
            for layer in pdf_res.discovery_layers:
                result.source_status.append({
                    "source": f"  ↳ {layer.get('layer')}",
                    "items":  layer.get("urls_found", 0),
                    "cached": False,
                    "error":  layer.get("error") or "",
                })
            logger.info(f"orchestrator: pdf_harvest → {pdf_res.summary()}")
        except Exception as exc:
            logger.warning(f"orchestrator: pdf_harvest failed: {exc}")
            result.warnings.append(f"PDF harvest skipped: {exc}")
            result.source_status.append({
                "source": "pdf_harvest", "items": 0, "cached": False,
                "error": str(exc),
            })
    else:
        result.source_status.append({
            "source": "pdf_harvest", "items": 0, "cached": False,
            "error": "DISABLED — tick '📄 Harvest annual-report PDFs' OR upload "
                     "PDFs in '📤 Upload AR PDFs (auto-fetch fallback)'",
        })

    # ── 5. Build pool with multi-source consensus ───────────────────────────
    progress_cb("Building data pool…", 0.50)
    pool, provenance, consensus_meta = _build_pool(source_results, result.sector)
    # Note: _canonical_dedup_pool no longer needed — consensus does this in one pass
    result.pool_size = len(pool)
    # Stash consensus_meta on the result for the UI / downstream
    result.consensus_meta = consensus_meta
    logger.info(f"orchestrator: pool built — {len(pool)} unique labels across "
                f"{sum(len(v) for v in pool.values())} (label,year) cells")

    # ── 5b. Deterministic aggregation (audit 2026-05-11) ────────────────────
    # Run the deterministic aggregation layer over the pool BEFORE any LLM
    # stage. This computes GP, EBITDA, EBIT, PBT, PAT, Total Assets, Total
    # Equity & Liabilities, margins, etc. from leaf cells that are already
    # in the pool. Every derived value is tagged source="derived_agg"
    # (priority 93 — above pdf, below raw_excel) and CANNOT overwrite an
    # existing pool entry; if computed disagrees with pre-existing it's
    # recorded as a discrepancy for human review.
    try:
        from pipeline.aggregation_layer import canonicalize, run_aggregation

        # Build a (canon, period) → value view of the pool. First non-empty
        # canonical label per year wins.
        canon_pool: dict[tuple[str, str], float] = {}
        canon_to_label: dict[str, str] = {}  # remember a representative label
        all_periods: set[str] = set()
        for label, year_map in pool.items():
            c = canonicalize(label)
            if not c:
                continue
            canon_to_label.setdefault(c, label)
            for yr, v in year_map.items():
                all_periods.add(yr)
                if (c, yr) not in canon_pool:
                    canon_pool[(c, yr)] = float(v)

        report = run_aggregation(canon_pool, sorted(all_periods))

        # Inject derived values back into the pool. Use the representative
        # label if one exists; otherwise the canonical concept name.
        n_injected = 0
        for canon, period, val in report.derivations:
            lbl = canon_to_label.get(canon, canon.replace("_", " ").title())
            if lbl not in pool:
                pool[lbl] = {}
            if period not in pool[lbl]:
                pool[lbl][period] = float(val)
                n_injected += 1
                provenance.setdefault((lbl.lower(), period), []).append(
                    CellProvenance(
                        sheet="", label=lbl, year=period, value=float(val),
                        source="derived_agg",
                        confidence="HIGH" if not report.discrepancies else "MEDIUM",
                        extracted_as=lbl,
                        notes=f"aggregation_layer rule={canon}",
                    )
                )

        logger.info(
            f"orchestrator: aggregation_layer → {len(report.derivations)} "
            f"derivations, {n_injected} injected into pool, "
            f"{len(report.discrepancies)} discrepancies"
        )
        if report.discrepancies:
            for canon, period, existing, computed in report.discrepancies[:5]:
                logger.warning(
                    f"  agg-discrepancy: {canon}@{period}: "
                    f"existing={existing:.2f} vs computed={computed:.2f}"
                )
            result.warnings.append(
                f"Aggregation found {len(report.discrepancies)} cells where "
                f"computed values disagree with extracted values — check logs."
            )
    except Exception as exc:
        logger.warning(f"orchestrator: aggregation_layer failed: {exc}")
        result.warnings.append(f"aggregation_layer failed: {exc}")

    if not pool:
        result.warnings.append("No data could be sourced — all connectors returned empty")

    # ── 6. Per-sheet processing ──────────────────────────────────────────────
    progress_cb("Mapping & deriving values…", 0.60)
    # filled[stmt_type][sheet][label][year] = value
    filled: dict[str, dict[str, dict[str, dict[str, float]]]] = {}

    for sheet_name, plan in plans.items():
        sheet = template_model.sheets.get(sheet_name)
        if not sheet:
            continue
        if not sheet.row_index:
            logger.info(f"orchestrator: sheet '{sheet_name}' has no labelled rows — skip")
            continue

        # Restrict to year columns this sheet actually has
        sheet_years = _filter_year_columns(years, sheet.year_cols)
        # Add quarterly columns if any
        sheet_q_cols = [yr for yr in sheet.year_cols if "QF" in yr]
        sheet_years_all = list(sheet_years) + sheet_q_cols
        if not sheet_years_all:
            continue

        result.sheets_processed.append(sheet_name)
        template_labels = list(sheet.row_index.keys())

        for stmt in plan.write:
            # Map template → pool
            matches = _map_template_to_pool(
                template_labels=template_labels,
                pool=pool,
                sector=result.sector,
                sheet_name=f"{sheet_name} ({stmt})",
                years=sheet_years_all,
                client=client,
                model=model,
            )

            # Build the sheet's filled dict from matches
            sheet_filled: dict[str, dict[str, float]] = {}
            for tlbl, plbl in matches.items():
                for yr, v in pool.get(plbl, {}).items():
                    if yr not in sheet_years_all:
                        continue
                    sheet_filled.setdefault(tlbl, {})[yr] = v
                    # Provenance for the chosen value
                    src_list = provenance.get((plbl.lower(), yr), [])
                    src = src_list[0].source if src_list else "unknown"
                    conf = src_list[0].confidence if src_list else "MEDIUM"
                    result.provenance.append(CellProvenance(
                        sheet=sheet_name, label=tlbl, year=yr, value=float(v),
                        source=src, confidence=conf, extracted_as=plbl,
                        notes=f"stmt={stmt}",
                    ))

            # Derivation pass
            n_derived = _derive_pass(
                template_labels, sheet_filled, sheet_years_all, result.sector,
            )
            if n_derived:
                logger.info(f"orchestrator: derived {n_derived} values for "
                            f"'{sheet_name}' [{stmt}]")
                # Record provenance for derived cells
                for tlbl, yrs in sheet_filled.items():
                    for yr, v in yrs.items():
                        if not any(p.sheet == sheet_name and p.label == tlbl and p.year == yr
                                   for p in result.provenance):
                            result.provenance.append(CellProvenance(
                                sheet=sheet_name, label=tlbl, year=yr, value=float(v),
                                source="derived", confidence="MEDIUM",
                                notes=f"derived; stmt={stmt}",
                            ))

            # ── Agent gap-fill (optional) ────────────────────────────────────
            # Only runs if (a) caller asked for it, (b) we have an LLM client,
            # (c) the sheet still has a meaningful fraction of empty cells,
            # and (d) the agent recognises the sheet type (P&L / BS / CF).
            # GUARD: if pool is empty AND no PDF text, the agent has nothing
            # to search. Don't burn LLM calls — tell the user instead.
            agent_can_help = (len(pool) > 0 or len(harvested_pdf_text) > 100)
            if use_agent_gapfill and client is not None and not agent_can_help:
                if not result.warnings or "agent skipped" not in " ".join(result.warnings):
                    result.warnings.append(
                        "Agent skipped — pool is empty AND no PDF text "
                        "available. The agent has nothing to search. "
                        "Upload an AR PDF (📤 in the Smart Pipeline expander) "
                        "or wait 30+ min for source rate-limits to clear."
                    )
            if use_agent_gapfill and client is not None and agent_can_help:
                # Initialise full-coverage view: every template label × every year.
                # Agent sees what's empty so it can decide what to chase.
                full_view: dict[str, dict[str, float | None]] = {}
                for tlbl in template_labels:
                    full_view[tlbl] = {yr: sheet_filled.get(tlbl, {}).get(yr)
                                       for yr in sheet_years_all}
                total_cells = sum(len(yv) for yv in full_view.values())
                empty_cells = sum(1 for yv in full_view.values()
                                  for v in yv.values() if v is None)
                empty_pct = empty_cells / total_cells if total_cells else 0.0

                if empty_pct >= agent_min_empty_pct:
                    try:
                        from pipeline.financial_agent import (
                            run_financial_agent, MAX_LABELS as _AGENT_MAX,
                        )
                        new_fills = 0
                        agent_log: list[dict] = []
                        # Loop in batches: agent only sees MAX_LABELS at a time;
                        # we keep going until either no progress or all empty
                        # labels exhausted.
                        for batch_iter in range(8):    # hard cap on batches
                            empty_now = [
                                lbl for lbl in template_labels
                                if any(full_view.get(lbl, {}).get(yr) is None
                                       for yr in sheet_years_all)
                            ]
                            if not empty_now:
                                break
                            before_filled = {
                                (lbl, yr) for lbl, yv in full_view.items()
                                for yr, v in yv.items() if v is not None
                            }
                            logger.info(
                                f"orchestrator: invoking agent on "
                                f"'{sheet_name}' [{stmt}] batch {batch_iter+1} — "
                                f"{len(empty_now)} labels with gaps"
                            )
                            agent_view, batch_log = run_financial_agent(
                                sheet_name=f"{sheet_name} ({stmt})",
                                sheet_filled=full_view,
                                template_labels=empty_now,    # only the gaps
                                years=sheet_years_all,
                                sector=result.sector,
                                pool=pool,
                                pdf_text=harvested_pdf_text,
                                nse_symbol=company.nse_symbol or "",
                                client=client,
                                model=model,
                                dropped_collector=result.agent_dropped,
                                row_context=getattr(sheet, "row_context", {}),
                            )
                            agent_log.extend(batch_log)
                            # Count fills landed in THIS batch
                            batch_fills = 0
                            for tlbl, yv in agent_view.items():
                                for yr, v in yv.items():
                                    if v is None:
                                        continue
                                    if (tlbl, yr) in before_filled:
                                        continue
                                    full_view.setdefault(tlbl, {})[yr] = float(v)
                                    sheet_filled.setdefault(tlbl, {})[yr] = float(v)
                                    batch_fills += 1
                                    result.provenance.append(CellProvenance(
                                        sheet=sheet_name, label=tlbl, year=yr,
                                        value=float(v),
                                        source="agent", confidence="MEDIUM",
                                        notes=f"agent gap-fill; stmt={stmt}",
                                    ))
                            new_fills += batch_fills
                            # Stop if no progress OR we've exhausted empty labels
                            if batch_fills == 0:
                                break
                            if len(empty_now) <= _AGENT_MAX:
                                break
                        # Always record activity, even if 0 fills landed
                        result.agent_activity.append({
                            "sheet": sheet_name, "stmt": stmt,
                            "empty_before": empty_cells,
                            "tool_calls": len(agent_log),
                            "fills_landed": new_fills,
                            "tools_used": [a.get("tool") for a in agent_log],
                        })
                        if new_fills:
                            logger.info(
                                f"orchestrator: agent added {new_fills} "
                                f"values to '{sheet_name}' [{stmt}] "
                                f"(via {len(agent_log)} tool calls)"
                            )
                            # Re-derive in case agent fills enable more derivations
                            _derive_pass(template_labels, sheet_filled,
                                         sheet_years_all, result.sector)
                        else:
                            logger.warning(
                                f"orchestrator: agent ran on '{sheet_name}' "
                                f"[{stmt}] but landed 0 fills "
                                f"(check pdf_text / pool / api key)"
                            )
                    except Exception as agent_err:
                        logger.warning(
                            f"orchestrator: agent gap-fill failed for "
                            f"'{sheet_name}' [{stmt}]: {agent_err}"
                        )
                        result.warnings.append(
                            f"Agent gap-fill failed on {sheet_name} ({stmt}): "
                            f"{agent_err}"
                        )

            # ── Validation with retry: if accounting identities fail, try
            #    alternative pool candidates for the suspect cell ───────────
            v_flagged, v_fixes = _validate_with_retry(
                sheet_name=f"{sheet_name} ({stmt})",
                sheet_filled=sheet_filled,
                template_labels=template_labels,
                years=sheet_years_all,
                pool=pool,
                sector=result.sector,
                matches=matches,
            )
            if v_fixes:
                logger.info(f"orchestrator: validation-retry fixed {v_fixes} "
                            f"cells on '{sheet_name}' [{stmt}]")
            if v_flagged:
                # Track flagged for the writer to color RED
                if not hasattr(result, "validation_flagged"):
                    result.validation_flagged = set()
                for lbl, yr in v_flagged:
                    result.validation_flagged.add((sheet_name, lbl, yr))

            # Validation report (final state — may still have issues if no
            # candidate could resolve them)
            vrep = validate_sheet(
                sheet_name=f"{sheet_name} ({stmt})",
                filled=sheet_filled,
                years=sheet_years_all,
                sector=result.sector,
                symbol=company.nse_symbol or "",
            )
            result.validation_reports[f"{sheet_name}::{stmt}"] = vrep
            n_fixes = apply_auto_fixes(sheet_filled, vrep, require_empty=True)
            if n_fixes:
                result.auto_fixes_applied += n_fixes
                logger.info(f"orchestrator: applied {n_fixes} auto-fixes on "
                            f"'{sheet_name}' [{stmt}]")
                for fix in vrep.auto_fixes:
                    result.provenance.append(CellProvenance(
                        sheet=sheet_name, label=fix.label, year=fix.year,
                        value=fix.new_value, source="auto_fix",
                        confidence="MEDIUM",
                        notes=fix.reason,
                    ))

            # Re-derive after auto-fix in case fixes unblock more derivations
            _derive_pass(template_labels, sheet_filled, sheet_years_all, result.sector)

            filled.setdefault(stmt, {})[sheet_name] = sheet_filled

    # ── 7. Write to template ─────────────────────────────────────────────────
    progress_cb("Writing template…", 0.85)
    cells_written = _write_filled(
        template_path=template_path,
        output_path=output_path,
        filled=filled,
        template_model=template_model,
        overwrite_existing=overwrite_existing,
        result=result,
    )
    result.cells_written = cells_written

    # ── 8. Persist HIGH-confidence values to historical DB ──────────────────
    # Fuels future YoY validation. Only HIGH-confidence (≥2 sources agree)
    # values are saved to prevent bad extractions from poisoning the DB.
    try:
        from pipeline.historical_db import save_run as _save_run
        # Build values_by_canonical from all written cells
        values_by_canonical: dict[str, dict[str, float]] = {}
        for stmt_t, sheets_d in filled.items():
            for sn, lbl_map in sheets_d.items():
                for tlbl, yrs in lbl_map.items():
                    canon = _kb_canonical_for_pool_label(tlbl, result.sector)
                    if not canon:
                        continue
                    for yr, v in yrs.items():
                        if v is None:
                            continue
                        try:
                            fv = float(v)
                        except (TypeError, ValueError):
                            continue
                        values_by_canonical.setdefault(canon, {}).setdefault(yr, fv)
        n_persisted = _save_run(
            company.nse_symbol or company.name,
            values_by_canonical,
            sector=result.sector,
            name=company.name,
            only_high_confidence=True,
            confidence_meta=result.consensus_meta,
        )
        if n_persisted:
            logger.info(f"orchestrator: persisted {n_persisted} HIGH-confidence "
                        f"values to historical DB")
    except Exception as exc:
        logger.warning(f"orchestrator: historical DB save failed: {exc}")

    result.elapsed_sec = time.time() - t0
    progress_cb("Done", 1.0)

    return result


def _write_filled(
    template_path: str,
    output_path: str,
    filled: dict[str, dict[str, dict[str, dict[str, float]]]],
    template_model,
    overwrite_existing: bool,
    result: OrchestratorResult,
) -> int:
    """
    Write the per-stmt-type filled values into the workbook using the existing
    cell-address resolution from template_model. Non-destructive: never
    overwrites formulas or pre-filled cells (unless overwrite_existing=True).
    Color-codes each written cell by confidence (green/yellow/red).
    """
    import openpyxl
    from openpyxl.styles import PatternFill
    wb = openpyxl.load_workbook(template_path, data_only=False)

    # Confidence color palette — Excel-friendly pastels
    GREEN  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # HIGH
    YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # MEDIUM
    RED    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # FLAGGED / LOW
    GREY   = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")  # derived

    # Build provenance lookup: (sheet, label_lower, year) → primary source
    prov_by_cell: dict[tuple[str, str, str], str] = {}
    for p in result.provenance:
        prov_by_cell.setdefault(
            (p.sheet, (p.label or "").lower(), p.year), p.source,
        )

    def _confidence_for(sheet_name: str, tlbl: str, year: str,
                         val: float) -> tuple[str, "PatternFill"]:
        """Resolve confidence and color for a written cell."""
        # 0. HIGHEST PRIORITY: validation gate flagged it as unfixable
        if (sheet_name, tlbl, year) in result.validation_flagged:
            return ("FLAGGED", RED)
        # 1. Provenance source
        src = prov_by_cell.get((sheet_name, tlbl.lower(), year), "")
        # 2. Consensus by canonical
        canon = _kb_canonical_for_pool_label(tlbl, result.sector)
        cm = result.consensus_meta.get((canon, year)) if canon else None
        if src == "agent":
            return ("MEDIUM-agent", YELLOW)
        if src == "auto_fix":
            return ("MEDIUM-derived", GREY)
        if src == "derived":
            return ("MEDIUM-derived", GREY)
        if cm:
            cf = cm.get("confidence")
            if cf == "HIGH":
                return ("HIGH", GREEN)
            if cf == "FLAGGED":
                return ("FLAGGED", RED)
            return ("MEDIUM", YELLOW)
        return ("MEDIUM", YELLOW)

    cells_written = 0
    cells_skipped_formula = 0
    cells_skipped_filled = 0
    confidence_counts = {"HIGH": 0, "MEDIUM": 0, "FLAGGED": 0,
                         "MEDIUM-agent": 0, "MEDIUM-derived": 0}

    for stmt, sheets in filled.items():
        for sheet_name, label_map in sheets.items():
            sheet = template_model.sheets.get(sheet_name)
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else None
            if not sheet or ws is None:
                continue
            for tlbl, yrs in label_map.items():
                for yr, val in yrs.items():
                    if val is None:
                        continue
                    addr = template_model.get_cell_address(
                        sheet_name, tlbl, yr, statement_type=stmt,
                    )
                    if not addr:
                        continue
                    key = (addr.row, addr.col)
                    if key in sheet.formula_cells:
                        cells_skipped_formula += 1
                        continue
                    if key in sheet.filled_cells and not overwrite_existing:
                        cells_skipped_filled += 1
                        continue
                    try:
                        cell = ws.cell(addr.row, addr.col)
                        cell.value = float(val)
                        # Apply confidence color
                        conf_tier, fill = _confidence_for(sheet_name, tlbl, yr, val)
                        cell.fill = fill
                        confidence_counts[conf_tier] = (
                            confidence_counts.get(conf_tier, 0) + 1
                        )
                        cells_written += 1
                    except Exception as exc:
                        logger.warning(f"write failed {sheet_name} {tlbl} {yr}: {exc}")
    result.confidence_counts = confidence_counts

    result.cells_skipped_formula = cells_skipped_formula
    result.cells_skipped_filled = cells_skipped_filled

    # Best-effort save
    try:
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
    except Exception as exc:
        result.warnings.append(f"Failed to save {output_path}: {exc}")

    return cells_written


__all__ = [
    "run_pipeline",
    "OrchestratorResult",
    "CellProvenance",
]
