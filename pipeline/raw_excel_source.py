"""
pipeline/raw_excel_source.py
────────────────────────────
Use a previously-produced *_raw_data.xlsx as a HIGH-priority source for the
orchestrator's pool. This lets the user bypass live source pulls when they
already have clean raw data from a previous run.

Format expected
───────────────
A workbook with one or more sheets, each containing:
  - Row 1: title (ignored)
  - Row 2: unit / preamble (ignored)
  - Row N: header with columns ["Particulars", "F2026", "F2025", "F2024", …]
           (also accepted: "Label", "Item")
  - Subsequent rows: label in column 1, numeric values in year columns.

Garbage labels (PDF text-extraction noise) are filtered.
Per-sheet section context is preserved for richer downstream mapping.

Returns: list[WebDataResult] tagged with source="raw_excel" so it sits at
the top of the priority stack in _build_pool.
"""

from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Optional

import openpyxl

from web.base import WebDataResult

logger = logging.getLogger(__name__)

_YEAR_PAT = re.compile(r"^[Ff]\d{4}$|^\d{4}$|^FY\d{2,4}$|^[1-4]Q[Ff]?\d{2,4}$",
                       re.IGNORECASE)


def _normalise_year(s: str) -> Optional[str]:
    s = str(s).strip().upper()
    if not s:
        return None
    if re.fullmatch(r"F\d{4}", s):
        return s
    if re.fullmatch(r"\d{4}", s) and 2000 <= int(s) <= 2050:
        return f"F{s}"
    m = re.fullmatch(r"FY(\d{2,4})", s)
    if m:
        y = m.group(1)
        return f"F{int(y) + 2000}" if len(y) == 2 else f"F{y}"
    if re.fullmatch(r"[1-4]Q[F]?\d{2,4}", s):
        # Already template-canonical-ish (1QF2025) — return as-is
        return s.replace("FY", "F").replace("Q F", "QF")
    return None


def _find_header_row(ws) -> Optional[int]:
    for r in range(1, min(15, ws.max_row + 1)):
        cnt = 0
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v and _normalise_year(str(v)):
                cnt += 1
        if cnt >= 2:
            return r
    return None


def load_raw_excel(path: str | Path, trust: str = "fallback") -> list[WebDataResult]:
    """
    Parse a *_raw_data.xlsx and return WebDataResult items ready for the pool.

    `trust` controls how aggressively to filter:
      • "trusted"   — accept everything (only use if you know your raw is clean)
      • "fallback"  — apply garbage-label + magnitude filters (default;
                      treats raw Excel as a HINT not ground truth)
      • "strict"    — additional checks: drop anything that looks like a
                      year, share count, or non-financial number
    """
    p = Path(path)
    if not p.exists():
        logger.warning(f"raw_excel: not found: {p}")
        return []

    try:
        wb = openpyxl.load_workbook(str(p), data_only=True)
    except Exception as exc:
        logger.warning(f"raw_excel: failed to open {p}: {exc}")
        return []

    # Lazy import — keep this module decoupled
    from pipeline.orchestrator import _is_garbage_label

    out: list[WebDataResult] = []
    n_garbage_total = 0
    n_suspect_total = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row = _find_header_row(ws)
        if not header_row:
            continue

        # Build column → year map
        year_cols: dict[int, str] = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(header_row, c).value
            if not v:
                continue
            y = _normalise_year(str(v))
            if y:
                year_cols[c] = y
        if not year_cols:
            continue

        # First pass — collect all values per row for magnitude/YoY checks
        rows_data: list[tuple[str, dict[str, float]]] = []
        n_garbage = 0
        for r in range(header_row + 1, ws.max_row + 1):
            label = ws.cell(r, 1).value
            if not label:
                continue
            label = str(label).strip()
            if not label:
                continue
            if trust != "trusted" and _is_garbage_label(label):
                n_garbage += 1
                continue
            row_yr_vals: dict[str, float] = {}
            for c, yr in year_cols.items():
                v = ws.cell(r, c).value
                if v in (None, "", "-", "—"):
                    continue
                try:
                    fv = float(v)
                except (TypeError, ValueError):
                    continue
                if fv == 0:
                    continue
                # Unit sniff
                if abs(fv) > 1e10:
                    fv = fv / 1e7
                row_yr_vals[yr] = fv
            if row_yr_vals:
                rows_data.append((label, row_yr_vals))

        n_garbage_total += n_garbage

        # Second pass — sanity-check each row before emitting
        n_kept = 0
        n_suspect = 0
        for label, yr_vals in rows_data:
            if trust != "trusted":
                # Suspect: looks like a year (4-digit 19xx-20xx) appearing as a value
                yr_like_count = sum(1 for v in yr_vals.values()
                                    if 1900 <= abs(v) <= 2100)
                if yr_like_count > 0 and "year" not in label.lower():
                    n_suspect += 1
                    continue
                # Suspect: looks like a share count (very large round number)
                # if all values > 1e6 and label doesn't say "shares" / "units"
                if (all(abs(v) > 1e6 for v in yr_vals.values())
                        and not any(t in label.lower()
                                    for t in ("share", "unit", "no.", "count"))):
                    n_suspect += 1
                    continue
                # Suspect: YoY swings > 100x — not just scale, fundamentally wrong row
                if trust == "strict":
                    sorted_vals = sorted((y, v) for y, v in yr_vals.items()
                                         if abs(v) > 0.01)
                    swing_violation = False
                    for i in range(1, len(sorted_vals)):
                        prev_v = sorted_vals[i-1][1]
                        curr_v = sorted_vals[i][1]
                        if abs(prev_v) > 1.0:
                            ratio = abs(curr_v) / abs(prev_v)
                            if ratio > 100 or ratio < 0.01:
                                swing_violation = True
                                break
                    if swing_violation:
                        n_suspect += 1
                        continue

            for yr, fv in yr_vals.items():
                out.append(WebDataResult(
                    source="raw_excel", raw_field=label,
                    template_field="", value=fv, year=yr,
                    confidence="MEDIUM",   # not HIGH — raw Excel is suspect
                    unit_applied=f"raw_excel:{sheet_name} (trust={trust})",
                ))
                n_kept += 1

        n_suspect_total += n_suspect
        logger.info(f"raw_excel: '{sheet_name}' → {n_kept} cells "
                    f"({n_garbage} garbage labels, {n_suspect} suspect rows skipped)")

    logger.info(
        f"raw_excel: total {len(out)} pool items from {p.name} "
        f"(filtered {n_garbage_total} garbage labels + "
        f"{n_suspect_total} suspect rows)"
    )
    return out


__all__ = ["load_raw_excel"]
