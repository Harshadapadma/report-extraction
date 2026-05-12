"""
pipeline/raw_extract.py
───────────────────────
Dump full P&L / Balance Sheet / Cash Flow tables VERBATIM from uploaded PDFs.

No template mapping, no LLM, no pattern matching. Just:
  1. For each PDF, classify and build TOC
  2. Locate BS / P&L / CF pages via TOC
  3. Extract the tables (pdfplumber tables + prose-line fallback)
  4. Write each to its own sheet in the output workbook with:
       • Font: Calibri 11
       • Number format: accounting with 2 decimals

Output sheet names: "<short_pdf_id> - <Scope> <Statement>" e.g.
  "AR_22_23 - Standalone BS"
  "AR_22_23 - Standalone P&L"
  "AR_22_23 - Consolidated CF"
  "Q4FY25_C - P&L"
"""
from __future__ import annotations

import logging
import re
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

from pipeline.llm_fill import (
    _pdf_to_markdown, _build_toc, PDFClassification,
)


def _extract_tables_via_fitz(pdf_path: str, page_idxs: list[int]
                              ) -> list[list[list[str]]]:
    """Extract tables from specific pages using PyMuPDF's find_tables().
    Returns list of tables (each as list of rows of strings). Preserves
    label column much better than pdfplumber on magazine-layout PDFs.
    """
    out: list[list[list[str]]] = []
    try:
        import fitz   # PyMuPDF
    except ImportError:
        return out
    try:
        doc = fitz.open(pdf_path)
    except Exception:
        return out
    try:
        for pn in page_idxs:
            if pn < 0 or pn >= len(doc):
                continue
            page = doc[pn]
            try:
                finder = page.find_tables()
                tables = finder.tables if hasattr(finder, "tables") else list(finder)
            except Exception:
                tables = []
            for tbl in tables:
                try:
                    rows = tbl.extract()
                    if not rows or len(rows) < 2:
                        continue
                    # Normalise: ensure strings, strip None
                    norm = []
                    for r in rows:
                        norm.append([
                            ("" if c is None else str(c).strip().replace("\n", " "))
                            for c in r
                        ])
                    # Filter: need ≥3 rows AND meaningful content
                    n_nums = sum(
                        1 for r in norm for c in r
                        if c and re.search(r"\d", c)
                    )
                    if len(norm) >= 3 and n_nums >= 3:
                        out.append(norm)
                except Exception as exc:
                    logger.debug(f"fitz table extract failed page {pn}: {exc}")
    finally:
        doc.close()
    return out

logger = logging.getLogger(__name__)


@dataclass
class RawExtractResult:
    output_path: str = ""
    sheets_written: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    elapsed_sec: float = 0.0


# Categorize TOC titles into statement types
_STMT_PATTERNS = [
    (r"\bbalance sheet\b|financial position", "Balance Sheet"),
    (r"statement of (?:profit|p&l)|p\s*&\s*l|\bp&l\b", "P&L"),
    (r"cash flow", "Cash Flow"),
    (r"statement of changes in equity|equity reconciliation", "Equity"),
    (r"segment (?:reporting|information|disclosure)|"
     r"primary segment|secondary segment|business segment", "Segment Reporting"),
]


def _classify_statement(title: str) -> Optional[str]:
    """Map a TOC title to a statement type. Returns None if not a statement page."""
    tl = (title or "").lower()
    for pat, stmt in _STMT_PATTERNS:
        if re.search(pat, tl):
            return stmt
    return None


def _short_pdf_id(name: str) -> str:
    """Generate a short sheet-name-friendly ID from PDF filename."""
    # Strip extension, remove temp prefixes, shorten
    s = name.replace(".pdf", "")
    s = re.sub(r"^(upload_|tmp[a-z0-9_]+_)", "", s)
    # Try to find year markers
    yr_match = re.findall(r"(\d{4})", s)
    short = re.sub(r"[^A-Za-z0-9]+", "_", s)[:25]
    return short


def _extract_md_tables_from_page(md_page_text: str) -> list[list[list[str]]]:
    """Parse markdown tables from a single page's md text.

    Also recovers labels for rows where the first column is empty by
    looking at the most recent non-empty cell in the same column position
    across earlier rows of the same table (carry-forward). This handles
    tables where every other row leaves the label blank to visually group
    sub-items under a parent."""
    tables: list[list[list[str]]] = []
    lines = md_page_text.split("\n")
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        if (line.startswith("|") and line.endswith("|")
                and i + 1 < len(lines)
                and re.match(r"\|[\s\-:|]+\|", lines[i + 1].strip())):
            tbl: list[list[str]] = []
            header = [c.strip() for c in line.strip("|").split("|")]
            tbl.append(header)
            j = i + 2
            while j < len(lines) and lines[j].strip().startswith("|"):
                row = [c.strip() for c in lines[j].strip().strip("|").split("|")]
                tbl.append(row)
                j += 1
            # Post-process: fill empty label column by looking at the most-
            # recent non-empty value in the same column. This recovers
            # labels that pymupdf4llm leaves blank for continuation rows.
            if len(tbl) >= 3:
                # Decide which column is the "label" column. Usually col 0,
                # but if col 0 is all-numeric and col 1 is mostly text, use col 1.
                def _is_numeric(s):
                    s = (s or "").strip().replace(",", "").replace(" ", "")
                    if not s: return False
                    s = s.strip("()").lstrip("-")
                    try:
                        float(s); return True
                    except ValueError:
                        return False
                col0_text_count = sum(1 for r in tbl[1:]
                                       if r and r[0] and not _is_numeric(r[0]))
                if col0_text_count >= 2:
                    label_col = 0
                else:
                    # Maybe labels are in col 1
                    has_col1 = all(len(r) > 1 for r in tbl[1:])
                    if has_col1:
                        col1_text_count = sum(1 for r in tbl[1:]
                                               if r[1] and not _is_numeric(r[1]))
                        label_col = 1 if col1_text_count > col0_text_count else 0
                    else:
                        label_col = 0
                # Carry-forward labels for empty cells
                last_label = ""
                for r in tbl[1:]:
                    if label_col < len(r):
                        cell = (r[label_col] or "").strip()
                        if cell:
                            last_label = cell
                        elif last_label:
                            r[label_col] = last_label
            if len(tbl) >= 2:
                tables.append(tbl)
            i = j
            continue
        i += 1
    return tables


def _try_parse_num(cell: str):
    """Parse cell text → float if numeric, else return None."""
    if not cell or not cell.strip():
        return None
    s = cell.strip().replace(",", "").replace(" ", "").replace("₹", "")
    is_neg = "(" in s and ")" in s
    s = s.replace("(", "").replace(")", "").lstrip("-").strip()
    if not s:
        return None
    try:
        v = float(s)
        return -v if is_neg else v
    except ValueError:
        return None


def _safe_sheet_name(name: str, existing: set) -> str:
    """Excel sheet names: ≤31 chars, no [ ] : / \\ ? *. Make unique."""
    s = re.sub(r"[\[\]:/\\?*]", "_", name)[:31]
    base = s
    n = 1
    while s in existing:
        suffix = f"_{n}"
        s = (base[:31 - len(suffix)] + suffix)
        n += 1
    return s


def raw_extract_statements(
    pdf_paths: list[str],
    output_path: str,
) -> RawExtractResult:
    """Extract P&L, BS, CF tables verbatim from each PDF and write to formatted Excel."""
    t0 = time.time()
    res = RawExtractResult(output_path=output_path)

    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # delete default sheet
    existing_sheets: set[str] = set()

    CALIBRI11 = Font(name="Calibri", size=11)
    CALIBRI11_BOLD = Font(name="Calibri", size=11, bold=True)
    HEADER_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1",
                               fill_type="solid")
    THIN_BORDER = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    # Accounting number format with 2 decimals (Indian/INR style)
    ACCOUNTING_FMT = '_-* #,##0.00_-;-* #,##0.00_-;_-* "-"??_-;_-@_-'

    # ── FAST INLINE PREP ─────────────────────────────────────────────────
    # _classify_pdf does pdfplumber work we don't need for raw extraction.
    # Just call _pdf_to_markdown (cached) and detect boundaries inline.
    def _prep_pdf(pp: str) -> tuple[Optional[PDFClassification], str, dict]:
        try:
            md = _pdf_to_markdown(pp)
        except Exception as exc:
            logger.warning(f"_pdf_to_markdown failed for {pp}: {exc}")
            return None, "", {}
        name = Path(pp).name
        # Split md into per-page chunks
        md_pages: dict[int, str] = {}
        chunks = re.split(r"(?m)^## Page (\d+)\n", md)
        for k in range(1, len(chunks) - 1, 2):
            try:
                pn = int(chunks[k]) - 1
                md_pages[pn] = chunks[k + 1]
            except ValueError:
                continue
        # Inline boundary detection from md content
        sa_markers = (
            "standalone balance sheet", "standalone statement of profit",
            "standalone cash flow", "audit of the standalone financial",
            "report on the audit of the standalone",
            "notes to the standalone financial",
        )
        co_markers = (
            "consolidated balance sheet", "consolidated statement of profit",
            "consolidated cash flow", "audit of the consolidated financial",
            "report on the audit of the consolidated",
            "notes to the consolidated financial",
        )
        std_s = cons_s = None
        for pn in sorted(md_pages.keys()):
            tl = md_pages[pn][:3000].lower()
            if std_s is None and any(k in tl for k in sa_markers):
                std_s = pn
            if cons_s is None and any(k in tl for k in co_markers):
                cons_s = pn
        toc = _build_toc(md_pages, std_s, cons_s, None)
        # Infer stmt_type + period_type from filename
        nl = name.lower()
        if "annual" in nl or "integrated" in nl or " ar " in nl:
            period = "ANNUAL"
        elif "quarter" in nl or "q1" in nl or "q2" in nl or "q3" in nl or "q4" in nl:
            period = "QUARTERLY"
        elif "presentation" in nl or "investor" in nl:
            period = "INVESTOR_PRESENTATION"
        else:
            period = "UNKNOWN"
        if "standalone" in nl:
            stmt = "STANDALONE"
        elif "consolidated" in nl:
            stmt = "CONSOLIDATED"
        else:
            stmt = "BOTH"
        cls = PDFClassification(pp, name, stmt, period, "", 0, "", toc)
        return cls, md, md_pages

    # Parallelize PDF prep — biggest win for multi-AR runs
    import concurrent.futures as _cf
    prepped: list[tuple[PDFClassification, str, dict]] = []
    with _cf.ThreadPoolExecutor(max_workers=min(5, len(pdf_paths))) as pool:
        futures = {pool.submit(_prep_pdf, pp): pp for pp in pdf_paths}
        for fut in _cf.as_completed(futures):
            pp = futures[fut]
            try:
                cls, md, md_pages = fut.result()
                if cls is None:
                    res.warnings.append(f"Failed to prep {pp}")
                    continue
                prepped.append((cls, md, md_pages))
            except Exception as exc:
                res.warnings.append(f"Prep failed for {pp}: {exc}")

    # Sort by name so output is deterministic
    prepped.sort(key=lambda x: x[0].name)
    logger.info(f"raw_extract: prepped {len(prepped)} PDFs in parallel")

    for cls, md, md_pages in prepped:
        toc = cls.toc or []
        if not toc:
            res.warnings.append(f"{cls.name}: no TOC, skipping")
            continue

        # Decide whether this report has BOTH Standalone AND Consolidated
        # sections. Logic: if TOC has at least one CONSOLIDATED entry, it's
        # a dual-section report → emit Standalone + Consolidated separately.
        # If no Consolidated entry → single-section report (most QRs, some
        # PSUs, parent-only filings) → emit ONE sheet per statement type
        # using cls.stmt_type as the scope.
        has_consolidated = any(sc == "CONSOLIDATED" for _, _, sc in toc)
        has_standalone   = any(sc == "STANDALONE"   for _, _, sc in toc)
        dual_section = has_consolidated and has_standalone

        # Group: which pages contain BS, P&L, CF, by scope
        # toc entry: (page_num_1indexed, title, scope)
        grouped: dict[tuple[str, str], list[int]] = {}   # (scope, stmt) → [page_idx]
        for pn1, title, scope in toc:
            stmt = _classify_statement(title)
            if not stmt:
                continue
            if dual_section:
                # Skip pages outside the financial-statement zones (MD&A
                # 5-year history, Post-AGM appendix etc.) — they have BS
                # labels but aren't the actual financial statement.
                if scope not in ("STANDALONE", "CONSOLIDATED"):
                    continue
                sc_label = "Standalone" if scope == "STANDALONE" else "Consolidated"
            else:
                # Single-section: use PDF-level stmt_type as scope label.
                # Don't split. Don't drop pages — every page that has a
                # statement classification is included.
                if cls.stmt_type in ("STANDALONE", "CONSOLIDATED"):
                    sc_label = cls.stmt_type.title()
                else:
                    sc_label = ""   # no scope prefix
            grouped.setdefault((sc_label, stmt), []).append(pn1 - 1)

        # Fallback: TOC didn't classify anything (rare). Use raw page scan.
        if not grouped:
            if cls.stmt_type in ("STANDALONE", "CONSOLIDATED"):
                for pn, page_text in md_pages.items():
                    tl = page_text[:1500].lower()
                    stmt = None
                    if "balance sheet" in tl: stmt = "Balance Sheet"
                    elif "statement of profit" in tl: stmt = "P&L"
                    elif "cash flow" in tl: stmt = "Cash Flow"
                    if stmt:
                        sc_label = cls.stmt_type.title()
                        grouped.setdefault((sc_label, stmt), []).append(pn)

        if not grouped:
            res.warnings.append(
                f"{cls.name}: no BS/P&L/CF pages identified"
            )
            continue

        pdf_id = _short_pdf_id(cls.name)

        # ── REGROUP: one sheet per statement type. Each sheet stacks
        # Standalone block (top) + Consolidated block (bottom).
        by_stmt: dict[str, dict[str, list[int]]] = {}   # stmt → {sc_label → page_idxs}
        for (sc_label, stmt), page_idxs in grouped.items():
            by_stmt.setdefault(stmt, {})[sc_label] = page_idxs

        # Render order within each sheet: Standalone first, Consolidated second
        # ("" scope last, only used when dual_section is False)
        scope_order = ["Standalone", "Consolidated", ""]

        def _write_block(ws, row_cursor: int, sc_label: str,
                          tables: list, stmt_name: str) -> int:
            """Write a [sc_label + stmt_name] block at row_cursor. Returns next row."""
            if not tables:
                return row_cursor
            # Block header
            block_title = f"{sc_label} {stmt_name}".strip()
            ws.cell(row_cursor, 1, value=block_title).font = CALIBRI11_BOLD
            ws.cell(row_cursor, 1).fill = HEADER_FILL
            row_cursor += 1
            for ti, tbl in enumerate(tables):
                if ti > 0:
                    row_cursor += 1   # gap between sub-tables
                header = tbl[0]
                for ci, h in enumerate(header, start=1):
                    cell = ws.cell(row_cursor, ci, value=h)
                    cell.font = CALIBRI11_BOLD
                    cell.fill = HEADER_FILL
                    cell.border = THIN_BORDER
                    cell.alignment = Alignment(
                        horizontal="center" if ci > 1 else "left",
                        wrap_text=True, vertical="center",
                    )
                row_cursor += 1
                for row in tbl[1:]:
                    for ci, cell_text in enumerate(row, start=1):
                        v_num = _try_parse_num(cell_text)
                        if v_num is not None and ci > 1:
                            cell = ws.cell(row_cursor, ci, value=v_num)
                            cell.number_format = ACCOUNTING_FMT
                        else:
                            cell = ws.cell(row_cursor, ci, value=cell_text)
                        cell.font = CALIBRI11
                        cell.border = THIN_BORDER
                        if ci == 1:
                            cell.alignment = Alignment(horizontal="left",
                                                       vertical="center")
                        else:
                            cell.alignment = Alignment(horizontal="right",
                                                       vertical="center")
                    row_cursor += 1
            return row_cursor + 2   # blank gap between blocks

        for stmt, scopes_pages in by_stmt.items():
            # Make a single sheet for this statement type
            sheet_name_raw = f"{pdf_id} - {stmt}".strip()
            sheet_name = _safe_sheet_name(sheet_name_raw, existing_sheets)
            existing_sheets.add(sheet_name)
            ws = wb.create_sheet(sheet_name)

            # Top-level title (PDF filename)
            ws.cell(1, 1, value=cls.name).font = CALIBRI11_BOLD
            ws.cell(1, 1).fill = HEADER_FILL
            row_cursor = 3   # leave a blank row

            wrote_any = False
            for sc_label in scope_order:
                if sc_label not in scopes_pages:
                    continue
                page_idxs = scopes_pages[sc_label]
                tables_collected: list[list[list[str]]] = []

                # PRIMARY: PyMuPDF find_tables() — preserves label column
                # on magazine-layout PDFs that mangle pdfplumber.
                pdf_path = getattr(cls, "path", None) or cls.name
                if pdf_path and os.path.exists(pdf_path):
                    fitz_tables = _extract_tables_via_fitz(
                        pdf_path, sorted(page_idxs)
                    )
                    tables_collected.extend(fitz_tables)

                # FALLBACK: parse from cached .md if fitz returns nothing
                if not tables_collected:
                    for pn in sorted(page_idxs):
                        tables = _extract_md_tables_from_page(md_pages.get(pn, ""))
                        for tbl in tables:
                            n_numeric = sum(1 for r in tbl for c in r
                                            if _try_parse_num(c) is not None)
                            if len(tbl) >= 3 and n_numeric >= 5:
                                tables_collected.append(tbl)

                if not tables_collected:
                    continue
                row_cursor = _write_block(ws, row_cursor, sc_label,
                                           tables_collected, stmt)
                wrote_any = True

            if not wrote_any:
                # No real tables on any scope's pages; remove the empty sheet
                wb.remove(ws)
                continue

            # Auto-size columns
            for ci in range(1, ws.max_column + 1):
                max_len = 0
                for r in range(1, ws.max_row + 1):
                    v = ws.cell(r, ci).value
                    if v is not None:
                        max_len = max(max_len, len(str(v)))
                ws.column_dimensions[get_column_letter(ci)].width = min(
                    max(14, max_len + 2), 50
                )

            res.sheets_written.append(sheet_name)
            logger.info(f"raw_extract: wrote {sheet_name} "
                        f"(scopes: {list(scopes_pages.keys())})")

    if not res.sheets_written:
        # Make sure workbook has at least one sheet
        wb.create_sheet("Empty")
        res.warnings.append("No financial statement tables extracted.")

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    res.elapsed_sec = time.time() - t0
    return res


__all__ = ["raw_extract_statements", "RawExtractResult"]
